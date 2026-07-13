"""
konteyner_qosh.py — Xitoy装箱单 (Truba/Profil) va出货清单 (List) parser
Konteynerlar yo'lda ro'yxatini shakllantiradi.

Parser ikki faylni o'qib, ISO konteyner raqami bo'yicha birlashtiradi.
Chiqish: [{"iso": "TCKU2238508", "sana": "21.03.2026", "items": [("Tovar nomi", 100), ...], "manba": "aralash"}]
"""
import re
import math
import json
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import openpyxl

# ── Admin tomonidan bir marta to'g'irlangan Xitoy spec → inventar nomi ──────
# Xitoy fayllari hech qachon "ideal" bo'lmaydi (turli marka yozilishi,
# almashgan tartib, kamdan-kam o'lchamlar va h.k.) — shu sababli har bir
# noodatiy holat uchun kodga yangi-yangi "taxmin" mantiq qo'shavermaslik
# uchun, admin draft Excelda bir marta qo'lda to'g'irlagan nom keyingi
# safar XUDDI SHU xom spec+marka chiqqanda AVTOMATIK ishlatiladi — qayta
# hech qanday evristika (dumaloq/kvadrat, stenka jadvali va h.k.) ishga
# tushmaydi. Bu saqlangan holat oddiy JSON faylda turadi (config.py:
# XITOY_TUZATISH_FILE), botni qayta ishga tushirish orasida ham saqlanadi.
_tuzatishlar_kesh: dict | None = None


def _tuzatishlar_fayli() -> Path:
    try:
        from config import XITOY_TUZATISH_FILE
        return XITOY_TUZATISH_FILE
    except Exception:
        return Path(__file__).resolve().parent / "bot_holat" / "xitoy_tuzatishlar.json"


def _tuzatishlarni_yukla() -> dict:
    global _tuzatishlar_kesh
    if _tuzatishlar_kesh is None:
        fpath = _tuzatishlar_fayli()
        try:
            _tuzatishlar_kesh = json.loads(fpath.read_text(encoding="utf-8"))
        except Exception:
            _tuzatishlar_kesh = {}
    return _tuzatishlar_kesh


def _tuzatish_saqla(kalit: str, nom: str) -> None:
    """Bitta spec→nom bog'lanishini doimiy saqlaydi (JSON faylga yozadi)."""
    if not kalit or not nom:
        return
    tuzatishlar = _tuzatishlarni_yukla()
    if tuzatishlar.get(kalit) == nom:
        return  # allaqachon xuddi shunday saqlangan
    tuzatishlar[kalit] = nom
    try:
        fpath = _tuzatishlar_fayli()
        fpath.parent.mkdir(parents=True, exist_ok=True)
        fpath.write_text(
            json.dumps(tuzatishlar, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    except Exception:
        pass  # saqlab bo'lmasa ham dastur ishlashda davom etadi


def _tuzatish_kaliti(turi: str, spec: str, marka, rang: str = "") -> str:
    """Xom Xitoy spec+marka(+rang)dan barqaror kalit yasaydi."""
    marka_s = str(marka).strip() if marka is not None else ""
    return f"{turi}|{str(spec).strip()}|{marka_s}|{str(rang).strip()}"


def _xom_keydan_avtomatik_nom(raw_key: str, inv_set: set) -> str | None:
    """
    Xom kalitdan (masalan 'TP|0.85*50.8*5800|201|') hozirgi EVRISTIKA
    (dumaloq/kvadrat aniqlash, stenka jadvali va h.k.) qanday nom
    berishini qayta hisoblaydi — bu admin bergan nom bilan solishtirish
    uchun kerak: FARQ bo'lsagina haqiqiy tuzatish deb hisoblanadi va
    saqlanadi, aks holda har bir (allaqachon to'g'ri) qatorni ham
    "tuzatish" deb abadiy muzlatib qo'yish xavfi bor edi.
    """
    try:
        turi, spec, marka, rang = raw_key.split("|", 3)
    except ValueError:
        return None
    if turi == "TP":
        return _inventarga_moslashtir(_truba_spec_to_name(spec, marka, inv_set))
    if turi == "LIST":
        return _inventarga_moslashtir(_list_spec_to_name(spec, marka, rang))
    return None


def _tuzatishdan_top(kalit: str) -> str | None:
    return _tuzatishlarni_yukla().get(kalit)


# ── Inventarga moslashtirish (Buyurtma/parsers.py bilan bir xil mexanizm) ────
# Bu yerda o'zimiznikini emas, "Yo'lga konteyner qo'shish"dan mustaqil
# ravishda allaqachon ishlab turgan Buyurtma-oqimidagi (parsers.py)
# "haqiqiy inventarga moslashtirish" mexanizmini ISHLATAMIZ — statik
# jadval o'rniga, nomni real inventar ro'yxati bilan solishtirib, mavjud
# bo'lgan aniq yozuvga moslaydi (masalan stenka 0,65↔0,7, uzunlik
# 5,8м↔6м chegara holatlarini hal qiladi). Aynan ko'chirilmagan — faqat
# import qilib, ustiga snap qilish qadami sifatida qo'shilgan.
def _inventarga_moslashtir(nom: str | None) -> str | None:
    if not nom:
        return nom
    try:
        from parsers import _fix_oddiy_nom, _get_inventar_set
        inv_set = _get_inventar_set()

        # 2026-07-11 (tuzatildi, Huzayfa: "yo'lga konteyner qo'shish"ni ham
        # ko'rib chiq): ILGARI Лист nomlari bu yerdan ALOHIDA o'tkazilardi
        # (_inventardan_moslashtir() to'g'ridan-to'g'ri, _fix_oddiy_nom()ni
        # butunlay chetlab o'tib) — sabab: _fix_oddiy_nom() birinchi qadami
        # sifatida common.normalize_product_name() ni chaqiradi, u "Лист-3,0"
        # kabi nomlarga bo'shliq qo'shib yuboradi ("Лист- 3,0"), bu esa
        # ko'pchilik bo'shliqsiz inventar yozuvlari bilan ANIQ moslikni
        # buzardi. LEKIN _fix_oddiy_nom() O'ZINING ICHIDA ham xuddi shu
        # bo'shliqqa-chidamli _inventardan_moslashtir() bilan tugaydi — shu
        # sabab bypass qilish HECH NARSA cheklamas edi, faqat _fix_oddiy_nom()
        # ichidagi QO'SHIMCHA qadamni (Xitoy "-0,05mm" Лист qalinlik
        # konvensiyasi -- 0,75->0,8 / 0,95->1,0 / 1,45->1,5 / 2,95->3,0 kabi)
        # yo'qotib qo'yardi. Natijada _list_spec_to_name()ning o'z jadvali
        # ("1,45->1,45 istisno" kabi, Труба/Профиль uchun to'g'ri, lekin
        # Лист uchun XATO -- inventarda "Лист-1,45" umuman yo'q, faqat
        # "Лист-1,5") tuzatilmasdan qolib, haqiqiy mavjud tovar "notanish"
        # deb chiqib ketardi. Endi bypasssiz -- _fix_oddiy_nom() Труба/
        # Профиль uchun ham, Лист uchun ham bir xil, TO'LIQ (bo'shliq +
        # +0,05 snap + kanonik) yo'l bilan ishlaydi.
        return _fix_oddiy_nom(nom, inv_set)
    except Exception:
        # parsers.py mavjud bo'lmasa yoki xato bo'lsa — o'zgarishsiz qaytaramiz
        return nom


# ISO konteyner raqami: 4 harf + 7 raqam
# DIQQAT: \b (so'z chegarasi) ISHLATILMAYDI — chunki Python regex'da
# xitoycha ierogliflar ham "so'z belgisi" (\w) hisoblanadi, shuning uchun
# masalan "柜号GLDU5169925" (ajratuvchisiz, ':' yoki bo'shliqsiz yozilgan)
# holatda 号→G o'tishida chegara TOPILMAYDI va ISO umuman aniqlanmay qolar
# edi. Shu sababli lotin harf/raqamdan boshqa narsa bo'lishini talab
# qiluvchi lookaround ishlatiladi — bu CJK belgilardan keyin ham to'g'ri
# ishlaydi.
_ISO_RE = re.compile(r'(?<![A-Za-z0-9])([A-Z]{4}\d{7})(?![A-Za-z0-9])')

# ── Yordamchi funksiyalar ─────────────────────────────────────────────────────

def _iso(text) -> str | None:
    """ISO konteyner raqamini matndan ajratib oladi."""
    m = _ISO_RE.search(str(text))
    return m.group(1) if m else None


# ── Konteyner raqami (柜号) YO'Q, faqat yuk mashinasi bilan tashiladigan ────
# yetkazib berishlar (odatda "平板车"/flatbed — Xitoy ichki port-zavod
# tashuvi) uchun PSEUDO-ID — 2026-07-06'da topilgan JIDDIY xato: bunday
# bloklar (haqiqiy 柜号 bo'lmagani uchun) BUTUNLAY tashlab yuborilar edi,
# garchi ichida haqiqiy Труба/Профиль tovarlar bo'lsa ham (tekshirilgan:
# "晋ME5312"/"晋KH6090"/"晋KA1207" nomli 3 ta mashina yetkazib berishida
# jami o'nlab qatordan iborat haqiqiy 201-marka truba/profil mahsulotlari
# bor edi, ular hech qachon konteyner ro'yxatiga qo'shilmagan). Bundan
# tashqari, admin ILGARI aynan shu formatda ("ATY123", "AFY662" — tarixda
# mavjud, bot_holat/qoshilgan_konteynerlar.json) qo'lda pseudo-ID sifatida
# ishlatgan — shu konvensiyaga mos qilib qurilgan: viloyat ieroglifi
# tashlanadi, faqat lotin harf+raqam qismi qoladi ("晋ME5312" → "ME5312",
# "吉ATY123（平板车）" → "ATY123").
_MASHINA_RAQAM_RE = re.compile(r'车[号牌][:：]?\s*[一-鿿]([A-Z]{1,3}\d{3,6})')


def _mashina_raqam(text) -> str | None:
    """Mashina davlat raqamidan (车号/车牌) pseudo-ID ajratib oladi — faqat
    haqiqiy 柜号 topilmagan holatlarda, fallback sifatida ishlatiladi."""
    m = _MASHINA_RAQAM_RE.search(str(text))
    return m.group(1) if m else None


def _sana_format(d) -> str:
    """datetime → '21.03.2026'"""
    if isinstance(d, (datetime, date)):
        return d.strftime("%d.%m.%Y") if isinstance(d, datetime) else d.strftime("%d.%m.%Y")
    return str(d)


def _yaxlitla_stenka(s: str) -> str:
    """Xitoy stenka → inventar stenka (0.85 → 0,9 va h.k.)"""
    try:
        v = float(str(s).replace(',', '.'))
    except (ValueError, TypeError):
        return str(s).replace('.', ',')
    # Jadval real inventardagi BARCHA "ст X" qiymatlariga (0,28dan 3,5gacha)
    # asoslangan — 2026-07-06'da inventar bo'yicha to'liq tekshirilgan.
    TABLE = [
        (0.28, '0,28'), (0.31, '0,31'), (0.40, '0,4'),
        (0.50, '0,5'), (0.55, '0,6'), (0.60, '0,6'), (0.63, '0,65'),
        (0.65, '0,65'), (0.68, '0,7'), (0.70, '0,7'), (0.75, '0,8'),
        (0.80, '0,8'), (0.85, '0,9'), (0.90, '0,9'), (0.92, '0,92'),
        (0.95, '1,0'), (1.00, '1,0'), (1.05, '1,1'), (1.10, '1,1'),
        (1.20, '1,2'), (1.25, '1,3'), (1.30, '1,3'), (1.35, '1,35'),
        (1.40, '1,4'), (1.45, '1,45'), (1.50, '1,5'), (1.75, '1,75'),
        (1.80, '1,8'), (1.85, '1,9'), (1.90, '1,9'), (2.00, '2,0'),
        (2.40, '2,4'), (2.50, '2,5'), (2.80, '2,8'), (3.00, '3,0'),
        (3.50, '3,5'),
    ]
    # DIQQAT: eng yaqinini tanlashda suzuvchi nuqta noaniqligi ("float
    # noise") ba'zan noto'g'ri natija berardi — masalan 1.15 haqiqatda
    # 1.10 va 1.20'dan BIR XIL uzoqlikda (0,05) turishi kerak, lekin
    # ikkilik kasr taqribiyligi tufayli 1.10 "arzimas darajada" yaqinroq
    # chiqib, noto'g'ri "1,1" (bunday tovar INVENTARDA UMUMAN yo'q) qaytarar
    # edi — to'g'risi "1,2" (mavjud). Shu sababli masofa 6 xonagacha
    # yaxlitlanadi, teng bo'lganda esa KATTAROQ (yuqoriroq) qiymat afzal
    # qilinadi — bu Xitoy "-0,05 pastroq" konvensiyasiga ham mos keladi.
    best = min(TABLE, key=lambda x: (round(abs(x[0] - v), 6), -x[0]))
    return best[1]


def _uzunlik_str(uzunlik_mm: float) -> str:
    uzunlik_m = uzunlik_mm / 1000.0
    if uzunlik_m == int(uzunlik_m):
        uz_s = str(int(uzunlik_m))
    else:
        uz_s = f"{uzunlik_m:.1f}".replace('.', ',')
    return f"{uz_s} м"


def _marka_normallashtir(marka_raw) -> str:
    marka = str(int(float(marka_raw))) if isinstance(marka_raw, float) else str(marka_raw).strip()
    # Ba'zan "201/钛金" yoki "J4/201"/"201/J1" (maxsus "Ж-seriya" belgisi
    # bilan aralashgan) kabi qo'shma marka keladi — HAQIQIY marka kodini
    # ("201","304" va h.k.) matn ICHIDAN qidiramiz, faqat boshidan emas,
    # chunki "J4/201" holatida marka OXIRIDA turadi.
    m = re.search(r'\b(201|304|316|321|430)\b', marka)
    marka = m.group(1) if m else marka
    if marka not in ('201', '304', '316', '321', '430'):
        marka = '201'
    return marka


# ── Yumaloq (Ф-N) vs kvadrat (Пр. NxN) profil o'lchamlari — INVENTARDAN ──
# Muammo: Xitoy 3-qismli spec ("stenka*N*uzunlik") formatida N BUTUN son
# bo'lganda, bu N yumaloq trubaning diametri (Ф-N) HAM, kvadrat profilning
# tomoni (Пр. NхN) HAM bo'lishi mumkin — spec matnining o'zida farq yo'q.
# Avvalgi eвristika ("butun son bo'lsa — kvadrat, kasr bo'lsa — dumaloq")
# NOTO'G'RI edi: masalan Ф-16/Ф-19/Ф-22 kabi YUMALOQ trubalar odatda BUTUN
# son diametr bilan yoziladi, lekin "Пр. 16х16"/"Пр. 22х22" kabi kvadrat
# profillar INVENTARDA UMUMAN MAVJUD EMAS — shuning uchun bular doim
# noto'g'ri "profil" deb chiqib, hech qachon inventar bilan mos kelmasdi.
# Endi HAQIQIY inventar ro'yxatidan qaysi shakl (Ф- yoki Пр. NxN) real
# mavjudligi aniqlanadi va shunga qarab tanlanadi.
_ROUND_SQUARE_CACHE: dict = {}


def _round_square_dims(inv_set: set) -> tuple[set, set]:
    """Inventardan mavjud Ф-N (dumaloq) va Пр. NxN (kvadrat) o'lchamlarini ajratib oladi."""
    key = id(inv_set)
    if key in _ROUND_SQUARE_CACHE:
        return _ROUND_SQUARE_CACHE[key]
    # DIQQAT: ^ bilan boshida ANIQ tekshiriladi (re.search EMAS) — chunki
    # ba'zi o'lchamlar faqat "(Аркон) Пр. 19х19 ст ..." kabi BREND
    # PREFIKSI bilan mavjud, plain (prefiкssiz) holda esa umuman yo'q.
    # _truba_spec_to_name() esa doim PLAIN nom generatsiya qiladi — shuning
    # uchun faqat plain formatda mavjud o'lchamlar hisobga olinishi kerak,
    # aks holda (masalan 19) noto'g'ri "ikkalasi ham bor" deb topilib,
    # standart taxminga (kvadrat) tushib qolar edi — aslida faqat Ф-19
    # (dumaloq) plain holda mavjud.
    round_d, square_d = set(), set()
    for n in inv_set:
        m = re.match(r'^Ф-(\d+)\s+ст', n)
        if m:
            round_d.add(int(m.group(1)))
        m2 = re.match(r'^Пр\.?\s*(\d+)х(\d+)\s+ст', n)
        if m2 and m2.group(1) == m2.group(2):
            square_d.add(int(m2.group(1)))
    _ROUND_SQUARE_CACHE[key] = (round_d, square_d)
    return round_d, square_d


def _truba_spec_to_name(spec: str, marka_raw, inv_set: set = None) -> str | None:
    """
    Truba/Profil Xitoy speci → inventar nomi.
    '0.85*50.8*5800' + 201 → 'Ф-51 ст 0,9 (6 м) (201 марка)'  (3 qismli: yumaloq truba)
    '0.65*20*10*5800' + 201 → 'Пр. 20х10 ст 0,65 (5,8 м) (201 марка)'  (4 qismli: to'rtburchak profil)
    """
    spec  = str(spec).strip()
    marka = _marka_normallashtir(marka_raw)
    marka_sfx = f" ({marka} марка)"

    # 4 qismli format: stenka * eni * balandligi * uzunligi — Профиль
    # (to'rtburchak, eni != balandligi bo'lishi ham mumkin: masalan 20х10)
    m4 = re.match(r'^([\d\.]+)[*×x]([\d\.]+)[*×x]([\d\.]+)[*×x]([\d\.]+)$', spec)
    if m4:
        try:
            f1         = float(m4.group(1))
            f2         = float(m4.group(2))
            f3         = float(m4.group(3))
            uzunlik_mm = float(m4.group(4))
        except ValueError:
            return None
        # DIQQAT: ba'zi qatorlarda (odatda maxsus "Ж-seriya" marka bilan,
        # masalan marka="J4/201") o'lcham va stenka o'rni ALMASHTIRILGAN
        # holda keladi: "eni*balandligi*stenka*uzunlik" ("25*25*0.85*5800"),
        # oddiy "stenka*eni*balandligi*uzunlik" o'rniga. Buni haqiqiy
        # stenka HECH QACHON 10mm dan katta bo'lmasligidan bilib olamiz —
        # birinchi son >10 bo'lib, uchinchisi <=10 bo'lsa, aniq almashgan.
        if f1 > 10 and f3 <= 10:
            stenka_raw, a_raw, b_raw = f3, f1, f2
        else:
            stenka_raw, a_raw, b_raw = f1, f2, f3
        stenka = _yaxlitla_stenka(stenka_raw)
        a = int(a_raw) if a_raw == int(a_raw) else a_raw
        b = int(b_raw) if b_raw == int(b_raw) else b_raw
        return f"Пр. {a}х{b} ст {stenka} ({_uzunlik_str(uzunlik_mm)}){marka_sfx}"

    # 3 qismli format: stenka * diametr(yoki kvadrat tomoni) * uzunlik
    m = re.match(r'^([\d\.]+)[*×x]([\d\.]+)[*×x]([\d\.]+)$', spec)
    if not m:
        return None

    try:
        f1         = float(m.group(1))
        f2         = float(m.group(2))
        uzunlik_mm = float(m.group(3))
    except ValueError:
        return None

    # Xuddi shu almashgan-tartib holati 3 qismli formatda ham uchraydi:
    # masalan "38*0.85*6000" (diametr*stenka*uzunlik) — odatdagi
    # "stenka*diametr*uzunlik" o'rniga. Haqiqiy stenka hech qachon 10mm dan
    # katta bo'lmaydi, shuning uchun shu belgi orqali aniqlanadi.
    if f1 > 10 and f2 <= 10:
        stenka_raw, dim2 = f2, f1
    else:
        stenka_raw, dim2 = f1, f2

    stenka  = _yaxlitla_stenka(stenka_raw)
    uz_str  = _uzunlik_str(uzunlik_mm)

    if dim2 > 200:
        return None

    # dim2 kasr bo'lsa (masalan 50.8) — deyarli har doim yumaloq truba
    # (dyuymdan mm'ga o'girilgan diametr), bunda shak-shubhasiz.
    if dim2 != int(dim2):
        d_int = round(dim2)
        return f"Ф-{d_int} ст {stenka} ({uz_str}){marka_sfx}"

    d_int = int(dim2)

    # dim2 BUTUN son bo'lsa — inventardan qaysi shakl (dumaloq yoki
    # kvadrat) haqiqatan mavjudligini tekshiramiz.
    if inv_set:
        round_d, square_d = _round_square_dims(inv_set)
        round_ok  = d_int in round_d
        square_ok = d_int in square_d
        if round_ok and not square_ok:
            return f"Ф-{d_int} ст {stenka} ({uz_str}){marka_sfx}"
        if square_ok and not round_ok:
            return f"Пр. {d_int}х{d_int} ст {stenka} ({uz_str}){marka_sfx}"
        # Ikkalasi ham mavjud (haqiqatan noaniq) yoki ikkalasi ham topilmadi
        # — pastdagi standart taxminga o'tiladi.

    # Standart taxmin (inv_set berilmagan yoki noaniq holat): kvadrat profil.
    return f"Пр. {d_int}х{d_int} ст {stenka} ({uz_str}){marka_sfx}"


# Inventarda List doim shu standart razmerlarda saqlanadi (Xitoy xom
# spec'i deyarli har doim 1219x2438 yoki 1500x3000 keladi — bular mos
# ravishda 1220x2440 va 1500x3000'ga aylantiriladi). Yaqin (≤15mm farq)
# bo'lmasa xom qiymat o'zgarishsiz qoldiriladi (haqiqatan noodatiy o'lcham
# bo'lishi mumkin — keyin inventarda topilmasa admin darhol ko'radi).
_STANDART_LIST_RAZMER = [
    (1220, 2440), (1500, 3000), (1220, 2700), (1220, 3000),
    (1250, 2700), (1000, 2000), (1250, 2500),
]


def _yaxlitla_list_razmer(en_raw: float, boy_raw: float) -> tuple[int, int]:
    best = min(_STANDART_LIST_RAZMER,
               key=lambda wh: abs(wh[0] - en_raw) + abs(wh[1] - boy_raw))
    if abs(best[0] - en_raw) <= 15 and abs(best[1] - boy_raw) <= 15:
        return best
    return (int(round(en_raw)), int(round(boy_raw)))


# Xitoy "颜色" (rang/sirt) kodini inventar nomidagi rang yorlig'iga
# moslashtiradi. Inventarda BARCHA Лист yozuvlari albatta bitta rang/sirt
# yorlig'iga ega (masalan "(Матовый)") — shu sababli mos kod topilmasa ham
# eng ko'p uchraydigan standart qiymat ("Матовый") qo'llanadi, aks holda
# 3-qavatli qavs strukturasi (razmer/rang/marka) buzilib, inventar bilan
# HECH QACHON mos kelmay qolar edi.
_LIST_RANG_MAP = {
    '砂板':   'Матовый',
    '砂 板':  'Матовый',
    '8K':    'Глянцевый',
    '8K钛金': 'Голд',
    '精磨8K': 'Глянцевый',
    '8K黑钛': 'Кора',
}


def _list_spec_to_name(spec: str, marka_raw, rang: str = '') -> str | None:
    """
    List Xitoy speci → inventar nomi.
    '2.95*1219*2438' + 304 + '砂板' → 'Лист-3,0 (1220х2440) (Матовый) (304 марка)'

    MUHIM: inventardagi HAR BIR Лист yozuvi "Лист-" (chiziqcha bilan,
    bo'sh joy emas!), o'lchami standart qiymatga yaxlitilgan (1220х2440 /
    1500х3000 va h.k.) va DOIM rang/sirt yorlig'iga ega bo'ladi — bu 3 ta
    detal to'g'ri bo'lmasa, tovar hech qachon inventar bilan mos kelmaydi.
    """
    spec = str(spec).strip()
    marka = str(int(float(marka_raw))) if isinstance(marka_raw, float) else str(marka_raw).strip()
    if marka not in ('201', '304', '316', '321', '430'):
        marka = '201'

    m = re.match(r'^([\d\.]+)[*×x]([\d\.]+)[*×x]([\d\.]+)$', spec)
    if not m:
        return None

    try:
        qalinlik_raw = float(m.group(1))
        en_raw       = float(m.group(2))
        boy_raw      = float(m.group(3))
    except ValueError:
        return None

    # Qalinlik: Труба/Профиль bilan BIR XIL Xitoy-delta jadvalidan
    # o'tkaziladi (0,55→0,6 / 0,75→0,8 / 0,95→1,0 / 1,15→1,2 / 1,45→1,45
    # (istisno) / 1,95→2,0 / 2,95→3,0). Avval bu qadam BUTUNLAY yo'q edi —
    # xom Xitoy qiymati o'zgarishsiz yozilardi (masalan "Лист 2,95..."),
    # bu esa inventarda HECH QACHON topilmaydigan nom yaratardi (inventar
    # doim yaxlitlangan: 0,6 / 0,8 / 1,0 / 2,0 / 3,0 va h.k.).
    q_s = _yaxlitla_stenka(qalinlik_raw)

    en, boy = _yaxlitla_list_razmer(en_raw, boy_raw)

    rang_norm = str(rang).replace(' ', '').strip()
    rang_nomi = _LIST_RANG_MAP.get(rang_norm, 'Матовый')

    return f"Лист-{q_s} ({en}х{boy}) ({rang_nomi}) ({marka} марка)"


# ── Og'irlik hisoblash: har bir qator uchun — o'z formulamiz, aks holda ──
# Xitoyning O'ZI SHU QATOR uchun bergan haqiqiy og'irligi ─────────────────
def _item_vazn_kg(nom: str, miqdor: int, xitoy_line_kg, inv_set: set) -> float:
    """
    Bitta tovar QATORINING umumiy og'irligini (kg) hisoblaydi. "Professional"
    hisob-kitob shu yerda: dastur BIRINCHI NAVBATDA o'z formulasi (bitta
    dona/varaq vazni x miqdor, vazn_hisobla.tovar_vazni) bo'yicha hisoblaydi
    — bu inventar bilan ANIQ mos kelgan (tanish) tovarlar uchun. Noma'lum
    (inventarda topilmagan) tovar uchun formula noto'g'ri/hech qanday
    natija bermasligi mumkin — shu holatda Xitoyning O'ZI aynan SHU qator
    uchun bergan haqiqiy og'irligi (合计 emas, individual qator) ishlatiladi.

    Konteynerning umumiy vazni keyinchalik shu qator vaznlari yig'indisi
    sifatida hisoblanadi — Xitoyning "板厂重量" (qadoqlash/yashik og'irligi)
    qo'shilgan umumiy 合计 raqami ENDI ishlatilmaydi, chunki u haqiqiy
    mahsulot vaznidan sezilarli yuqori chiqadi (tekshirilgan: bitta
    konteynerda ~26% qadoqlash ulushi bor edi).
    """
    try:
        from vazn_hisobla import tovar_vazni as _tv
    except Exception:
        _tv = None

    if inv_set and nom in inv_set and _tv:
        unit_w = _tv(nom)
        if unit_w:
            return round(unit_w * miqdor, 2)
    try:
        return round(float(xitoy_line_kg), 2) if xitoy_line_kg else 0.0
    except (ValueError, TypeError):
        return 0.0


# ── Truba/Profil parser (装箱单) ──────────────────────────────────────────────

def _parse_truba_zhuangxiang(raw: bytes) -> dict:
    """
    装箱单 (Truba/Profil packing list) ni o'qiydi.
    Qaytaradi: {iso_no: {"sana": date, "items": [(tovar, miqdor, vazn_kg), ...]}}
    """
    try:
        from parsers import _get_inventar_set
        inv_set = _get_inventar_set()
    except Exception:
        inv_set = set()

    wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
    ws = wb.active

    # 1) 柜号 → qator_raqami. FAQAT merged celllardan emas — hujjatning
    # ba'zi bo'limlarida (masalan keyingi/yangiroq partiyalarda) "备注"
    # yorlig'i merge qilingan, lekin "柜号:XXXX" matnining o'zi ALOHIDA
    # (merge qilinmagan) katakda turadi. Faqat merged_cells.ranges'ni
    # tekshirish bunday hollarda ISO'ni butunlay o'tkazib yuborar edi —
    # shuning uchun endi HAR BIR katak tekshiriladi.
    kont_at = {}  # row_idx → iso_no
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            val = cell.value
            if val and '柜号' in str(val):
                iso = _iso(str(val))
                if iso:
                    kont_at[cell.row] = iso
                break

    # 1.5) Mashina (车号/车牌) → qator raqami — konteyner raqami (柜号) YO'Q,
    # faqat yuk mashinasi bilan tashiladigan bo'limlar uchun FALLBACK.
    mashina_at = {}  # row_idx → pseudo_id
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            val = cell.value
            if val and ('车号' in str(val) or '车牌' in str(val)):
                m_id = _mashina_raqam(str(val))
                if m_id:
                    mashina_at[cell.row] = m_id
                break

    # 2) Blok boshlarini topish (B ustun tarkibida '柜' bo'lsa) — bloklar
    # har xil nomlanishi mumkin: "柜1", "柜2", ... yoki "小柜4" (kichik柜),
    # "大柜1" (katta柜) va h.k. Faqat .startswith('柜') tekshiruvi "小柜"/
    # "大柜" kabi prefiksli bloklarni o'tkazib yuborar edi — natijada o'sha
    # konteynerlar Труба faylida umuman topilmagandek ko'rinib, Лист bilan
    # birlashtirilmay ("aralash" emas, faqat "list") qolib ketardi.
    #
    # MUHIM QO'SHIMCHA (2026-07-06, JIDDIY XATO TOPILDI): ba'zi bo'limlar
    # haqiqiy konteyner EMAS — faqat yuk mashinasida (车号/车牌, "平板车")
    # tashiladigan yetkazib berish, ular "柜N" o'rniga "挂车" yorlig'i bilan
    # boshlanadi ('柜' harfi YO'Q, chunki '挂'≠'柜' — vizual jihatdan
    # o'xshash, lekin BUTUNLAY BOSHQA belgi). Bu ilgari blok deb umuman
    # tan olinmasdi — natijada bunday bo'limning barcha qatorlari OLDINGI
    # (haqiqiy konteynerli) blokning "davomi" deb hisoblanib, o'sha
    # konteynerga NOTO'G'RI qo'shilib ketardi (haqiqiy misolda: CRXU3333908
    # nomli haqiqiy konteyner 6ta o'z tovaridan tashqari, unga umuman
    # aloqasi yo'q 3 ta mashina yetkazmasining 14 ta tovarini ham o'ziga
    # "yutib" olgan edi). Endi "挂车" ham blok boshlanishi sifatida tan
    # olinadi.
    blocks = []  # (start_row, gui_name, sana)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        b = row[1].value
        b_s = str(b).strip() if b else ""
        if b_s and (('柜' in b_s and len(b_s) <= 6) or b_s == '挂车'):
            sana = row[8].value if len(row) > 8 else None
            if isinstance(sana, datetime):
                sana = sana.date()
            blocks.append((row[0].row, b_s, sana))

    result = {}

    for i, (start_row, gui, sana) in enumerate(blocks):
        end_row = blocks[i + 1][0] if i + 1 < len(blocks) else ws.max_row + 1

        # Shu blok ichidagi ISO raqam (haqiqiy 柜号 ustuvor)
        iso = None
        for kr in sorted(kont_at):
            if start_row <= kr < end_row:
                iso = kont_at[kr]
                break
        if not iso:
            # Haqiqiy 柜号 topilmadi — mashina raqami (车号/车牌) bilan
            # pseudo-ID sifatida FALLBACK qilinadi (butunlay tashlab
            # yubormaslik uchun — ichida haqiqiy Труба/Профиль tovar
            # bo'lishi mumkin).
            for kr in sorted(mashina_at):
                if start_row <= kr < end_row:
                    iso = mashina_at[kr]
                    break
        if not iso:
            continue

        # Mahsulotlarni o'qish
        items = []
        in_products = False
        for ri in range(start_row, end_row):
            row_vals = [ws.cell(ri, c).value for c in range(1, 10)]
            b_val = row_vals[1]  # B ustun (0-indexed: index 1 = column B)
            # Sarlavha satrini aniqlash
            if b_val == '规格':
                in_products = True
                continue
            # "合计：" qatori — bu bloklarning YAKUNIY (板厂重量 qo'shilgan)
            # jami summasi, ENDI ishlatilmaydi — har bir qatorning o'z
            # og'irligi pastda alohida o'qiladi va yig'indisi ishlatiladi.
            if b_val and str(b_val).strip().startswith('合计'):
                continue
            if not in_products:
                continue
            spec        = row_vals[1]   # B = 规格
            marka       = row_vals[2]   # C = 材质
            zhishu      = row_vals[4]   # E = 支数 (jami dona soni)
            xitoy_kg    = row_vals[5]   # F = 重量(Kg) — SHU QATOR uchun
            if not spec or not marka:
                continue
            spec_s = str(spec).strip()
            if not re.match(r'^[\d\.]+[*×x]', spec_s):
                continue
            try:
                miqdor = int(float(zhishu)) if zhishu else 0
            except (ValueError, TypeError):
                miqdor = 0
            if miqdor <= 0:
                continue
            raw_key = _tuzatish_kaliti("TP", spec_s, marka)
            saqlangan = _tuzatishdan_top(raw_key)
            if saqlangan:
                # Admin avval bu ANIQ spec+marka uchun qo'lda tasdiqlagan
                # nom bor — hech qanday evristika ishlatilmaydi.
                nom_final = saqlangan
            else:
                nom = _inventarga_moslashtir(_truba_spec_to_name(spec_s, marka, inv_set))
                nom_final = nom or spec_s
            vazn_kg = _item_vazn_kg(nom_final, miqdor, xitoy_kg, inv_set)
            items.append((nom_final, miqdor, vazn_kg, raw_key))

        if iso not in result:
            result[iso] = {"sana": sana, "items": []}
        result[iso]["items"].extend(items)

    return result


# ── List parser (出货清单) ────────────────────────────────────────────────────

_XITOY_SANA_MATN_RE = re.compile(
    r'(\d{4})[年./-](\d{1,2})[月./-](\d{1,2})'  # 2026年3月20日 / 2026-3-20 / 2026.3.20
)


def _matndan_sana(v) -> "date | None":
    """
    '2026年3月20日' / '2026-3-20' / '2026.03.20' kabi matn ko'rinishidagi
    sanani date obyektiga aylantiradi. Mos kelmasa None.
    """
    m = _XITOY_SANA_MATN_RE.search(str(v))
    if not m:
        return None
    try:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return date(y, mo, d)
    except ValueError:
        return None


def _list_sana_qatormi(row: list) -> bool:
    """
    Bo'lim sanasi qatorini aniqlaydi: faqat col1 (B ustun) to'ldirilgan,
    boshqa hamma ustun bo'sh bo'lgan qator — masalan
    ['', 46180.0, '', '', '', '', '', '', '', ''].
    Bunday qatorlar hujjat ichida BIR NECHTA marta uchraydi — har biri
    o'zidan keyingi konteynerlar guruhi uchun YANGI sana belgilaydi
    (fayl bir nechta jo'natish sanasini bittalab hujjatga yig'ib
    qo'yilgan bo'lishi mumkin).

    Sana ODATDA float (Excel sana-seriali) ko'rinishida keladi, lekin ba'zi
    hujjatlarda matn ko'rinishida ("2026年3月20日" va h.k.) ham kelishi
    mumkin — shuning uchun ikkalasi ham qabul qilinadi.
    """
    if len(row) < 2:
        return False
    v = row[1]
    is_sana = isinstance(v, (int, float)) or (isinstance(v, str) and _matndan_sana(v))
    if not is_sana:
        return False
    boshqalar = [c for i, c in enumerate(row) if i != 1]
    return all(c == '' or c is None for c in boshqalar)


def _sana_qator_qiymatini_parse(v, datemode) -> "date | None":
    """_list_sana_qatormi True bergan qatordagi sana qiymatini date'ga o'giradi."""
    if isinstance(v, (int, float)):
        try:
            import xlrd
            return xlrd.xldate_as_datetime(v, datemode).date()
        except Exception:
            return None
    return _matndan_sana(v)


def _parse_list_chuhuo(raw: bytes) -> dict:
    """
    出货清单 (List packing list) ni o'qiydi.
    Qaytaradi: {iso_no: {"sana": date, "items": [(tovar, miqdor, vazn_kg), ...]}}

    MUHIM: bitta fayl ichida BIR NECHTA turli sanadagi jo'natish bo'limi
    bo'lishi mumkin — har bir "材质/颜色/规格..." sarlavha blokidan oldin
    o'sha bo'limga tegishli sana qatori keladi. Shuning uchun BUTUN faylga
    bitta sana yopishtirib bo'lmaydi — har bir konteyner o'zi joylashgan
    bo'limning sanasini olishi kerak.
    """
    try:
        import xlrd
    except ImportError:
        raise ImportError("xlrd o'rnatilmagan: pip install xlrd")

    try:
        from parsers import _get_inventar_set
        inv_set = _get_inventar_set()
    except Exception:
        inv_set = set()

    book = xlrd.open_workbook(file_contents=raw)
    sh   = book.sheets()[0]

    # Birinchi sarlavha satrini topish (ustun indekslarini aniqlash uchun)
    hdr_row = 2
    for ri in range(min(5, sh.nrows)):
        row = sh.row_values(ri)
        if '规格' in row or any('规格' in str(c) for c in row):
            hdr_row = ri
            break

    # Ustun indekslari
    hdrs = [str(c).strip() for c in sh.row_values(hdr_row)]
    mat_i  = next((i for i, h in enumerate(hdrs) if '材质' in h), 0)
    rang_i = next((i for i, h in enumerate(hdrs) if '颜色' in h), 1)
    gg_i   = next((i for i, h in enumerate(hdrs) if '规格' in h), 2)
    qty_i  = next((i for i, h in enumerate(hdrs) if '数量' in h), 3)
    bz_i   = next((i for i, h in enumerate(hdrs) if '备注' in h), 9)
    # Og'irlik (tonna) ustuni — "毛重" (yalpi og'irlik, kg)
    ogirlik_i = next((i for i, h in enumerate(hdrs) if '毛重' in h), None)
    if ogirlik_i is None:
        ogirlik_i = next((i for i, h in enumerate(hdrs) if '净重' in h), None)

    result    = {}
    cur_iso   = None
    cur_sana  = None   # hozirgi bo'limning sanasi — har yangi sana qatorida yangilanadi

    # DIQQAT: 0-qatordan boshlanadi (hdr_row+1 emas!) — hujjatning ENG
    # BIRINCHI sana qatori (masalan row 1) sarlavha qatoridan OLDIN keladi.
    # Faqat hdr_row+1 dan boshlasak, shu birinchi sanani o'tkazib yuborib,
    # birinchi bo'limdagi konteynerlar sanasiz (None) qolib ketardi.
    for ri in range(0, sh.nrows):
        row = sh.row_values(ri)

        # Bo'lim sanasi qatori — yangi jo'natish bo'limi boshlanmoqda
        if _list_sana_qatormi(row):
            yangi_sana = _sana_qator_qiymatini_parse(row[1], book.datemode)
            if yangi_sana:
                cur_sana = yangi_sana
            continue

        # Qayta-qayta chiqadigan sarlavha qatori ("材质 | 颜色名称 | 规格 ...")
        if row and str(row[0]).strip() == '材质':
            continue

        # "合计" (jami) qatori — bu bo'limning YAKUNIY (板厂重量/qadoqlash
        # qo'shilgan) summasi, ENDI ishlatilmaydi — har bir qatorning o'z
        # og'irligi pastda alohida o'qiladi.
        if row and str(row[0]).strip() == '合计':
            continue

        if len(row) <= bz_i:
            continue

        # Yangi konteyner bloki? Haqiqiy 柜号 topilmasa — 车号/车牌 (yuk
        # mashinasi) orqali pseudo-ID FALLBACK qilinadi, aks holda bunday
        # (konteyner raqamisiz, faqat mashina bilan tashiladigan) bo'lim
        # butunlay o'tkazib yuborilar edi (2026-07-06'da Труба faylida
        # topilgan xuddi shu turdagi jiddiy xato — bu yerda ham bir xil
        # fallback qo'llanadi, izchillik uchun).
        note = str(row[bz_i]).strip() if row[bz_i] else ''
        iso  = _iso(note) or _mashina_raqam(note)
        if iso:
            cur_iso = iso
            if cur_iso not in result:
                result[cur_iso] = {"sana": cur_sana, "items": []}

        if not cur_iso:
            continue

        spec  = str(row[gg_i]).strip() if row[gg_i] else ''
        marka = row[mat_i]
        rang  = str(row[rang_i]).strip() if rang_i < len(row) and row[rang_i] else ''

        if not re.match(r'^[\d\.]+[*×x]', spec):
            continue
        if any(k in spec for k in ('合计', '小计', '序号')):
            continue

        try:
            miqdor = int(float(row[qty_i])) if qty_i < len(row) and row[qty_i] else 0
        except (ValueError, TypeError):
            miqdor = 0
        if miqdor <= 0:
            continue

        xitoy_kg = row[ogirlik_i] if ogirlik_i is not None and ogirlik_i < len(row) else None
        raw_key = _tuzatish_kaliti("LIST", spec, marka, rang)
        saqlangan = _tuzatishdan_top(raw_key)
        if saqlangan:
            nom_final = saqlangan
        else:
            nom = _inventarga_moslashtir(_list_spec_to_name(spec, marka, rang))
            nom_final = nom or spec
        vazn_kg = _item_vazn_kg(nom_final, miqdor, xitoy_kg, inv_set)
        result[cur_iso]["items"].append((nom_final, miqdor, vazn_kg, raw_key))

    return result


# ── Aksessuar (Баласина/Чашка/Шар/Сокка va h.k.) konteyner fayli ────────────
# Bu tovarlar uchun Труба/Профиль/Лист kabi qat'iy formula (ст/уzunlik/规格)
# YO'Q — har bir nom o'ziga xos ("Баласина-18", "Найза №-05" kabi), Xitoy
# 装箱单/出货清单 formatidagi 柜号/规格/长度 ustunlari ham yo'q. Shuning uchun
# xodim tomonidan tayyorlanadigan JUDA SODDA jadval ishlatiladi:
#   T/r (No.) | Tovar nomi (Product Name) | Soni (Quantity, pcs)
# 2026-07-08: Huzayfa bilan kelishilgan — nom inventarda ANIQ topilmasa ham
# BLOKLANMAYDI, o'sha nom bilan "notanish" (⚠️) sifatida qo'shiladi (admin
# draft Excelda ko'rib to'g'irlashi yoki "baribir tasdiqlash"i mumkin) —
# xuddi Труба/Профиль oqimidagi kabi yumshoq ogohlantirish tartibi.

_AKSESSUAR_NOM_SIGNALS  = ('tovar', 'mahsulot', 'nomi', 'name', 'товар')
_AKSESSUAR_SONI_SIGNALS = ('soni', 'miqdor', 'quantity', 'qty', 'кол', 'сон')


def aksessuar_fayl_mi(raw: bytes) -> bool:
    """Fayl aksessuar (oddiy T/r|Tovar nomi|Soni) formatidami — TEZ tekshirish.
    Труба/Профиль/Лист fayllarida har doim Xitoycha ustun nomlari (柜号/规格/
    长度/库存 va h.k.) bo'ladi — aksessuar faylida esa faqat lotin/o'zbekcha
    sarlavhalar bo'ladi. Shu farq orqali ikkalasi bir-biridan ajratiladi.
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
        ws = wb.active
        nom_ok = False
        soni_ok = False
        for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), values_only=True):
            for c in row:
                if not c:
                    continue
                s = str(c).strip().lower()
                if any(sig in s for sig in _AKSESSUAR_NOM_SIGNALS):
                    nom_ok = True
                if any(sig in s for sig in _AKSESSUAR_SONI_SIGNALS):
                    soni_ok = True
                # Xitoycha 装箱单/出货清单 signali topilsa — bu AKSESSUAR EMAS
                if any(sig in str(c) for sig in ('柜号', '车号', '车牌', '规格', '长度', '库存', '品号', '颜色')):
                    return False
        return nom_ok and soni_ok
    except Exception:
        return False


def aksessuar_fayl_oqi(raw: bytes, iso: str, sana_s: str) -> dict:
    """Oddiy aksessuar konteyner faylini o'qib, xitoy_yuklar_oqi() bilan BIR
    XIL strukturadagi 'kont' dict qaytaradi (draft_excel_yarat/konteyner_xlsx_yarat
    kabi umumiy funksiyalar bilan to'g'ridan-to'g'ri ishlatilishi uchun).
    iso    — konteynerni ajratib turadigan nom (odatda yuborilgan fayl nomi).
    sana_s — 'DD.MM.YYYY' formatidagi sana (fayl nomida bo'lmasa — bugun).
    """
    wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    nom_i = None
    soni_i = None
    hdr_row = 0
    for ri, row in enumerate(rows[:10]):
        cand = [str(c).strip().lower() if c else "" for c in row]
        n_i = next((i for i, h in enumerate(cand)
                    if any(sig in h for sig in _AKSESSUAR_NOM_SIGNALS)), None)
        s_i = next((i for i, h in enumerate(cand)
                    if any(sig in h for sig in _AKSESSUAR_SONI_SIGNALS)), None)
        if n_i is not None and s_i is not None:
            nom_i, soni_i, hdr_row = n_i, s_i, ri
            break
    if nom_i is None or soni_i is None:
        return {"iso": iso, "sana": sana_s, "items": [], "manba": "aksessuar", "tonna": 0.0}

    items = []
    for row in rows[hdr_row + 1:]:
        if nom_i >= len(row) or soni_i >= len(row):
            continue
        nom_raw = row[nom_i]
        soni_raw = row[soni_i]
        if not nom_raw or str(nom_raw).strip() == "":
            continue
        try:
            miqdor = float(soni_raw) if soni_raw is not None else 0
        except (ValueError, TypeError):
            continue
        if miqdor <= 0:
            continue
        nom = _inventarga_moslashtir(str(nom_raw).strip()) or str(nom_raw).strip()
        items.append((nom, miqdor))

    return {"iso": iso, "sana": sana_s, "items": items, "manba": "aksessuar", "tonna": 0.0}


# ── 2026-07-13 (Huzayfa talabi): xom Xitoy装箱单/出货清单 o'rniga -- foydalanuvchi
# o'zi tayyorlaydigan, TARJIMA QILINMAYDIGAN (Xitoycha ierogliflarsiz) BITTA
# Excel fayl. 1-varaq odatda Труба/Профиль, 2-varaq odatda Лист -- lekin
# varaq NOMIGA emas, har bir varaqdagi USTUN sarlavhalariga qarab o'qiladi
# (shu sabab tartib/nom farq qilsa ham ishlaydi). Ustunlar (har ikkala
# varaqda bir xil): Konteyner raqami | Tovar nomi | Tovar miqdori |
# Yuklatilgan sana.
#
# MUHIM: bot endi Xitoycha spec kodini (masalan "0.85*50.8*5800") o'zi
# TARJIMA qilishga urinmaydi -- "Tovar nomi" ustuni allaqachon o'qiladigan
# ko'rinishda (masalan "Ф-32 ст 3,0 (6 м) (304 марка)") deb kutiladi. LEKIN
# baribir _inventarga_moslashtir() (== Buyurtma/Yuklatish rejasi'dagi
# _fix_oddiy_nom() bilan BIR XIL chuqur solishtiruv mexanizmi) orqali
# haqiqiy inventar bilan ANIQ moslashtiriladi -- bo'shliq/vergul-nuqta/
# "-0,05mm" kabi farqlar avtomatik tuzatiladi, haqiqatan notanish bo'lsa
# ⚠️ bilan draft Excelda (o'zgarishsiz qoladigan dizaynda) ko'rsatiladi.
#
# Bir xil "Konteyner raqami" bilan kelgan qatorlar -- ikkala varaqdan
# (Труба/Профиль VA Лист) bo'lsa ham -- BITTA konteynerga birlashtiriladi.
_TRF_KONT_SIG = ('konteyner', 'контейнер', 'контеинер')
_TRF_NOM_SIG  = ('tovar nomi', 'товар номи', 'товарноми', 'tovar', 'товар')
_TRF_MIQ_SIG  = ('miqdor', 'миқдор', 'микдор')
_TRF_SANA_SIG = ('sana', 'сана')


def tayyor_royhat_fayl_oqi(raw: bytes) -> list[dict]:
    """
    Tayyor (Xitoycha tarjimasiz) ro'yxat faylini o'qiydi -- ustunlar:
    Konteyner raqami | Tovar nomi | Tovar miqdori | Yuklatilgan sana.
    Bir yoki bir necha varaq bo'lishi mumkin (odatda Труба/Профиль +
    Лист) -- barchasi bir xil formatda o'qiladi, "Konteyner raqami"
    bo'yicha birlashtiriladi.

    Qaytaradi: xitoy_yuklar_oqi() bilan BIR XIL struktura --
    [{"iso":.., "sana":.., "items":[(tovar,miqdor,vazn_kg), ...],
      "manba":.., "tonna":..}, ...]
    """
    try:
        from parsers import _get_inventar_set
        inv_set = _get_inventar_set()
    except Exception:
        inv_set = set()
    try:
        from vazn_hisobla import tovar_vazni as _tv
    except Exception:
        _tv = None

    wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)

    kont_map: dict[str, dict] = {}   # iso -> {"sana": date|None, "items": [...]}

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        hdr_row = kont_i = nom_i = miq_i = sana_i = None
        for ri in range(min(5, len(rows))):
            row = rows[ri]
            if not row:
                continue
            cand = [str(c).strip().lower() if c else "" for c in row]
            k_i = next((i for i, h in enumerate(cand) if any(s in h for s in _TRF_KONT_SIG)), None)
            n_i = next((i for i, h in enumerate(cand) if any(s in h for s in _TRF_NOM_SIG)), None)
            m_i = next((i for i, h in enumerate(cand) if any(s in h for s in _TRF_MIQ_SIG)), None)
            if k_i is not None and n_i is not None and m_i is not None and k_i != n_i:
                s_i = next((i for i, h in enumerate(cand) if any(s in h for s in _TRF_SANA_SIG)), None)
                hdr_row, kont_i, nom_i, miq_i, sana_i = ri, k_i, n_i, m_i, s_i
                break
        if hdr_row is None:
            continue  # bu varaq mos formatda emas — o'tkazib yuboriladi

        for row in rows[hdr_row + 1:]:
            if not row or kont_i >= len(row) or nom_i >= len(row):
                continue
            iso_raw = row[kont_i]
            nom_raw = row[nom_i]
            if not iso_raw or not str(iso_raw).strip():
                continue
            if not nom_raw or not str(nom_raw).strip():
                continue
            iso = str(iso_raw).strip()

            miq_raw = row[miq_i] if miq_i is not None and miq_i < len(row) else 0
            try:
                miqdor = int(float(str(miq_raw).replace(',', '.').replace(' ', '')))
            except (ValueError, TypeError):
                continue
            if miqdor <= 0:
                continue

            sana_d = None
            if sana_i is not None and sana_i < len(row):
                sana_val = row[sana_i]
                if isinstance(sana_val, datetime):
                    sana_d = sana_val.date()
                elif isinstance(sana_val, date):
                    sana_d = sana_val
                elif sana_val:
                    sana_d = _sana_parse(str(sana_val).strip())

            nom_clean = str(nom_raw).strip()
            nom_final = _inventarga_moslashtir(nom_clean) or nom_clean

            unit_w = _tv(nom_final) if _tv else None
            vazn_kg = round(unit_w * miqdor, 2) if unit_w else 0.0

            entry = kont_map.setdefault(iso, {"sana": None, "items": []})
            if sana_d and (entry["sana"] is None or sana_d > entry["sana"]):
                entry["sana"] = sana_d
            entry["items"].append((nom_final, miqdor, vazn_kg))

    result = []
    for iso, data in kont_map.items():
        items = data["items"]
        kats = {_tovar_kategoriya(it[0]) for it in items}
        truba_bor = bool(kats & {"Труба", "Профиль"})
        list_bor  = "Лист" in kats
        if truba_bor and list_bor:
            manba = "aralash"
        elif truba_bor:
            manba = "truba"
        elif list_bor:
            manba = "list"
        else:
            manba = "tahrirlangan"
        ogirlik = sum(it[2] for it in items if len(it) > 2)
        result.append({
            "iso":   iso,
            "sana":  _sana_format(data["sana"]) if data["sana"] else "?",
            "items": items,
            "manba": manba,
            "tonna": round(ogirlik / 1000, 2) if ogirlik else 0.0,
        })

    return sorted(result, key=lambda k: k["iso"])


# ── Asosiy birlashtirish funksiyasi (eski, xom Xitoy装箱单/出货清单 uchun) ──────

def xitoy_yuklar_oqi(truba_raw: bytes, list_raw: bytes) -> list[dict]:
    """
    Ikki faylni o'qib, konteynerlar ro'yxatini qaytaradi.
    Qaytaradi: [
      {
        "iso":   "TCKU2238508",
        "sana":  "21.03.2026",
        "items": [("Ф-51 ст 0,9 (6 м) (201 марка)", 100, 6.37), ...],
        "manba": "aralash" | "truba" | "list",
        "tonna": 12.34,   # = items vaznlari yig'indisi / 1000 (Xitoy 合计 EMAS)
      }, ...
    ]
    """
    truba_map = _parse_truba_zhuangxiang(truba_raw)
    list_map  = _parse_list_chuhuo(list_raw)

    all_isos = set(truba_map) | set(list_map)
    result   = []

    for iso in sorted(all_isos):
        t_data = truba_map.get(iso)
        l_data = list_map.get(iso)

        if t_data and l_data:
            manba = "aralash"
            # Лист (出货清单) sanasi ustuvor — Труба (装箱单) faylida har bir
            # blok o'zining (odatda 1 kun oldingi, "qadoqlash") sanasi bilan
            # keladi, bu tarixiy yozuvlar (xitoy_parsed) bilan mos kelmay,
            # bir xil konteynerni "yangi" deb noto'g'ri aniqlab qo'yishi
            # mumkin edi. Лист sanasi arxivdagi ko'plab guruhlangan sanalar
            # bilan mos ekani tasdiqlangan.
            sana  = l_data["sana"] or t_data["sana"]
            items = t_data["items"] + l_data["items"]
        elif t_data:
            manba = "truba"
            sana  = t_data["sana"]
            items = t_data["items"]
        else:
            manba = "list"
            sana  = l_data["sana"]
            items = l_data["items"]

        sana_s = _sana_format(sana) if sana else "?"

        # Konteyner umumiy vazni — DASTUR o'zi hisoblagan (yoki tanish
        # bo'lmagan tovarlar uchun Xitoyning shu QATOR uchun bergan)
        # og'irliklar yig'indisi. Xitoyning "合计" (板厂重量 qo'shilgan
        # umumiy) raqami endi ISHLATILMAYDI — bu haqiqiy mahsulot
        # vaznidan sezilarli yuqori chiqadi (qadoqlash ulushi tufayli).
        ogirlik = sum(it[2] for it in items if len(it) > 2)

        result.append({
            "iso":   iso,
            "sana":  sana_s,
            "items": items,
            "manba": manba,
            "tonna": round(ogirlik / 1000, 2) if ogirlik else 0.0,
        })

    return result


# ── Mavjud konteynerlar bilan solishtirish ────────────────────────────────────

def yangi_konteynerlar(yuklar: list[dict], kont_dir: Path) -> list[dict]:
    """
    Allaqachon saqlangan xlsx fayllar bilan solishtiradi — ISO + sana
    (kelish sanasi) kombinatsiyasi bo'yicha. Faqat aniq shu ISO+sana juftligi
    bo'lgan konteyner "mavjud" deb hisoblanadi — chunki jismoniy konteyner
    raqamlari (ISO) dunyoda qayta-qayta ishlatiladi, shuning uchun bitta ISO
    boshqa sanada butunlay YANGI yuk bilan qaytib kelishi normal holat.
    """
    existing = set()
    if kont_dir.exists():
        for f in kont_dir.glob("*.xlsx"):
            # CRXU1561318_07.06.2026.xlsx yoki _..._D.xlsx
            stem = f.stem[:-2] if f.stem.endswith("_D") else f.stem
            iso, _, sana = stem.partition("_")
            if iso:
                existing.add((iso, sana))

    return [k for k in yuklar if (k["iso"], k["sana"]) not in existing]


# ── Sana chegarasi bo'yicha filtrlash ──────────────────────────────────────────

def _sana_parse(s: str) -> "date | None":
    """'21.03.2026' → date(2026,3,21). Noto'g'ri/bo'sh bo'lsa None."""
    try:
        return datetime.strptime(str(s).strip(), "%d.%m.%Y").date()
    except (ValueError, TypeError):
        return None


def oxirgi_malum_sana(kont_dir: Path, tarix: set = None) -> "date | None":
    """
    Tizimda hozircha ma'lum bo'lgan ENG OXIRGI (eng yangi) konteyner
    sanasini qaytaradi — kont_dir dagi fayl nomlaridan va (agar berilsa)
    tarix jurnalidan (fayli o'chirilgan konteynerlar ham hisobga olinishi
    uchun). Hech narsa topilmasa None.
    """
    sanalar: list[date] = []
    if kont_dir.exists():
        for f in kont_dir.glob("*.xlsx"):
            stem = f.stem[:-2] if f.stem.endswith("_D") else f.stem
            _, _, sana_s = stem.partition("_")
            d = _sana_parse(sana_s)
            if d:
                sanalar.append(d)
    if tarix:
        for kalit in tarix:
            _, _, sana_s = str(kalit).partition("|")
            d = _sana_parse(sana_s)
            if d:
                sanalar.append(d)
    return max(sanalar) if sanalar else None


def faqat_sanadan_keyingi(yuklar: list[dict], oxirgi: "date | None") -> list[dict]:
    """
    Faqat `oxirgi` sanadan KEYINGI (undan keyingi kunlardagi) yuklarnigina
    qoldiradi. `oxirgi` None bo'lsa (tizimda hali hech narsa yo'q) — hammasi
    o'tkaziladi. Sanasi aniqlanmagan ("?") yozuvlar ham o'tkaziladi — admin
    ko'rib chiqsin.

    DIQQAT (2026-07-06 YANGILANDI): bu funksiya GLOBAL chegara ishlatadi —
    barcha ISO'lar uchun bitta umumiy "oxirgi sana". Ilgari bu "eski-lekin-
    hech-qachon-qoshilmagan" konteynerlarni abadiy yashirib qo'yishi mumkin
    degan xavotir bilan asosiy filtrdan chetlashtirilgan edi. Lekin amalda
    teskari muammo chiqdi: Xitoy fayli kumulyativ (mart/aprel kabi juda
    eski yozuvlarni ham saqlab qoladi) va ayrim eski yozuvlar (masalan
    yangi payqalgan mashina-raqam psevdo-ID'lari) ISO sifatida "hech qachon
    uchramagan" bo'lib chiqib, sanasidan qat'iy nazar noto'g'ri "yangi" deb
    qo'shilib ketardi. Shu sababli endi bu funksiya `handlers.py`da BIRINCHI
    QADAM sifatida qayta ishga tushirildi (kont_list bosqichida, xom Xitoy
    faylini o'qigandan keyin, `iso_boyicha_yangilarini_ajrat()`dan OLDIN) —
    tizimdagi eng oxirgi ma'lum sanadan eskilarini butunlay kesib tashlaydi.
    """
    if oxirgi is None:
        return yuklar
    natija = []
    for k in yuklar:
        d = _sana_parse(k["sana"])
        if d is None or d > oxirgi:
            natija.append(k)
    return natija


def _mavjud_sanalar_iso_boyicha(kont_dir: Path, tarix: set = None) -> dict:
    """ISO → tizimda shu ISO uchun ma'lum bo'lgan barcha sanalar ro'yxati."""
    natija: dict = {}
    if kont_dir.exists():
        for f in kont_dir.glob("*.xlsx"):
            stem = f.stem[:-2] if f.stem.endswith("_D") else f.stem
            iso, _, sana_s = stem.partition("_")
            d = _sana_parse(sana_s)
            if iso and d:
                natija.setdefault(iso, []).append(d)
    if tarix:
        for kalit in tarix:
            iso, _, sana_s = str(kalit).partition("|")
            d = _sana_parse(sana_s)
            if iso and d:
                natija.setdefault(iso, []).append(d)
    return natija


def iso_boyicha_yangilarini_ajrat(yuklar: list[dict], kont_dir: Path, tarix: set = None) -> list[dict]:
    """
    IKKINCHI QADAM filtr — HAR BIR ISO alohida tekshiriladi (2026-07-06:
    endi bu `handlers.py`da `faqat_sanadan_keyingi()` GLOBAL sana filtridan
    KEYIN chaqiriladi, undan oldin emas — qarang shu funksiya haqidagi izoh):

      • Bu ISO tizimda (fayllar + tarix) UMUMAN uchramagan bo'lsa → YANGI
        (masalan avval unutilib qolgan/hech qachon qo'shilmagan konteyner —
        sanasi qancha eski bo'lishidan qat'iy nazar ko'rsatiladi).
      • Bu ISO bor, lekin yangi yukning sanasi undan KATTA (keyingi) bo'lsa
        → YANGI (jismoniy konteyner qayta ishlatilib, haqiqatan ham yangi
        yuk bilan qaytib kelgan).
      • Aks holda (sanasi mavjud eng katta sanadan katta emas) → allaqachon
        qayd etilgan, o'tkazib yuboriladi.
    """
    mavjud = _mavjud_sanalar_iso_boyicha(kont_dir, tarix)
    natija = []
    for k in yuklar:
        eski_sanalar = mavjud.get(k["iso"])
        if not eski_sanalar:
            natija.append(k)
            continue
        d = _sana_parse(k["sana"])
        if d is None or d > max(eski_sanalar):
            natija.append(k)
    return natija


# ── Konteyner xlsx yaratish ───────────────────────────────────────────────────

def konteyner_xlsx_yarat(kont: dict, kont_dir: Path) -> Path:
    """
    Bitta konteyner uchun xlsx fayl yaratadi.
    Fayl nomi: TCKU2238508_21.03.2026.xlsx
    Ustunlar: № | Mahsulot nomi | Mahsulot soni | Vazn_kg

    DIQQAT: ustun nomlari/tartibi ATAYLAB mavjud xitoy_parsed fayllari bilan
    BIR XIL qilib olindi (masalan eski AFY662_30.05.2026.xlsx: '№' |
    'Mahsulot nomi' | 'Mahsulot soni') — main.py'ning _parse_konteyner_fayli()
    funksiyasi aynan shu tuzilishni kutadi (mahsulot_col = ustun[1],
    miqdor_col = ustun[2:] ichidan birinchi RAQAMLI ustun). Avval bu yerda
    "Товар"/"Миқдор" (2 ustun, № siz) yozilardi — bu asosiy formatga MOS
    KELMASDI.

    "Vazn_kg" ustuni — har bir qatorning og'irligi (kg) BIR MARTA, shu
    konteyner qo'shilayotgan paytda hisoblanib, shu yerga YOZIB QO'YILADI
    (main.py buni keyinchalik "Контейнерлар" varag'iga o'tkazadi). Shu
    tufayli "Yo'ldagi konteynerlar" hisoboti har safar ochilganda tovar
    nomidan qayta hisoblashga hojat qolmaydi — faqat shu ustunni yig'adi.
    """
    kont_dir.mkdir(parents=True, exist_ok=True)
    sana_fayl = kont["sana"].replace('.', '-') if kont["sana"] != "?" else "nomalum"
    # Fayl nomi: ISO_DD-MM-YYYY.xlsx
    fname = f"{kont['iso']}_{kont['sana']}.xlsx"
    fpath = kont_dir / fname

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Yuk"
    ws.append(["№", "Mahsulot nomi", "Mahsulot soni", "Vazn_kg"])
    for i, item in enumerate(kont["items"], start=1):
        tovar, miqdor = item[0], item[1]
        vazn = item[2] if len(item) > 2 else ""
        ws.append([i, tovar, miqdor, vazn])

    # Ustun kengligi
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12

    wb.save(fpath)
    return fpath


# ── Tasdiqlashdan oldingi QISQA xulosa (chatga, Excel emas) ─────────────────

def _manba_tavsif(k: dict) -> str:
    """
    Konteyner qaysi fayl(lar)dan qanday tovar turlari topilib
    birlashtirilganini ANIQ ko'rsatadi — masalan:
    'Труба(Ф,Пр) ⟷ Лист' (ikkalasidan ham topilib birlashdi) yoki
    'faqat Труба(Ф) — Лист faylida topilmadi'.
    Bu ISO'ning nega/qanday birlashgani (yoki birlashmagani) darhol
    ko'rinishi uchun.
    """
    kats = {_tovar_kategoriya(it[0]) for it in k["items"]}
    truba_turlari = []
    if "Труба" in kats:
        truba_turlari.append("Ф")
    if "Профиль" in kats:
        truba_turlari.append("Пр")
    truba_bor = bool(truba_turlari)
    list_bor  = "Лист" in kats
    truba_str = f"Труба({','.join(truba_turlari)})" if truba_turlari else "Труба"

    if truba_bor and list_bor:
        return f"🔀 {truba_str} ⟷ Лист — ikkala fayldan topilib BITTA konteyner sifatida birlashdi"
    if truba_bor:
        return f"🔩 faqat {truba_str} topildi — Лист faylida bu ISO yo'q edi"
    if list_bor:
        return "📄 faqat Лист topildi — Труба faylida bu ISO yo'q edi"
    return "✏️ tahrirlangan"


def qisqa_xulosa(yangilar: list[dict]) -> str:
    """
    Excel yuborilishidan oldin ko'rsatiladigan QISQA xulosa — har bir
    konteyner uchun 2 qator: (1) qaysi fayl(lar)dan topilib qanday
    birlashgani, (2) sanasi/tonnaji/inventar ogohlantirishi.
    Admin shu orqali "bittada" nima yangi qo'shilayotgani, ikki fayl
    to'g'ri birlashtirilganmi va qayerda e'tibor kerakligini ko'radi.
    """
    try:
        from parsers import _get_inventar_set
        inv = _get_inventar_set()
    except Exception:
        inv = set()

    jami_tonna = round(sum(k.get("tonna", 0.0) for k in yangilar), 2)
    lines = [f"🆕 *{len(yangilar)} ta yangi konteyner* (jami {jami_tonna} t):\n"]
    for k in yangilar:
        nomos = sum(1 for it in k["items"] if inv and it[0] not in inv) if inv else 0
        ogohlantirish = f"  ⚠️ {nomos} ta tovar inventarda topilmadi" if nomos else ""
        lines.append(f"*{k['iso']}* — {_manba_tavsif(k)}")
        lines.append(f"   📅 {k['sana']}   ⚖️ {k.get('tonna', 0)} t{ogohlantirish}\n")
    return "\n".join(lines).rstrip()


def notanish_soni(yangilar: list[dict]) -> int:
    """
    Barcha konteynerlar bo'yicha JAMI (inventarda topilmagan, ⚠️ belgili)
    tovar qatorlari sonini qaytaradi — tasdiqlashdan oldin adminga
    alohida, aniq ogohlantirish ko'rsatish uchun (qisqa_xulosa()dagi
    har-konteyner ko'rsatkichi bilan bir xil mantiq, lekin yig'ilgan).
    """
    try:
        from parsers import _get_inventar_set
        inv = _get_inventar_set()
    except Exception:
        inv = set()
    if not inv:
        return 0
    return sum(
        1 for k in yangilar for it in k.get("items", []) if it[0] not in inv
    )


# ── Tekshirish uchun chiroyli (Yo'ldagi konteynerlar uslubidagi) Excel ───────

def _tovar_kategoriya(nom: str) -> str:
    """Tovar nomidan kategoriyasini aniqlaydi (Труба/Профиль/Лист)."""
    n = str(nom or "")
    if n.startswith("Ф-"):
        return "Труба"
    if n.startswith("Пр."):
        return "Профиль"
    if "Лист" in n:
        return "Лист"
    return ""


_HDR_ISO_SANA_RE   = re.compile(r'Yuklangan:\s*([\d.]+)')
_HDR_TONNA_RE      = re.compile(r'Tonna:\s*([\d.,]+)\s*t')
# Sarlavha qatoridagi ID'ni POZITSIYA bo'yicha ("🆕  <ID>    │ ...") ajratib
# oladi — _iso() dan farqli o'laroq, bu HAR QANDAY ID formatini (haqiqiy
# ISO HAM, mashina-asosidagi pseudo-ID — "ME5312", "ATY123" HAM) qo'llab-
# quvvatlaydi, chunki draft_excel_yarat() qanday "iso" bergan bo'lsa,
# xuddi shuni qayta o'qiy olishi shart (round-trip).
_HDR_ID_RE         = re.compile(r'🆕\s+(\S+)\s+│')


def draft_excel_yarat(yangilar: list[dict]) -> BytesIO:
    """
    Parslangan (yangi) konteynerlarni ADMIN tekshirishi uchun — "Yo'ldagi
    konteynerlar" Excelimiz bilan BIR XIL uslubda (rangli sarlavha bloklar,
    Tovar nomi/Miqdor/Kategoriya ustunlari, JAMI qatori) tayyorlaydi, ustiga
    har bir konteynerning umumiy TONNAsi ham sarlavhada ko'rsatiladi.

    Admin shu faylni ko'rib xato/to'g'riligini tekshiradi; xato bo'lsa xuddi
    shu faylni (Tovar nomi/Miqdor ustunlarini) tahrirlab qayta yuborishi
    mumkin — draft_excel_oqi() buni qayta o'qib oladi.
    """
    from yolda_excel import (
        _fill, _font, _align, _set_row, BORDER_THIN,
        CLR_YOLDA_BG, CLR_YOLDA_TEXT, CLR_YOLDA_SUB, CLR_COL_HDR_TEXT,
        CLR_TOTAL_BG, ROW_CLR_DARK, ROW_CLR_LIGHT, CAT_ORDER,
    )

    # Inventarda TOPILMAGAN tovarlarni qizil rangda ajratib ko'rsatish
    # uchun — admin Excelni ko'rib "notanish tovar" qaysi qatorda ekanini
    # darhol ko'radi, qidirib yurmaydi.
    try:
        from parsers import _get_inventar_set
        inv_set = _get_inventar_set()
    except Exception:
        inv_set = set()
    CLR_NOMOS_BG   = "FDEBD0"   # yumshoq sariq — inventarda topilmagan qator foni
    CLR_NOMOS_TEXT = "B9770E"   # to'q sariq/jigarrang matn

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Yangi konteynerlar"

    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    # G — ICHKI texnik ustun (xom Xitoy spec kaliti). Admin buni
    # O'ZGARTIRMASLIGI kerak — Tovar nomi ustunini to'g'irlab qayta
    # yuborilganda, bot aynan shu kalit orqali "bu xom spec uchun to'g'ri
    # nom shu" deb ESLAB QOLADI va keyingi safar qayta so'ramaydi.
    ws.column_dimensions["G"].width = 4
    ws.column_dimensions["G"].hidden = True

    # ── Umumiy xulosa qatori (eng tepada) ────────────────────────────────
    jami_tonna = round(sum(k.get("tonna", 0.0) for k in yangilar), 2)
    xulosa = ws.cell(row=1, column=1, value=(
        f"🆕  {len(yangilar)} ta yangi konteyner topildi    │    "
        f"Jami: {jami_tonna} t"
    ))
    xulosa.font = _font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.row_dimensions[1].height = 22
    cur_row = 3

    for kont in yangilar:
        iso   = kont["iso"]
        sana  = kont["sana"]
        tonna = kont.get("tonna", 0.0)
        manba_lbl = {"aralash": "🔀 Труба+Лист", "truba": "🔩 Труба",
                     "list": "📄 Лист", "tahrirlangan": "✏️ Тахрирланган"}.get(kont.get("manba"), "")

        # ── Sarlavha qatori ────────────────────────────────────────────
        hdr_text = (
            f"🆕  {iso}    │    Yuklangan: {sana}    │    "
            f"Tonna: {tonna} t    │    {len(kont['items'])} ta tovar"
            + (f"    │    {manba_lbl}" if manba_lbl else "")
        )
        hdr_cell = ws.cell(row=cur_row, column=1, value=hdr_text)
        hdr_cell.fill      = _fill(CLR_YOLDA_BG)
        hdr_cell.font      = _font(bold=True, color=CLR_YOLDA_TEXT, size=12)
        hdr_cell.alignment = _align(h="left", v="center")
        hdr_cell.border    = BORDER_THIN
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=6)
        ws.row_dimensions[cur_row].height = 24
        blok_start = cur_row
        cur_row += 1

        # ── Ustun sarlavhalari ─────────────────────────────────────────
        col_headers = ["Tovar nomi", "Miqdor", "Kategoriya", "Vazn (kg)", "", "", ""]
        col_fills   = [_fill(CLR_YOLDA_SUB)] * 6 + [None]
        col_fonts   = [_font(bold=True, color=CLR_COL_HDR_TEXT)] * 6 + [None]
        col_aligns  = [
            _align(h="left",   v="center", indent=1),
            _align(h="center", v="center"),
            _align(h="left",   v="center", indent=1),
        ] + [_align()] * 4
        _set_row(ws, cur_row, col_headers, fills=col_fills, fonts=col_fonts,
                  aligns=col_aligns, height=18)
        cur_row += 1

        # ── Tovar qatorlari (Лист → Труба → Профиль tartibida) ──────────
        items_sorted = sorted(
            kont["items"],
            key=lambda it: CAT_ORDER.get(_tovar_kategoriya(it[0]), 99)
        )
        jami_miq = 0
        for i, item in enumerate(items_sorted):
            tovar, miqdor = item[0], item[1]
            vazn_kg = item[2] if len(item) > 2 else 0.0
            raw_key = item[3] if len(item) > 3 else ""
            kat = _tovar_kategoriya(tovar)
            jami_miq += miqdor
            nomos = bool(inv_set) and tovar not in inv_set
            if nomos:
                row_bg = _fill(CLR_NOMOS_BG)
                tovar_disp = f"⚠️ {tovar}"
                tovar_font = _font(size=10, bold=True, color=CLR_NOMOS_TEXT)
            else:
                row_clr = ROW_CLR_DARK if i % 2 == 0 else ROW_CLR_LIGHT
                row_bg = _fill(row_clr)
                tovar_disp = tovar
                tovar_font = _font(size=10)
            values  = [tovar_disp, miqdor, kat, vazn_kg, "", "", raw_key]
            fills   = [row_bg] * 6 + [None]
            fonts   = [tovar_font, _font(size=10, bold=True),
                       _font(size=10, color="444444")] + [_font(size=10)] * 4
            aligns  = [
                _align(h="left",   v="center", wrap=True, indent=1),
                _align(h="center", v="center"),
                _align(h="left",   v="center", indent=1),
            ] + [_align()] * 4
            _set_row(ws, cur_row, values, fills=fills, fonts=fonts,
                      aligns=aligns, height=17)
            cur_row += 1

        # ── JAMI qatori ──────────────────────────────────────────────────
        jami_vals  = ["JAMI:", jami_miq, f"{tonna} t", "", "", "", ""]
        jami_fills = [_fill(CLR_TOTAL_BG)] * 6 + [None]
        jami_fonts = [_font(bold=True, size=10), _font(bold=True, size=10),
                      _font(bold=True, size=10, color="444444")] + [_font(size=10)] * 4
        jami_aligns = [
            _align(h="right",  v="center"),
            _align(h="center", v="center"),
            _align(h="left",   v="center", indent=1),
        ] + [_align()] * 4
        _set_row(ws, cur_row, jami_vals, fills=jami_fills, fonts=jami_fonts,
                  aligns=jami_aligns, height=16)
        blok_end = cur_row
        cur_row += 1

        # ── Blok atrofini qalin chegara bilan o'rash ─────────────────────
        from openpyxl.styles import Border, Side
        brd_s = Side(style="medium", color="1A3A5C")
        brd_i = Side(style="thin", color="CCCCCC")
        for r in range(blok_start, blok_end + 1):
            for c in range(1, 7):
                cell = ws.cell(row=r, column=c)
                cell.border = Border(
                    left=brd_s if c == 1 else brd_i,
                    right=brd_s if c == 6 else brd_i,
                    top=brd_s if r == blok_start else brd_i,
                    bottom=brd_s if r == blok_end else brd_i,
                )

        ws.row_dimensions[cur_row].height = 8
        cur_row += 1

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def draft_excel_oqi(raw: bytes) -> list[dict]:
    """
    draft_excel_yarat() natijasini (admin tahrirlagan yoki tahrirlamagan
    holda) qayta o'qib, xitoy_yuklar_oqi() bilan bir xil formatga qaytaradi:
    [{"iso", "sana", "items": [(tovar, miqdor, vazn_kg), ...],
      "manba": "tahrirlangan", "tonna": float}]

    Har bir konteyner blokining sarlavha qatoridan ("🆕  ISO │ Yuklangan: ...
    │ Tonna: ... t │ ...") ISO/sana o'qiladi, undan keyingi
    Tovar-nomi/Miqdor/Vazn(kg) qatorlari "JAMI:" qatorigacha yig'iladi.

    DIQQAT: tonna sarlavhadagi (eski, admin tahrirlashdan OLDINGI) qiymatdan
    emas — "Vazn (kg)" ustunidagi (yoki admin uni o'chirib qo'ygan bo'lsa,
    qayta hisoblangan) qiymatlar yig'indisidan olinadi, aks holda admin
    miqdorni o'zgartirsa ham eski (endi noto'g'ri) tonna qolib ketardi.
    """
    try:
        from parsers import _get_inventar_set
        inv_set = _get_inventar_set()
    except Exception:
        inv_set = set()

    wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
    ws = wb.active

    natija: list[dict] = []
    cur_iso = cur_sana = None
    cur_items: list = []
    in_items = False

    def _blokni_yopish():
        if cur_iso and cur_items:
            jami_kg = sum(it[2] for it in cur_items)
            natija.append({
                "iso": cur_iso, "sana": cur_sana or "?", "items": list(cur_items),
                "manba": "tahrirlangan", "tonna": round(jami_kg / 1000, 2),
            })

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        a = row[0]
        a_s = str(a).strip() if a is not None else ""

        # Sarlavha qatori — yangi konteyner bloki boshlanadi
        if a_s.startswith("🆕") and "Yuklangan:" in a_s:
            _blokni_yopish()
            # Avval haqiqiy ISO formatini sinab ko'ramiz (eng ishonchli),
            # topilmasa — pozitsiya bo'yicha (mashina-asosidagi pseudo-ID
            # kabi holatlar uchun, masalan "ME5312") FALLBACK qilinadi.
            cur_iso   = _iso(a_s)
            if not cur_iso:
                m_id = _HDR_ID_RE.search(a_s)
                cur_iso = m_id.group(1) if m_id else None
            m_sana    = _HDR_ISO_SANA_RE.search(a_s)
            cur_sana  = m_sana.group(1) if m_sana else "?"
            cur_items = []
            in_items  = False
            continue

        if a_s == "Tovar nomi":
            in_items = True
            continue

        if a_s == "JAMI:":
            in_items = False
            continue

        if not in_items or not a_s or not cur_iso:
            continue

        # "⚠️ " belgisi faqat KO'RSATISH uchun (inventarda topilmagan
        # tovarni ajratib ko'rsatadi) — qayta o'qishda olib tashlanadi,
        # aks holda tovar nomi ichida qolib, doim "notanish" bo'lib qolar edi.
        if a_s.startswith("⚠️"):
            a_s = a_s.replace("⚠️", "", 1).strip()

        miq = row[1] if len(row) > 1 else None
        try:
            miq_i = int(float(miq)) if miq not in (None, "") else 0
        except (ValueError, TypeError):
            miq_i = 0
        if miq_i <= 0:
            continue

        # "Vazn (kg)" ustuni (D, index 3) — admin uni saqlab qoldirgan bo'lsa
        # o'qiladi; bo'sh/o'chirilgan bo'lsa (masalan admin yangi qator
        # qo'shgan) — o'z formulamiz bilan qayta hisoblanadi (tanish
        # bo'lmasa 0, chunki bu bosqichda Xitoyning original qator vazni
        # endi qo'limizda yo'q).
        vazn_raw = row[3] if len(row) > 3 else None
        try:
            vazn_kg = float(vazn_raw) if vazn_raw not in (None, "") else None
        except (ValueError, TypeError):
            vazn_kg = None
        if vazn_kg is None:
            vazn_kg = _item_vazn_kg(a_s, miq_i, None, inv_set)

        # G ustuni (index 6) — YASHIRIN texnik kalit (xom Xitoy spec).
        # Admin shu qatorda YOZGAN nom ("Tovar nomi", a_s) hozirgi
        # AVTOMATIK (evristika) natijadan FARQ QILSAGINA — bu chindan
        # ham qo'lda tuzatish deb hisoblanib, doimiy saqlanadi. Farq
        # bo'lmasa (admin hech narsani o'zgartirmagan, shunchaki qayta
        # yuborgan) — saqlanmaydi, aks holda har bir allaqachon to'g'ri
        # qator ham abadiy "muzlab" qolib, kelajakda inventar o'zgarsa
        # ham eski javobni qaytaraverar edi.
        raw_key = row[6] if len(row) > 6 and row[6] else None
        if raw_key:
            raw_key = str(raw_key).strip()
            avtomatik = _xom_keydan_avtomatik_nom(raw_key, inv_set)
            if avtomatik is not None and avtomatik != a_s:
                _tuzatish_saqla(raw_key, a_s)

        cur_items.append((a_s, miq_i, vazn_kg, raw_key or ""))

    _blokni_yopish()
    return natija


# ── Preview matn ─────────────────────────────────────────────────────────────

def preview_matn(yangilar: list[dict], oxirgi_sana: "date | None" = None) -> str:
    """
    Foydalanuvchiga ko'rsatiladigan qisqacha xabar.
    """
    oxirgi_qator = (
        f"📅 _Tizimdagi eng oxirgi ma'lum sana: {oxirgi_sana.strftime('%d.%m.%Y')} "
        f"— shundan keyingi yuklar qidirildi._\n\n"
        if oxirgi_sana else ""
    )
    if not yangilar:
        return oxirgi_qator + "✅ Barcha konteynerlar allaqachon ro'yxatda."

    aralash = [k for k in yangilar if k["manba"] == "aralash"]
    faqat_t = [k for k in yangilar if k["manba"] == "truba"]
    faqat_l = [k for k in yangilar if k["manba"] == "list"]

    lines = [oxirgi_qator + f"🆕 *{len(yangilar)} ta yangi konteyner topildi:*\n"]

    if aralash:
        lines.append("🔀 *Birlashtirilgan (Трубa + Лист):*")
        for k in aralash:
            tonna_qator = f" — {k['tonna']} t" if k.get("tonna") else ""
            lines.append(f"  • `{k['iso']}` — {k['sana']} — {len(k['items'])} ta tovar{tonna_qator}")

    if faqat_t:
        lines.append("\n🔩 *Faqat Труба/Профиль:*")
        for k in faqat_t:
            tonna_qator = f" — {k['tonna']} t" if k.get("tonna") else ""
            lines.append(f"  • `{k['iso']}` — {k['sana']} — {len(k['items'])} ta tovar{tonna_qator}")

    if faqat_l:
        lines.append("\n📄 *Faqat Лист:*")
        for k in faqat_l:
            tonna_qator = f" — {k['tonna']} t" if k.get("tonna") else ""
            lines.append(f"  • `{k['iso']}` — {k['sana']} — {len(k['items'])} ta tovar{tonna_qator}")

    lines.append("\n✅ Tasdiqlaysizmi?")
    return "\n".join(lines)
