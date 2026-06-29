"""
excel_styles.py
===============
Barcha bot-generated Excel fayllari uchun umumiy uslub (style) moduli.
Import qiling:
    from excel_styles import *
    from excel_styles import FONT, CLR, fill, font, border, align, CAT_COLORS
"""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── Font ──────────────────────────────────────────────────────────────────────
FONT_NAME   = "Calibri"
FONT_SIZE   = 11       # asosiy
FONT_SIZE_H = 12       # sarlavha
FONT_SIZE_T = 13       # yirik sarlavha

# ── Asosiy ranglar ────────────────────────────────────────────────────────────
CLR_MAIN_HDR  = "1A3A5C"   # asosiy sarlavha (to'q ko'k)
CLR_SUB_HDR   = "1F618D"   # kichik sarlavha (o'rta ko'k)
CLR_COL_HDR   = "2C4A6E"   # ustun nomlari
CLR_SEPARATOR = "2C3E50"   # bloklar orasidagi ajratgich (to'q)
CLR_TOTAL     = "D5DBDB"   # jami qator (kulrang)
CLR_WHITE     = "FFFFFF"
CLR_BLACK     = "000000"
CLR_RED       = "C0392B"
CLR_ORANGE    = "E67E22"
CLR_GREEN     = "1E8449"
CLR_BLUE      = "1F618D"

# ── Kategoriya ranglari (toq_bg, yengil_bg, text_color) ──────────────────────
CAT_COLORS = {
    "Лист":     ("D6EAF8", "EBF5FB", "2E86C1"),
    "Труба":    ("D5F5E3", "EAFAF1", "1E8449"),
    "Профиль":  ("FEF9E7", "FDFEFE", "B7950B"),
    "Баласина": ("F9EBEA", "FDEDEC", "922B21"),
    "Стойка":   ("F5EEF8", "FAF5FF", "7D3C98"),
    "_other":   ("F2F3F4", "FDFEFE", "5D6D7E"),
}

# ── Marka fon ranglari ────────────────────────────────────────────────────────
MARKA_BG = {
    "201": "E8F4FD",
    "304": "E8F8E8",
    "430": "FFFBEA",
    "316": "F3E8FF",
    "321": "FFE8E8",
    "":    "F8F8F8",
}

# ── Holat ranglari ────────────────────────────────────────────────────────────
HOLAT_COLOR = {
    "КРИТИК": CLR_RED,
    "ПАСТ":   CLR_ORANGE,
    "НОРМА":  CLR_GREEN,
}

# ── Yordamchi funksiyalar ─────────────────────────────────────────────────────
def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def font(bold: bool = False, size: int = FONT_SIZE,
         color: str = CLR_BLACK, italic: bool = False) -> Font:
    return Font(name=FONT_NAME, size=size, bold=bold,
                color=color, italic=italic)


def border(style: str = "thin", color: str = "CCCCCC") -> Border:
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def border_medium(color: str = "888888") -> Border:
    s = Side(style="medium", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def align(h: str = "left", v: str = "center",
          wrap: bool = False, indent: int = 0) -> Alignment:
    return Alignment(horizontal=h, vertical=v,
                     wrap_text=wrap, indent=indent)


def holat_color(holat: str) -> str:
    """Holat matni bo'yicha rang qaytaradi."""
    s = str(holat)
    for k, v in HOLAT_COLOR.items():
        if k in s:
            return v
    return "555555"


def separator_row(ws, row: int, n_cols: int = 6, height: int = 4) -> int:
    """To'q ajratgich qator (bloklar orasida)."""
    sep = fill(CLR_SEPARATOR)
    for ci in range(1, n_cols + 1):
        ws.cell(row=row, column=ci, value="").fill = sep
    ws.row_dimensions[row].height = height
    return row + 1


def header_row(ws, row: int, text: str, n_cols: int = 6,
               bg: str = CLR_MAIN_HDR, size: int = FONT_SIZE_T,
               height: int = 28) -> int:
    """Asosiy sarlavha qatori (merge qilingan)."""
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill      = fill(bg)
    c.font      = font(bold=True, size=size, color=CLR_WHITE)
    c.alignment = align(h="center")
    ws.row_dimensions[row].height = height
    return row + 1


def col_headers_row(ws, row: int, headers: list, widths: list = None,
                    bg: str = CLR_COL_HDR, height: int = 18) -> int:
    """Ustun sarlavhalari qatori."""
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill      = fill(bg)
        c.font      = font(bold=True, size=FONT_SIZE, color=CLR_WHITE)
        c.alignment = align(h="center")
        c.border    = border()
    ws.row_dimensions[row].height = height
    if widths:
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return row + 1
