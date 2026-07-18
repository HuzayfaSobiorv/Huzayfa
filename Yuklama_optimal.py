# -*- coding: utf-8 -*-
"""
Yuklama_optimal.py  — ABC asosida optimal konteyner yuklash.

12m konteyner: faqat 6.0 m Truба/Profil, jami 28 t, kamida 3 xil tovar
6m  konteyner: Truба/Profil (5.8 m) <=11 t  +  List <=18 t, jami <=28 t
Ustunlik: ABC sinf (A->B->C) -> Holat (KRITIK->PAST) -> Kam miqdori
"""
import re, sys, logging
from pathlib import Path

_THIS_DIR = Path(__file__).resolve().parent
if str(_THIS_DIR) not in sys.path:
    sys.path.insert(0, str(_THIS_DIR))

import pandas as pd
logger = logging.getLogger(__name__)

LIMIT_TOTAL        = 28_000
LIMIT_TRUBA_PROFIL = 11_000   # 6m konteyner uchun (backward compat)
LIMIT_LIST         = 18_000   # 6m konteyner uchun (backward compat)
LIMITS_BY_TYPE = {
    "12m": {"truba_profil": 28_000, "list":      0},
    "6m":  {"truba_profil": 11_000, "list": 18_000},
}
ABC_CAP_PCT = {"A": 0.40, "B": 0.60, "C": 1.00}
ABC_RANK    = {"A": 0, "B": 1, "C": 2}


def _holat_rank(holat: str) -> int:
    s = str(holat)
    if any(x in s for x in ("КРИТ", "CRIT")):   return 0
    if any(x in s for x in ("ПАСТ", "PAST")):   return 1
    if any(x in s for x in ("НОРМ", "NORM")):   return 2
    return 3


def get_length(name: str) -> float | None:
    m = re.search(r'\(([\d,\.]+)\s*[мm]\)', str(name))
    return float(m.group(1).strip().replace(",", ".")) if m else None


def get_category(name: str) -> str:
    s = str(name).strip()
    if re.match(r'^(\([^)]*\)\s*)?\u0424-\d+', s):   return "\u0422\u0440\u0443\u0431\u0430"
    if re.match(r'^(\([^)]*\)\s*)?\u041f\u0440\.\s*\d+', s): return "\u041f\u0440\u043e\u0444\u0438\u043b\u044c"
    if s.startswith("\u041b\u0438\u0441\u0442"):         return "\u041b\u0438\u0441\u0442"
    return "\u0411\u043e\u0448\u049b\u0430"


def yuk_turi(name: str, cat: str) -> str:
    if cat == "\u041b\u0438\u0441\u0442":
        return "6m"
    return "12m" if get_length(name) == 6.0 else "6m"


def _yangi_yuk(turi: str) -> dict:
    return {"turi": turi, "truba_profil_kg": 0.0,
            "list_kg": 0.0, "jami_kg": 0.0, "items": [], "_item_kg": {}}


def abc_map_yuklash() -> dict:
    """
    2026-07-10: ABC manbasi endi YAGONA -- Minimal_zaxiralar/Min_Zaxira.xlsx
    ning "ABC" ustuni (Huzayfa bilan kelishilgan: u yerni qo'lda tahrirlaydi,
    boshliq bilan ko'rib chiqadi). Eski alohida abc_lookup.xlsx endi
    ishlatilmaydi -- ikki fayl orasida farq/eskirish xavfi bo'lmasin.
    """
    try:
        import openpyxl
        f = _THIS_DIR / "Minimal_zaxiralar" / "Min_Zaxira.xlsx"
        if not f.exists():
            logger.warning("Min_Zaxira.xlsx topilmadi: %s", f); return {}
        wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
        ws = wb["Min_Zaxira"] if "Min_Zaxira" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {}
        header = [str(h).strip() if h else "" for h in rows[0]]
        tovar_i = next((i for i, h in enumerate(header) if "Товар" in h or "tovar" in h.lower()), None)
        abc_i   = next((i for i, h in enumerate(header) if h.strip().upper() == "ABC"), None)
        if tovar_i is None or abc_i is None:
            logger.warning("Min_Zaxira.xlsx: Товар/ABC ustunlari topilmadi (%s)", header)
            return {}
        res = {}
        for row in rows[1:]:
            if not row or tovar_i >= len(row) or abc_i >= len(row):
                continue
            tovar = row[tovar_i]
            abc   = row[abc_i]
            if not tovar or not isinstance(tovar, str) or abc is None:
                continue
            # 2026-07-14: qo'lda tahrirda uchraydigan variantlar
            # normallashtiriladi — kichik harf (a/b/c) va KIRILLcha А/В/С
            # (rus klaviaturada terilsa lotincha bilan BIR XIL ko'rinadi,
            # farqini ko'z bilan sezib bo'lmaydi!). Ilgari bunday qiymatlar
            # jimgina rad etilib, tovar "C" (eng past ustunlik) deb qolardi.
            abc_s = str(abc).strip().upper()
            abc_s = {"А": "A", "В": "B", "С": "C"}.get(abc_s, abc_s)  # kirill→lotin
            if abc_s not in ("A", "B", "C"):
                continue
            res[tovar.strip()] = abc_s
        wb.close()
        logger.info("abc_map (Min_Zaxira dan): %d ta tovar", len(res))
        return res
    except Exception as e:
        logger.error("abc_map_yuklash: %s", e); return {}


def tovar_vazni(tovar_nomi: str) -> float | None:
    try:
        from services import tovar_vazni_pb
        return tovar_vazni_pb(tovar_nomi)
    except Exception:
        return None


def _truba_diametr(name: str) -> float | None:
    m = re.search(r'\u0424-([\d,\.]+)', str(name))
    if not m:
        return None
    try:
        return float(m.group(1).replace(',', '.'))
    except ValueError:
        return None


def _profil_min_dim(name: str) -> float | None:
    m = re.search(r'(\d+)[\u0445x](\d+)', str(name))
    if not m:
        return None
    return min(float(m.group(1)), float(m.group(2)))


def _list_qalinlik(name: str) -> float | None:
    m = re.match(r'^\u041b\u0438\u0441\u0442-?\s*([\d,\.]+)', str(name).strip())
    if not m:
        return None
    try:
        return float(m.group(1).replace(',', '.'))
    except ValueError:
        return None


def _kichikmi(tovar: str, cat: str) -> bool:
    """
    2026-07-10 (Huzayfa bilan kelishilgan qoida): mayda-chuda tovarlar
    bo'lib-bo'lib, kam miqdorda (hatto 1 dona) yuklanib "uyalarli" ko'rinish
    bermasin -- katta/qalin tovarlar uchun kam miqdor (20-50 dona) esa
    normal (ular og'ir/qimmat, kam sonda ham katta hajm/qiymat beradi).
      Труба:   diametr <= 51mm       -> kichik
      Профиль: eng kichik tomoni <50mm -> kichik
      Лист:    qalinlik <5.0mm        -> kichik
    """
    if cat == "\u0422\u0440\u0443\u0431\u0430":
        d = _truba_diametr(tovar)
        return d is not None and d <= 51
    if cat == "\u041f\u0440\u043e\u0444\u0438\u043b\u044c":
        d = _profil_min_dim(tovar)
        return d is not None and d < 50
    if cat == "\u041b\u0438\u0441\u0442":
        q = _list_qalinlik(tovar)
        return q is not None and q < 5.0
    return False


MIN_KICHIK_DONA = 10  # 2026-07-18 (Huzayfa): 2 -> 10, mayda tovar 10 donadan kam bo'lsa yuklanmaydi
MIN_KICHIK_PARCHA = 20  # qisman joylashtirishda bundan kichik "bo'lak" qabul qilinmaydi
# 2026-07-18 (Huzayfa, "kulgili 1-2 dona" muammosi): vazn-asosli minimal qator.
# Qator (bitta konteynerdagi bitta tovar yozuvi) "mayda" hisoblanadi, agar
# DONA ham (< MIN_KICHIK_PARCHA), VAZN ham (< MIN_QATOR_KG) past bo'lsa --
# og'ir tovar (mas. 2 dona x 150 kg list ~ 300 kg) kam donada ham normal,
# yengil tovar esa faqat sezilarli partiyada yuklanadi.
MIN_QATOR_KG = 150.0


def _mayda_qatormi(dona: float, vazn_dona: float) -> bool:
    """Qator ham dona, ham vazn bo'yicha chegaradan past bo'lsa -- "mayda"."""
    return dona < MIN_KICHIK_PARCHA and dona * vazn_dona < MIN_QATOR_KG


def optimallashtir(
    kerak_df: pd.DataFrame,
    mavjud_df: pd.DataFrame,
    abc_map: dict | None = None,
    max_yuklar: int = 20,
    xitoy_vazn: dict | None = None,
) -> tuple[list[dict], list[dict], list[dict], list[dict]]:
    """kerak_df: Товар|Холат|Кам  mavjud_df: Товар|Миқдор
    xitoy_vazn: {tovar_nomi: 1_dona_kg} — vazn_lookup da yo'q tovarlar uchun fallback
    Qaytaradi: (yuklar, qolgan, kerak_yoq, vazn_yoq)
      qolgan    — limit yoki minimal-miqdor qoidasi tufayli yuklanmadi
      kerak_yoq — hozircha doim bo'sh ro'yxat (2026-07-11: Кам-cheklovi
                  bekor qilindi -- Xitoyda tayyor bo'lgan HAMMASI yana
                  yuklanadi, "juda cheklovchi/qulay emas" bo'lib chiqdi)
      vazn_yoq  — 2026-07-18: vazni HECH QAYERDAN topilmagan tovarlar
                  (avval jimgina tashlab ketilardi -- endi Excelda
                  ogohlantirish bloki bilan ko'rsatiladi)
    """
    if abc_map is None:
        abc_map = abc_map_yuklash()
    mavjud = dict(zip(mavjud_df["Товар"], mavjud_df["Миқдор"]))
    kerak = kerak_df.copy()
    kerak["_ar"] = kerak["Товар"].map(lambda t: ABC_RANK.get(abc_map.get(str(t).strip(), "C"), 2))
    kerak["_hr"] = kerak["Холат"].map(_holat_rank)
    kerak = kerak.sort_values(["_ar", "_hr", "Кам"], ascending=[True, True, False])

    yuklar: list[dict] = []
    qolgan: list[dict] = []
    kerak_yoq: list[dict] = []   # 2026-07-11: cheklov bekor qilindi, doim bo'sh
    vazn_yoq: list[dict] = []    # 2026-07-18: vazni topilmagan tovarlar

    for _, row in kerak.iterrows():
        tovar      = row["Товар"]
        kerak_dona = int(row["Кам"])
        bor_dona   = int(mavjud.get(tovar, 0))
        if bor_dona <= 0:
            continue
        cat  = get_category(tovar)
        turi = yuk_turi(tovar, cat)
        if cat in ("\u0411\u043e\u0448\u049b\u0430",):
            # Aksessuar/boshqa toifa -- vazn ataylab hisoblanmaydi (loyiha qarori)
            continue
        vazn_dona = tovar_vazni(tovar)
        # Xitoy faylidan fallback: vazn_lookup da yo'q tovar uchun
        if (not vazn_dona or vazn_dona <= 0) and xitoy_vazn:
            vazn_dona = xitoy_vazn.get(str(tovar).strip())
        if not vazn_dona or vazn_dona <= 0:
            # 2026-07-18: 3-zaxira -- nomdan formula bilan hisoblash
            # (vazn_hisobla, Xitoy stenkasi -0.05 konventsiyasi bilan)
            try:
                from vazn_hisobla import tovar_vazni as _vazn_formula
                vazn_dona = _vazn_formula(tovar, xitoy=True)
            except Exception:
                vazn_dona = None
        if not vazn_dona or vazn_dona <= 0:
            # 2026-07-18: avval jimgina tashlab ketilardi (foydalanuvchi
            # tovar "sirli yo'qolgan"ini bilmasdi) -- endi ro'yxatga tushadi
            vazn_yoq.append({"tovar": tovar, "dona": bor_dona})
            continue
        kichik = _kichikmi(tovar, cat)
        if kichik and bor_dona < MIN_KICHIK_DONA:
            # 2026-07-10: mayda tovar juda oz -- yuklanmaydi, "qolgan"
            # ro'yxatiga tushadi (keyingi safar to'planganda yuklanadi)
            qolgan.append({"tovar": tovar, "dona": bor_dona,
                           "vazn_kg": round(bor_dona * vazn_dona, 2)})
            continue
        if _mayda_qatormi(bor_dona, vazn_dona):
            # 2026-07-18: butun zaxira ham dona, ham vazn bo'yicha mayda
            # (mas. katta tovardan 1-2 dona) -- kulgili qator yaratmaymiz
            qolgan.append({"tovar": tovar, "dona": bor_dona,
                           "vazn_kg": round(bor_dona * vazn_dona, 2)})
            continue
        key     = "list_kg"          if cat == "\u041b\u0438\u0441\u0442" else "truba_profil_kg"
        lim_key = "list"             if cat == "\u041b\u0438\u0441\u0442" else "truba_profil"
        abc_s   = abc_map.get(str(tovar).strip(), "C")
        # 2026-07-10 (tuzatildi): ABC diversifikatsiya cheklovi (cap_pct)
        # BARCHA tovar uchun o'chirildi -- avval faqat "kichik" tovarlarga
        # qo'llanilgandi, lekin "katta" tovarlarda ham xuddi shu muammo
        # chiqdi (masalan jami 30 dona "Лист-5,95" 1+2+27 bo'lib uch joyga
        # bo'linib ketgan edi). Endi HAMMA tovar uchun cheklovsiz (1.0) --
        # mayda bo'lak yaratmaslik MIN_KICHIK_PARCHA tekshiruvi orqali
        # ta'minlanadi (pastda), diversifikatsiya emas.
        cap_pct = 1.0
        # 2026-07-11 (qaytarildi -- Huzayfa: "juda cheklovchi bo'ldi"):
        # Кам bilan cheklash sinovdan o'tmadi -- juda ko'p tovar
        # yuklanmay qolib ketardi. Xitoyda nechta tayyor bo'lsa,
        # yana hammasi yuklanadi (Кам faqat tartib/ustunlik uchun).
        qoldi   = bor_dona

        while qoldi > 0:
            for yuk in yuklar:
                if yuk["turi"] != turi:
                    continue
                slot_limit  = LIMITS_BY_TYPE[turi][lim_key]
                if slot_limit <= 0:
                    continue
                already_kg  = yuk["_item_kg"].get(tovar, 0.0)
                qolgan_cap  = slot_limit * cap_pct - already_kg
                qolgan_slot = slot_limit            - yuk[key]
                qolgan_jami = LIMIT_TOTAL           - yuk["jami_kg"]
                sig_kg      = min(qolgan_cap, qolgan_slot, qolgan_jami)
                sig_dona    = int(sig_kg // vazn_dona)
                if sig_dona <= 0:
                    continue
                if sig_dona < qoldi and _mayda_qatormi(sig_dona, vazn_dona):
                    # 2026-07-10 (2026-07-18 vazn-asosli qilindi): bu
                    # konteynerda faqat MAYDA bo'lak (dona ham, vazn ham past)
                    # sig'sa -- bu yerga tashlab qo'yilmaydi, boshqa (yoki
                    # yangi) konteyner qidiriladi. Endi og'ir tovarning kichik
                    # (lekin >= MIN_QATOR_KG) bo'lagi qabul qilinadi --
                    # to'ldirish yaxshilanadi, qirindi baribir bloklanadi.
                    continue
                dona = min(qoldi, sig_dona)
                og   = round(dona * vazn_dona, 2)
                yuk["items"].append({"tovar": tovar, "dona": dona, "vazn_kg": og, "abc": abc_s})
                yuk[key]               = round(yuk[key]  + og, 2)
                yuk["jami_kg"]         = round(yuk["jami_kg"] + og, 2)
                yuk["_item_kg"][tovar] = round(already_kg + og, 2)
                qoldi        -= dona
                mavjud[tovar] = mavjud.get(tovar, 0) - dona
                if qoldi == 0:
                    break
                if _mayda_qatormi(qoldi, vazn_dona):
                    # 2026-07-18 (DUM-QOIDASI, "kulgili 1-2 dona" tuzatmasi):
                    # asosiy qism joylashdi, qolgan "dum" mayda -- uni BOSHQA
                    # konteynerga sochmaymiz (avval shu yerda 83+2 bo'lib
                    # ikkinchi konteynerda yolg'iz "2 dona" qator chiqardi).
                    break
            if qoldi == 0:
                break
            if qoldi < bor_dona and _mayda_qatormi(qoldi, vazn_dona):
                # Dum "qolgan" ro'yxatiga -- keyingi safar to'planganda yuklanadi
                qolgan.append({"tovar": tovar, "dona": qoldi,
                               "vazn_kg": round(qoldi * vazn_dona, 2)})
                break
            if len(yuklar) >= max_yuklar:
                qolgan.append({"tovar": tovar, "dona": qoldi,
                               "vazn_kg": round(qoldi * vazn_dona, 2)})
                break
            yuklar.append(_yangi_yuk(turi))

    return yuklar, qolgan, kerak_yoq, vazn_yoq


def konteyner_xulosa(yuklar: list[dict]) -> dict:
    res = {"12m": 0, "6m": 0, "jami_kg": 0.0}
    for y in yuklar:
        res[y["turi"]] += 1
        res["jami_kg"] += y["jami_kg"]
    res["jami_t"] = round(res["jami_kg"] / 1000, 2)
    return res


def chop_natija(yuklar: list[dict], qolgan: list[dict] | None = None):
    for i, yuk in enumerate(yuklar, 1):
        print("=" * 68)
        print("KONTEYNER #%d  [%s]  --  Jami: %.0f kg  (TP: %.0f  List: %.0f)"
              % (i, yuk["turi"], yuk["jami_kg"],
                 yuk["truba_profil_kg"], yuk["list_kg"]))
        print("-" * 68)
        for it in yuk["items"]:
            print("  [%s] %-52s %5d d  %7.0f kg"
                  % (it["abc"], it["tovar"][:52], it["dona"], it["vazn_kg"]))
    if qolgan:
        print("=" * 68)
        print("QOLGAN (keyingi oyga): %d tovar" % len(qolgan))
        for it in qolgan:
            print("  %-52s %5d d  %7.0f kg"
                  % (it["tovar"][:52], it["dona"], it["vazn_kg"]))
    xs = konteyner_xulosa(yuklar)
    print("=" * 68)
    print("JAMI: %d konteyner  (12m: %d  6m: %d)  --  %.1f tonna"
          % (len(yuklar), xs["12m"], xs["6m"], xs["jami_t"]))


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s -- %(message)s")
    abc = abc_map_yuklash()
    print("ABC: %d ta  (A=%d B=%d C=%d)" % (
        len(abc), sum(v=="A" for v in abc.values()),
        sum(v=="B" for v in abc.values()),
        sum(v=="C" for v in abc.values())))
    tests = [
        "Ф-51 ст 0,9 (5,8 м) (201 марка)",
        "Ф-51 ст 0,9 (6 м) (201 марка)",
        "Пр. 30х30 ст 2,0 (6 м) (201 марка)",
    ]
    for t in tests:
        v = tovar_vazni(t)
        print(f"  {t}: {v} kg")
