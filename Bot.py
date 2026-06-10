"""
NEJAVIYKA Telegram bot — v3 (Reply Keyboard)
=============================================
ASOSIY O'ZGARISH:
  Navigatsiya ↓ pastki klaviaturaga ko'chirildi (Reply Keyboard)
  — Tugmalar ekran pastida doim turadi
  — Chat tarixini TO'LDIRMAYDI (faqat fayllar va natijalar ko'rinadi)
  — Professional ko'rinish

  Til tanlash: Inline (bir martalik)
  Navigatsiya: Reply Keyboard (doim pastda)
  Kontent:     Fayllar va xabarlar — odatdagidek

Ishga tushirish:
  pip install "python-telegram-bot[job-queue]" pandas openpyxl
  BOT_TOKEN ni qo'ying → python Bot.py
"""

import json, logging
import pandas as pd
from datetime import datetime
from io import BytesIO
from pathlib import Path

from telegram import (
    Update,
    InlineKeyboardButton, InlineKeyboardMarkup,
    ReplyKeyboardMarkup, KeyboardButton,
)
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, ContextTypes, filters,
)

# ============================================================
# SOZLAMALAR
# ============================================================
BOT_TOKEN  = "8812102771:AAFXFLA2VxRDXJh4EaJqLfF3uyKgxvcNEOk"
BASE_DIR   = Path(__file__).parent
DATA_FILE  = BASE_DIR / "NEJAVIYKA_POWER_BI.xlsx"

logging.basicConfig(
    format="%(asctime)s — %(name)s — %(levelname)s — %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

VARAQLAR = ["Труба", "Профиль", "Лист", "Баласина", "Стойка", "Аксессуар"]
CH_KEY   = {"asosiy": "ch_asosiy", "sex": "ch_sex", "osh": "ch_osh"}
CAT_SHEET = {
    "ТРУБА": "Труба", "ПРОФИЛЬ": "Профиль",
    "ЛИСТ": "Лист",  "ЛИСТ РУЛОН": "Лист",
    "БАЛАСИНА": "Баласина", "СТОЙКА": "Стойка",
}
AKSESSUAR_KATS = {"ШАР","ОТВОД","СОККА","ЧАШКА","СОВУН","КУЗИКОРИН","БОШҚА"}

# Navigatsiya: qaysi ekrandan "Orqaga" qayerga boradi
BACK_MAP = {
    "order":         "main",
    "order_channel": "order",
    "load":          "main",
    "load_channel":  "load",
    "status":        "main",
    "settings":      "main",
    "search":        "main",
}

# ============================================================
# TARJIMA
# ============================================================
STR = {
"cyr": {
    "main_title":    "📦 *НЕРЖАВЕЙКА*\n\nАсосий менюдасиз:",
    "b_order":       "📥 Буюртма йиғиш",
    "b_load":        "🚛 Контейнер юклаш",
    "b_status":      "📈 Ҳолат",
    "b_search":      "🔍 Қидируш",
    "b_settings":    "⚙️ Созламалар",
    "order_title":   "📥 *Буюртма йиғиш*\n\nҚайси канал учун?",
    "ch_asosiy":     "🏢 Асосий (омборлар ва филиаллар)",
    "ch_sex":        "🏭 Цех (ишлаб чиқариш)",
    "ch_osh":        "🇰🇬 Ўш (Қирғизистон)",
    "order_ch_title":"Амални танланг:",
    "b_kamomat":     "📊 Камоматни кўриш",
    "b_excel":       "📄 Буюртма Excel олиш",
    "b_tasdiq":      "✅ Тасдиқланган буюртмани юклаш",
    "kamomat_title": "📊 *{ch}* — камомат:",
    "kamomat_line":  "🔴 Критик: *{n}* та товар\n✅ Буюртма берилди: {b} та\n⏳ Ҳали берилмади: {p} та",
    "kamomat_yoq":   "✅ Ҳозирча камомат йўқ.",
    "excel_caption": "📄 Камомат рўйхати ({ch}).\n\n☝️ Буюртма устунини тўлдиринг.\nТайёр бўлгач — *Тасдиқланган буюртмани юклаш* тугмасини босинг.",
    "tasdiq_prompt": "✅ *Тасдиқланган буюртма — {ch}*\n\nЎзгартирилган Excel файлни юборинг.\nБот реал буюртма сифатида сақлайди.",
    "tasdiq_ok":     "✅ Буюртма қабул қилинди: {n} та товар.\nКейин камоматда берилмаганлар кўринади.",
    "tasdiq_err":    "❌ *Нотўғри формат*\n\n{xato}\n\nКеракли формат:\n— 6 та варақ: Труба · Профиль · Лист · Баласина · Стойка · Аксессуар\n— Ҳар бирида: №, Товар, Буюртма, Изоҳ",
    "load_title":    "🚛 *Контейнер юклаш*\n\nҚайси канал учун юк тайёрланади?",
    "load_ch_title": "🚛 *{ch}*\n\nХитой тайёр рўйхатини юборинг (.xlsx).",
    "b_upload":      "📎 Хитой Excelини юклаш",
    "xitoy_ok":      "📎 Хитой файли қабул қилинди — {n} та товар аниқланди.\n_(Юклаш плани — кейинги босқич)_",
    "xitoy_err":     "❌ Файлни ўқишда хато: {xato}",
    "status_title":  "📈 *Ҳолат*",
    "b_kritik":      "🔴 Бугунги Критик",
    "b_yolda":       "🚢 Йўлдаги контейнерлар",
    "kritik_stub":   "🔴 *Бугунги Критик*\n\n_(ҳозирча скелет)_",
    "yolda_title":   "🚢 *Йўлдаги контейнерлар ({n} та):*\n\n",
    "yolda_yoq":     "✅ Ҳозирча йўлда контейнер йўқ.",
    "yolda_header":  "Контейнер     | Юкланган   | Келиш      | Қолди",
    "yolda_sep":     "--------------|------------|------------|----------",
    "kech_label":    "{k} кун кеч ⚠️",
    "kun_label":     "{k} кун 🚢",
    "search_title":  "🔍 *Қидируш*\n\nТовар номини ёзинг (мисол: _51 0,9_ ёки _лист 0.6_).\nОрқага тугмасини босиш — қидирувни бекор қилади.",
    "search_stub":   "🔍 *{q}* — натижалар:\n\n_(ҳозирча скелет — мантиқ кейин уланади)_",
    "settings_title":"⚙️ *Созламалар*",
    "b_lang":        "🌐 Тилни ўзгартириш",
    "back":          "⬅️ Орқага",
    "data_yoq":      "⚠️ Маълумот топилмади.\nАввал NEJAVIYKA\\_v3.py ни ишга туширинг.",
    "fayl_kutilmadi":"❓ Файл кутилмаган эди. Менюдан амални танланг.",
},
"lat": {
    "main_title":    "📦 *NERJAVEYKA*\n\nAsosiy menyudasiz:",
    "b_order":       "📥 Buyurtma yig'ish",
    "b_load":        "🚛 Konteyner yuklash",
    "b_status":      "📈 Holat",
    "b_search":      "🔍 Qidiruv",
    "b_settings":    "⚙️ Sozlamalar",
    "order_title":   "📥 *Buyurtma yig'ish*\n\nQaysi kanal uchun?",
    "ch_asosiy":     "🏢 Asosiy (omborlar va filiallar)",
    "ch_sex":        "🏭 Tsex (ishlab chiqarish)",
    "ch_osh":        "🇰🇬 O'sh (Qirg'iziston)",
    "order_ch_title":"Amalni tanlang:",
    "b_kamomat":     "📊 Kamomatni ko'rish",
    "b_excel":       "📄 Buyurtma Excel olish",
    "b_tasdiq":      "✅ Tasdiqlangan buyurtmani yuklash",
    "kamomat_title": "📊 *{ch}* — kamomat:",
    "kamomat_line":  "🔴 Kritik: *{n}* ta tovar\n✅ Buyurtma berildi: {b} ta\n⏳ Hali berilmadi: {p} ta",
    "kamomat_yoq":   "✅ Hozircha kamomat yo'q.",
    "excel_caption": "📄 Kamomat ro'yxati ({ch}).\n\n☝️ Buyurtma ustunini to'ldiring.\nTayyor bo'lgach — *Tasdiqlangan buyurtmani yuklash* tugmasini bosing.",
    "tasdiq_prompt": "✅ *Tasdiqlangan buyurtma — {ch}*\n\nO'zgartirilgan Excel faylni yuboring.\nBot real buyurtma sifatida saqlaydi.",
    "tasdiq_ok":     "✅ Buyurtma qabul qilindi: {n} ta tovar.\nKeyingi kamomatda berilmaganlar ko'rinadi.",
    "tasdiq_err":    "❌ *Noto'g'ri format*\n\n{xato}\n\nKerakli format:\n— 6 ta varaq: Труба · Профиль · Лист · Баласина · Стойка · Аксессуар\n— Har birida: №, Товар, Буюртма, Изоҳ",
    "load_title":    "🚛 *Konteyner yuklash*\n\nQaysi kanal uchun yuk tayyorlanadi?",
    "load_ch_title": "🚛 *{ch}*\n\nXitoy tayyor ro'yxatini yuboring (.xlsx).",
    "b_upload":      "📎 Xitoy Excelini yuklash",
    "xitoy_ok":      "📎 Xitoy fayli qabul qilindi — {n} ta tovar aniqlandi.\n_(Yuklash plani — keyingi bosqich)_",
    "xitoy_err":     "❌ Faylni o'qishda xato: {xato}",
    "status_title":  "📈 *Holat*",
    "b_kritik":      "🔴 Bugungi Kritik",
    "b_yolda":       "🚢 Yo'ldagi konteynerlar",
    "kritik_stub":   "🔴 *Bugungi Kritik*\n\n_(hozircha skelet)_",
    "yolda_title":   "🚢 *Yo'ldagi konteynerlar ({n} ta):*\n\n",
    "yolda_yoq":     "✅ Hozircha yo'lda konteyner yo'q.",
    "yolda_header":  "Konteyner     | Yuklangan  | Kelish     | Qoldi",
    "yolda_sep":     "--------------|------------|------------|----------",
    "kech_label":    "{k} kun kech ⚠️",
    "kun_label":     "{k} kun 🚢",
    "search_title":  "🔍 *Qidiruv*\n\nTovar nomini yozing (misol: _51 0,9_ yoki _list 0.6_).\nOrqaga tugmasi — qidiruvni bekor qiladi.",
    "search_stub":   "🔍 *{q}* — natijalar:\n\n_(hozircha skelet — mantiq keyin ulanadi)_",
    "settings_title":"⚙️ *Sozlamalar*",
    "b_lang":        "🌐 Tilni o'zgartirish",
    "back":          "⬅️ Orqaga",
    "data_yoq":      "⚠️ Ma'lumot topilmadi.\nAvval NEJAVIYKA\\_v3.py ni ishga tushiring.",
    "fayl_kutilmadi":"❓ Fayl kutilmagan edi. Menyudan amalni tanlang.",
},
}


# ============================================================
# YORDAMCHI
# ============================================================
def t(lang: str, key: str) -> str:
    return STR.get(lang, STR["cyr"]).get(key, key)


def rkb(*rows, one_time: bool = False) -> ReplyKeyboardMarkup:
    """Reply keyboard (pastki klaviatura) yaratadi."""
    return ReplyKeyboardMarkup(
        [[KeyboardButton(lbl) for lbl in row] for row in rows],
        resize_keyboard=True,
        one_time_keyboard=one_time,
    )


def ikb(*rows) -> InlineKeyboardMarkup:
    """Inline keyboard (faqat til tanlash uchun)."""
    return InlineKeyboardMarkup(
        [[InlineKeyboardButton(lbl, callback_data=cb) for lbl, cb in row]
         for row in rows]
    )


# ============================================================
# KLAVIATURALAR (ekran bo'yicha)
# ============================================================
def main_kb(lang: str) -> ReplyKeyboardMarkup:
    return rkb(
        [t(lang, "b_order"),   t(lang, "b_load")],
        [t(lang, "b_status"),  t(lang, "b_search")],
        [t(lang, "b_settings")],
    )

def order_kb(lang: str) -> ReplyKeyboardMarkup:
    return rkb(
        [t(lang, "ch_asosiy")],
        [t(lang, "ch_sex")],
        [t(lang, "ch_osh")],
        [t(lang, "back")],
    )

def order_channel_kb(lang: str) -> ReplyKeyboardMarkup:
    return rkb(
        [t(lang, "b_kamomat")],
        [t(lang, "b_excel")],
        [t(lang, "b_tasdiq")],
        [t(lang, "back")],
    )

def load_kb(lang: str) -> ReplyKeyboardMarkup:
    return rkb(
        [t(lang, "ch_asosiy")],
        [t(lang, "ch_sex")],
        [t(lang, "ch_osh")],
        [t(lang, "back")],
    )

def load_channel_kb(lang: str) -> ReplyKeyboardMarkup:
    return rkb(
        [t(lang, "b_upload")],
        [t(lang, "back")],
    )

def status_kb(lang: str) -> ReplyKeyboardMarkup:
    return rkb(
        [t(lang, "b_kritik")],
        [t(lang, "b_yolda")],
        [t(lang, "back")],
    )

def settings_kb(lang: str) -> ReplyKeyboardMarkup:
    return rkb(
        [t(lang, "b_lang")],
        [t(lang, "back")],
    )

def search_kb(lang: str) -> ReplyKeyboardMarkup:
    return rkb([t(lang, "back")])


def til_ikb() -> InlineKeyboardMarkup:
    """Til tanlash uchun inline keyboard."""
    return ikb(
        [("Кирилл (Ўзбекча)", "lang:cyr")],
        [("Lotin (O'zbekcha)", "lang:lat")],
    )


# ============================================================
# EKRAN QURUVCHI VA NAVIGATSIYA
# ============================================================
def build_screen(screen: str, lang: str, context) -> tuple:
    """(matn, klaviatura) qaytaradi."""
    kanal = context.user_data.get("kanal", "asosiy")

    if screen == "main":
        return t(lang, "main_title"), main_kb(lang)
    if screen == "order":
        return t(lang, "order_title"), order_kb(lang)
    if screen == "order_channel":
        ch = t(lang, CH_KEY[kanal])
        return f"{ch}\n\n{t(lang, 'order_ch_title')}", order_channel_kb(lang)
    if screen == "load":
        return t(lang, "load_title"), load_kb(lang)
    if screen == "load_channel":
        ch = t(lang, CH_KEY[kanal])
        return t(lang, "load_ch_title").format(ch=ch), load_channel_kb(lang)
    if screen == "status":
        return t(lang, "status_title"), status_kb(lang)
    if screen == "settings":
        return t(lang, "settings_title"), settings_kb(lang)
    if screen == "search":
        return t(lang, "search_title"), search_kb(lang)
    return t(lang, "main_title"), main_kb(lang)


async def go_screen(msg, context, screen: str, kanal: str | None = None):
    """Yangi ekranga o'tadi va reply keyboard ni yangilaydi."""
    lang = context.user_data.get("lang", "cyr")
    if kanal:
        context.user_data["kanal"] = kanal
    context.user_data["screen"] = screen
    text, kb = build_screen(screen, lang, context)
    await msg.reply_text(text, reply_markup=kb, parse_mode="Markdown")


def get_action(lang: str, screen: str, text: str):
    """
    Ekran + bosilgan matn → amal kalit so'zi.
    Tuple: (keyingi_ekran, kanal) — kanal ekraniga o'tish.
    String: amal nomi.
    None:  noma'lum.
    """
    MAP = {
        "main": {
            t(lang, "b_order"):    "order",
            t(lang, "b_load"):     "load",
            t(lang, "b_status"):   "status",
            t(lang, "b_search"):   "search",
            t(lang, "b_settings"): "settings",
        },
        "order": {
            t(lang, "ch_asosiy"): ("order_channel", "asosiy"),
            t(lang, "ch_sex"):    ("order_channel", "sex"),
            t(lang, "ch_osh"):    ("order_channel", "osh"),
        },
        "order_channel": {
            t(lang, "b_kamomat"): "kamomat",
            t(lang, "b_excel"):   "excel",
            t(lang, "b_tasdiq"):  "tasdiq",
        },
        "load": {
            t(lang, "ch_asosiy"): ("load_channel", "asosiy"),
            t(lang, "ch_sex"):    ("load_channel", "sex"),
            t(lang, "ch_osh"):    ("load_channel", "osh"),
        },
        "load_channel": {
            t(lang, "b_upload"): "upload",
        },
        "status": {
            t(lang, "b_kritik"): "kritik",
            t(lang, "b_yolda"):  "yolda",
        },
        "settings": {
            t(lang, "b_lang"): "lang_pick",
        },
    }
    return MAP.get(screen, {}).get(text)


# ============================================================
# MA'LUMOT FUNKSIYALARI (o'zgarishsiz)
# ============================================================
def xlsx_mavjud() -> bool:
    return DATA_FILE.exists()


def kamomat_olish(kanal: str) -> pd.DataFrame:
    try:
        df = pd.read_excel(DATA_FILE, sheet_name="Инвентар")
        for col in ["Қолдиқ","Мин_Захира","Фарқ","Кун_Етади"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        kritik = df[df["Холат"] == "🔴 КРИТИК"].copy()
        if "Тур" in kritik.columns:
            return (kritik[kritik["Тур"] == "ЦЕХ🏭"]
                    if kanal == "sex"
                    else kritik[kritik["Тур"] != "ЦЕХ🏭"]).reset_index(drop=True)
        return kritik.reset_index(drop=True)
    except Exception as e:
        logger.error(f"kamomat_olish: {e}")
        return pd.DataFrame()


def buyurtma_yuklash(kanal: str) -> dict | None:
    p = BASE_DIR / f"buyurtma_{kanal}.json"
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except:
        return None


def buyurtma_saqlash(kanal: str, items: list):
    p = BASE_DIR / f"buyurtma_{kanal}.json"
    p.write_text(json.dumps(
        {"sana": datetime.now().strftime("%d.%m.%Y %H:%M"),
         "kanal": kanal, "buyurtmalar": items},
        ensure_ascii=False, indent=2), encoding="utf-8")


def kamomat_stats(kanal: str) -> dict:
    df  = kamomat_olish(kanal)
    if df.empty:
        return {"n": 0, "b": 0, "p": 0}
    buy     = buyurtma_yuklash(kanal)
    ordered = {i["tovar"] for i in buy.get("buyurtmalar",[])} if buy else set()
    tovs    = df["Товар"].tolist() if "Товар" in df.columns else []
    b       = sum(1 for x in tovs if x in ordered)
    return {"n": len(df), "b": b, "p": len(df) - b}


def kamomat_excel_yarat(kanal: str, lang: str) -> BytesIO | None:
    df  = kamomat_olish(kanal)
    if df.empty:
        return None
    buy     = buyurtma_yuklash(kanal)
    ordered = {i["tovar"] for i in buy.get("buyurtmalar",[])} if buy else set()
    cols    = [c for c in ["Товар","Категория","Қолдиқ","Мин_Захира","Фарқ","Кун_Етади"]
               if c in df.columns]
    out     = df[cols].copy() if cols else df.copy()
    if "Товар" in out.columns:
        berildi = "Берилди ✅" if lang=="cyr" else "Berildi ✅"
        kutilmq = "Кутилмоқда ⏳" if lang=="cyr" else "Kutilmoqda ⏳"
        out["Буюртма_Ҳолати"] = out["Товар"].apply(
            lambda x: berildi if x in ordered else kutilmq)
    if "Фарқ" in out.columns:
        out["Таклиф_Миқдор"] = out["Фарқ"].apply(
            lambda x: max(0,-int(x)) if pd.notna(x) and x<0 else 0)
    out.insert(0,"№",range(1,len(out)+1))
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as w:
        out.to_excel(w, sheet_name="Камомат" if lang=="cyr" else "Kamomat", index=False)
    bio.seek(0)
    return bio


def draft_excel_yarat(kanal: str) -> BytesIO:
    import openpyxl
    df       = kamomat_olish(kanal)
    kat_items = {v: [] for v in VARAQLAR}
    if not df.empty and "Категория" in df.columns and "Товар" in df.columns:
        for _, row in df.iterrows():
            kat    = str(row.get("Категория","БОШҚА"))
            tovar  = str(row.get("Товар",""))
            farq   = row.get("Фарқ", 0)
            miqdor = max(0,-int(farq)) if pd.notna(farq) and farq<0 else 0
            sheet  = CAT_SHEET.get(kat) or ("Аксессуар" if kat in AKSESSUAR_KATS else None)
            if sheet:
                kat_items[sheet].append((tovar, miqdor))
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for varaq in VARAQLAR:
        ws = wb.create_sheet(varaq)
        ws.append(["№","Товар","Буюртма","Изоҳ"])
        for n,(tovar,miqdor) in enumerate(kat_items.get(varaq,[]),1):
            ws.append([n, tovar, miqdor if miqdor>0 else "", ""])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def buyurtma_tekshir(fayl_bytes: bytes):
    import openpyxl
    try:
        wb  = openpyxl.load_workbook(BytesIO(fayl_bytes))
        yoq = set(VARAQLAR) - set(wb.sheetnames)
        if yoq:
            return False, f"Varaqlar topilmadi: {', '.join(sorted(yoq))}", None
        items = []
        for varaq in VARAQLAR:
            ws      = wb[varaq]
            rows    = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(c).strip() if c else "" for c in rows[0]]
            if "Товар" not in headers:
                return False, f"'{varaq}' varaqida 'Товар' topilmadi", None
            if "Буюртма" not in headers:
                return False, f"'{varaq}' varaqida 'Буюртма' topilmadi", None
            ti, bi  = headers.index("Товар"), headers.index("Буюртма")
            for row in rows[1:]:
                tovar = row[ti] if ti<len(row) else None
                bval  = row[bi] if bi<len(row) else None
                if not tovar or bval in (None,""):
                    continue
                try:
                    m = float(bval)
                    if m > 0:
                        items.append({"tovar":str(tovar).strip(),"miqdor":m,"varaq":varaq})
                except (ValueError,TypeError):
                    return False, f"'{varaq}' varaqida raqam bo'lmagan buyurtma qiymati", None
        return True, None, items
    except Exception as e:
        return False, f"Faylni ochishda xato: {type(e).__name__}", None


def konteynerlar_olish() -> pd.DataFrame:
    try:
        df    = pd.read_excel(DATA_FILE, sheet_name="Контейнерлар")
        yolda = df[df["Холат"] != "КЕЛДИ ✅"].copy()
        if "Контейнер" in yolda.columns:
            yolda = yolda.drop_duplicates(subset="Контейнер")
        if "Кун_Қолди" in yolda.columns:
            yolda["Кун_Қолди"] = pd.to_numeric(yolda["Кун_Қолди"],errors="coerce").fillna(999)
            yolda = yolda.sort_values("Кун_Қолди")
        return yolda.reset_index(drop=True)
    except Exception as e:
        logger.error(f"konteynerlar_olish: {e}")
        return pd.DataFrame()


def konteynerlar_text(df: pd.DataFrame, lang: str) -> str:
    if df.empty:
        return t(lang, "yolda_yoq")
    lines = [t(lang,"yolda_header"), t(lang,"yolda_sep")]
    for _, row in df.iterrows():
        k      = str(row.get("Контейнер","?"))[:13].ljust(13)
        yukl   = str(row.get("Юкланган_Сана","?"))[:10].ljust(10)
        kelish = str(row.get("Келиш_Санаси","?"))[:10].ljust(10)
        holat  = str(row.get("Холат",""))
        kq     = row.get("Кун_Қолди", None)
        kk     = row.get("Кечикиш_Кун", 0)
        if "КЕЧИКДИ" in holat:
            qoldi = t(lang,"kech_label").format(k=int(kk) if pd.notna(kk) else "?")
        elif pd.notna(kq):
            qoldi = t(lang,"kun_label").format(k=int(float(kq)))
        else:
            qoldi = "?"
        lines.append(f"| {k} | {yukl} | {kelish} | {qoldi}")
    return "\n".join(lines)


# ============================================================
# KONTENT HANDLERLARI (reply keyboard bilan soddavor)
# ============================================================
async def kamomat_ko_rish(msg, context, kanal: str, lang: str):
    ch = t(lang, CH_KEY[kanal])
    if not xlsx_mavjud():
        await msg.reply_text(t(lang,"data_yoq"), parse_mode="Markdown")
        return
    stats = kamomat_stats(kanal)
    if stats["n"] == 0:
        await msg.reply_text(t(lang,"kamomat_yoq"))
        return
    await msg.reply_text(
        t(lang,"kamomat_title").format(ch=ch) + "\n\n" +
        t(lang,"kamomat_line").format(**stats),
        parse_mode="Markdown",
    )
    bio = kamomat_excel_yarat(kanal, lang)
    if bio:
        await msg.reply_document(document=bio, filename=f"Kamomat_{kanal}.xlsx")


async def draft_buyurtma_yubor(msg, context, kanal: str, lang: str):
    ch  = t(lang, CH_KEY[kanal])
    bio = draft_excel_yarat(kanal)
    await msg.reply_document(
        document=bio,
        filename=f"Buyurtma_taklif_{kanal}.xlsx",
        caption=t(lang,"excel_caption").format(ch=ch),
        parse_mode="Markdown",
    )


async def yolda_ko_rish(msg, context, lang: str):
    if not xlsx_mavjud():
        await msg.reply_text(t(lang,"data_yoq"), parse_mode="Markdown")
        return
    df    = konteynerlar_olish()
    title = t(lang,"yolda_title").format(n=len(df))
    table = konteynerlar_text(df, lang)
    await msg.reply_text(title + table, parse_mode="Markdown")


# ============================================================
# TELEGRAM HANDLERLARI
# ============================================================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    lang = context.user_data.get("lang")
    if lang:
        await go_screen(update.message, context, "main")
    else:
        await update.message.reply_text(
            "🇺🇿 Тилни танланг / Tilni tanlang:",
            reply_markup=til_ikb(),
        )


async def callback_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Faqat til tanlash uchun (inline keyboard)."""
    query = update.callback_query
    await query.answer()
    if query.data.startswith("lang:"):
        lang = query.data.split(":")[1]
        context.user_data["lang"]   = lang
        context.user_data["screen"] = "main"
        await query.edit_message_text(
            "✅ Тил танланди!" if lang == "cyr" else "✅ Til tanlandi!"
        )
        await go_screen(query.message, context, "main")
    elif query.data == "lang_pick":
        await query.message.reply_text("🇺🇿 Тилни танланг:", reply_markup=til_ikb())


async def text_keldi(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Barcha matn xabarlarini boshqaradi — navigatsiya va qidiruv."""
    lang   = context.user_data.get("lang", "cyr")
    screen = context.user_data.get("screen", "main")
    text   = update.message.text.strip()
    msg    = update.message

    # Fayl kutilayotgan bo'lsa, tugma bosish = bekor qilish
    kut = context.user_data.get("kutilmoqda")
    if isinstance(kut, tuple):
        context.user_data.pop("kutilmoqda", None)

    # Orqaga (universal — har qanday ekranda ishlaydi)
    if text == t(lang, "back"):
        parent = BACK_MAP.get(screen, "main")
        await go_screen(msg, context, parent)
        return

    # Amal topish
    action = get_action(lang, screen, text)

    if action is None:
        # Qidiruv rejimi — har qanday matn = qidiruv so'rovi
        if screen == "search":
            await msg.reply_text(
                t(lang,"search_stub").format(q=text),
                parse_mode="Markdown",
            )
            # TODO: real qidiruv mantiqini shu yerga ulang
        return

    # Kanal ekraniga o'tish
    if isinstance(action, tuple):
        next_screen, kanal = action
        await go_screen(msg, context, next_screen, kanal=kanal)
        return

    # Oddiy navigatsiya
    if action in ("order","load","status","settings","search","main"):
        await go_screen(msg, context, action)
        return

    # Buyurtma amallari
    kanal = context.user_data.get("kanal","asosiy")

    if action == "kamomat":
        await kamomat_ko_rish(msg, context, kanal, lang)

    elif action == "excel":
        await draft_buyurtma_yubor(msg, context, kanal, lang)

    elif action == "tasdiq":
        context.user_data["kutilmoqda"] = ("buyurtma_tasdiq", kanal)
        ch = t(lang, CH_KEY[kanal])
        await msg.reply_text(
            t(lang,"tasdiq_prompt").format(ch=ch),
            parse_mode="Markdown",
        )

    elif action == "upload":
        context.user_data["kutilmoqda"] = ("xitoy_fayl", kanal)
        ch = t(lang, CH_KEY[kanal])
        await msg.reply_text(
            t(lang,"load_ch_title").format(ch=ch),
            parse_mode="Markdown",
        )

    elif action == "kritik":
        await msg.reply_text(t(lang,"kritik_stub"), parse_mode="Markdown")

    elif action == "yolda":
        await yolda_ko_rish(msg, context, lang)

    elif action == "lang_pick":
        await msg.reply_text("🇺🇿 Тилни танланг:", reply_markup=til_ikb())


async def fayl_keldi(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Yuklangan .xlsx fayllarni qayta ishlaydi."""
    lang = context.user_data.get("lang","cyr")
    kut  = context.user_data.get("kutilmoqda")
    doc  = update.message.document

    if not doc.file_name.lower().endswith(".xlsx"):
        await update.message.reply_text("⚠️ Faqat .xlsx fayl qabul qilinadi.")
        return

    if isinstance(kut, tuple) and kut[0] == "buyurtma_tasdiq":
        kanal    = kut[1]
        fayl_obj = await doc.get_file()
        bio      = BytesIO()
        await fayl_obj.download_to_memory(bio)
        ok, xato, items = buyurtma_tekshir(bio.getvalue())
        if not ok:
            await update.message.reply_text(
                t(lang,"tasdiq_err").format(xato=xato),
                parse_mode="Markdown",
            )
            return
        buyurtma_saqlash(kanal, items or [])
        context.user_data.pop("kutilmoqda", None)
        await update.message.reply_text(
            t(lang,"tasdiq_ok").format(n=len(items or [])),
            parse_mode="Markdown",
        )
        # TODO: tasdiqlangan buyurtmani zanjirga ulash

    elif isinstance(kut, tuple) and kut[0] == "xitoy_fayl":
        fayl_obj = await doc.get_file()
        bio      = BytesIO()
        await fayl_obj.download_to_memory(bio)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(bio)
            n  = sum(1 for ws in wb.worksheets
                     for row in ws.iter_rows(min_row=2, values_only=True)
                     if any(c is not None for c in row))
            context.user_data.pop("kutilmoqda", None)
            await update.message.reply_text(
                t(lang,"xitoy_ok").format(n=n),
                parse_mode="Markdown",
            )
            # TODO: parse_china_konteyner.py bilan to'liq tahlil
        except Exception as e:
            await update.message.reply_text(t(lang,"xitoy_err").format(xato=str(e)[:80]))
    else:
        await update.message.reply_text(t(lang,"fayl_kutilmadi"))


async def kunlik_xabar(context: ContextTypes.DEFAULT_TYPE):
    """Har kuni ertalab 09:00 da avtomatik kamomat xabari."""
    # TODO: kritik tovarlarni aniqlash + barcha foydalanuvchilarga yuborish
    pass


# ============================================================
# ISHGA TUSHIRISH
# ============================================================
def main():
    app = Application.builder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CallbackQueryHandler(callback_handler))
    app.add_handler(MessageHandler(filters.Document.ALL, fayl_keldi))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, text_keldi))

    if app.job_queue:
        from datetime import time
        app.job_queue.run_daily(kunlik_xabar, time=time(hour=9, minute=0))

    logger.info("Бот ишга тушди (v3 — Reply Keyboard)...")
    app.run_polling()


if __name__ == "__main__":
    main()