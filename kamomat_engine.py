"""
kamomat_engine.py — Kamomat tahlili motori
==========================================
Mantiq:
  - KRITIK + PAST tovarlarni aniqlash
  - Kunlik sotuv = min_zaxira / KUNLIK_SOTUV_BOLISH (common.py dan olinadi, 2026-07-09)
  - Zanjir simulyatsiyasi: konteynerlar sana bilan hisobga olinadi
  - Tartibli Excel: kategoriya -> o'lcham -> qalinlik -> uzunlik -> marka
  - Rangli Excel: har kategoriya o'z rang oilasida, juft/toq qatorlar

Bot.py import qiladi:
  from kamomat_engine import kamomat_stats_v2, kamomat_excel_v2
"""

import re, math, logging
import pandas as pd
from io import BytesIO
from pathlib import Path
from datetime import datetime

logger = logging.getLogger(__name__)

# ============================================================
# KONSTANTALAR — kelish kunini common.py dan olamiz
# ============================================================
from common import KELISH_KUNI, KUNLIK_SOTUV_BOLISH, BUYURTMA_SIKL_KUN
KUNLAR = KUNLIK_SOTUV_BOLISH  # 2026-07-09: endi 30 (eski kod o'zgarmasin, nomi qoldi)

CAT_ORDER = {
    "ТРУБА": 1, "ПРОФИЛЬ": 2, "ЛИСТ": 3, "ЛИСТ РУЛОН": 4,
    "БАЛАСИНА": 5, "СТОЙКА": 6, "СОККА": 7,
    "ШАР": 8, "ОТВОД": 9, "ЧАШКА": 10,
    "СОВУН": 11, "КУЗИКОРИН": 12, "БОШҚА": 13,
}

# Excel ranglari: h=header, a=toq qator, b=juft qator
CAT_COLORS = {
    "ТРУБА":      {"h": "9DC3E6", "a": "DEEBF7", "b": "BDD7EE"},
    "ПРОФИЛЬ":    {"h": "A9D18E", "a": "E2F0D9", "b": "C6E0B4"},
    "ЛИСТ":       {"h": "FFD966", "a": "FFF2CC", "b": "FFE699"},
    "ЛИСТ РУЛОН": {"h": "FFD966", "a": "FFF2CC", "b": "FFE699"},
    "БАЛАСИНА":   {"h": "C9B1D9", "a": "F0E5F7", "b": "E2CCEF"},
    "СТОЙКА":     {"h": "81C9C9", "a": "E0F2F2", "b": "C2E6E6"},
    "СОККА":      {"h": "F4B183", "a": "FCE4D6", "b": "F8CBAD"},
}
CAT_COLORS_DEF = {"h": "BFBFBF", "a": "F2F2F2", "b": "E8E8E8"}


# ============================================================
# SARALASH KALITI
# ============================================================
def tovar_sort_key(nom: str, kat: str) -> tuple:
    """
    Tovar nomidan saralash kalitini chiqaradi.
    Trubalarda: diametr -> qalinlik -> uzunlik -> marka
    Profillarda: en×boy -> qalinlik -> uzunlik -> marka
    Listlarda:  marka -> format -> qalinlik
    """
    cat_n = CAT_ORDER.get(kat, 99)
    n = str(nom)

    def _f(pat, default=99.0):
        m = re.search(pat, n)
        return float(m.group(1).replace(',', '.')) if m else default

    def _i(pat, default=999):
        m = re.search(pat, n)
        return int(m.group(1)) if m else default

    def _marka_n(text):
        m = re.search(r'\((\d+)\s*марка\)', text)
        return {"201": 1, "304": 2, "430": 3, "316": 4}.get(
            m.group(1) if m else "", 9)

    if kat == "ТРУБА":
        return (cat_n,
                _i(r'Ф-(\d+)'),
                _f(r'ст\s+(\d+[,.]?\d*)'),
                _f(r'\((\d+[,.]?\d*)\s*м\)'),
                _marka_n(n))

    elif kat == "ПРОФИЛЬ":
        s = re.search(r'(\d+)х(\d+)', n)
        w = int(s.group(1)) if s else 999
        h = int(s.group(2)) if s else 999
        return (cat_n, w, h,
                _f(r'ст\s+(\d+[,.]?\d*)'),
                _f(r'\((\d+[,.]?\d*)\s*м\)'),
                _marka_n(n))

    elif kat in ("ЛИСТ", "ЛИСТ РУЛОН"):
        fmt_m = re.search(r'\((\d+)х', n)
        return (cat_n,
                _marka_n(n),
                int(fmt_m.group(1)) if fmt_m else 9999,
                _f(r'Лист-\s*(\d+[,.]?\d*)'))

    return (cat_n, n)


# ============================================================
# ZANJIR SIMULYATSIYASI
# ============================================================
def zanjir_sim(qoldiq: float, min_z: float,
               konteynerlar: list, yaxlitla: bool = True,
               kunlik_override: float | None = None,
               horizon_override: float | None = None) -> dict:
    """
    Kun-ma-kun zanjir simulyatsiyasi (2026-07-14 qayta yozildi).

    kunlik_override (2026-07-21, Huzayfa bilan kelishildi): berilsa va
    > 0 bo'lsa, "kunlik" min_z/KUNLAR o'rniga shu qiymatdan olinadi.
    FOYDALANISH: Tsex kanalini Asosiy/O'sh uchun tuzatilgan global
    KUNLIK_SOTUV_BOLISH (45) o'zgarishidan HIMOYA qilish uchun —
    Generate_Asosiy_order.calculate() kanal=="sex" bo'lsa, min_z/30
    (eski, o'zgarmagan divisor) ni shu parametr orqali majburlaydi.
    Boshqa hech narsa o'zgarmaydi (nishon, horizon, lost-sales clamp —
    barchasi bir xil, faqat "kunlik" manbasi almashadi). None yoki
    <=0 bo'lsa — global KUNLIK_SOTUV_BOLISH asosidagi eski xatti-harakat
    saqlanadi.

    horizon_override (2026-07-23, Huzayfa bilan kelishildi -- Ф-51 real
    misolida "bugun buyurtma bersak, 70 kundan keyin kelib boshlaydi"
    tekshiruvi asosida): berilsa va >0 bo'lsa, gorizont KELISH_KUNI(55)
    o'rniga shu qiymatdan olinadi. SABAB: KELISH_KUNI (common.py)
    BOSHQA joylarda (main.py'ning "allaqachon yuklangan konteyner qachon
    yetib keladi" hisobi, tavsiya.py) HAM ishlatiladi -- u yerda 55 kun
    TO'G'RI (bu FAQAT dengiz yo'li, konteyner ALLAQACHON yuklangan bo'lgani
    uchun). Lekin YANGI buyurtma uchun gorizont bundan KATTAROQ bo'lishi
    kerak -- chunki yangi buyurtma hali tayyorlanmagan/yuklanmagan (qo'shimcha
    ~15 kun: tayyorlash+yuklash). Shu ikki tushunchani (allaqachon yo'ldagi
    konteyner transit vaqti VS yangi buyurtma umumiy kutish vaqti) ADASH-
    TIRMASLIK uchun umumiy KELISH_KUNI konstantasi O'ZGARTIRILMAYDI --
    faqat BUYURTMA HISOB-KITOBI (Generate_Asosiy_order.calculate, kanal
    Asosiy/O'sh) shu parametr orqali kengaytirilgan gorizontni (70) ishlatadi.
    None yoki <=0 bo'lsa -- eski KELISH_KUNI(55) saqlanadi.

    Qoldiq kuniga (min_z / KUNLAR=30) kamayadi, har konteyner o'z kunida
    miqdorini qo'shadi. Gorizont — KELISH_KUNI (55) kun: bugun berilgan
    YANGI buyurtma yetib kelguncha bo'lgan davr.

    ESKI XATO (2026-07-14 gacha): izohda "55 kun tekshiruv" deyilgan
    bo'lsa-da, kod oxirgi konteynerdan keyin faqat KUNLAR (30) kun
    tekshirardi — oxirgi konteyner erta kelsa (masalan 10-kun), 40–55
    kunlardagi tanqislik KO'RINMAY qolib, buyurtma kam chiqardi.

    BUYURTMA MANTIG'I — "order-up-to" (2026-07-14, Huzayfa bilan kelishildi):
      * Trigger: gorizont ichida qoldiq min_z dan pastga tushsaGINA buyurtma
        taklif qilinadi. Ilgari "min'dan 100 ta pastga tushibdi — 100 ta
        buyur" kabi MAYDA (50–200 talik) takliflar har kuni chiqib turardi.
      * Hajm: buyurtma kelganda (55-kun) zaxira
            nishon = min_z + kunlik * BUYURTMA_SIKL_KUN (30)
        darajasiga chiqadigan qilib hisoblanadi — bir yo'la ~1 oylik hajm,
        keyin bu tovar ~1 oy Excel'da umuman chiqmaydi.
      * 55 kundan KEYIN keladigan konteynerlar simulyatsiyaga kirmaydi
        (yangi buyurtma bilan birga/keyin keladi), lekin buyurtma hajmini
        kamaytirishda hisobga olinadi (ikki marta buyurmaslik uchun).

    konteynerlar: [(kun_qoldi: int, miqdor: float), ...]

    Qaytaradi:
      uzilish_kun — None yoki bugundan necha kun (gorizont ichida)
      min_nuqta   — gorizont ichidagi eng past qoldiq
      taklif      — order-up-to buyurtma (50 ga yaxlit)
      taklif_A/B  — eski nomlar mosligi uchun (A = eski min-farq usuli,
                    B = taklif bilan bir xil)
      xavf        — 'KRITIK' | 'PAST' | 'NORMA' | 'MEYOR_YOQ'
    """
    EMPTY = dict(uzilish_kun=None, min_nuqta=qoldiq,
                 taklif_A=0, taklif_B=0, taklif=0, xavf="MEYOR_YOQ")
    if min_z <= 0:
        return EMPTY

    kunlik  = float(kunlik_override) if kunlik_override and kunlik_override > 0 \
              else min_z / float(KUNLAR)
    horizon = float(horizon_override) if horizon_override and horizon_override > 0 \
              else float(KELISH_KUNI)
    kont    = sorted(konteynerlar, key=lambda x: x[0])

    joriy       = float(qoldiq)
    joriy_kun   = 0.0
    uzilish_kun = None

    # Bugun (kun<=0) kelayotgan konteynerlar min_nuqta O'RNATILISHIDAN
    # OLDIN qo'shiladi — aks holda "konteyner tushayotgan kuni" bir lahzalik
    # past qoldiq noto'g'ri trigger otilishiga sabab bo'lardi.
    bugun_kelganlar = [(k, m) for k, m in kont if k <= 0]
    for _, m in bugun_kelganlar:
        joriy += m
    kont = [(k, m) for k, m in kont if k > 0]

    min_nuqta = joriy

    def _uzilish_hisobla(oldin: float, kun_boshi: float) -> int:
        # min_z chizig'i kesib o'tilgan kunni topish
        return int(kun_boshi + max(0.0, (oldin - min_z) / kunlik))

    for kun_q, miqdor in kont:
        if kun_q > horizon:
            break  # 55+ kunda keladi — pastda alohida hisobga olinadi
        if kun_q <= joriy_kun:
            joriy += miqdor
            continue
        gap         = kun_q - joriy_kun
        joriy_oldin = joriy
        joriy      -= kunlik * gap
        min_nuqta   = min(min_nuqta, joriy)
        if joriy < min_z and uzilish_kun is None:
            uzilish_kun = _uzilish_hisobla(joriy_oldin, joriy_kun)
        if joriy < 0:
            # 2026-07-18 (Huzayfa bilan kelishildi -- "lost sales" modeli):
            # qoldiq 0 ga tushgach sotuv TO'XTAYDI -- minus to'planmaydi.
            # Ilgari minus yig'ilib (mas. -9000), o'sha "yo'qotilgan sotuv"
            # ham buyurtmaga qo'shilib, 2 baravar shishgan zakaz chiqardi
            # (real hodisa: Пр. 20х20 ст 0,7 -- 20 400 o'rniga ~11 200
            # to'g'ri). Mijoz tovar yo'qligida boshqa joydan oladi, u
            # sotuv "qarz" bo'lib kutmaydi. min_nuqta/uzilish_kun clamp
            # OLDIDAN yozildi -- KRITIK aniqlash o'zgarmagan.
            joriy = 0.0
        joriy    += miqdor
        joriy_kun = kun_q

    # Dumini GORIZONTGACHA yetkazish — 2026-07-14 dagi asosiy tuzatish
    # (ilgari bu yerda "oxirgi konteyner + 30 kun" edi).
    if joriy_kun < horizon:
        joriy_oldin = joriy
        joriy      -= kunlik * (horizon - joriy_kun)
        min_nuqta   = min(min_nuqta, joriy)
        if joriy < min_z and uzilish_kun is None:
            uzilish_kun = _uzilish_hisobla(joriy_oldin, joriy_kun)
        if joriy < 0:
            joriy = 0.0   # 2026-07-18: lost-sales clamp (yuqoridagi izohga qarang)

    qoldiq_gorizont = joriy                                   # 55-kun prognozi
    kech_jami  = sum(m for k, m in kont if k > horizon)       # 55+ kunda keladiganlar
    yolda_jami = sum(m for _, m in kont)

    def _50(x: float) -> int:
        return int(math.ceil(x / 50)) * 50 if x > 0 else 0

    # Order-up-to taklif
    taklif = 0.0
    if min_nuqta < min_z:
        nishon = min_z + kunlik * BUYURTMA_SIKL_KUN
        taklif = max(0.0, nishon - (qoldiq_gorizont + kech_jami))
    # 2026-07-18 (Huzayfa): yaxlitla=False -- qalin Лист (>=1,5) uchun
    # buyurtma 50 ga yaxlitlanmaydi, aniq son ochiq qoldiriladi.
    taklif_50 = _50(taklif) if yaxlitla else (int(math.ceil(taklif)) if taklif > 0 else 0)

    # Xavf darajasi
    if uzilish_kun is not None:
        xavf = "KRITIK"   # gorizont ichida uzilish — yangi buyurtma ham ulgurmaydi
    elif taklif_50 > 0:
        xavf = "PAST"     # hali uzilmaydi, lekin buyurtma vaqti keldi
    else:
        xavf = "NORMA"

    return dict(
        uzilish_kun=uzilish_kun,
        min_nuqta=int(round(min_nuqta)),
        taklif_A=_50(max(0.0, min_z - (qoldiq + yolda_jami))),
        taklif_B=taklif_50,
        taklif=taklif_50,
        xavf=xavf,
    )


# ============================================================
# KAMOMAT STATISTIKASI
# ============================================================
def kamomat_stats_v2(data_file: Path, kanal: str,
                     buyurtma_yuklash_fn) -> dict:
    """
    KRITIK + PAST tovarlar sonini va buyurtma holatini qaytaradi.
    Qaytaradi: {n, kritik, past, b (berildi), p (pending)}
    """
    try:
        df = pd.read_excel(data_file, sheet_name="Инвентар")
        if "Тур" in df.columns:
            df = (df[df["Тур"] == "ЦЕХ🏭"] if kanal == "sex"
                  else df[df["Тур"] != "ЦЕХ🏭"])

        kritik = int((df["Холат"] == "🔴 КРИТИК").sum())
        past   = int((df["Холат"] == "🟡 ПАСТ").sum())
        jami   = kritik + past

        if jami == 0:
            return {"n": 0, "kritik": 0, "past": 0, "b": 0, "p": 0}

        buy     = buyurtma_yuklash_fn(kanal)
        ordered = {i["tovar"] for i in buy.get("buyurtmalar", [])} if buy else set()
        kamomat = df[df["Холат"].isin(["🔴 КРИТИК", "🟡 ПАСТ"])]
        b       = sum(1 for t in kamomat.get("Товар", pd.Series()).tolist()
                      if t in ordered)
        return {"n": jami, "kritik": kritik, "past": past,
                "b": b, "p": jami - b}
    except Exception as e:
        logger.error(f"kamomat_stats_v2: {e}")
        return {"n": 0, "kritik": 0, "past": 0, "b": 0, "p": 0}


# 2026-07-27 (Huzayfa: "faqat truba profil listda qo'llaymiz"): узилиш
# xavfi ro'yxati faqat shu kategoriyalar bilan cheklanadi — balясина/
# stoyka/aksessuar va h.k. bu ro'yxatga kirmaydi.
UZILISH_XAVFI_KATEGORIYALAR = {"ТРУБА", "ПРОФИЛЬ", "ЛИСТ", "ЛИСТ РУЛОН"}


def _nol_kutish_hisobla(qoldiq: float, kunlik: float, konteynerlar: list,
                          horizon: float) -> tuple:
    """
    2026-07-27 (Huzayfa, TO'G'RILANGAN LOGIKA — "kelajakda konteyner bor,
    ammo u kelguncha 0 ga tushadigan tovarlarni yig'ishi kerak"):
    zanjir_sim()dagi "min_z chizig'idan pastga tushish" EMAS — bu yerda
    tovar REAL 0 GA qachon tushishini va o'sha paytdan keyin ENG YAQIN
    QAYSI YO'LDAGI KONTEYNER uni qutqarishini (qachon yetib kelishini)
    hisoblaymiz.

    Kun-ma-kun (konteyner voqealari orasida chiziqli): `qoldiq` har kuni
    `kunlik`ga kamayadi, har konteyner o'z kunida miqdorini qo'shadi.

    Qaytaradi: (nol_kuni, kutish_kun)
      nol_kuni   — necha kundan keyin qoldiq REAL nolga tushadi (0 —
                   allaqachon shunday, ya'ni bugun ham kam/tugagan).
      kutish_kun — nol_kuni'dan keyin ENG YAQIN yo'ldagi konteyner
                   yetib keladigan kun (gorizont ichida). Agar
                   nol_kuni topilmasa YOKI topilgan-u lekin gorizont
                   ichida qutqaradigan konteyner bo'lmasa — None.

    MUHIM: agar (nol_kuni, kutish_kun) ikkalasi ham aniq bo'lmasa —
    bu funksiya chaqiruvchisi bu tovarni RO'YXATGA QO'SHMASLIGI kerak:
    "konteyner umuman yo'q" yoki "hech qachon 0ga tushmaydi" — ikkalasi
    ham BOSHQA muammo, bu hisobotning maqsadi emas.
    """
    if kunlik <= 0:
        return None, None

    kont = sorted([(k, m) for k, m in konteynerlar if k > 0], key=lambda x: x[0])
    # allaqachon yetib kelgan/kechikkan, hali КЕЛДИ belgilanmagan (k<=0)
    # bo'lsa — zanjir_sim bilan bir xil naqsh, boshlang'ich qoldiqqa qo'shiladi
    joriy = float(qoldiq) + sum(m for k, m in konteynerlar if k <= 0)

    joriy_kun = 0.0
    nol_kuni  = 0 if joriy <= 0 else None

    for kun_q, miqdor in kont:
        if kun_q > horizon:
            break
        if nol_kuni is not None:
            return nol_kuni, kun_q   # shu konteyner qutqaradi
        gap         = kun_q - joriy_kun
        joriy_oldin = joriy
        joriy      -= kunlik * gap
        if joriy <= 0:
            nol_kuni = int(joriy_kun + max(0.0, joriy_oldin / kunlik))
            return nol_kuni, kun_q   # xuddi shu konteyner "kelguncha" kutadi
        joriy    += miqdor
        joriy_kun = kun_q

    if nol_kuni is None:
        # gorizont oxirigacha hisoblab ko'ramiz — agar shu yerda ham
        # tushmasa, tovar bu hisobotga umuman kirmaydi
        gap         = horizon - joriy_kun
        joriy_oldin = joriy
        joriy      -= kunlik * gap
        if joriy <= 0:
            nol_kuni = int(joriy_kun + max(0.0, joriy_oldin / kunlik))
    return nol_kuni, None   # konteyner topilmadi (yo'q yoki gorizontdan tashqarida)


UZILISH_XAVFI_FOIZ_CHEGARA = 0.70  # 2026-08-05 (Huzayfa): qoldiq/min_zaxira
# shu chegaradan PAST (yoki teng) bo'lsa, konteyner yo'lda bor-yo'qligidan
# qat'iy nazar "amalda tugagan" deb ro'yxatga qo'shiladi (pastga qarang).


def uzilish_xavfi_royxat(data_file: Path, buyurtma_yuklash_fn) -> list[dict]:
    """
    2026-07-27 (Huzayfa, oxirgi tuzatilgan mantiq), 2026-08-05'da
    KENGAYTIRILDI: ikki MUSTAQIL mezondan BIRI bajarilsa, tovar ro'yxatga
    qo'shiladi:
      1. **Konteyner-vaqt mezoni** (eski, o'zgarmagan): tovarga YO'LDA
         (ma'lum, taqvimi bor) konteyner BOR, LEKIN u yetib kelguncha
         tovar REAL 0 ga tushib qoladi.
      2. **Foiz-chuqurlik mezoni** (2026-08-05, Huzayfa: "2000 min-1500
         qoldiq uzilgan deyilmaydi, ammo 7500 min-1000 qoldiq katta
         teshik — bu tugagan deyiladi"): `qoldiq / min_z <=
         UZILISH_XAVFI_FOIZ_CHEGARA` (70%) bo'lsa — KONTEYNER yo'lda
         bor-yo'qligidan QAT'IY NAZAR qo'shiladi. Sabab: min_zaxiradan
         juda chuqur pasayish amalda "tugagan"ga teng, garchi son hali
         aniq 0 bo'lmasa ham.
    Ikkalasi ham yo'q bo'lsa (yoki min_z<=0/ABC mos kelmasa) — chiqarib
    tashlanadi.

    IKKI QO'SHIMCHA FILTR (Huzayfa aniq talab qildi):
      1. FAQAT Труба/Профиль/Лист(+Лист рулон) kategoriyalari —
         UZILISH_XAVFI_KATEGORIYALAR.
      2. FAQAT ABC toifasi A yoki B bo'lgan tovarlar — manba
         Yuklama_optimal.py::abc_map_yuklash() (Minimal_zaxiralar/
         Min_Zaxira.xlsx "ABC" ustuni, Huzayfa qo'lda tahrirlaydi —
         konteyner yuklashda ham AYNAN shu manba ishlatiladi). ABC
         belgilanmagan tovar "C" deb hisoblanadi va RO'YXATGA KIRMAYDI.

    "Kunlik" sarf — butun tizimda ishlatiladigan standart formula
    (min_z / KUNLAR, Tsex uchun eski /30), gorizont — Асосий/Ош 70 kun,
    Цех 55 kun (Buyurtma/Kamomat bilan bir xil).

    Natija nol_kuni (real 0 ga tushish kuni) bo'yicha o'sish tartibida
    saralanadi — nol_kuni aniqlanmagan (faqat foiz-mezoni bilan qo'shilgan,
    gorizont ichida 0ga tushmaydigan) qatorlar oxirida, o'z ichida foiz
    bo'yicha (eng chuqur teshik birinchi).

    Qaytaradi: [{"tovar", "kategoriya", "kanal", "kanal_nomi", "abc",
                 "nol_kuni" (yoki None), "kutish_kun" (yoki None),
                 "qoldiq", "min_z", "foiz", "yolda_jami",
                 "buyurtma_berilgan", "buyurtma_miqdor"}, ...]
    """
    try:
        inv = pd.read_excel(data_file, sheet_name="Инвентар")
        for col in ["Қолдиқ", "Мин_Захира", "Асосий_Қолдиқ", "Цех_Қолдиқ",
                    "Ош_Қолдиқ", "Асосий_Захира", "Цех_Захира", "Ош_Захира"]:
            if col in inv.columns:
                inv[col] = pd.to_numeric(inv[col], errors="coerce").fillna(0)
    except Exception as e:
        logger.error(f"uzilish_xavfi_royxat inv: {e}")
        return []

    if "Категория" in inv.columns:
        inv = inv[inv["Категория"].astype(str).str.upper().isin(UZILISH_XAVFI_KATEGORIYALAR)].copy()
    if inv.empty:
        return []

    from Yuklama_optimal import abc_map_yuklash, _abc_olish
    abc_map = abc_map_yuklash()

    # Konteyner ma'lumotlari — kamomat_excel_v2 bilan bir xil naqsh,
    # kanaldan MUSTAQIL (bitta marta, tsikldan tashqarida quriladi).
    kont_map: dict[str, list] = {}
    try:
        kont = pd.read_excel(data_file, sheet_name="Контейнерлар")
        kont = kont[kont["Холат"] != "КЕЛДИ ✅"].copy()
        for col in ["Кун_Қолди", "Миқдор"]:
            if col in kont.columns:
                kont[col] = pd.to_numeric(kont[col], errors="coerce").fillna(0)
        for _, r in kont.iterrows():
            tovar = str(r.get("Товар", ""))
            kq    = float(r.get("Кун_Қолди", 0))
            mq    = float(r.get("Миқдор", 0))
            if tovar and mq > 0:
                kont_map.setdefault(tovar, []).append((kq, mq))
    except Exception as e:
        logger.warning(f"uzilish_xavfi_royxat konteyner: {e}")

    KANAL_NOMI = {"asosiy": "Асосий", "sex": "Цех", "osh": "Ош"}
    natija: list[dict] = []
    for kanal in ("asosiy", "sex", "osh"):
        df = inv
        if "Тур" in df.columns:
            df = (df[df["Тур"] == "ЦЕХ🏭"] if kanal == "sex"
                  else df[df["Тур"] != "ЦЕХ🏭"])

        qoldiq_col = {"sex": "Цех_Қолдиқ", "osh": "Ош_Қолдиқ"}.get(kanal, "Асосий_Қолдиқ")
        if qoldiq_col not in df.columns:
            qoldiq_col = "Қолдиқ"
        minz_col = {"sex": "Цех_Захира", "osh": "Ош_Захира"}.get(kanal, "Асосий_Захира")
        if minz_col not in df.columns:
            minz_col = "Мин_Захира"

        buy     = buyurtma_yuklash_fn(kanal)
        ordered = {i["tovar"] for i in buy.get("buyurtmalar", [])} if buy else set()
        # 2026-08-05 (Huzayfa: "byurtma ustunida agar berilgan bo'lsa soni"):
        # shunchaki bor/yo'qligi emas, aniq buyurilgan MIQDORni ham kerak.
        ordered_miqdor: dict = {}
        for i in (buy.get("buyurtmalar", []) if buy else []):
            nomi = i.get("tovar")
            if nomi:
                ordered_miqdor[nomi] = ordered_miqdor.get(nomi, 0) + float(i.get("miqdor", 0))

        _horizon = 70.0 if kanal != "sex" else float(KELISH_KUNI)
        for _, row in df.iterrows():
            tovar = str(row.get("Товар", ""))
            if not tovar:
                continue
            abc = _abc_olish(abc_map, tovar)
            if abc not in ("A", "B"):
                continue
            qoldiq = float(row.get(qoldiq_col, 0))
            min_z  = float(row.get(minz_col, 0))
            if min_z <= 0:
                continue
            foiz   = qoldiq / min_z
            kont_l = kont_map.get(tovar, [])
            kunlik = min_z / float(KUNLAR)
            # nol_kuni/kutish_kun konteyner bo'lmasa ham hisoblanadi (burn-rate
            # asosida) — konteyner-vaqt mezoni uchun, kont_l bo'sh bo'lsa ham
            # ishlaydi (_nol_kutish_hisobla bo'sh ro'yxatni qo'llab-quvvatlaydi).
            nol_kuni, kutish_kun = _nol_kutish_hisobla(qoldiq, kunlik, kont_l, _horizon)
            kont_mezoni = nol_kuni is not None and kutish_kun is not None
            foiz_mezoni = foiz <= UZILISH_XAVFI_FOIZ_CHEGARA
            if not (kont_mezoni or foiz_mezoni):
                continue
            yolda_jami = sum(m for _k, m in kont_l)   # gorizontdan qat'iy nazar — JAMI
            natija.append({
                "tovar":            tovar,
                "kategoriya":       str(row.get("Категория", "")),
                "kanal":            kanal,
                "kanal_nomi":       KANAL_NOMI[kanal],
                "abc":              abc,
                "nol_kuni":         int(nol_kuni) if nol_kuni is not None else None,
                "kutish_kun":       (int(kutish_kun) - int(nol_kuni)) if kont_mezoni else None,
                "qoldiq":           int(qoldiq),
                "min_z":            int(min_z),
                "foiz":             round(foiz * 100, 1),
                "yolda_jami":       int(yolda_jami),
                "buyurtma_berilgan": tovar in ordered,
                "buyurtma_miqdor":  int(ordered_miqdor.get(tovar, 0)),
            })

    natija.sort(key=lambda r: (
        r["nol_kuni"] if r["nol_kuni"] is not None else 99999,
        r["foiz"],
    ))
    return natija


def uzilish_xavfi_excel(data_file: Path, lang: str, buyurtma_yuklash_fn) -> BytesIO | None:
    """
    2026-07-27 (Huzayfa, TO'G'RILANGAN MANTIQ): uzilish_xavfi_royxat()
    natijasini rangli, TARTIBLI Excel qilib beradi. Har bir qator —
    tovarga konteyner YO'LDA (ma'lum), lekin u yetib kelguncha tovar
    REAL 0 ga tushib qoladi:
      - Tovar nomlari CHALKASHMAYDI — avval Труба, keyin Профиль, keyin
        Лист(+Лист рулон) (CAT_ORDER), har kategoriya ICHIDA o'lcham
        bo'yicha (diametr/en×boy/qalinlik kichikdan kattaga —
        tovar_sort_key(), butun tizimda ishlatiladigan standart).
      - Kategoriyalar ANIQ AJRALIB TURADI — "Yo'ldagi yuklar" (yolda_
        excel.py) uslubida: ajratuvchi sarlavha qatori + o'z rang
        oilasida ochroq/тоқроқ (juft/toq) qatorlar, har katak border bilan.
      - Ikki "kun" ustuni bor (Huzayfa ikkala misolini ham alohida
        so'ragan edi):
          "0 гача (кун)"  — necha kundan keyin tovar REAL 0 ga tushadi
                             (0 — allaqachon shunday). Faqat foiz-mezoni
                             bilan qo'shilgan (gorizont ichida 0ga
                             tushmaydigan) qatorlarda "—" ko'rinadi.
          "Кутиш (кун)"   — 0 ga tushgandan keyin ma'lum konteyner
                             yetib kelguncha necha kun kutadi. Konteyner
                             yo'q/gorizontdan tashqarida bo'lsa "—".
    2026-08-05 qo'shildi: "Йўлда_Жами" ustuni (shu tovar uchun YO'LDA
    bo'lgan JAMI miqdor, gorizontdan qat'iy nazar) va "Буюртма_Ҳолати"
    endi buyurilgan bo'lsa aniq MIQDORni ham ko'rsatadi.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    royxat = uzilish_xavfi_royxat(data_file, buyurtma_yuklash_fn)
    if not royxat:
        return None

    # Kategoriya bo'yicha (Труба->Профиль->Лист), ICHIDA o'lcham bo'yicha
    # (kichikdan kattaga) — tovar_sort_key(), butun tizimda (Kamomat,
    # Buyurtma Excel) ishlatiladigan standart o'lcham tartibi.
    royxat = sorted(royxat, key=lambda r: (
        CAT_ORDER.get(r["kategoriya"], 99),
        tovar_sort_key(r["tovar"], r["kategoriya"]),
    ))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Узилиш хавфи" if lang == "cyr" else "Uzilish xavfi"

    if lang == "cyr":
        hdrs = ["№", "Товар", "Канал", "ABC", "Қолдиқ", "Мин_Захира",
                "Йўлда_Жами", "0 гача (кун)", "Кутиш (кун)", "Буюртма_Ҳолати"]
        berildi, kutilmq, kun_suffix = "Берилди ✅", "Кутилмоқда ⏳", "кун"
    else:
        hdrs = ["№", "Tovar", "Kanal", "ABC", "Qoldiq", "Min_Zaxira",
                "Yolda_Jami", "0 gacha (kun)", "Kutish (kun)", "Buyurtma_Holati"]
        berildi, kutilmq, kun_suffix = "Berildi ✅", "Kutilmoqda ⏳", "kun"
    # 2026-07-27 (Huzayfa: "shrift juda kichkina, katak ham kichkina"):
    # ustun kengliklari, shrift o'lchami va qator balandligi kattalashtirildi.
    # 2026-08-05: "Йўлда_Жами" ustuni qo'shildi (+1 ustun).
    col_w = [6, 54, 12, 7, 12, 14, 12, 14, 14, 24]
    NCOL  = len(hdrs)

    def fill(hex_: str): return PatternFill("solid", fgColor=hex_)
    def font(sz=13, bold=False, color="000000"): return Font(size=sz, bold=bold, color=color)
    def aln(h="left", v="center", wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    thin = Side(style="thin", color="B0B0B0")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(hdrs)
    ws.row_dimensions[1].height = 34
    for i, cell in enumerate(ws[1], 1):
        cell.fill      = fill("C00000")
        cell.font      = font(14, True, "FFFFFF")
        cell.alignment = aln("center", "center", True)
        cell.border    = border_all
        ws.column_dimensions[cell.column_letter].width = col_w[i - 1]
    ws.freeze_panes = "A2"

    cur_kat   = None
    kat_cnt   = 0
    excel_row = 1
    n         = 0

    for r in royxat:
        kat = r["kategoriya"]

        # Kategoriya ajratuvchi qator — xuddi yolda_excel/kamomat_excel
        # uslubida, "Yo'ldagi yuklar" bilan bir xil "blok" tuyg'usi uchun
        if kat != cur_kat:
            cur_kat = kat
            kat_cnt = 0
            colors  = CAT_COLORS.get(kat, CAT_COLORS_DEF)
            excel_row += 1
            ws.merge_cells(start_row=excel_row, start_column=1,
                            end_row=excel_row, end_column=NCOL)
            sep = ws.cell(row=excel_row, column=1, value=f"  {kat}")
            sep.fill      = fill(colors["h"])
            sep.font      = font(14, True, "1F1F1F")
            sep.alignment = aln("left", "center")
            ws.row_dimensions[excel_row].height = 30

        n       += 1
        kat_cnt += 1
        colors   = CAT_COLORS.get(kat, CAT_COLORS_DEF)
        row_clr  = colors["a"] if kat_cnt % 2 == 1 else colors["b"]
        # 2026-08-05 (Huzayfa: "byurtma ustunida agar berilgan bo'lsa soni,
        # bo'lmasa kutilmoqda"): miqdor bilan birga ko'rsatiladi.
        if r["buyurtma_berilgan"]:
            holat_txt = f"{berildi} ({r['buyurtma_miqdor']})"
        else:
            holat_txt = kutilmq
        # foiz-mezoni bilan qo'shilgan (konteyner gorizont ichida
        # qutqarmaydigan/yo'q) qatorlarda nol_kuni/kutish_kun aniqlanmagan
        # bo'lishi mumkin — "—" bilan ko'rsatiladi.
        nol_txt    = "—" if r["nol_kuni"] is None else f"{r['nol_kuni']} {kun_suffix}"
        kutish_txt = "—" if r["kutish_kun"] is None else f"{r['kutish_kun']} {kun_suffix}"

        excel_row += 1
        ws.append([
            n, r["tovar"], r["kanal_nomi"], r["abc"], r["qoldiq"],
            r["min_z"], r["yolda_jami"], nol_txt, kutish_txt, holat_txt,
        ])
        ws.row_dimensions[excel_row].height = 26
        for col_i in range(1, NCOL + 1):
            cell = ws.cell(row=excel_row, column=col_i)
            cell.fill      = fill(row_clr)
            cell.font      = font(13, bold=(col_i in (8, 9)))
            cell.alignment = aln("center" if col_i != 2 else "left")
            cell.border    = border_all

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ============================================================
# EXCEL GENERATSIYASI
# ============================================================
def kamomat_excel_v2(data_file: Path, kanal: str,
                     lang: str, buyurtma_yuklash_fn) -> BytesIO | None:
    """
    Tartibli, rangli kamomat Excel.
    - KRITIK + PAST tovarlar
    - Kategoriya bo'yicha saralanadi (TRUБА -> PROFIL -> LIST ...)
    - Har kategoriya o'z rang oilasida (juft/toq qatorlar)
    - Kategoriya separator qatori (ajralib turadi)
    - Ustunlar: №, Товар, Холат, Қолдиқ, Йўлда, Мин_Захира,
                Кун_Хавф, Буюртма_Ҳолати, Таклиф_Миқдор
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    # ── Ma'lumot o'qish ──────────────────────────────────────
    try:
        inv = pd.read_excel(data_file, sheet_name="Инвентар")
        for col in ["Қолдиқ", "Мин_Захира", "Йўлда_Жами", "Фарқ",
                    "Асосий_Қолдиқ", "Цех_Қолдиқ", "Ош_Қолдиқ",
                    "Асосий_Захира", "Цех_Захира", "Ош_Захира"]:
            if col in inv.columns:
                inv[col] = pd.to_numeric(inv[col], errors="coerce").fillna(0)
    except Exception as e:
        logger.error(f"kamomat_excel_v2 inv: {e}")
        return None

    # 2026-07-24 (Huzayfa: "Asosiyda buyurtma yozsak, u O'shning/Tsexning
    # qoldig'ini hisobga olmasligi kerak" -- Buyurtma Excel'da tuzatilgan
    # gap shu yerda ham bor edi): endi Қолдиқ VA Мин_Захира ikkalasi ham
    # kanalga mos ustundan o'qiladi (Generate_Asosiy_order.py::load_data()
    # bilan bir xil naqsh). Fallback -- eski (kanal ustunlarisiz) Power BI
    # fayli uchun -- umumiy ustunlarga tushadi.
    qoldiq_col = {"sex": "Цех_Қолдиқ", "osh": "Ош_Қолдиқ"}.get(kanal, "Асосий_Қолдиқ")
    if qoldiq_col not in inv.columns:
        qoldiq_col = "Қолдиқ"
    minz_col = {"sex": "Цех_Захира", "osh": "Ош_Захира"}.get(kanal, "Асосий_Захира")
    if minz_col not in inv.columns:
        minz_col = "Мин_Захира"

    # Kanal filtri
    if "Тур" in inv.columns:
        inv = (inv[inv["Тур"] == "ЦЕХ🏭"] if kanal == "sex"
               else inv[inv["Тур"] != "ЦЕХ🏭"])

    df = inv[inv["Холат"].isin(["🔴 КРИТИК", "🟡 ПАСТ"])].copy()
    if df.empty:
        return None

    # ── Konteyner ma'lumotlari ────────────────────────────────
    kont_map: dict[str, list] = {}
    try:
        kont = pd.read_excel(data_file, sheet_name="Контейнерлар")
        kont = kont[kont["Холат"] != "КЕЛДИ ✅"].copy()
        for col in ["Кун_Қолди", "Миқдор"]:
            if col in kont.columns:
                kont[col] = pd.to_numeric(kont[col], errors="coerce").fillna(0)
        for _, r in kont.iterrows():
            tovar = str(r.get("Товар", ""))
            kq    = float(r.get("Кун_Қолди", 0))
            mq    = float(r.get("Миқдор", 0))
            if tovar and mq > 0:
                kont_map.setdefault(tovar, []).append((kq, mq))
    except Exception as e:
        logger.warning(f"Konteyner o'qilmadi: {e}")

    # ── Buyurtma holati ────────────────────────────────────────
    buy     = buyurtma_yuklash_fn(kanal)
    ordered = {i["tovar"] for i in buy.get("buyurtmalar", [])} if buy else set()
    berildi = "Берилди ✅" if lang == "cyr" else "Berildi ✅"
    kutilmq = "Кутилмоқда ⏳" if lang == "cyr" else "Kutilmoqda ⏳"

    # ── Saralash ───────────────────────────────────────────────
    if "Категория" not in df.columns:
        df["Категория"] = "БОШҚА"
    df["_c"]   = df["Категория"].apply(lambda x: CAT_ORDER.get(str(x), 99))
    df["_s"]   = df.apply(lambda r: tovar_sort_key(
                    str(r.get("Товар", "")), str(r.get("Категория", ""))), axis=1)
    df = df.sort_values(["_c", "_s"]).reset_index(drop=True)

    # ── Zanjir simulyatsiyasi ──────────────────────────────────
    # 2026-07-23: Buyurtma Excel bilan bir xil natija chiqishi uchun --
    # Asosiy/O'sh 70 kunlik gorizont, Tsex eski 55da (alohida mavzu,
    # hozircha tegilmadi).
    _horizon_ov = None if kanal == "sex" else 70
    sims = []
    for _, row in df.iterrows():
        tovar  = str(row.get("Товар", ""))
        qoldiq = float(row.get(qoldiq_col, 0))
        min_z  = float(row.get(minz_col, 0))
        kont_l = kont_map.get(tovar, [])
        sims.append(zanjir_sim(qoldiq, min_z, kont_l, horizon_override=_horizon_ov))

    # ── Excel ──────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Камомат" if lang == "cyr" else "Kamomat"

    # Sarlavhalar
    if lang == "cyr":
        hdrs = ["№", "Товар", "Холат", "Қолдиқ", "Йўлда_Жами",
                "Мин_Захира", "Кун_Хавф", "Буюртма_Ҳолати", "Таклиф_Миқдор"]
    else:
        hdrs = ["№", "Tovar", "Holat", "Qoldiq", "Yolda_Jami",
                "Min_Zaxira", "Kun_Xavf", "Buyurtma_Holati", "Taklif_Miqdor"]

    col_w = [5, 46, 13, 11, 12, 13, 11, 19, 16]
    NCOL  = len(hdrs)

    # Stil yordamchilari
    def fill(hex_: str):
        return PatternFill("solid", fgColor=hex_)

    def font(sz=10, bold=False, color="000000"):
        return Font(size=sz, bold=bold, color=color)

    def aln(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    thin = Side(style="thin", color="D0D0D0")
    border_thin = Border(bottom=thin)

    # Header qatori
    ws.append(hdrs)
    ws.row_dimensions[1].height = 28
    hdr_fill = fill("1F4E79")
    hdr_font = font(11, True, "FFFFFF")
    hdr_aln  = aln("center", "center", True)
    for i, cell in enumerate(ws[1], 1):
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = hdr_aln
        ws.column_dimensions[cell.column_letter].width = col_w[i - 1]

    ws.freeze_panes = "A2"

    # ── Qatorlar ──────────────────────────────────────────────
    cur_kat    = None
    kat_cnt    = 0    # har kategoriya ichida qator hisobi
    excel_row  = 2
    n          = 0

    for idx, row in df.iterrows():
        kat    = str(row.get("Категория", "БОШҚА"))
        tovar  = str(row.get("Товар", ""))
        holat  = str(row.get("Холат", ""))
        # 2026-07-24: ko'rsatiladigan Қолдиқ/Мин_Захира ham sims ro'yxatini
        # yasashda ishlatilgan XUDDI SHU kanalga mos ustundan olinadi —
        # aks holda Excel'da ko'rinadigan son bilan Таклиф_Миқдор (sims
        # asosida) mos kelmay qolardi (xuddi C-ustuni bugidagi kabi).
        qoldiq = int(row.get(qoldiq_col, 0))
        yolda  = int(row.get("Йўлда_Жами", 0))
        min_z  = int(row.get(minz_col, 0))
        sim    = sims[idx]

        # Kategoriya separator
        if kat != cur_kat:
            cur_kat = kat
            kat_cnt = 0
            colors  = CAT_COLORS.get(kat, CAT_COLORS_DEF)

            ws.append([kat] + [""] * (NCOL - 1))
            ws.merge_cells(
                start_row=excel_row, start_column=1,
                end_row=excel_row, end_column=NCOL
            )
            sep_cell           = ws.cell(row=excel_row, column=1)
            sep_cell.value     = f"  {kat}"
            sep_cell.fill      = fill(colors["h"])
            sep_cell.font      = font(11, True, "1F1F1F")
            sep_cell.alignment = aln("left", "center")
            ws.row_dimensions[excel_row].height = 22
            excel_row += 1

        # Ma'lumot qatori
        n      += 1
        kat_cnt += 1
        colors  = CAT_COLORS.get(kat, CAT_COLORS_DEF)
        row_clr = colors["a"] if kat_cnt % 2 == 1 else colors["b"]

        # Kун_Хавф qiymati
        uzilish = sim.get("uzilish_kun")
        if uzilish is None:
            kun_xavf = "—"
        elif uzilish <= 0:
            kun_xavf = "XOZIR ❗" if lang == "lat" else "ҲОЗИР ❗"
        else:
            kun_xavf = f"{uzilish} kun" if lang == "lat" else f"{uzilish} кун"

        taklif       = sim.get("taklif", 0)
        b_holat      = berildi if tovar in ordered else kutilmq

        ws.append([
            n, tovar, holat, qoldiq, yolda,
            min_z, kun_xavf, b_holat,
            taklif if taklif > 0 else ""
        ])

        row_fill = fill(row_clr)
        for ci, cell in enumerate(ws[excel_row], 1):
            cell.fill   = row_fill
            cell.font   = font(10)
            cell.border = border_thin

            if ci == 1:                       # №
                cell.alignment = aln("center")
            elif ci == 2:                     # Товар
                cell.alignment = aln("left")
            elif ci == 3:                     # Холат
                cell.alignment = aln("center")
                if "КРИТИК" in holat:
                    cell.font = font(10, True, "C00000")
                elif "ПАСТ" in holat:
                    cell.font = font(10, True, "7B3F00")
            elif ci in (4, 5, 6, 9):         # Raqamlar
                cell.alignment = aln("right")
            elif ci == 7:                     # Кун_Хавф
                cell.alignment = aln("center")
                if uzilish is not None and uzilish <= 15:
                    cell.font = font(10, True, "C00000")
                elif uzilish is not None and uzilish <= 30:
                    cell.font = font(10, False, "7B3F00")
            else:
                cell.alignment = aln("center")

        ws.row_dimensions[excel_row].height = 17
        excel_row += 1

    # Umumiy info (oxirda)
    excel_row += 1
    info_text = (
        f"Sana: {datetime.now().strftime('%d.%m.%Y %H:%M')}  |  "
        f"Kanal: {kanal}  |  "
        f"Jami: {n} ta  |  "
        f"Kunlik = min/{KUNLIK_SOTUV_BOLISH}"
    )
    ws.append([info_text])
    ws.merge_cells(
        start_row=excel_row, start_column=1,
        end_row=excel_row, end_column=NCOL
    )
    info_cell           = ws.cell(row=excel_row, column=1)
    info_cell.font      = font(9, False, "808080")
    info_cell.alignment = aln("left")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ============================================================
# ZANJIR GRAFIK
# ============================================================

# ============================================================
# GRAFIK
# ============================================================

# ============================================================
# GRAFIK
# ============================================================
def grafik_chiz(tovar: str, qoldiq: float, min_z: float,
                konteynerlar: list, kunlar: int = 0) -> "BytesIO | None":
    """
    Piecewise-linear stock trace:
      - qoldiq kunlik ravishda pasayadi
      - konteyner kelgan kunda qoldiq keskin ko'tariladi
      - har bir konteyner uchun alohida label (rotated, linea bo'yida)
      - bir kunda bir nechta: bitta chiziq, labellar stacked
    """
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.ticker as ticker
        from collections import defaultdict
    except ImportError:
        return None

    kunlik = min_z / float(KUNLIK_SOTUV_BOLISH) if min_z > 0 else 1.0

    # ── Konteynerlarni kun bo'yicha guruhlash ──────────────────
    kont_sorted = sorted(konteynerlar, key=lambda x: x[0])
    kont_by_day = defaultdict(list)
    for kq, mq in kont_sorted:
        d = max(0, int(round(kq)))
        kont_by_day[d].append(float(mq))

    # ── X oralig'i ─────────────────────────────────────────────
    max_day = max(kont_by_day.keys()) if kont_by_day else 0
    if kunlar == 0:
        kunlar = max(max_day + KELISH_KUNI + 10, 90)
    kunlar = min(kunlar, 200)

    # ── Piecewise-linear trace ─────────────────────────────────
    # Har bir segment: linear tushish, keyin konteynerda sakrash
    tx, ty = [0], [float(qoldiq)]
    cur = float(qoldiq)
    prev_d = 0

    for d in sorted(kont_by_day.keys()):
        if d > kunlar:
            continue
        # tushish davri
        elapsed = d - prev_d
        cur_before = max(cur - kunlik * elapsed, 0)
        if d > prev_d:
            tx.append(d)
            ty.append(cur_before)
        # sakrash
        cur = cur_before + sum(kont_by_day[d])
        tx.append(d)
        ty.append(cur)
        prev_d = d

    # oxirgi nuqtagacha tushish
    remaining = kunlar - prev_d
    end_y = max(cur - kunlik * remaining, 0)
    tx.append(kunlar)
    ty.append(end_y)

    # ── Y o'qi chegarasi ───────────────────────────────────────
    y_max = max(max(ty) * 1.10, min_z * 1.35)

    # ── Uzilish kuni (chiziq min_z dan pastga tushgan kun) ─────
    uzilish_day = None
    for i in range(1, len(tx)):
        x0, y0 = tx[i-1], ty[i-1]
        x1, y1 = tx[i], ty[i]
        if y0 >= min_z and y1 < min_z and x1 > x0:
            # interpolatsiya
            frac = (y0 - min_z) / (y0 - y1)
            uzilish_day = int(x0 + frac * (x1 - x0))
            break
        elif y0 < min_z and uzilish_day is None:
            uzilish_day = int(x0)
            break

    # ── Ranglar ────────────────────────────────────────────────
    BG    = "#0F1923"
    GRID  = "#1A2B3C"
    C_STK = "#3498DB"
    C_MIN = "#E74C3C"
    C_KNT = "#2ECC71"
    C_UZL = "#E74C3C"

    fig, ax = plt.subplots(figsize=(13, 5))
    fig.patch.set_facecolor(BG)
    ax.set_facecolor(BG)

    # ── Uzilish zonasi ─────────────────────────────────────────
    if uzilish_day is not None and uzilish_day < kunlar:
        ax.axvspan(uzilish_day, kunlar, alpha=0.12, color=C_UZL, zorder=1)

    # ── Min zaxira ─────────────────────────────────────────────
    ax.axhline(y=min_z, color=C_MIN, linewidth=2.0,
               linestyle="--", alpha=1.0, zorder=3)
    ax.text(kunlar * 0.01, min_z * 1.02,
            "Min: {:,}".format(int(min_z)).replace(",", " "),
            color=C_MIN, fontsize=8, fontweight="bold", va="bottom", zorder=6)

    # ── Qoldiq chizig'i ────────────────────────────────────────
    ax.plot(tx, ty, color=C_STK, linewidth=2.2, zorder=4)
    ax.plot(0, qoldiq, "o", color=C_STK, markersize=5, zorder=5)

    # ── Konteyner chiziqlari + rotated labellar ────────────────
    total_kont = sum(len(v) for v in kont_by_day.values())
    sorted_days = sorted(kont_by_day.keys())

    for d in sorted_days:
        if d > kunlar:
            continue
        containers = kont_by_day[d]

        # qoldiq darajasi d kunda (sakrashdan OLDIN)
        elapsed = d - (sorted_days[sorted_days.index(d) - 1]
                       if sorted_days.index(d) > 0 else 0)
        # tx/ty dan d nuqtasini topamiz (sakrashdan oldingi qiymat)
        y_before = 0
        for i, x in enumerate(tx):
            if x == d:
                y_before = ty[i]   # sakrashdan oldingi qiymat
                break
            if x > d:
                # interpolatsiya
                x0, y0 = tx[i-1], ty[i-1]
                x1, y1 = tx[i], ty[i]
                if x1 > x0:
                    y_before = y0 + (y1 - y0) * (d - x0) / (x1 - x0)
                break

        ax.axvline(x=d, color=C_KNT, linewidth=1.8, alpha=0.9, zorder=4)

        # Har bir konteyner: label chiziq ustida, boshqa y darajasida
        # rotation=90 → matn vertikal, har biri o'z qoldiq segmentida
        cum_y = y_before
        for j, mq in enumerate(containers):
            label_y = cum_y + mq * 0.5   # segmentning o'rtasi
            qty_str = "+{:,}".format(int(mq)).replace(",", " ")
            ax.text(d, label_y, qty_str,
                    color=C_KNT, fontsize=8, fontweight="bold",
                    rotation=90, ha="center", va="center",
                    zorder=6,
                    bbox=dict(boxstyle="round,pad=0.1",
                              fc="#0F1923", ec="none", alpha=0.7))
            cum_y += mq

    # ── Uzilish belgisi ────────────────────────────────────────
    if uzilish_day is not None and 0 < uzilish_day < kunlar:
        ax.axvline(x=uzilish_day, color=C_UZL, linewidth=1.5,
                   linestyle=":", alpha=0.9, zorder=5)
        ax.text(uzilish_day + kunlar * 0.005, min_z * 0.5,
                "UZILISH {}k".format(uzilish_day),
                color=C_UZL, fontsize=8, fontweight="bold",
                va="center", zorder=6)

    # ── O'qlar — real sanalar bilan ────────────────────────────
    import calendar as _cal
    from datetime import date as _date, timedelta as _td

    _bugun = _date.today()

    # Har oyning 10, 20 va oxirgi kuni da tick
    _ticks, _labels = [], []
    _prev_month = None
    for _off in range(kunlar + 1):
        _d = _bugun + _td(days=_off)
        _last = _cal.monthrange(_d.year, _d.month)[1]
        if _d.day in (10, 20, _last):
            _ticks.append(_off)
            # Oy o'zgarganda nomi ham ko'rsatiladi
            _ay = ["Yan","Fev","Mar","Apr","May","Iyun",
                   "Iyul","Avg","Sen","Okt","Noy","Dek"][_d.month - 1]
            if _d.month != _prev_month:
                _labels.append(f"{_d.day}\n{_ay}")
                _prev_month = _d.month
            else:
                _labels.append(str(_d.day))

    ax.set_xlim(0, kunlar)
    ax.set_xticks(_ticks)
    ax.set_xticklabels(_labels, fontsize=7, color="#8899AA")
    ax.tick_params(axis="x", which="major", length=4, color="#334455")

    ax.set_ylim(0, y_max)
    ax.yaxis.set_major_formatter(ticker.FuncFormatter(
        lambda x, _: "{:,}".format(int(x)).replace(",", " ")))

    ax2 = ax.twinx()
    ax2.set_ylim(0, y_max)
    ax2.set_facecolor(BG)
    ax2.tick_params(colors="#8899AA", labelsize=8)
    ax2.yaxis.set_major_formatter(ticker.FuncFormatter(
        lambda x, _: "{:.0f}".format(x / kunlik) if kunlik > 0 else "0"))
    ax2.set_ylabel("Kunlar", color="#8899AA", fontsize=8)
    for sp in ax2.spines.values():
        sp.set_color(GRID)

    ax.set_xlabel(
        "Kunlik: {:,} dona".format(int(kunlik)).replace(",", " "),
        color="#8899AA", fontsize=9)
    ax.set_ylabel("Qoldiq (dona)", color="#8899AA", fontsize=9)
    ax.tick_params(colors="#8899AA", labelsize=8)
    for sp in ax.spines.values():
        sp.set_color(GRID)
    ax.grid(True, alpha=0.15, color=GRID, linestyle="-")

    short = tovar if len(tovar) <= 58 else tovar[:55] + "..."
    ax.set_title(short, color="#FFFFFF", fontsize=10, pad=10)

    handles = [
        plt.Line2D([0], [0], color=C_STK, linewidth=2.2, label="Qoldiq"),
        plt.Line2D([0], [0], color=C_MIN, linewidth=2.0,
                   linestyle="--", label="Min zaxira"),
        plt.Line2D([0], [0], color=C_KNT, linewidth=1.8,
                   label="Konteyner ({} ta)".format(total_kont)),
    ]
    ax.legend(handles=handles, facecolor="#1A2A3A",
              labelcolor="#CCDDEE", fontsize=8.5,
              loc="upper right", framealpha=0.9)

    plt.tight_layout(pad=1.0)

    bio = BytesIO()
    plt.savefig(bio, format="png", dpi=120,
                bbox_inches="tight", facecolor=BG)
    plt.close(fig)
    bio.seek(0)
    return bio
