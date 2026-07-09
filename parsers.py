"""
parsers.py — Xitoy Excel fayl parserlari
  - _yaxlitla_stenka : xitoy stenka formatini inventar formatiga o'giradi
  - _inventar_snap   : inventar to'plamiga snap (0,65→0,7, 1,35→1,4 va h.k.)
  - _china_spec_to_inventar : xitoy spec stringidan inventar nomini yaratadi
  - _parse_truba_profil_xitoy : Труба/Профиль format xitoy Excel
  - _parse_list_xitoy         : Лист format xitoy Excel
  - xitoy_ostatka_oqi         : asosiy entry point
"""
import math, re
from io import BytesIO

from common import normalize_product_name
from config import BASE_DIR, DATA_FILE, XITOY_NOM_MAP

# Inventar set — lazy cache
_inventar_set_cache: set | None = None

def _yaxlitla_stenka(s: str) -> str:
    """
    Xitoy stenka qiymatini inventar formatiga aylantiradi.
    STENKA_DELTA: xitoy 0.85mm = inventar 0.9mm (0.55→0.6 ham shunday).
    Boshqa *.X5 qiymatlar inventarda aniq saqlangan (1.45→1,45; 1.35→1,35).
    """
    # STENKA_DELTA ga ega qiymatlar (xitoy → inventar)
    _ANIQ = {
        '0.55': '0,6',  '0,55': '0,6',
        '0.85': '0,9',  '0,85': '0,9',
        '0.65': '0,65', '0,65': '0,65',
        '0.60': '0,6',  '0,60': '0,6',
        '0.70': '0,7',  '0,70': '0,7',
        '1.05': '1,1',  '1,05': '1,1',
        '2.95': '3,0',  '2,95': '3,0',
        '1.95': '2,0',  '1,95': '2,0',
        '2.00': '2,0',  '3.00': '3,0',
    }
    s_clean = str(s).strip()
    if s_clean in _ANIQ:
        return _ANIQ[s_clean]
    s_dot = s_clean.replace(',', '.')
    try:
        val = float(s_dot)
    except ValueError:
        return s_clean.replace('.', ',')
    if '.' in s_dot:
        dec = s_dot.split('.')[1]
        dec_stripped = dec.rstrip('0')
        if not dec_stripped:
            return f"{int(val)},0"
        elif len(dec_stripped) == 1:
            return f"{val:.1f}".replace('.', ',')
        else:
            return f"{val:.{len(dec_stripped)}f}".replace('.', ',')
    else:
        return f"{val:.1f}".replace('.', ',')


# ── Inventar snap: lazy cache ──────────────────────────────────────────────────
_inventar_set_cache: set | None = None

def _norm_inv(s: str) -> str:
    """Inventar nomini normallashtiradi: (6м) → (6 м), (5,8м) → (5,8 м)"""
    return re.sub(r'\((\d+(?:[,\.]\d+)?)м\)', r'(\1 м)', s.strip())

def _get_inventar_set() -> set:
    """Inventar tovar nomlari to'plamini bir marta yuklaydi (kesh)."""
    global _inventar_set_cache
    if _inventar_set_cache is None:
        import openpyxl as _ox
        try:
            _wb = _ox.load_workbook(DATA_FILE, read_only=True, data_only=True)
            _ws = _wb.active
            _inventar_set_cache = {
                _norm_inv(str(r[0])) for r in _ws.iter_rows(values_only=True)
                if r[0] and str(r[0]).strip()
            }
        except Exception:
            _inventar_set_cache = set()
    return _inventar_set_cache


# ── Kanonik (bo'shliq-farqiga chidamli) inventar moslashtirish ──────────────
# TOPILGAN MUAMMO (2026-07-06): inventarning o'zida ba'zi qatorlar noodatiy
# bo'shliq bilan yozilgan — masalan haqiqiy qator "Лист- 2,0 (1500х3000)
# (Глянцевый) (304 марка)" (chiziqchadan KEYIN bo'shliq bor), lekin bizning
# generator har doim "Лист-2,0..." (bo'shliqsiz) hosil qiladi — natijada bu
# ANIQ mavjud tovar "notanish" (⚠️) deb noto'g'ri belgilanardi, chunki aniq
# satr solishtirish (`nom in inventar_set`) muvaffaqiyatsiz bo'lardi.
#
# Yechim: solishtirish uchun "kanonik" (barcha bo'shliqlar bittaga siqilgan,
# chiziqchadan keyingi bo'shliq olib tashlangan) shaklga o'tkaziladi — LEKIN
# moslik topilganda ASL (inventardagi, o'zgarishsiz — hatto g'alati bo'shliq
# bilan bo'lsa ham) matn qaytariladi, chunki chiqishda biz DOIM inventarda
# qanday yozilgan bo'lsa xuddi shundayligicha ko'rsatishimiz kerak.
_inventar_kanonik_cache: dict | None = None


def _kanonik_nom(s: str) -> str:
    """Solishtirish uchun — ketma-ket bo'shliqlarni bittaga siqadi va
    chiziqchadan keyingi bo'shliqni olib tashlaydi. FAQAT solishtirishda
    ishlatiladi, asl inventar matnini o'zgartirmaydi."""
    s = re.sub(r'\s+', ' ', str(s).strip())
    s = re.sub(r'-\s+', '-', s)
    return s


def _get_inventar_kanonik_map() -> dict:
    """kanonik_nom → inventardagi ASL (o'zgarishsiz) matn."""
    global _inventar_kanonik_cache
    if _inventar_kanonik_cache is None:
        _inventar_kanonik_cache = {}
        for orig in _get_inventar_set():
            _inventar_kanonik_cache.setdefault(_kanonik_nom(orig), orig)
    return _inventar_kanonik_cache


def _inventardan_moslashtir(nom: str, inventar_set: set | None = None) -> str:
    """
    Nomni inventar bilan ANIQ solishtiradi; topilmasa — kanonik (bo'shliqqa
    chidamli) solishtiradi. Ikkalasida ham topilsa — inventardagi ASL matn
    qaytariladi (hatto g'alati bo'shliq bilan bo'lsa ham); aks holda
    o'zgarishsiz `nom` qaytariladi.
    """
    if inventar_set is None:
        inventar_set = _get_inventar_set()
    if not nom:
        return nom
    if nom in inventar_set:
        return nom
    kmap = _get_inventar_kanonik_map()
    return kmap.get(_kanonik_nom(nom), nom)


def _inventar_snap(nom: str, inventar_set: set) -> str:
    """
    Xitoy tovar nomini inventar to'plamiga snap qiladi.
    Qoida:
      0,65 → avvalo '0,7' qidir (yaxlitlangan), yo'q bo'lsa '0,65' (aniq)
      1,35 → avvalo '1,35' qidir (aniq), yo'q bo'lsa '1,4' (yaxlitlangan)
      1,45 → avvalo '1,45' qidir (aniq), yo'q bo'lsa '1,5' (yaxlitlangan)
    Sabab: Xitoy 1.35 desa, bizda 1.35 bor bo'lsa unga ulaymiz,
           yo'q bo'lsa 1.4 bormi qidiramiz (Xitoy 1.35 = bizning 1.4).
    """
    if not inventar_set:
        return nom
    m = re.search(r'ст\s+([\d,]+)', nom)
    if not m:
        return nom
    stenka = m.group(1)
    _SNAP_ORDER = {
        '0,65': ['0,7',  '0,65'],   # 0,65 aniq bo'lmasa 0,7 ga yaxlitla
        '1,35': ['1,35', '1,4' ],   # avval 1,35 qidir, topilmasa 1,4
        '1,45': ['1,45', '1,5' ],   # avval 1,45 qidir, topilmasa 1,5
    }
    candidates = _SNAP_ORDER.get(stenka)
    if candidates:
        # Marka va prefix ajratamiz — uzunlik farqi bo'lsa ham topilsin
        marka_m = re.search(r'\(\d+ марка\)$', nom.strip())
        marka_sfx = marka_m.group(0) if marka_m else ''   # "(201 марка)"
        for c in candidates:
            new_nom = nom[:m.start(1)] + c + nom[m.end(1):]
            if new_nom in inventar_set:
                return new_nom
            # Uzunlik farqi bo'lishi mumkin (masalan xitoy 5,8м, inventar 6м)
            stenka_end = m.start(1) + len(c)
            pfx_end = new_nom.find('(', stenka_end)
            if pfx_end != -1:
                prefix = new_nom[:pfx_end].rstrip()
                if marka_sfx:
                    match = next((x for x in inventar_set
                                  if x.startswith(prefix + ' (') and x.endswith(marka_sfx)), None)
                else:
                    match = next((x for x in inventar_set
                                  if x.startswith(prefix + ' (')), None)
                if match:
                    return match
    # (Ж-X) suffixni olib tashlab qidiramiz — har qanday stenka uchun
    jm = re.search(r'\s*\(Ж-\d+\)\s*$', nom)
    if jm:
        nom_no_j = nom[:jm.start()].rstrip()
        if nom_no_j in inventar_set:
            return nom_no_j
    # Hech biri topilmadi — asl nom (yangi tovar)
    return nom


def _china_spec_to_inventar(spec: str, length: str) -> str | None:
    """
    Хитой Труба/Профиль спецификациясини инвентар номига айлантиради.
    'φ51 cT 0.85' + '5.8M' → 'Ф-51 ст 0,9 (5,8 м) (201 марка)'
    'KB 20x20 CT 0.65' + '5.8M' → 'Пр. 20х20 ст 0,7 (5,8 м) (201 марка)'
    'φ25 cT 0.85 Apkoh' + '6M' → '(Аркон) Ф-25 ст 0,9 (6 м) (201 марка)'
    """
    spec = str(spec).strip()

    # Brend
    brend = ''
    if 'Apkoh' in spec or 'Аркон' in spec:
        brend = 'Аркон'
    elif 'GOLD钛金' in spec:
        brend = 'Голд'
    elif '哑光黑' in spec:
        brend = 'Кора'

    # Marka (default 201)
    marka = '201'
    for m_val in ('304', '316', '321', '430'):
        if re.search(rf'\b{m_val}\b', spec):
            marka = m_val
            break

    # Suffix (J1, J4 ...)
    jm = re.search(r'[（(]J(\d+)[）)]', spec)
    suffix = f' (Ж-{jm.group(1)})' if jm else ''

    # Uzunlik: 5.62M → 5,6 м   5.8M → 5,8 м   6M → 6 м
    L_raw = re.sub(r'[Mm]$', '', str(length).strip()).strip()
    try:
        L_f = round(float(L_raw), 1)
        L_s = f"{L_f:.1f}"
        if L_s.endswith('.0'):
            L_s = L_s[:-2]          # "6.0" → "6"
        L_str = L_s.replace('.', ',') + " м"
    except (ValueError, TypeError):
        L_str = L_raw.replace('.', ',') + " м"
    brend_pfx = f'({brend}) ' if brend else ''
    marka_sfx = f' ({marka} марка)'

    # Труба: φ51 cT 0.85 / φ16 cT 0,65 / Φ 16 cT 0.65
    m = re.match(r'^[φΦ]\s*(\d+)\s+[cC][tT]\s+([\d,\.]+)', spec)
    if m:
        stenka = _yaxlitla_stenka(m.group(2))
        return f'{brend_pfx}Ф-{m.group(1)} ст {stenka} ({L_str}){marka_sfx}{suffix}'

    # Профиль: KB 20x20 CT 0.65  /  KB 30x30CT 0.85
    m = re.match(r'^KB\s+(\d+)[xXхх×]\s*(\d+)\s*[cC][tT]\s*([\d,\.]+)', spec)
    if m:
        stenka = _yaxlitla_stenka(m.group(3))
        return f'{brend_pfx}Пр. {m.group(1)}х{m.group(2)} ст {stenka} ({L_str}){marka_sfx}{suffix}'

    # Профиль D型管 (yarim aval): KB D型管30*15*1.05 → (Ярим овал) Пр. 30х15 ст 1,1
    m = re.match(r'^KB\s+D型管(\d+)\*(\d+)\*([\d\.]+)', spec)
    if m:
        stenka_f = math.ceil(round(float(m.group(3)) * 10, 4)) / 10
        stenka   = f"{stenka_f:.1f}".replace('.', ',')
        return f'(Ярим овал) Пр. {m.group(1)}х{m.group(2)} ст {stenka} ({L_str}){marka_sfx}{suffix}'

    return None


def _parse_truba_profil_xitoy(rows: list, ombor_i: int,
                               gui_ge_i: int, len_i: int,
                               zakaz_i: int | None = None,
                               vazn_i: int | None = None) -> tuple:
    """克力木 Труба/Профиль format: 规格 + 长度 + 库存 Ombor
    zakaz_i — '订单 Uzbek' ustuni (K): jami buyurtma miqdori.
              Agar mavjud bo'lsa, ombor_i (L) dan ustun tutiladi, chunki
              K = tayyor (L) + tayyorlanayotgan, ya'ni K ≥ L va K+L qilinmaydi.
    vazn_i  — '支重' ustuni (kg): 1 dona (6m) og'irligi.
    Qaytaradi: (known, unknown, ombor_known, vazn_map)
      known       — K ustun (jami zakaz, buyurtma hisoblash uchun)
      unknown     — inventarda topilmagan yoki parser tanimagan spec+uzunlik ro'yxati
      ombor_known — L ustun (hozir tayyor/yuklatishga tayyor, yuklatish rejasi uchun)
      vazn_map    — {inventar_nom: 1_dona_kg} (Xitoy faylidan olingan haqiqiy vazn)
    """
    from common import normalize_product_name
    # K ustun (zakaz_i) mavjud bo'lsa uni ishlatamiz, aks holda L (ombor_i)
    qty_i = zakaz_i if zakaz_i is not None else ombor_i
    known       = {}   # K ustun (jami zakaz)
    ombor_known = {}   # L ustun (tayyor/ready)
    unknown     = []
    vazn_map    = {}   # {nom: 1_dona_kg} Xitoydan
    for row in rows[1:]:
        max_i = max(qty_i, gui_ge_i, len_i)
        if zakaz_i is not None:
            max_i = max(max_i, ombor_i)
        if len(row) <= max_i:
            continue
        spec    = row[gui_ge_i]
        length  = row[len_i]
        miqdor  = row[qty_i]
        ombor_v = row[ombor_i] if len(row) > ombor_i else None
        if not spec or not length:
            continue
        spec_s  = str(spec).strip()
        len_raw = str(length).strip()
        if any(k in spec_s for k in ('小计', '合计', '序号')):
            continue
        if not re.match(r'^(?:[φΦ]|KB)', spec_s):
            continue
        # Xitoy unit vazni (支重 ustuni)
        xitoy_vazn = None
        if vazn_i is not None and len(row) > vazn_i:
            try:
                v = row[vazn_i]
                xv = float(v) if v else 0
                if xv > 0:
                    xitoy_vazn = xv
            except (ValueError, TypeError):
                pass
        try:
            mq = float(miqdor) if miqdor else 0
        except (ValueError, TypeError):
            continue
        try:
            ombor_mq = float(ombor_v) if ombor_v else 0
        except (ValueError, TypeError):
            ombor_mq = 0
        if mq <= 0 and ombor_mq <= 0:
            continue

        # 1. XITOY_NOM_MAP dan qidirish
        map_key = f"{spec_s} [{len_raw}]"
        if map_key in XITOY_NOM_MAP:
            nom = normalize_product_name(XITOY_NOM_MAP[map_key])
            if mq > 0:
                known[nom] = known.get(nom, 0) + mq
            if ombor_mq > 0:
                ombor_known[nom] = ombor_known.get(nom, 0) + ombor_mq
            if xitoy_vazn and nom not in vazn_map:
                vazn_map[nom] = xitoy_vazn
            continue

        # 2. Avtomatik parser
        inventar_nom = _china_spec_to_inventar(spec_s, len_raw)
        if inventar_nom:
            # K ustun (buyurtma) — snap YO'Q, nom o'zgarmaydi
            nom_k = normalize_product_name(inventar_nom)
            # L ustun (yuklatish) — snap BOR, inventarga moslanadi
            nom_l = normalize_product_name(_inventar_snap(inventar_nom, _get_inventar_set()))
            if mq > 0:
                known[nom_k] = known.get(nom_k, 0) + mq
            if ombor_mq > 0:
                ombor_known[nom_l] = ombor_known.get(nom_l, 0) + ombor_mq
            # Vaznni ikkala nom uchun ham saqlaymiz
            if xitoy_vazn:
                if nom_l not in vazn_map:
                    vazn_map[nom_l] = xitoy_vazn
                if nom_k not in vazn_map:
                    vazn_map[nom_k] = xitoy_vazn
        else:
            if mq > 0 or ombor_mq > 0:
                unknown.append(map_key)
    return known, unknown, ombor_known, vazn_map


def _parse_list_xitoy(rows: list) -> dict:
    """克力木 Лист format.
    Sarlavhalar har xil satrda bo'lishi mumkin (0-5), shu sababli skan qilamiz.
    """
    from common import normalize_product_name

    # Sarlavha satrini topish
    hdr_i = None
    h0 = []
    best_score = 0
    for ri in range(min(10, len(rows))):
        candidate = [str(c).strip() if c else "" for c in rows[ri]]
        score = sum(1 for h in candidate if any(sig in h for sig in ('品号', '数量', '规格', '颜色')))
        if score > best_score:
            best_score = score
            h0 = candidate
            hdr_i = ri
    if hdr_i is None:
        return {}, []

    pin_hao_i = next((i for i, h in enumerate(h0) if '品号' in h), None)
    yan_se_i  = next((i for i, h in enumerate(h0) if '颜色' in h), None)
    gui_ge_i  = next((i for i, h in enumerate(h0) if '规格' in h), None)
    shu_i     = next((i for i, h in enumerate(h0) if '数量' in h), None)
    if gui_ge_i is None or shu_i is None:
        return {}, []

    RANG = {
        '8K钛金': 'Голд', '精磨8K': 'Глянцевый', '8K黑钛': 'Кора',
        '8K': 'Глянцевый', '砂板（工业面）': 'Матовый', '砂板': 'Матовый',
    }
    TOTAL_KEYS = ('小计', '合计', '序号', '规格', '品号')
    known = {}
    unknown = []
    for row in rows[hdr_i + 1:]:
        nlen = len(row)
        gui_ge = row[gui_ge_i] if gui_ge_i < nlen else None
        if not gui_ge:
            continue
        gui_ge_s = str(gui_ge).strip()
        if any(k in gui_ge_s for k in TOTAL_KEYS):
            continue
        m = re.match(r'^([\d\.]+)[*×xX]([\d\.]+)[*×xX]([\d\.]+)', gui_ge_s)
        if not m:
            # Haqiqiy mahsulot satrini tekshirish: pin_hao kodi bor bo'lsa — noma'lum
            ph_check = row[pin_hao_i] if pin_hao_i is not None and pin_hao_i < nlen else None
            if ph_check and str(ph_check).strip():
                unknown.append(gui_ge_s)
            continue
        try:
            mq = float(row[shu_i]) if shu_i < nlen and row[shu_i] else 0
        except (ValueError, TypeError):
            continue
        if mq <= 0:
            continue
        # Qalinlik: 1.00 → '1,0', 0.60 → '0,6'
        s = f"{float(m.group(1)):.2f}".rstrip('0')
        if s.endswith('.'):
            s += '0'
        qalinlik_s = s.replace('.', ',')
        # O'lcham (10 ga yaxlitlash)
        en  = math.ceil(float(m.group(2)) / 10) * 10
        boy = math.ceil(float(m.group(3)) / 10) * 10
        # Rang
        yan_se_val = row[yan_se_i] if yan_se_i is not None and yan_se_i < nlen else None
        yan_se = str(yan_se_val).strip() if yan_se_val else ''
        rang = ''
        for key in ('8K钛金', '精磨8K', '8K黑钛', '8K', '砂板（工业面）', '砂板'):
            if key in yan_se:
                rang = RANG[key]
                break
        # Marka
        marka = ''
        ph_val = row[pin_hao_i] if pin_hao_i is not None and pin_hao_i < nlen else None
        if ph_val:
            pm = re.match(r'^(\d+)', str(ph_val))
            if pm:
                marka = pm.group(1)
        name = f"Лист-{qalinlik_s} ({en}х{boy})"
        if rang:
            name += f" ({rang})"
        if marka:
            name += f" ({marka} марка)"
        nom = normalize_product_name(name)
        known[nom] = known.get(nom, 0) + mq
    return known, unknown



def _fix_oddiy_nom(nom: str, inventar_set: set) -> str:
    """Oddiy format (Товар|Миқдор) tovar nomini inventar nomiga moslashtiradi.
    1. normalize_product_name
    2. Stenka Xitoy→inventar konvertatsiya (_yaxlitla_stenka)
    3. _inventar_snap (0,65↔0,7 / 1,35↔1,4 / 1,45↔1,5 kabi borderline qiymatlar)
    4. Prefix qidiruv — uzunlik "(6 м)" kabi farqlarni hal qiladi
       Har bir stenka kandidati uchun ham urinadi (0,65 uchun 0,7 va 0,65 ikkalasini ham).
    """
    from common import normalize_product_name
    nom = normalize_product_name(str(nom).strip())
    if not inventar_set:
        return nom
    if nom in inventar_set:
        return nom

    # Stenka konvertatsiya: 0.85→0,9 / 2.95→3,0 va hokazo
    m_st = re.search(r'(ст\s+)([\d,\.]+)', nom)
    if m_st:
        orig_st = m_st.group(2)
        new_st = _yaxlitla_stenka(orig_st)
        if new_st != orig_st.replace('.', ',') and new_st != orig_st:
            nom = nom[:m_st.start(2)] + new_st + nom[m_st.end(2):]
    if nom in inventar_set:
        return nom

    # Borderline stenka uchun bir nechta kandidat sinab ko'ramiz
    # (0,65 → avval 0,7, keyin 0,65; 1,35 → avval 1,35, keyin 1,4; etc.)
    _CANDIDATES: dict[str, list[str]] = {
        '0,65': ['0,7',  '0,65'],
        '0,70': ['0,7'],
        '1,35': ['1,35', '1,4'],
        '1,45': ['1,45', '1,5'],
    }
    marka_m  = re.search(r'\(\d+\s*марка\)$', nom.strip())
    marka_sfx = marka_m.group(0) if marka_m else ''
    m_st2 = re.search(r'ст\s+([\d,\.]+)', nom)
    st_val = m_st2.group(1) if m_st2 else None
    # Kandidatlar: borderline bo'lsa ko'p, aks holda faqat mavjud qiymat
    cands = _CANDIDATES.get(st_val, [st_val] if st_val else [])

    def _prefix_search(test_nom: str) -> str | None:
        """test_nom ni avval aniq, keyin prefix qidiruv bilan topadi."""
        if test_nom in inventar_set:
            return test_nom
        m3 = re.search(r'ст\s+[\d,\.]+', test_nom)
        if not m3:
            return None
        after = test_nom[m3.end():]
        pp    = after.find('(')
        prefix = test_nom[:m3.end() + pp].rstrip() if pp >= 0 else test_nom[:m3.end()].rstrip()
        if marka_sfx:
            return next((x for x in inventar_set
                         if x.startswith(prefix + ' (') and x.endswith(marka_sfx)), None)
        else:
            return next((x for x in inventar_set
                         if x.startswith(prefix + ' (')), None)

    for c in cands:
        candidate_nom = nom if not m_st2 else (
            nom[:m_st2.start(1)] + c + nom[m_st2.end(1):]
        )
        found = _prefix_search(candidate_nom)
        if found:
            return found

    # SO'NGGI CHORA: aniq/stenka-kandidat qidiruvlarning barchasi
    # muvaffaqiyatsiz bo'lsa ham, inventarning O'ZIDA g'alati bo'shliq bilan
    # yozilgan bo'lishi mumkin (masalan "Лист- 2,0..." chiziqchadan keyin
    # bo'shliq bilan, 2026-07-06'da haqiqiy inventardan topilgan) — kanonik
    # (bo'shliqqa chidamli) solishtirish shuni ham ushlab qoladi. Topilsa —
    # INVENTARDAGI ASL matn qaytariladi, biznikidan tuzatilgani emas.
    kanonik_topilgan = _inventardan_moslashtir(nom, inventar_set)
    if kanonik_topilgan != nom:
        return kanonik_topilgan

    return nom


_AI_NOM_SIGNALS  = ('tovar', 'mahsulot')
_AI_ZAKAZ_SIGNALS = ('zakaz', '(k)')
_AI_TAYYOR_SIGNALS = ('tayyor', '(l)')
_AI_JAMI_SIGNALS = ('jami', 'итого', 'всего', 'total', 'сумма')


def ai_ostatka_fayl_mi(fayl_bytes: bytes) -> bool:
    """
    Fayl Claude (AI yordamchi) tomonidan OLDINDAN TARJIMA qilingan ostatka
    formatidami? Belgisi: har qanday varaqda "Tovar nomi" + ("Zakaz"/"(K)"
    yoki "Tayyor"/"(L)") lotin sarlavhalari bor — xom Xitoycha ierogliflar
    (品号/规格/库存/订单 va h.k.) YO'Q, chunki AI ularni allaqachon inventar
    formatiga o'tkazib bergan.
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(BytesIO(fayl_bytes), data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True):
                if not row:
                    continue
                cand = [str(c).strip().lower() if c else "" for c in row]
                has_nom = any(any(sig in h for sig in _AI_NOM_SIGNALS) for h in cand)
                has_kl  = any(any(sig in h for sig in _AI_ZAKAZ_SIGNALS + _AI_TAYYOR_SIGNALS) for h in cand)
                if has_nom and has_kl:
                    return True
        return False
    except Exception:
        return False


def ai_ostatka_fayl_oqi(fayl_bytes: bytes) -> tuple:
    """
    Claude (AI yordamchi Chat 1) tomonidan tarjima qilingan ostatka faylini
    o'qiydi — bitta yoki bir nechta varaq (mas. "Труба-Профиль" + "Лист"),
    ustunlar: Tovar nomi | Zakaz (K) | Tayyor (L) | 1 dona vazni (kg) | Izoh.

    Qoidalar (Huzayfa 2026-07-09 tasdiqlagan):
      - Bir xil nom bir necha qatorda uchrasa (turli partiya/narx/sana) —
        BU XATO EMAS, K/L QO'SHIB (summalab) olinadi.
      - "Jami"/"Итого" kabi jamlovchi qatorlar o'tkazib yuboriladi.
      - NOANIQ (nom ichida "NOANIQ:" yoki Izoh ustunida "NOANIQ"/"DIQQAT")
        belgilangan qatorlar HAM qabul qilinadi (bloklanmaydi), lekin
        unknown_list ga qo'shiladi — foydalanuvchiga alohida ogohlantirish
        sifatida ko'rsatiladi.
      - Har bir nom HAQIQIY inventar bilan _fix_oddiy_nom() orqali
        solishtiriladi (AI tarjimasiga ko'r-ko'rona ishonilmaydi) — bu
        bo'shliq/format farqlarini (masalan "Лист- 0,8" vs "Лист-0,8")
        avtomatik tuzatadi. Inventarda topilmasa — YANGI TOVAR bo'lishi
        mumkin (bloklanmaydi), faqat unknown_list ga qo'shiladi.

    Qaytaradi: (ok, xato, xitoy_map, unknown_list, ombor_map, vazn_map)
      — xitoy_ostatka_oqi() bilan AYNAN BIR XIL shakl (drop-in almashtirish).
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(BytesIO(fayl_bytes), data_only=True)
    except Exception as e:
        return False, f"Faylni ochishda xato: {type(e).__name__}: {e}", {}, [], {}, {}

    inv_set = _get_inventar_set()
    xitoy_map: dict = {}
    ombor_map: dict = {}
    vazn_map: dict = {}
    unknown: list = []

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        hdr_row = nom_i = k_i = l_i = vazn_i = izoh_i = None
        for ri in range(min(5, len(rows))):
            row = rows[ri]
            if not row:
                continue
            cand = [str(c).strip().lower() if c else "" for c in row]
            n_i = next((i for i, h in enumerate(cand) if any(sig in h for sig in _AI_NOM_SIGNALS)), None)
            kk_i = next((i for i, h in enumerate(cand) if any(sig in h for sig in _AI_ZAKAZ_SIGNALS)), None)
            ll_i = next((i for i, h in enumerate(cand) if any(sig in h for sig in _AI_TAYYOR_SIGNALS)), None)
            if n_i is not None and (kk_i is not None or ll_i is not None):
                hdr_row, nom_i, k_i, l_i = ri, n_i, kk_i, ll_i
                vazn_i = next((i for i, h in enumerate(cand) if 'vazn' in h or 'kg' in h), None)
                izoh_i = next((i for i, h in enumerate(cand) if 'izoh' in h or 'comment' in h), None)
                break
        if hdr_row is None:
            continue  # bu varaq AI-ostatka formatiga mos emas — o'tkazib yuboriladi

        for row in rows[hdr_row + 1:]:
            if not row or nom_i >= len(row):
                continue
            nom_raw = row[nom_i]
            if not nom_raw or not str(nom_raw).strip():
                continue
            nom_s = str(nom_raw).strip()
            low = nom_s.lower()
            if any(k in low for k in _AI_JAMI_SIGNALS):
                continue  # "Jami"/"Итого" qatori

            izoh_val = ""
            if izoh_i is not None and izoh_i < len(row) and row[izoh_i]:
                izoh_val = str(row[izoh_i]).strip()
            is_noaniq = 'NOANIQ' in nom_s.upper() or 'NOANIQ' in izoh_val.upper() or 'DIQQAT' in izoh_val.upper()

            # "NOANIQ: {xom matn}" yoki "[NOANIQ ...]" qismini nomdan tozalash
            nom_clean = re.sub(r'^NOANIQ:\s*', '', nom_s, flags=re.IGNORECASE).strip()
            nom_clean = re.sub(r'\s*[\[\(]NOANIQ.*?[\]\)]\s*', '', nom_clean, flags=re.IGNORECASE).strip()
            if not nom_clean:
                nom_clean = nom_s

            def _to_float(v):
                if v is None or v == "":
                    return 0.0
                try:
                    return float(str(v).replace(',', '.').replace(' ', ''))
                except (ValueError, TypeError):
                    return 0.0

            k_val = _to_float(row[k_i]) if k_i is not None and k_i < len(row) else 0.0
            l_val = _to_float(row[l_i]) if l_i is not None and l_i < len(row) else 0.0
            if k_val <= 0 and l_val <= 0:
                continue
            v_val = _to_float(row[vazn_i]) if vazn_i is not None and vazn_i < len(row) else 0.0

            # Inventar bilan solishtirish — bo'shliq/format farqlariga chidamli
            nom_final = _fix_oddiy_nom(nom_clean, inv_set)
            if is_noaniq or nom_final not in inv_set:
                if nom_final not in unknown:
                    unknown.append(nom_final)

            if k_val > 0:
                xitoy_map[nom_final] = xitoy_map.get(nom_final, 0) + k_val
            if l_val > 0:
                ombor_map[nom_final] = ombor_map.get(nom_final, 0) + l_val
            if v_val and nom_final not in vazn_map:
                vazn_map[nom_final] = v_val

    if not xitoy_map and not ombor_map:
        return False, (
            "Faylda tovar topilmadi. Ustunlarni tekshiring: 'Tovar nomi' "
            "+ 'Zakaz (K)'/'Tayyor (L)' bo'lishi shart."
        ), {}, [], {}, {}

    return True, None, xitoy_map, unknown, ombor_map, vazn_map


def xitoy_ostatka_oqi(fayl_bytes: bytes) -> tuple:
    """
    Xitoy ostatka Excel faylini o'qiydi.
    Qaytaradi: (ok, xato, known_map, unknown_list)
      known_map    — {inventar_nomi: miqdor}
      unknown_list — parser tanimagan yoki XITOY_NOM_MAP da yo'q spec'lar
    """
    import openpyxl
    try:
        wb   = openpyxl.load_workbook(BytesIO(fayl_bytes), data_only=True)
        ws   = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return False, "Fayl bo'sh", {}, [], {}, {}

        # Sarlavha qatorini aniqlash: eng ko'p ustun kalit so'zi bo'lgan qatorni tanlaymiz
        # (birinchi signal emas, chunki "库存明细" kabi sarlavhalar ham signal beradi)
        ALL_SIGNALS = ('品号', '颜色', '数量', '规格', '长度', '库存', 'Ombor', 'Uzbek', '订单')
        # Ko'p-signal signallari faqat haqiqiy header qatorida bo'ladi
        STRICT_SIGNALS = ('品号', '颜色', '数量', '规格', '长度', '库存', 'Ombor', 'Uzbek', '订单')
        h0 = []; hdr_row = 0; first_nonempty = None
        best_score = 0
        for ri in range(min(10, len(rows))):
            candidate = [str(c).strip() if c else "" for c in rows[ri]]
            has_any = any(c for c in candidate)
            if not has_any:
                continue
            if first_nonempty is None:
                first_nonempty = (ri, candidate)
            # Signal soni — nechta alohida katak signal o'z ichiga oladi
            score = sum(1 for h in candidate if any(sig in h for sig in STRICT_SIGNALS))
            if score > best_score:
                best_score = score
                h0 = candidate
                hdr_row = ri
        # Agar hech qanday signal topilmasa — birinchi bo'sh bo'lmagan qator
        if best_score == 0 and first_nonempty is not None:
            hdr_row, h0 = first_nonempty

        # ── 1. Лист xitoy format
        # 品号=mahsulot kodi, 颜色=rang, 数量=miqdor, 规格=spec (t*en*boy formatida)
        LIST_SIGNALS = ('品号', '颜色', '数量')
        is_list_fmt = any(sig in h for h in h0 for sig in LIST_SIGNALS)
        # Agar 规格 bor lekin 长度 yo'q bo'lsa ham Лист format (t*en*boy formatli)
        has_gui_ge  = any('规格' in h for h in h0)
        has_chang_du= any('长度' in h for h in h0)
        if not is_list_fmt and has_gui_ge and not has_chang_du:
            is_list_fmt = True
        if is_list_fmt:
            known, unknown = _parse_list_xitoy(rows)
            if known or unknown:
                # Лист uchun alohida ombor yo'q — known == ombor
                return True, None, known, unknown, dict(known), {}
            # Debug: nima topilganini ko'rsatish
            debug = f"h0={h0[:5]} | score={best_score} | rows={len(rows)}"
            return False, f"Лист formatida ma'lumot topilmadi.\nDebug: {debug[:200]}", {}, [], {}, {}

        # ── 2. Труба/Профиль xitoy format
        ombor_i  = next((i for i, h in enumerate(h0) if 'Ombor' in h or '库存' in h), None)
        zakaz_i  = next((i for i, h in enumerate(h0) if 'Uzbek' in h or '订单' in h), None)
        gui_ge_i = next((i for i, h in enumerate(h0) if '规格' in h), None)
        len_i    = next((i for i, h in enumerate(h0) if '长度' in h), None)
        vazn_i   = next((i for i, h in enumerate(h0) if '支重' in h), None)
        if ombor_i is not None and gui_ge_i is not None and len_i is not None:
            known, unknown, ombor_known, vazn_map = _parse_truba_profil_xitoy(
                rows, ombor_i, gui_ge_i, len_i, zakaz_i, vazn_i
            )
            if known or unknown or ombor_known:
                return True, None, known, unknown, ombor_known, vazn_map
            return False, "Труба/Профиль formatida ma'lumot topilmadi", {}, [], {}, {}

        # ── 3. Oddiy format: Товар | Миқдор
        h0_lower = [h.lower() for h in h0]
        tovar_i  = next((i for i, h in enumerate(h0_lower) if 'товар' in h or 'tovar' in h), None)
        miqdor_i = next((i for i, h in enumerate(h0_lower)
                         if any(k in h for k in ('миқдор', 'miqd', 'кол', 'qty', 'amount', 'сон'))), None)
        if tovar_i is None:
            return False, "'Tovar' ustuni topilmadi", {}, [], {}, {}
        if miqdor_i is None:
            miqdor_i = 1 if len(h0) > 1 and tovar_i != 1 else (0 if tovar_i != 0 else None)
        if miqdor_i is None:
            return False, "'Miqdor' ustuni topilmadi", {}, [], {}, {}

        # Inventar to'plamini yuklaymiz
        inv_set = _get_inventar_set()
        known = {}
        for row in rows[hdr_row + 1:]:
            tovar = row[tovar_i] if tovar_i < len(row) else None
            mq    = row[miqdor_i] if miqdor_i < len(row) else None
            if not tovar: continue
            try:
                mv = float(mq) if mq is not None else 0
                if mv > 0:
                    nom = _fix_oddiy_nom(str(tovar), inv_set)
                    known[nom] = known.get(nom, 0) + mv
            except (ValueError, TypeError):
                continue
        return True, None, known, [], {}, {}
    except Exception as e:
        return False, str(e), {}, [], {}, {}
