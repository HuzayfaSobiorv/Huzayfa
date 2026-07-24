"""
common.py — Umumiy yordamchi modul
====================================
Bu fayl NEJAVIYKA_v3.py va min_cache_tizimi.py tomonidan import qilinadi.
Normalize, kategoriya, himoya foizi kabi BARCHA umumiy mantiq shu yerda.

MUHIM: Ikki faylda bir xil normalize ishlatilishi shart, aks holda
       cache kalitlari qoldiq kalitlariga mos kelmaydi.
====================================
"""

import re
import math
import os
import json
import pandas as pd
from datetime import datetime
import warnings

warnings.filterwarnings('ignore')


# ============================================================
# ATOMIC JSON YOZISH — holat fayllari buzilmasligi uchun
# ============================================================

def atomic_json_write(path, data, **dumps_kwargs) -> None:
    """JSON ni ATOMIC yozadi: avval vaqtinchalik .tmp faylga, keyin
    os.replace bilan asl fayl ustiga almashtiradi.

    Nima uchun: bot yozish o'rtasida o'chsa (svet, crash, Ctrl+C),
    to'g'ridan write_text yarim yozilgan/buzilgan JSON qoldiradi va
    buyurtma/tarix ma'lumotlari yo'qoladi. os.replace esa OS darajasida
    atomic — fayl yo eski, yo to'liq yangi holatda bo'ladi.

    Ishlatish: atomic_json_write(p, data, ensure_ascii=False, indent=2)
    """
    dumps_kwargs.setdefault("ensure_ascii", False)
    path = str(path)
    tmp = f"{path}.tmp{os.getpid()}"
    try:
        with open(tmp, "w", encoding="utf-8") as f:
            f.write(json.dumps(data, **dumps_kwargs))
            f.flush()
            os.fsync(f.fileno())
        os.replace(tmp, path)
    except BaseException:
        # tmp qoldiq qolmasin
        try:
            os.remove(tmp)
        except OSError:
            pass
        raise

# ============================================================
# KONSTANTALAR — faqat shu yerda, boshqa joyda takrorlanmaydi
# ============================================================

KELISH_KUNI     = 55    # ◄ ASOSIY: konteyner yetib kelish vaqti (kun) — FAQAT SHU YERDA O'ZGARTIR
YOLDA_KUN       = KELISH_KUNI   # alias (eski kod uchun)

# 2026-07-09: Huzayfa bilan kelishilgan yagona qoida — "kunlik sotuv"ni
# min_zaxiradan qayta chiqarishda HAMMA YERDA shu songa bo'linadi (buyurtma,
# kamomat holat, qidiruv/grafik — bittasi ham boshqacha son ishlatmasin).
# DIQQAT: bu KELISH_KUNI (55, haqiqiy yetib kelish vaqti) bilan ADASHTIRILMASIN
# — u alohida, jismoniy konstanta, o'zgarishsiz qoladi.
#
# 2026-07-21 (Huzayfa bilan kelishildi — "juda ko'p buyurtma beryapti"
# shikoyati tekshirilgach): tarix/*.xlsx dagi haqiqiy kunlik sotuv bilan
# solishtirilganda 30ga bo'lish deyarli hamma tovarda real sotuvdan
# 1,6-2,9 baravar (ba'zan ko'proq) katta "kunlik" chiqarardi -- natijada
# buyurtmalar shishib ketardi. Har tovar uchun ALOHIDA real-tarix hisobi
# (real_kunlik_sotuv.py) sinovdan o'tkazildi va ishlagan bo'lsa-da,
# Huzayfa buni ORTGA QAYTARDI (murakkab/xavfli, git revertlandi) va
# o'rniga ODDIY, universal tuzatishni tanladi: 30 o'rniga 45ga bo'lish.
# Bu 5ta sinov tovarida o'rtacha xatoni ~117%dan ~45%gacha kamaytiradi
# (mukammal emas, lekin bir qatorlik, past xavfli tuzatish).
KUNLIK_SOTUV_BOLISH = 45
# 2026-07-14 (Huzayfa bilan kelishildi, "order-up-to" mantiq): buyurtma
# berilganda zaxira min ustiga yana necha KUNlik savdoga yetadigan qilib
# to'ldiriladi. Katta son = kam, lekin yirik buyurtmalar. Bu kamomat_engine
# .zanjir_sim'da ishlatiladi — mayda (50-200 talik) takliflar o'rniga
# tovar oyiga ~1 marta, ~1 oylik hajmda chiqadi.
BUYURTMA_SIKL_KUN = 30
FAST_KUN        = 20    # tezkor (fayl nomi "F_" bilan boshlanadigan) konteyner yetib kelish vaqti (kun)
# "12 metrlik" konteyner — ISO'siz, ko'pincha mashina-raqami (masalan
# "ME5312") bilan yozilgan, tarkibidagi tovarlarning KO'PCHILIGI 6 metrlik
# (Труба/Профиль uzunligi "(6 м)") bo'lgan yetkazmalar (2026-07-08 qo'shildi
# — qarang main.py::_konteyner_12m_mi()).
M12_KUN         = 45
TREND_KUN       = KELISH_KUNI   # necha kun barqaror bo'lsa min oshiriladi
SLIDING_WINDOW  = 30    # oxirgi necha kunni kuzatish
SEZON_OYNA      = 7     # sezon taqqoslash oynasi (kun)
MIN_KUZATUV_KUN = 7     # yangi tovar uchun kamida necha kun kuzatish kerak

# Kategoriya bo'yicha himoya yostig'i (buffer foizi)
HIMOYA_FOIZ: dict[str, float] = {
    'ТРУБА':        0.50,
    'ПРОФИЛЬ':      0.50,
    'ЛИСТ':         0.15,
    'ЛИСТ РУЛОН':   0.15,
}
HIMOYA_DEFAULT = 0.50

# Kategoriya tartib raqami (Power BI da saralash uchun)
KATEGORIYA_TARTIB: dict[str, int] = {
    'ТРУБА (201)':    1,  'ТРУБА (304)':    2,  'ТРУБА (430)':    3,
    'ТРУБА (316)':    4,  'ТРУБА (Бесшовный)': 5, 'ТРУБА':         6,
    'ПРОФИЛЬ (201)':  7,  'ПРОФИЛЬ (304)':  8,  'ПРОФИЛЬ (430)':  9,
    'ПРОФИЛЬ (316)': 10,  'ПРОФИЛЬ':       11,
    'ЛИСТ (201)':    12,  'ЛИСТ (304)':    13,  'ЛИСТ (430)':    14,
    'ЛИСТ (316)':    15,  'ЛИСТ':          16,
    'БАЛАСИНА':      17,  'СТОЙКА':        18,  'ЧАШКА':         19,
    'ОТВОД':         20,  'СОККА':         21,  'ШАР':           22,
    'СОВУН':         23,  'БОШҚА':         24,  'ЛИСТ РУЛОН':    99,
}

# Hisobot uchun filtrlanadigan kalit so'zlar
KERAKSIZ_QISM = [
    'тумба', 'ХС №', 'заказ', 'Гардероб', 'консол',
    'ясаб', 'учун', 'дар', 'берилди', 'вешалка', 'ди',
    'стеллаж', 'ЛС-', 'забор', 'Ошхона', 'подставка',
    'Сори', 'степлер', 'кронштейн', 'Стул', 'Сушилка',
    'Табуретка', 'Ю/С', 'гич', 'ЛК-', 'Казиро', 'Германтин', 'тувак', 'Декоративни',
    '100х100-Чашка-02 (Голд)',
    '100х100-Чашка-01 (Голд)',
    '100х100-Чашка-01 (Глянцевый)',
    '100х100-Чашка-02 (Глянцевый)',
    '100х100-Чашка-01 (Матовый)',
    '100х100-Чашка-02 (Матовый)',
    'Декоративний', 'Детский качале', 'Стелаж', 'Умвалик', 'Баласина (Балик)', 'Брак баласина',
    'Фланс','S-гул ','БП','Дрель','Дуга (Сотув булми)','Зажим ', 'Заклёпка','Итого','КР-001 (210х100)',
    'Кабр','Казерог ','Каска','Катта гул','Кислород балон','Контейнер','Краска. Голд','Кресло', 
    'Кук материал','Кучага хизмат сварка хизмати','Ламинад Д=1700','Латок','Ловия',
    'Малочни материал','Масжид','Маска','(Сотув булими)',
    'Мойка 2600х80х40 (Сотув булмига)','Мешалка','Ножка 900Х650Х45 (Сотув булмига) Голд','Нукус сахна','сахна','Обувница (карши)Голд',
    'карши','бухоро','Ойна (кора)','Петля',' (марям)','урожайнига','Тахта','Очки №1','Перфоратор',
    'Перегаротка 210х50 (Ташкент) Голд','Перегароткa',
    'Петля','Перчатка','Тележка','Фанер',
]

KERAKSIZ_ALON = ['стол']   # to'liq so'z sifatida qidiriladi

# Har doim o'tsin — filtrlanmasin
KERAKSIZ_ISTISNO = [
    'стул ойокчаси №04',
    'стул ойокчаси №05',
]


# ============================================================
# NORMALIZE — bu funksiya ikkala faylda bir xil ishlatiladi
# ============================================================

def normalize_product_name(name: str) -> str:
    if pd.isna(name):
        return name

    name = str(name).strip()
    name = ' '.join(name.split())

    def _add_m(match):
        content = match.group(1)
        if 'м' in content:
            return match.group(0)
        if any(c.isdigit() for c in content):
            return f"({content} м)"
        return match.group(0)

    name = re.sub(r'\(([0-9,\.]+)\)(?!\s*марка)', _add_m, name)
    # 2026-07-11 (Huzayfa: "51 0.9 304 bizda bor, bot tanimayapti"): Tarix
    # qatorlarida ba'zan uzunlik "м" raqamga YOPISHIB yoziladi (masalan
    # "(6м)" -- bo'shliqsiz), yuqoridagi _add_m buni ko'rmaydi ('м' allaqachon
    # bor deb hisoblab o'zgartirmaydi) -- natijada bu nom haqiqiy inventar
    # ("6 м", bo'shliq bilan) bilan mos kelmay, ЯНГИ deb noto'g'ri
    # belgilanardi. Endi raqam bilan "м" orasiga bo'shliq qo'yiladi.
    name = re.sub(r'(\d)м\)', r'\1 м)', name)
    name = re.sub(r'м\s+\)', 'м)', name)
    name = re.sub(r'Лист-(\d)', r'Лист- \1', name)
    name = ' '.join(name.split())
    return name


# ============================================================
# KATEGORIYA ANIQLASH
# ============================================================

def get_category(name) -> str:
    if pd.isna(name):
        return 'БОШҚА'

    s  = str(name).strip()
    sl = s.lower()

    if re.match(r'^(\([^)]*\)\s*)?Ф-\d+', s) and 'ст' in s and re.search(r'\(\d+.*м\)', s):
        return 'ТРУБА'
    if re.match(r'^.*Пр\.\s+\d+х\d+', s) and 'ст' in s and re.search(r'\(\d+.*м\)', s):
        return 'ПРОФИЛЬ'
    if re.match(r'^Лист-\s*\d+', s) and re.search(r'\(\d+х\d+\)', s):
        return 'ЛИСТ'
    if re.match(r'^Лист\s+рулон', s):
        return 'ЛИСТ РУЛОН'

    keyword_map = [
        ('баласина', 'БАЛАСИНА'),
        ('стойка',   'СТОЙКА'),
        ('сокка',    'СОККА'),
        ('шар',      'ШАР'),
        ('отвод',    'ОТВОД'),
        ('кузикорин','КУЗИКОРИН'),
        ('чашка',    'ЧАШКА'),
        ('совун',    'СОВУН'),
    ]
    for keyword, category in keyword_map:
        if keyword in sl:
            return category

    return 'БОШҚА'


def get_marka(name) -> str:
    if pd.isna(name):
        return ''
    match = re.search(r'\((\d+)\s*марка\)', str(name))
    return match.group(1) if match else ''


def get_category_with_marka(name) -> str:
    kat   = get_category(name)
    marka = get_marka(name)
    if marka and kat in ('ТРУБА', 'ПРОФИЛЬ', 'ЛИСТ', 'ЛИСТ РУЛОН'):
        return f"{kat} ({marka})"
    return kat


def get_category_order(kategoriya: str) -> int:
    return KATEGORIYA_TARTIB.get(kategoriya, 99)


# ============================================================
# HIMOYA FOIZI VA MIN HISOBLASH
# ============================================================

def get_himoya_foiz(kategoriya: str) -> float:
    return HIMOYA_FOIZ.get(kategoriya, HIMOYA_DEFAULT)


def yaxlitla_50(qiymat: float) -> int:
    if qiymat <= 0:
        return 0
    return math.ceil(qiymat / 50) * 50


def hisobla_min_zaxira(kunlik_istemol: float, kategoriya: str) -> int:
    foiz = get_himoya_foiz(kategoriya)
    return yaxlitla_50(kunlik_istemol * YOLDA_KUN * (1 + foiz))


def min_dan_kunlik_chiqar(min_zaxira: float, kategoriya: str) -> float:
    foiz = get_himoya_foiz(kategoriya)
    denom = YOLDA_KUN * (1 + foiz)
    if denom <= 0:
        return 0.0
    return round(min_zaxira / denom, 3)


# ============================================================
# QOLDIQ FAYLINI YUKLASH
# ============================================================

def parse_qoldiq_str(text) -> int:
    if pd.isna(text):
        return 0
    parts = str(text).strip().split('/')
    raw = parts[0].strip().replace(',', '').replace(' ', '')
    try:
        return int(raw) if raw and raw != '-' else 0
    except ValueError:
        return 0


def load_qoldiq_file(filepath: str) -> pd.DataFrame:
    if _yangi_tarix_formati_mi(filepath):
        return _load_qoldiq_yangi_format(filepath)
    df = pd.read_excel(filepath, header=4)
    col_count = len(df.columns)
    mid_cols = [f'_Col{i}' for i in range(1, col_count - 2)]
    df.columns = ['Mahsulot'] + mid_cols + ['Qoldiq_Str', 'Qoldiq_Summa']
    df = df[df['Mahsulot'].notna()].copy()
    df = df[df['Mahsulot'] != 'Товар'].copy()
    df['Qoldiq_Dona']         = df['Qoldiq_Str'].apply(parse_qoldiq_str)
    df['Qoldiq_Summa']        = pd.to_numeric(df['Qoldiq_Summa'], errors='coerce').fillna(0)
    df['Mahsulot_Normalized'] = df['Mahsulot'].apply(normalize_product_name)
    # 2026-07-24: eski format faqat BITTA (butun kompaniya) jami beradi —
    # kanal bo'yicha ajratib bo'lmaydi. Shu sabab uchala kanal ham xuddi
    # shu qiymatni oladi (bu YANGI regressiya EMAS — tizim ilgari ham
    # doim shu bitta jami sonni ishlatgan; yangi formatda esa endi
    # HAQIQIY ajratish mumkin bo'ladi, pastdagi _load_qoldiq_yangi_format'ga
    # qarang).
    df['Asosiy_Qoldiq_Dona'] = df['Qoldiq_Dona']
    df['Cex_Qoldiq_Dona']    = df['Qoldiq_Dona']
    df['Osh_Qoldiq_Dona']    = df['Qoldiq_Dona']
    return df[['Mahsulot', 'Mahsulot_Normalized', 'Qoldiq_Dona', 'Qoldiq_Summa',
               'Asosiy_Qoldiq_Dona', 'Cex_Qoldiq_Dona', 'Osh_Qoldiq_Dona']].copy()


# ============================================================
# YANGI TARIX FORMATI (2026-07-24 dan) — har filial/ombor
# alohida ustun juftligida ("кор /дона" + "жами"), bitta faylda.
# Huzayfa bilan kelishilgan YAKUNIY qoida (2026-07-24, uchinchi bosqich —
# har kanal ENDI O'Z ALOHIDA jismoniy zaxirasidan hisoblanishi kerak,
# UMUMIY jamlama YETARLI EMAS ekani aniqlangach — "Asosiyda buyurtma
# yozsak, u O'shning/Tsexning qoldig'ini hisobga olmasligi kerak"):
#   - "Ош..." bilan boshlanadigan ustunlar         -> Ош kanaliga
#   - "Промзона (Хомашё)" va
#     "Промзона Транзит (склад)"                   -> Tsex kanaliga
#   - Мебел/Стул/Тумба/Универсал цехlari, "Цех склад (Основной)",
#     "Темур Склад", "Инвентар"/"Инвентарлар омбори", "* аппарат"
#     (masalan Голд аппарат — qurilma hisobi, filial emas),
#     "Сотув булими" (ichki bo'lim hisobi), "* сервис" (xizmat markazi),
#     "Лазер Промзона..." va "* Таййор махсулот"    -> HECH QAYSI kanalga
#     kirmaydi (umuman hisoblanmaydi)
#   - qolgan barcha oddiy filial/ombor ustunlari
#     (Транзит bilan birga)                         -> Асосий kanaliga
#
# Natijada har mahsulot uchun 3 ta ALOHIDA son chiqadi: Asosiy_Qoldiq_Dona/
# Cex_Qoldiq_Dona/Osh_Qoldiq_Dona — main.py bularni Power BI fayliga
# (Асосий_Қолдиқ/Цех_Қолдиқ/Ош_Қолдиқ) chiqaradi, Generate_Asosiy_order.py/
# kamomat_engine.py/ui.py esa endi HAR BIRI FAQAT O'Z kanaliga mos ustundan
# o'qiydi (xuddi Асосий_Захира/Цех_Захира/Ош_Захира kabi). Umumiy
# (kompaniya darajasidagi) Qoldiq_Dona/Qoldiq_Summa ham saqlanadi — faqat
# main.py'dagi umumiy "Холат" ko'rsatkichi uchun, buyurtma hisobiga
# ta'sir qilmaydi.
#
# MUHIM (Huzayfa: "ertaga yangi ustunlar qo'shilib qolishi mumkin,
# dastur chalg'imasligi kerak"): tuzatish ATAYLAB aniq fayl nomlari
# emas, KALIT SO'ZLARGA asoslangan — shu bilan masalan ertaga "Бухоро
# Транзит" degan YANGI ustun qo'shilsa, u hech qaysi skip-kalitga mos
# kelmagani uchun avtomatik TO'G'RI (Асосийga hisobga olinadi holatda)
# ishlaydi, kod o'zgartirish shart emas. Qo'shimcha xavfsizlik pardasi
# sifatida _yangi_lokatsiyalarni_tekshir() — ILGARI ko'rilmagan ustun
# nomi chiqsa shunchaki OGOHLANTIRADI (main.py konsolida), hisoblashni
# to'xtatmaydi — admin format o'zgarishini "sezishi" uchun.
# ============================================================

_YANGI_FORMAT_SKIP_KALIT = (
    'стул', 'тумба', 'универсал', 'мебел',   # mebel sexlari
    'цех склад',                              # Цех склад (Основной)
    'темур склад',                            # shaxsiy/nomlangan ombor
    'инвентар',                               # Инвентар + Инвентарлар омбори
    'аппарат',                                # Голд аппарат — qurilma hisobi
    'сотув булими',                           # ichki savdo bo'limi hisobi
    'сервис',                                 # xizmat markazi (masalan Шахрихон сервис)
    'лазер',                                  # Лазер Промзона* — Tsexga kirmaydi
    'таййор махсулот',                        # tayyor mahsulot omborlari — kirmaydi
)

_LOKATSIYA_KUZATUV_FAYL = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), 'bot_holat', 'tarix_lokatsiyalari.json'
)


def _yangi_tarix_formati_mi(filepath: str) -> bool:
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        ws = wb.active
        return ws.max_column is not None and ws.max_column > 20
    finally:
        wb.close()


def _lokatsiya_kanali(nom) -> str | None:
    """Yangi tarix faylidagi bitta filial/ombor ustuni qaysi kanalga
    tegishli: 'asosiy' | 'sex' | 'osh' | None (hech qaysiga kirmaydi,
    umuman hisoblanmaydi). Tartib muhim: avval skip-kalitlar tekshiriladi
    (шу bilan "Лазер Промзона..."/"* Таййор махсулот" — nomida "промзона"
    bo'lsa ham — Tsexga emas, None'ga tushadi)."""
    s = str(nom).strip().lower()
    if any(k in s for k in _YANGI_FORMAT_SKIP_KALIT):
        return None
    if s.startswith('ош'):
        return 'osh'
    if 'промзона' in s:
        return 'sex'
    return 'asosiy'


# ============================================================
# FILIAL RO'YXATI (2026-07-24, "har filial o'z qoldig'ini ko'rsin"
# funksiyasi uchun) — Huzayfa bilan kelishilgan YAKUNIY ro'yxat:
# 15 ta Асосий filial + Ош + Промзона (Tsex), oxirida Промзона.
# Har filial uchun склад+транзит ustunlari BIRGA hisoblanadi
# (Huzayfa: "Albatta tranzitlar ham qoshib hisoblansin").
# MUHIM: bu ro'yxat _lokatsiya_kanali() bilan mos kelishi SHART —
# har bir xom ustun nomi shu yerda ANIQ bitta filialga tegishli
# bo'lishi kerak (testda tekshiriladi).
# ============================================================

FILIAL_GURUHLARI: dict[str, list[str]] = {
    'Нержавейка склад (Основной)': [
        'Нержавейка склад (Основной)', 'Нержавейка склад (Основной) Транзит',
    ],
    'Сой':          ['Сой (Нержавейка)', 'Сой (Нержавейка) Транзит'],
    'Шахрихон':     ['Шахрихон Нержавейка', 'Шахрихон Нержавейка Транзит'],
    'Тошкент':      ['Тошкент (Нержавейка)', 'Тошкент (Нержавейка) Транзит'],
    'Наманган':     ['Наманган (Нержавейка)', 'Наманган (Нержавейка) Транзит'],
    'Урожайный':    ['Урожайный (Нержавейка)', 'Урожайный (Нержавейка) Транзит'],
    'Маргилон':     ['Маргилон (Нержавейка)', 'Маргилон (Нержавейка) Транзит'],
    'Самарканд':    ['Самарканд (Нержавейка)'],
    'Бухоро':       ['Бухоро (Нержавейка)', 'Бухоро (Нержавейка) Транзит'],
    'Сергили база': ['Сергили база', 'Сергили база Транзит'],
    'Карши':        ['Карши (Нержавека)', 'Карши (Нержавейка) Транзит'],
    'Хоразм':       ['Хоразм (Нержавейка)', 'Хоразм (Нержавейка) Транзит'],
    '37 Склад':     ['37 Склад (Нержавейка)', '37 Склад (Нержавейка) транзит'],
    'Оптом Тошкент':['Оптом Тошкент склад', 'Оптом Тошкент (Транзит)'],
    'Гиздувон':     ['Гиздувон (Нержавейка)', 'Гиздувон (Нержавейка) Транзит'],
    'Ош':           ['Ош (Нержавейка ) склад', 'Ош (Нержавейка ) транзит'],
    'Промзона':     ['Промзона (Хомашё)', 'Промзона Транзит (склад)'],
}

# Ro'yxat tartibda — Huzayfa: "eng oxirida Promzona"
FILIALLAR: list[str] = [f for f in FILIAL_GURUHLARI.keys()]


def _filial_nomi(nom) -> str | None:
    """Xom ustun nomi (masalan 'Урожайный (Нержавейка) Транзит') qaysi
    KANONIK filialga tegishli ekanini topadi. MUHIM: avval _lokatsiya_kanali()
    orqali "bu ustun umuman hisobga olinadimi" tekshiriladi — aks holda
    masalan "Шахрихон сервис (склад)" (skip-kalit "сервис") yoki "Лазер
    Промзона..." (skip-kalit "лазер") kabi ATAYLAB chetlangan ustunlar
    nomida "Шахрихон"/"Промзона" so'zi bor deb NOTO'G'RI filialga
    qo'shilib ketardi."""
    if _lokatsiya_kanali(nom) is None:
        return None
    s = str(nom).strip().lower()
    # 1) Avval ANIQ (FILIAL_GURUHLARI'dagi xom nom bilan) moslikni qidiramiz
    for filial, xom_nomlar in FILIAL_GURUHLARI.items():
        if any(str(x).strip().lower() == s for x in xom_nomlar):
            return filial
    # 2) Topilmasa (masalan ertaga "Бухоро Транзит 2" kabi yangi variant
    #    qo'shilsa) — filial NOMINING O'ZI ustun nomida bor-yo'qligini
    #    tekshiramiz. "Ош" bundan mustasno — 2 harfli bo'lgani uchun
    #    ("Тошкент" kabi so'zlar ichida ham uchraydi) faqat QATOR BOSHIDA
    #    mos deb hisoblanadi, xuddi _lokatsiya_kanali() dagi kabi.
    for filial in FILIAL_GURUHLARI:
        if filial == 'Ош':
            if s.startswith('ош'):
                return filial
            continue
        if filial.lower() in s:
            return filial
    return None


def filial_qoldiqlarini_chiqar(filepath: str) -> dict:
    """Yangi (filial-bo'yicha) tarix faylidan HAR BIR filial uchun ALOHIDA
    qoldiqni chiqaradi. Qaytadi: {Mahsulot_Normalized: {Filial: qoldiq_dona}}.
    Eski (9 ustunli) format uchun BO'SH dict qaytaradi — u faylda
    filial-bo'yicha buzilish umuman yo'q, ajratib bo'lmaydi.

    2026-07-24 (Huzayfa: "har filial faqat o'zinikini ko'rsin" funksiyasi
    uchun) — buyurtma hisob-kitobi (Асосий/Цех/Ош jamlama, load_qoldiq_file())
    bilan ATAYLAB ARALASHTIRILMAYDI: bu alohida, faqat QIDIRUV ekranida
    ko'rsatish uchun ishlatiladigan yengil natija."""
    if not _yangi_tarix_formati_mi(filepath):
        return {}

    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    try:
        ws = wb.active

        # Har filial uchun ustun juftliklari
        filial_ustunlar: dict[str, list] = {f: [] for f in FILIALLAR}
        c = 11
        while c < 121:
            nom = ws.cell(6, c).value
            if nom is not None:
                filial = _filial_nomi(nom)
                if filial:
                    filial_ustunlar[filial].append((c, c + 1))
            c += 2

        natija: dict[str, dict[str, float]] = {}
        for r in range(8, ws.max_row + 1):
            mahsulot = ws.cell(r, 2).value
            if mahsulot is None:
                continue
            mahsulot_str = str(mahsulot).strip()
            if not mahsulot_str or mahsulot_str == 'Товар':
                continue

            qatorlar = {}
            for filial, ustunlar in filial_ustunlar.items():
                dona = 0
                for kor_col, _jami_col in ustunlar:
                    kor_val = ws.cell(r, kor_col).value
                    if kor_val not in (None, ''):
                        dona += parse_qoldiq_str(kor_val)
                if dona:
                    qatorlar[filial] = dona
            if qatorlar:
                key = normalize_product_name(mahsulot_str)
                natija[key] = qatorlar
    finally:
        wb.close()
    return natija


def _yangi_lokatsiyalarni_tekshir(joriy_nomlar: list) -> None:
    """Yangi tarix faylida ILGARI ko'rilmagan ustun (filial/ombor) nomi
    paydo bo'lsa — konsolga OGOHLANTIRISH chiqaradi (hisoblashni
    TO'XTATMAYDI, standart qoida bo'yicha davom etadi). Maqsad: format
    o'zgarganda admin buni JIM emas, KO'RIB bilsin (Huzayfa talabi)."""
    try:
        eski = set()
        if os.path.exists(_LOKATSIYA_KUZATUV_FAYL):
            with open(_LOKATSIYA_KUZATUV_FAYL, encoding='utf-8') as f:
                eski = set(json.load(f))
        joriy = set(joriy_nomlar)
        yangi = joriy - eski
        if eski and yangi:
            print(f"  ⚠️  Tarix faylida YANGI lokatsiya ustun(lar)i topildi "
                  f"(avval ko'rilmagan): {sorted(yangi)}")
            print(f"      Standart qoida bo'yicha avtomatik HISOBGA OLINADI "
                  f"(chiqarib tashlash kerak bo'lsa — common.py::"
                  f"_YANGI_FORMAT_SKIP_KALIT ro'yxatiga qo'shing).")
        atomic_json_write(_LOKATSIYA_KUZATUV_FAYL, sorted(joriy | eski))
    except Exception:
        pass  # kuzatuv ixtiyoriy — asosiy hisoblashga hech qanday ta'sir qilmasin


def _load_qoldiq_yangi_format(filepath: str) -> pd.DataFrame:
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    try:
        ws = wb.active

        # Har kanal uchun alohida ustun juftliklari (кор_col, jami_col)
        kanal_ustunlar = {'asosiy': [], 'sex': [], 'osh': []}
        barcha_nomlar  = []
        c = 11
        while c < 121:   # 121-122 = Итого, alohida, shu yerda hisoblanmaydi
            nom = ws.cell(6, c).value
            if nom is not None:
                barcha_nomlar.append(str(nom).strip())
                kanal = _lokatsiya_kanali(nom)
                if kanal:
                    kanal_ustunlar[kanal].append((c, c + 1))
            c += 2
        _yangi_lokatsiyalarni_tekshir(barcha_nomlar)

        rows = []
        for r in range(8, ws.max_row + 1):
            mahsulot = ws.cell(r, 2).value
            if mahsulot is None:
                continue
            mahsulot_str = str(mahsulot).strip()
            if not mahsulot_str or mahsulot_str == 'Товар':
                continue

            dona  = {'asosiy': 0, 'sex': 0, 'osh': 0}
            summa = {'asosiy': 0.0, 'sex': 0.0, 'osh': 0.0}
            for kanal, ustunlar in kanal_ustunlar.items():
                for kor_col, jami_col in ustunlar:
                    kor_val = ws.cell(r, kor_col).value
                    if kor_val not in (None, ''):
                        dona[kanal] += parse_qoldiq_str(kor_val)
                    jami_val = ws.cell(r, jami_col).value
                    if isinstance(jami_val, (int, float)):
                        summa[kanal] += jami_val

            rows.append({
                'Mahsulot':           mahsulot_str,
                'Asosiy_Qoldiq_Dona': dona['asosiy'],
                'Cex_Qoldiq_Dona':    dona['sex'],
                'Osh_Qoldiq_Dona':    dona['osh'],
                'Qoldiq_Summa':       summa['asosiy'] + summa['sex'] + summa['osh'],
            })
    finally:
        wb.close()

    df = pd.DataFrame(rows, columns=[
        'Mahsulot', 'Asosiy_Qoldiq_Dona', 'Cex_Qoldiq_Dona', 'Osh_Qoldiq_Dona', 'Qoldiq_Summa',
    ])
    df['Mahsulot_Normalized'] = df['Mahsulot'].apply(normalize_product_name)
    # Qoldiq_Dona — UMUMIY (uch kanal jami), faqat main.py'dagi kompaniya
    # darajasidagi "Холат" ko'rsatkichi uchun; buyurtma hisob-kitobi endi
    # Asosiy_Qoldiq_Dona/Cex_Qoldiq_Dona/Osh_Qoldiq_Dona'dan foydalanadi.
    df['Qoldiq_Dona'] = (df['Asosiy_Qoldiq_Dona'] + df['Cex_Qoldiq_Dona']
                          + df['Osh_Qoldiq_Dona'])
    return df[['Mahsulot', 'Mahsulot_Normalized', 'Qoldiq_Dona', 'Qoldiq_Summa',
               'Asosiy_Qoldiq_Dona', 'Cex_Qoldiq_Dona', 'Osh_Qoldiq_Dona']].copy()


# ============================================================
# FILTR — keraksiz tovarlar
# ============================================================

def keraksizmi(nom) -> bool:
    if pd.isna(nom):
        return False
    nom_lower = str(nom).lower()

    # Istisno — bu nomlar har doim o'tsin, filtrlanmasin
    if any(i in nom_lower for i in KERAKSIZ_ISTISNO):
        return False

    # Qism sifatida kirgan bo'lsa
    for qism in KERAKSIZ_QISM:
        if qism.lower() in nom_lower:
            return True

    # To'liq so'z sifatida
    for alon in KERAKSIZ_ALON:
        if re.search(r'\b' + re.escape(alon.lower()) + r'\b', nom_lower):
            return True

    return False


# ============================================================
# FAYL SANA YORDAMCHISI
# ============================================================

def fayl_sanasi(fayl_yoli: str) -> datetime:
    filename = os.path.basename(fayl_yoli).replace('.xlsx', '')
    try:
        return datetime.strptime(filename, '%d.%m.%Y')
    except ValueError:
        return datetime.fromtimestamp(os.path.getmtime(fayl_yoli))