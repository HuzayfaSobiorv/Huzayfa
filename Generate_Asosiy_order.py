"""
generate_asosiy_order.py  v3
============================
Faqat 3 ustun: Tovar nomi | Uzunlik | Buyurtma
- Tovar nomi: 15pt bold qora
- Uzunlik: qizil
- Buyurtma: qizil
- Ranglar rasmga mos
"""

import re
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.comments import Comment

# ============================================================
# 1. SOZLAMALAR
# ============================================================
BASE_DIR         = Path(__file__).resolve().parent
POWER_BI_FILE    = BASE_DIR / "chiqish" / "NEJAVIYKA_POWER_BI.xlsx"
MIN_ZAXIRA_FILE  = BASE_DIR / "Minimal_zaxiralar" / "Min_Zaxira.xlsx"
DELIVERY_DAYS    = 55
STENKA_DELTA     = 0.05

PB_SHEET_INVENTAR = "Инвентар"
PB_COL_TOVAR      = "Товар"
PB_COL_QOLDIQ     = "Қолдиқ"            # UMUMIY (3 kanal jami) — faqat fallback
# 2026-07-24 (Huzayfa: "Asosiyda buyurtma yozsak, u O'shning/Tsexning
# qoldig'ini hisobga olmasligi kerak"): har kanal ENDI FAQAT o'z jismoniy
# zaxirasidan hisoblanadi — xuddi Асосий_Захира/Цех_Захира/Ош_Захира kabi.
PB_COL_QOLDIQ_ASOSIY = "Асосий_Қолдиқ"
PB_COL_QOLDIQ_CEX    = "Цех_Қолдиқ"
PB_COL_QOLDIQ_OSH    = "Ош_Қолдиқ"
PB_COL_YOLDA      = "Йўлда_Жами"
PB_COL_MIN        = "Мин_Захира"        # Асосий+Цех+Ош (jamlangan) — faqat backward-compat fallback
PB_COL_ASOSIY     = "Асосий_Захира"     # faqat "asosiy" kanali uchun
PB_COL_CEX        = "Цех_Захира"        # faqat "sex" (Цех) kanali uchun
PB_COL_OSH        = "Ош_Захира"         # faqat "osh" kanali uchun
MZ_SHEET          = "Min_Zaxira"
MZ_COL_TOVAR      = "Товар"
MZ_COL_MIN        = "Мин_Захира"

# ============================================================
# 2. RANGLAR (rasmga mos)
# ============================================================
# Marka fon ranglari - rasmga qarab aniqlanadi:
# 201 marка → och ko'k (#CCE5FF)
# 304 marka → och yashil (#D4EDDA)  
# 430 marka → och sariq (#FFF3CD)
# Голд → deyarli oq (#FAFAFA)
# Ustun sarlavha → to'q (#2C3E50 kabi)
# Banner → to'q (#1A252F kabi)
# Surface row → o'rta kulrang

# Hamma mahsulot qatorlari — bir xil och ko'k, alternating
ROW_CLR_DARK  = "D6EAF8"   # toq qator  (Excel qator raqami juft)
ROW_CLR_LIGHT = "EBF5FB"   # yengil qator

CATEGORY_BG  = "1A3A5C"   # asosiy sarlavha (yuklatish bilan bir xil)
COL_HDR_BG   = "2C4A6E"   # ustun sarlavhasi
SURFACE_BG   = "4A6FA5"   # surface (Матовый/Голд) satri
MARKA_BG     = "B8CCE4"   # marka ajratuvchi (201 / 304) — och ko'k

RED   = "C0392B"
BLACK = "000000"
WHITE = "FFFFFF"

FONT_SZ   = 15
FONT_NAME = "Calibri"

# ============================================================
# 3. KATEGORIYALAR
# ============================================================
# Varaq tartibi:
# Безшовный труба — Лисдан keyin (foydalanuvchi talabi)
# Қўзиқорин — alohida varaq emas, Чашка varag'iga birlashtiriladi
CATEGORIES  = [
    "Труба",
    "Профиль",
    "Лист",
    "Безшовный труба",
    "Баласина",
    "Стойка",
    "Шар",
    "Соқка",
    "Отвод",
    "Чашка",       # Қўзиқорин ham shu varaqqa kiradi
    "Полировка",   # 2026-07-21: Совун (sovun) ham shu varaqqa kiradi endi
]
SORTED_CATS = {"Труба","Профиль","Лист"}

# 2026-07-14 (Huzayfa qoidasi — mayda tovarlar Excelni to'ldirib yubormasin):
#   * Mayda truba  = diametri 51 dan KICHIK (Ф-51 ning o'zi mayda EMAS)
#   * Mayda profil = ikkala o'lchami 50 dan kichik (50х50 ning o'zi EMAS,
#     80х40 ham emas — bitta tomoni 50+)
#   * Bunday tovar buyurtmasi MAYDA_LIMIT dan OSHSAgina Excelga chiqadi
#     (keyingi safar ehtiyoj yig'ilib oshgach o'zi chiqadi, yo'qolmaydi)
#   * Безшовный труба bundan MUSTASNO (alohida kategoriya, 50 ta ham
#     buyurtsa bo'ladi); Лист va boshqa kategoriyalarga daxli yo'q.
MAYDA_LIMIT = 200

# 2026-07-18 (Huzayfa qoidasi): YIRIK/og'ir tovarlar buyurtmasi 50 ga
# YAXLITLANMAYDI — aniq son ochiq yoziladi (og'ir/qimmat mahsulot,
# yaxlitlash katta pulga aylanadi). Qamrov:
#   * Лист    — qalinligi 1,5 va undan yuqori (marka muhim emas)
#   * Труба   — diametri 76 dan KATTA (Ф-76 ning o'zi yaxlitlanADI)
#   * Профиль — biror tomoni 60 va undan katta (60х40, 60х60, 80х80, 100х100...)
# Безшовный труба alohida kategoriya — unga tegilmaydi (yaxlitlanadi).
LIST_YAXLIT_QALINLIK = 1.5
TRUBA_YAXLIT_DIA     = 76
PROFIL_YAXLIT_TOMON  = 60

def _list_yaxlitlanmasmi(tovar) -> bool:
    """True — yirik tovar (qalin Лист / katta Труба / katta Профиль),
    buyurtmasi yaxlitlanmasin. Kategoriya bilan himoyalangan — _dia_f/_d1_f
    topilmasa 9999 qaytarishi noto'g'ri trigger bermasin."""
    t   = str(tovar)
    kat = get_category(t)
    if kat == "Лист":
        m = re.search(r'Лист-\s*([\d,\.]+)', t)
        if not m:
            return False
        try:
            return float(m.group(1).replace(',', '.')) >= LIST_YAXLIT_QALINLIK
        except ValueError:
            return False
    if kat == "Труба":
        d = _dia_f(t)
        return d != 9999.0 and d > TRUBA_YAXLIT_DIA
    if kat == "Профиль":
        d1, d2 = _d1_f(t), _d2_f(t)
        return d1 != 9999.0 and max(d1, d2) >= PROFIL_YAXLIT_TOMON
    return False

def mayda_buyurtma_limiti(tovar, kategoriya) -> int:
    """Tovar uchun minimal buyurtma chegarasi: mayda truba/profil -> 200,
    qolganlarga 0 (ya'ni cheklovsiz)."""
    if kategoriya == 'Труба' and _dia_f(tovar) < 51:
        return MAYDA_LIMIT
    if kategoriya == 'Профиль' and max(_d1_f(tovar), _d2_f(tovar)) < 50:
        return MAYDA_LIMIT
    return 0

SURFACE_ORDER = ["","Матовый","Глянцевый","Чёрный","Голд","Цветной"]
MARKA_ORDER   = ["201","304","430","316","321",""]

# Faqat 3 ustun (A-C) — bannerlar/sarlavhalar shu 3 ustunni egallaydi
NCOLS      = 3
COL_NAMES  = ["Tovar nomi", "Uzunlik", "Buyurtma"]
COL_WIDTHS = [50, 12, 14]

# 2026-07-13 (Huzayfa: "K va L ustunlarini ham ko'rsatib turish kerak"):
# D ustuni BO'SH qoladi (ajratuvchi bo'shliq), E va F ustunlarida esa
# Xitoy ostatka faylidagi K (Zakaz) va L (Tayyor) qiymatlari -- FAQAT
# KO'RSATISH uchun.
EXTRA_COL_ZAKAZ  = 5   # E
# 2026-07-21 (Huzayfa: "Ostatkadan keyin Yo'lda, keyin Qoldiq, keyin Min
# Zaxira tursin, Buyurtma ustuni Excel formula bo'lsin — Min Zaxirani
# o'zgartirsam Buyurtma o'zi qayta hisoblansin"): tartib qayta qurildi va
# Min Zaxira ustuni qo'shildi. E/F/G/H — FAQAT ko'rsatish uchun EMAS
# endi, C ustunidagi formula shulardan foydalanadi (pastga qarang,
# write_product va EXCEL_TAKLIF_FORMULA).
EXTRA_COL_YOLDA  = 6   # F
EXTRA_COL_QOLDIQ = 7   # G
EXTRA_COL_MINZ   = 8   # H
# 2026-07-23 (Huzayfa: "Ф-51 ni buyurtmasi 1650, real 8600 bo'lishi kerak
# edi -- ANIQLANDI, JIDDIY BUG, keyin "Taxminiy" ustuni ham "umuman
# tushunarsiz, olib tashla" deb BEKOR QILINDI): C ustunidagi JONLI
# FORMULA "butun Yo'lda GORIZONT OXIRIDA bir yo'la keladi" deb soddalash-
# tirilgan taxmin bilan ishlaydi -- real konteyner sanalarini bila
# olmagani uchun ba'zida son marta xato chiqargan (Ф-51: real 8600,
# formula 1650). Alohida "Taxminiy*" ustunga ko'chirilgan edi, lekin bu
# ham chalkashtirdi -- ENDI BUTUNLAY OLIB TASHLANDI. C ustuni FAQAT real
# zanjir_sim natijasi (statik, aniq son). Excel ichida H (Min Zaxira)ni
# o'zgartirish endi C ga TA'SIR QILMAYDI -- bunga qarshi yagona to'g'ri
# yo'l Min_Zaxira.xlsx faylida haqiqiy o'zgartirish qilib, botdan qayta
# "Buyurtma yig'ish" so'rash (shunda ANIQ, konteyner sanalarini hisobga
# olgan natija chiqadi).
EXTRA_HDRS       = {EXTRA_COL_ZAKAZ: "🇨🇳 Ostatka", EXTRA_COL_YOLDA: "Yo'lda",
                    EXTRA_COL_QOLDIQ: "Qoldiq", EXTRA_COL_MINZ: "Min Zaxira"}
EXTRA_WIDTHS     = {EXTRA_COL_ZAKAZ: 14, EXTRA_COL_YOLDA: 12,
                    EXTRA_COL_QOLDIQ: 12, EXTRA_COL_MINZ: 14}

# ============================================================
# 4. PARSING
# ============================================================
def get_category(name):
    s = str(name).strip()
    # Безшовный — Труба'dan OLDIN tekshirilsin; "Б/Ш" ham ushlansin.
    # 2026-07-14 (Huzayfa so'rovi bilan tekshirildi): inventarda bu tovarlar
    # "БеСшовный" (С bilan) yozilgan, eski regex esa faqat "БеЗш" (З) ni
    # qidirardi — shu IMLO FARQI tufayli 16 ta tovar hech qachon o'z
    # varag'iga tushmay, oddiy Труба ichida yurgan. Endi ikkala imlo ham.
    if re.search(r'Бе[сз]ш|Б/Ш', s, re.I):                        return 'Безшовный труба'
    if re.match(r'^(\([^)]*\)\s*)?Ф-\d+', s) and 'ст' in s:      return 'Труба'
    if re.search(r'Пр\.\s*\d+[хx]\d+', s) and 'ст' in s:         return 'Профиль'
    if re.match(r'^Лист-', s) or re.match(r'^Лист\s*-', s):       return 'Лист'
    if re.match(r'^Балас', s, re.I):                               return 'Баласина'
    if re.match(r'^Стойк', s, re.I):                               return 'Стойка'
    # Шар — Шаркона/Шарнир/Шаршара kirmaydi, istalgan joyda bo'lishi mumkin
    if re.search(r'\bШар\b', s, re.I) and not re.search(r'Шар(кона|нир|шара)', s, re.I):
                                                                    return 'Шар'
    # Соқка/Сокка — istalgan joyda (masalan: "32-Сокка")
    if re.search(r'Со[қк]к', s, re.I):                             return 'Соқка'
    # Отвод — istalgan joyda (masalan: "16-Отвод №-01", "114-Отвод 304 марка")
    if re.search(r'Отвод', s, re.I):                               return 'Отвод'
    # Чашка — istalgan joyda (masalan: "25х25-Чашка", "(Голд) 40х40-Чашка тепа-19")
    if re.search(r'Чашк', s, re.I):                                return 'Чашка'
    # Қўзиқорин/Кузикорин → Чашка varag'iga (masalan: "51-Кузикорин", "25-Кузикорин")
    if re.search(r'К[уўy]зи[қк]?орин|Қ[уў]зиқорин', s, re.I):    return 'Чашка'
    # 2026-07-21 (Huzayfa: "Polirovka sahifasiga barcha sovunlarni qo'sh"):
    # ESKI XATO -- kod "Мыло" (rus tilida sovun) so'zini qidirardi, lekin
    # bizning katalogda sovunlar HAMMASI "Совун" (o'zbekcha, Min_Zaxira.xlsx
    # tekshirilganda tasdiqlandi: "Совун- кичик", "Совун А-5" va h.k.) deb
    # yozilgan -- "Мыло" bilan boshlanadigan tovar UMUMAN yo'q, shu sabab
    # bitta ham sovun hech qachon 'Мыло' varag'iga (yoki boshqa hech qayerga
    # -- 'Мыло' CATEGORIES ro'yxatida yo'q edi) tushmagan, butunlay
    # KO'RINMAS bo'lib qolgan edi. Endi 'Совун' ham shu ro'yxatga qo'shildi
    # va alohida 'Мыло' varag'i o'rniga TO'G'RIDAN-TO'G'RI 'Полировка'га
    # yo'naltiriladi (Huzayfa qat'iy so'radi). Solishtirish endi katta-kichik
    # harfga sezgir EMAS (.lower()) -- ilgari "Шлифовка" katta harf bilan
    # yozilmagan variantlar ("Кичик шлифовка", "ШЛИФОВКАНИКИ") ham
    # tushirib qoldirilardi.
    if any(k.lower() in s.lower() for k in
           ['Намат', 'Мелкий', 'Грубый', 'Капрон', 'Шлифовка', 'Полировка',
            'Совун', 'Мыло']):
        return 'Полировка'
    return 'Бошқа'

def get_surface(name):
    b = re.match(r'^\(([^)]+)\)', name.strip())
    if b: return b.group(1)
    s = re.search(r'\((Матовый|Глянцевый|Чёрный|Голд|Цветной)\)', name)
    if s: return s.group(1)
    return ""

def get_marka(name):
    m = re.search(r'\((\d{3})\s*марка\)', name)
    if m: return m.group(1)
    m2 = re.search(r'\b(201|304|430|316|321)\b', name)
    if m2: return m2.group(1)
    return ""

def get_size_order(name, cat):
    """Лист uchun kichik/katta format ajratish (plain sheets uchun)."""
    if cat == 'Лист':
        d = re.search(r'\((\d+)[хx]', name)
        return 0 if d and int(d.group(1)) <= 1250 else 1
    return 0

def get_size_dims(name):
    """Лист o'lchamini olish: '(1220x2440)' → '1220х2440' (Kirill x)"""
    m = re.search(r'\((\d+[хx×]\d+)\)', str(name))
    if m:
        return m.group(1).replace('x', 'х').replace('×', 'х')
    return ""

def _size_dim_first(dims: str) -> int:
    """Birinchi o'lcham raqami (sort uchun): '1220х2440' → 1220"""
    m = re.match(r'(\d+)', str(dims))
    return int(m.group(1)) if m else 9999

def _parse_num(s):
    """Vergul yoki nuqtali raqam → float"""
    try:
        return float(str(s).replace(',', '.'))
    except (ValueError, TypeError):
        return 9999.0

def _dia_f(name):
    """Труба диаметрини float: 'Ф-9,5 ст ...' → 9.5, 'Ф-102 ...' → 102.0"""
    m = re.search(r'Ф-([\d,\.]+)', str(name))
    return _parse_num(m.group(1)) if m else 9999.0

def _d1_f(name):
    """Профиль birinchi o'lchami float: 'Пр. 25х13' → 25.0"""
    m = re.search(r'(\d+)[хx](\d+)', str(name))
    return float(m.group(1)) if m else 9999.0

def _d2_f(name):
    """Профиль ikkinchi o'lchami float: 'Пр. 25х13' → 13.0"""
    m = re.search(r'(\d+)[хx](\d+)', str(name))
    return float(m.group(2)) if m else 9999.0

def _stenka_f(name):
    """Stenka qalinligi float: 'ст 0,85' → 0.85; 'Лист-0,75' → 0.75"""
    m = re.search(r'ст\s*([\d,\.]+)', str(name))
    if m:
        return _parse_num(m.group(1))
    m2 = re.match(r'^Лист[\s-]*([\d,\.]+)', str(name).strip(), re.I)
    if m2:
        return _parse_num(m2.group(1))
    return 9999.0

def get_length(name):
    m = re.search(r'\(([\d,\.]+)\s*м\)', name)
    return f"{m.group(1)} м" if m else ""

def strip_length(name):
    return re.sub(r'\s*\([\d,\.]+\s*м\)', '', name).strip()

def get_stenka_china(name, cat):
    if cat not in {"Труба","Профиль","Лист"}:
        return ""
    t = re.search(r'ст\s*([\d,\.]+)', name)
    if t:
        val = float(t.group(1).replace(',', '.'))
        return str(round(val - STENKA_DELTA, 3)).replace('.', ',')
    lm = re.match(r'^Лист-([\d,\.]+)', name)
    if lm:
        val = float(lm.group(1).replace(',', '.'))
        return str(round(val - STENKA_DELTA, 3)).replace('.', ',')
    return ""

# ============================================================
# 5. MA'LUMOT
# ============================================================
def load_data(kanal: str = "asosiy"):
    from common import keraksizmi
    if not POWER_BI_FILE.exists():
        raise FileNotFoundError(f"Power BI fayli topilmadi: {POWER_BI_FILE}")

    pb = pd.read_excel(POWER_BI_FILE, sheet_name=None)
    if PB_SHEET_INVENTAR not in pb:
        raise ValueError(f"'{PB_SHEET_INVENTAR}' varaqi topilmadi. Mavjud varaqlar: {', '.join(pb)}")

    inv = pb[PB_SHEET_INVENTAR].copy()
    required = [PB_COL_TOVAR]
    missing = [c for c in required if c not in inv.columns]
    if missing:
        raise ValueError(f"Inventar varaqida ustunlar yetishmaydi: {', '.join(missing)}")

    # 2026-07-15 (Huzayfa: "ham sotuvga, ham sexga kerak bo'lgan tovar bor",
    # keyin: "Osh ham o'z alohida zaxirasiga ega bo'lishi kerak"): ILGARI
    # kanal filtri "Тур" ustuniga (Цех_Захира>0 bo'lsa BUTUN qator "ЦЕХ🏭"
    # deb belgilanib, Asosiy/Osh kanaldan MUTLAQO chiqarib tashlanardi, Osh
    # esa Asosiy bilan BITTA "Сотув_Захира"ni bo'lishardi) asoslangandi.
    # Endi HAR BIR kanal (Asosiy/Cex/Osh) FAQAT O'ZINING ustunidan
    # foydalanadi — bir tovar bir nechta kanalga kerak bo'lsa, har birida
    # o'z ulushicha, alohida-alohida chiqadi.
    min_col = {"sex": PB_COL_CEX, "osh": PB_COL_OSH}.get(kanal, PB_COL_ASOSIY)
    if min_col not in inv.columns:
        # 2026-07-18 (Huzayfa: "min zaxirasi 0 ga tenglar buyurtma
        # berilmasligi kerak"): ESKI backward-compat bu yerda Мин_Захира
        # (UCH KANAL JAMI) ustuniga tushardi -- natijada faqat Tsex/Osh'ga
        # kerak tovar (bu kanaldagi mini = 0) ASOSIY kanalda ham buyurtma
        # olib ketardi. Endi PB'da kanal ustuni bo'lmasa, min to'g'ridan-
        # to'g'ri Min_Zaxira.xlsx'ning KANALGA MOS ustunidan olinadi
        # (pastdagi merge) -- jami ustun ISHLATILMAYDI.
        min_col = None

    # 2026-07-24 (Huzayfa: "Asosiyda buyurtma yozsak, u O'shning/Tsexning
    # qoldig'ini hisobga olmasligi kerak" -- tasdiqlangan, real gap edi):
    # QOLDIQ ham ENDI kanalga mos ustundan olinadi, xuddi min_zaxira kabi.
    # Fallback: agar Power BI fayl ESKI (hali qayta yaratilmagan, kanal
    # ustunlari yo'q) bo'lsa -- umumiy "Қолдиқ"ga tushadi (backward-compat,
    # "Ma'lumotlarni yangilash" bosilguncha eski xatti-harakat saqlanadi).
    qoldiq_col = {"sex": PB_COL_QOLDIQ_CEX, "osh": PB_COL_QOLDIQ_OSH}.get(kanal, PB_COL_QOLDIQ_ASOSIY)
    if qoldiq_col not in inv.columns:
        qoldiq_col = PB_COL_QOLDIQ
    if qoldiq_col not in inv.columns:
        raise ValueError(f"Inventar varaqida qoldiq ustuni topilmadi ({qoldiq_col})")

    cols = [PB_COL_TOVAR, qoldiq_col]
    if PB_COL_YOLDA in inv.columns:
        cols.append(PB_COL_YOLDA)
    if min_col:
        cols.append(min_col)

    df = inv[cols].rename(columns={
        PB_COL_TOVAR: "tovar",
        qoldiq_col: "qoldiq",
        PB_COL_YOLDA: "yoldagi",
        **({min_col: "min_zaxira"} if min_col else {}),
    }).dropna(subset=["tovar"])

    if "yoldagi" not in df.columns:
        df["yoldagi"] = 0

    # ── 2026-07-18 (Huzayfa: "min zaxirasi 0 ga tenglar buyurtma
    # berilmasligi kerak"): Min_Zaxira.xlsx — YAKUNIY (authoritative) manba.
    # ILGARI min faqat PB Инвентар'dan olinardi; PB esa "Ma'lumotlarni
    # yangilash"dagina qayta quriladi — Huzayfa Min_Zaxira.xlsx'da min'ni
    # 0 qilsa ham, PB yangilanmaguncha buyurtma ESKI min bilan chiqaverardi.
    # Endi Min_Zaxira.xlsx'dagi KANALGA MOS ustun har doim PB qiymatining
    # USTIDAN yoziladi — faylni tahrirlagan zahoti kuchga kiradi. Faylda
    # yo'q tovarlar uchun PB qiymati (bo'lsa) saqlanadi.
    if MIN_ZAXIRA_FILE.exists():
        mz = pd.read_excel(MIN_ZAXIRA_FILE, sheet_name=MZ_SHEET)
        mz.columns = [str(c).strip() for c in mz.columns]  # 'Мин_Захира ' kabi bo'shliqlar
        mz_min_col = {"sex": "Цех_Захира", "osh": "Ош_Захира"}.get(kanal, "Асосий_Захира")
        if mz_min_col not in mz.columns:
            mz_min_col = MZ_COL_MIN   # juda eski Min_Zaxira formati uchun
        if MZ_COL_TOVAR in mz.columns and mz_min_col in mz.columns:
            # Nomlar IKKALA tomonda ham normallashtirilib kalit qilinadi --
            # bo'shliq/formatlash farqlari SOXTA nomuvofiqlik bermasin.
            from common import normalize_product_name as _norm_mz
            mzv = mz[[MZ_COL_TOVAR, mz_min_col]].dropna(subset=[MZ_COL_TOVAR]).copy()
            mzv["_k"]   = mzv[MZ_COL_TOVAR].astype(str).map(_norm_mz)
            # Bo'sh katak = 0 (bu kanalda meyor yo'q -> buyurtma berilmaydi)
            mzv["_min"] = pd.to_numeric(mzv[mz_min_col], errors="coerce").fillna(0)
            mz_map = mzv.groupby("_k")["_min"].max()
            _k   = df["tovar"].astype(str).map(_norm_mz)
            # 2026-07-18 (Huzayfa, rangli listlar hodisasi): Min_Zaxira.xlsx'da
            # YO'Q tovar uchun PB fallback ISHLATILMAYDI -- min=0 (buyurtma
            # berilmaydi). Sabab: PB'da eski/chiqarilgan tovarlar (Лист-0,41,
            # 0,61, 0,81 kabi) tarixiy Мин_Захира=50 bilan qolib ketgan va
            # fallback orqali buyurtma olib ketayotgan edi. Qoida endi qat'iy:
            # Min_Zaxira.xlsx = yagona katalog; faylda yo'q yoki 0 -> buyurtma yo'q.
            df["min_zaxira"] = _k.map(mz_map).fillna(0)

    if "min_zaxira" not in df.columns:
        raise FileNotFoundError(
            f"Min zaxira manbasi topilmadi: PB'da kanal ustuni ham yo'q, "
            f"Min_Zaxira.xlsx ham yo'q ({MIN_ZAXIRA_FILE})")

    for col in ["qoldiq", "yoldagi", "min_zaxira"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Keraksiz tovarlarni olib tashlash
    df = df[~df["tovar"].apply(keraksizmi)].copy()

    # 2026-07-09: zanjir-simulyatsiya (kamomat_engine.zanjir_sim) uchun —
    # har tovarning hali KELMAGAN konteynerlari, ANIQ kelish kuni bilan.
    # Eski kod faqat "yoldagi" JAMI sonini bilardi, aniq qachon kelishini
    # bilmasdi — shu sabab konteynerlar orasidagi bo'shliqni (masalan 40 kun
    # ichida 2-3 kun kelmay qolgan oraliqni) hisobga olmas edi.
    kont_map: dict[str, list] = {}
    PB_SHEET_KONT = "Контейнерлар"
    if PB_SHEET_KONT in pb:
        kdf = pb[PB_SHEET_KONT].copy()
        if "Холат" in kdf.columns:
            kdf = kdf[kdf["Холат"] != "КЕЛДИ ✅"]
        if {"Товар", "Кун_Қолди", "Миқдор"}.issubset(kdf.columns):
            kdf["Кун_Қолди"] = pd.to_numeric(kdf["Кун_Қолди"], errors="coerce").fillna(0)
            kdf["Миқдор"]    = pd.to_numeric(kdf["Миқдор"], errors="coerce").fillna(0)
            for _, r in kdf.iterrows():
                if r["Миқдор"] > 0:
                    kont_map.setdefault(str(r["Товар"]).strip(), []).append(
                        (float(r["Кун_Қолди"]), float(r["Миқдор"]))
                    )

    print(f"✅ Real ma'lumot ({kanal}): {len(df)} tovar")
    return df, kont_map

def demo_data():
    rows = [
        {"tovar":"Ф-19 ст 0,6 (5,8 м) (201 марка)",         "qoldiq":4000,"yoldagi":3000,"min_zaxira":7500},
        {"tovar":"Ф-19 ст 0,7 (5,8 м) (201 марка)",         "qoldiq":500, "yoldagi":0,   "min_zaxira":2000},
        {"tovar":"Ф-25 ст 0,8 (6 м) (201 марка)",           "qoldiq":300, "yoldagi":200, "min_zaxira":3000},
        {"tovar":"Ф-19 ст 1,2 (6 м) (201 марка)",           "qoldiq":800, "yoldagi":0,   "min_zaxira":1500},
        {"tovar":"Ф-25 ст 1,5 (6 м) (201 марка)",           "qoldiq":200, "yoldagi":0,   "min_zaxira":800},
        {"tovar":"Ф-19 ст 0,6 (5,8 м) (304 марка)",         "qoldiq":600, "yoldagi":400, "min_zaxira":1200},
        {"tovar":"Ф-25 ст 0,8 (6 м) (304 марка)",           "qoldiq":150, "yoldagi":0,   "min_zaxira":600},
        {"tovar":"Ф-19 ст 1,0 (6 м) (430 марка)",           "qoldiq":250, "yoldagi":100, "min_zaxira":500},
        {"tovar":"(Голд) Ф-19 ст 0,6 (5,6 м) (201 марка)",  "qoldiq":800, "yoldagi":0,   "min_zaxira":2000},
        {"tovar":"(Голд) Ф-25 ст 0,8 (6 м) (201 марка)",    "qoldiq":100, "yoldagi":200, "min_zaxira":600},
        {"tovar":"(Голд) Ф-19 ст 1,2 (6 м) (201 марка)",    "qoldiq":50,  "yoldagi":0,   "min_zaxira":400},
        {"tovar":"Пр. 20х20 ст 0,7 (6 м) (201 марка)",      "qoldiq":1200,"yoldagi":500, "min_zaxira":3000},
        {"tovar":"Пр. 25х25 ст 0,9 (6,2 м) (201 марка)",    "qoldiq":800, "yoldagi":300, "min_zaxira":2000},
        {"tovar":"Пр. 40х40 ст 1,2 (6 м) (201 марка)",      "qoldiq":300, "yoldagi":0,   "min_zaxira":800},
        {"tovar":"Пр. 20х20 ст 0,7 (6 м) (304 марка)",      "qoldiq":200, "yoldagi":0,   "min_zaxira":600},
        {"tovar":"Лист-0,6 (1220х2440) (Матовый) (201 марка)","qoldiq":150,"yoldagi":80,  "min_zaxira":300},
        {"tovar":"Лист-0,8 (1220х2440) (Матовый) (201 марка)","qoldiq":80, "yoldagi":50,  "min_zaxira":200},
        {"tovar":"Лист-1,0 (1220х2440) (Матовый) (201 марка)","qoldiq":40, "yoldagi":0,   "min_zaxira":150},
        {"tovar":"Лист-0,6 (1500х3000) (Матовый) (201 марка)","qoldiq":30, "yoldagi":20,  "min_zaxira":100},
        {"tovar":"Лист-0,6 (1220х2440) (Матовый) (304 марка)","qoldiq":20, "yoldagi":0,   "min_zaxira":80},
        {"tovar":"Лист-0,8 (1220х2440) (Матовый) (304 марка)","qoldiq":15, "yoldagi":0,   "min_zaxira":60},
        {"tovar":"Лист-0,6 (1220х2440) (Матовый) (430 марка)","qoldiq":60, "yoldagi":30,  "min_zaxira":120},
        {"tovar":"Лист-0,6 (1220х2440) (Глянцевый) (201 марка)","qoldiq":100,"yoldagi":0, "min_zaxira":200},
        {"tovar":"Лист-0,8 (1220х2440) (Глянцевый) (304 марка)","qoldiq":25,"yoldagi":0,  "min_zaxira":80},
        {"tovar":"Баласина-01",  "qoldiq":500, "yoldagi":0,   "min_zaxira":800},
        {"tovar":"Баласина-02",  "qoldiq":300, "yoldagi":200, "min_zaxira":600},
        {"tovar":"Стойка-01",    "qoldiq":200, "yoldagi":0,   "min_zaxira":400},
        {"tovar":"Соқка-19мм",   "qoldiq":1500,"yoldagi":0,   "min_zaxira":2000},
        {"tovar":"Чашка-01",     "qoldiq":2000,"yoldagi":0,   "min_zaxira":3000},
        {"tovar":"Намат №-01",   "qoldiq":20,  "yoldagi":0,   "min_zaxira":50},
        {"tovar":"Совун-01",     "qoldiq":50,  "yoldagi":0,   "min_zaxira":80},
    ]
    return pd.DataFrame(rows)

# ============================================================
# 6. HISOB
# ============================================================
def calculate(df, kont_map: dict | None = None, kanal: str = "asosiy"):
    """
    2026-07-09: buyurtma miqdori endi zanjir-simulyatsiya bilan hisoblanadi
    (kamomat_engine.zanjir_sim) — har tovarning YO'LDAGI konteynerlari ANIQ
    kelish kuni bilan hisobga olinadi, faqat jami son emas. Eski statik
    formula (min_zaxira/30 * 55 kun) konteynerlar orasidagi bo'shliqlarni
    (masalan real misolda 23-42 kun oralig'ida yangi yuk kelmasligini)
    umuman hisobga olmas edi — shu sabab almashtirildi.
    Kunlik sotuv qoidasi (Asosiy/O'sh — ikkalasi ham real sotuvga qarab
    ishlaydi): min_zaxira / common.KUNLIK_SOTUV_BOLISH (2026-07-21dan
    45) — kamomat_engine.zanjir_sim ichida.

    2026-07-21 (Huzayfa bilan kelishildi — "Tsex aralashmasin" qoidasi):
    Tsex (kanal=="sex") ehtiyoji kunlik sotuvdan EMAS, bir oylik
    FIKSIRLANGAN ehtiyojdan kelib chiqadi (Цех_Захира) — shuning uchun
    Asosiy/O'sh uchun tuzatilgan global KUNLIK_SOTUV_BOLISH (45) unga
    TEGMASLIGI kerak. Tsex uchun eski, o'zgarmagan divisor (30) shu
    yerda kunlik_override orqali MAJBURLANADI, global konstanta qanday
    o'zgarsa ham Tsex buyurtmasi buzilmaydi.
    """
    from kamomat_engine import zanjir_sim
    from common import KUNLIK_SOTUV_BOLISH as _GLOBAL_KUNLAR
    df = df.copy()
    for col in ["qoldiq","yoldagi","min_zaxira"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    kont_map = kont_map or {}
    TSEX_ESKI_DIVISOR = 30   # Tsex — qat'iy, global tuzatishdan mustasno
    # 2026-07-23 (Huzayfa bilan kelishildi, Ф-51 real misolida tekshirilgan):
    # YANGI buyurtma berilgandan yetib kelgunga qadar real muddat = ~10 kun
    # tayyorlash + ~5 kun yuklash + 55 kun dengiz yo'li = 70 kun. common.py
    # dagi KELISH_KUNI(55) esa BOSHQA joyda (main.py -- ALLAQACHON yuklangan
    # konteynerning transit vaqti, tavsiya.py) ham ishlatiladi va u yerda
    # 55 TO'G'RI (konteyner allaqachon yuklangan, faqat dengiz yo'li qoladi)
    # -- shu sabab umumiy konstanta o'zgartirilmaydi, faqat shu yerda (Asosiy/
    # O'sh buyurtma hisob-kitobi) horizon_override orqali 70 majburlanadi.
    # Tsex — alohida mavzu (Huzayfa: hozircha tegilmasin), eski 55da qoladi.
    YANGI_BUYURTMA_HORIZON = 70

    def _buyurtma(row):
        konts = kont_map.get(str(row["tovar"]).strip())
        if konts is None:
            # kont_map berilmagan/tovar topilmagan holat uchun (masalan
            # demo_data() bilan qo'lda test qilinganda) — "yoldagi" jamini
            # bitta konteyner sifatida DELIVERY_DAYS kunda keladi deb faraz
            # qilamiz, shunda ham funksiya ishlayveradi.
            konts = [(DELIVERY_DAYS, row["yoldagi"])] if row["yoldagi"] > 0 else []

        kunlik_override  = None
        horizon_override = None
        if kanal == "sex" and row["min_zaxira"] > 0:
            kunlik_override = row["min_zaxira"] / TSEX_ESKI_DIVISOR
        else:
            horizon_override = YANGI_BUYURTMA_HORIZON

        sim = zanjir_sim(row["qoldiq"], row["min_zaxira"], konts,
                         yaxlitla=not _list_yaxlitlanmasmi(row["tovar"]),
                         kunlik_override=kunlik_override,
                         horizon_override=horizon_override)
        return sim["taklif"]

    df["buyurtma"] = df.apply(_buyurtma, axis=1).astype(int)
    # 2026-07-22 (Huzayfa: "buyurtma berilmagan tovarlarni ham alohida
    # ko'rsatish kerak, ro'yxat oxirida ajratib"): ILGARI bu yerda
    # buyurtma==0 qatorlar BUTUNLAY o'chirilardi -- shu tovarlar qayerda
    # ekani umuman ko'rinmasdi. Endi FILTRLANMAYDI -- chaqiruvchi
    # (services.py::asosiy_styled_excel_yarat / build()) o'zi kerakli
    # joyda ajratadi: asosiy buyurtma ro'yxati uchun buyurtma>0, "kerak
    # emas" bo'limi uchun qolganlar.

    df["kategoriya"] = df["tovar"].apply(get_category)
    df["surface"]    = df["tovar"].apply(get_surface)
    df["marka"]      = df["tovar"].apply(get_marka)
    df["size_ord"]   = df.apply(lambda r: get_size_order(r["tovar"], r["kategoriya"]), axis=1)
    df["surf_ord"]   = df["surface"].apply(lambda s: SURFACE_ORDER.index(s) if s in SURFACE_ORDER else 99)
    df["mark_ord"]   = df["marka"].apply(lambda m: MARKA_ORDER.index(m) if m in MARKA_ORDER else 99)
    # Raqamli sort kalitlari (string sort muammosidan qochish uchun)
    df["_dia_f"]  = df["tovar"].apply(_dia_f)    # Труба diametri: Ф-9,5 → 9.5
    df["_d1_f"]   = df["tovar"].apply(_d1_f)     # Профиль: 25х13 → 25.0
    df["_d2_f"]   = df["tovar"].apply(_d2_f)     # Профиль: 25х13 → 13.0
    df["_st_f"]   = df["tovar"].apply(_stenka_f) # Stenka: ст 0,85 → 0.85
    return df

# ============================================================
# 7. EXCEL YOZISH
# ============================================================
def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _align(center=False, indent=0):
    return Alignment(horizontal="center" if center else "left",
                     vertical="center", wrap_text=False, indent=indent)


def get_row_bg(row_num: int) -> str:
    """Excel qator raqami bo'yicha alternating och ko'k rang."""
    return ROW_CLR_DARK if row_num % 2 == 0 else ROW_CLR_LIGHT


def rich_tovar_name(name: str) -> CellRichText:
    """
    Tovar nomi: Arial 15pt bold qora.
    Uzunlik qismini olib tashlaymiz (u alohida ustunda).
    Stenka STENKA_DELTA ga kamaytirilgan holda ko'rsatiladi (Xitoy buyurtma).
    """
    from vazn_hisobla import xitoy_nomi
    base = strip_length(xitoy_nomi(name))
    return CellRichText(
        TextBlock(InlineFont(rFont=FONT_NAME, b=True, sz=FONT_SZ, color=BLACK), base)
    )


def set_cols(ws):
    for i, (_, width) in enumerate(zip(COL_NAMES, COL_WIDTHS), 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for col_i, width in EXTRA_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_i)].width = width


def write_category_banner(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    c = ws.cell(row=row, column=1, value=f"  \U0001f4e6 {text.upper()}")
    c.font      = Font(name=FONT_NAME, bold=True, size=FONT_SZ, color=WHITE)
    c.fill      = _fill(CATEGORY_BG)
    c.alignment = _align(indent=1)
    c.border    = _border()
    ws.row_dimensions[row].height = 30


def write_col_headers(ws, row):
    for i, name in enumerate(COL_NAMES, 1):
        c = ws.cell(row=row, column=i, value=name)
        c.font      = Font(name=FONT_NAME, bold=True, size=FONT_SZ, color=WHITE)
        c.fill      = _fill(COL_HDR_BG)
        c.alignment = _align(center=(i > 1))
        c.border    = _border()
    # E/F -- "Zakaz"/"Tayyor" (Xitoy ostatka, faqat ko'rsatish uchun). D
    # ustuni ATAYLAB bo'sh qoldiriladi (ajratuvchi bo'shliq).
    for col_i, name in EXTRA_HDRS.items():
        c = ws.cell(row=row, column=col_i, value=name)
        c.font      = Font(name=FONT_NAME, bold=True, size=FONT_SZ, color=WHITE)
        c.fill      = _fill(COL_HDR_BG)
        c.alignment = _align(center=True)
        c.border    = _border()
    ws.row_dimensions[row].height = 26


def write_surface_row(ws, row, surface):
    label = f"\u2500\u2500 {surface} \u2500\u2500" if surface else "\u2500\u2500 \u041e\u0434\u0434\u0438\u0439 \u2500\u2500"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    c = ws.cell(row=row, column=1, value=label)
    c.font      = Font(name=FONT_NAME, bold=True, size=FONT_SZ - 1, color="222222")
    c.fill      = _fill(SURFACE_BG)
    c.alignment = _align(indent=1)
    c.border    = _border()
    ws.row_dimensions[row].height = 24


def write_marka_row(ws, row, marka: str):
    """Marka ajratuvchi qator: ── 201 марка ── yoki ── 304 марка ──"""
    label = f"── {marka} марка ──" if marka else "── Номаʼлум ──"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    c = ws.cell(row=row, column=1, value=label)
    c.font      = Font(name=FONT_NAME, bold=True, size=FONT_SZ - 1, color="1A3A5C")
    c.fill      = _fill(MARKA_BG)
    c.alignment = _align(indent=1)
    c.border    = _border()
    ws.row_dimensions[row].height = 22
    return row + 1


def write_list_group_row(ws, row, label: str):
    """Лист uchun (surface + marka + o'lcham) ajratuvchi: ── Матовый 201 (1220х2440) ──"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    c = ws.cell(row=row, column=1, value=label)
    c.font      = Font(name=FONT_NAME, bold=True, size=FONT_SZ - 1, color="FFFFFF")
    c.fill      = _fill(SURFACE_BG)
    c.alignment = _align(indent=1)
    c.border    = _border()
    ws.row_dimensions[row].height = 22
    return row + 1


def excel_taklif_formula(row: int, yaxlitla: bool, kanal: str = "asosiy") -> str:
    """
    2026-07-21 (Huzayfa: "Min Zaxirani o'zgartirsam Buyurtma o'zi
    qayta hisoblansin"): C ustunidagi (Buyurtma) qiymatni ZANJIR_SIM
    bilan BIR XIL arifmetikada, lekin EXCEL FORMULASI sifatida qayta
    quradi -- H (Min Zaxira) katakni tahrirlaganda, Excel o'zi C ni
    qayta hisoblaydi.

    DIQQAT — bu FAQAT taxminiy/ko'rib chiqish uchun (shuning uchun endi
    ALOHIDA "Taxminiy*" ustunga, I, yoziladi -- 2026-07-23, qarang
    write_product()), python dasturidagi aniq natijadan farq qilishi
    mumkin, chunki:
      * Bu yerda BUTUN "Yo'lda" miqdori GORIZONT OXIRIDA bir yo'la
        keladi deb faraz qilinadi -- aslida har konteyner O'Z SANASIDA
        keladi (bu ma'lumot Excelda yo'q, faqat jami son bor). Konteyner
        ERTAROQ kelsa, real dastur ODATDA KATTAROQ buyurtma chiqaradi
        (chunki depletatsiya vaqti uzunroq) -- bu formula esa "eng
        yaxshi holat"ni (hammasi oxirida keladi) faraz qiladi. Haqiqiy
        misolda (Ф-51 ст 0,9) bu ikkalasi 1650 va 8600 kabi FARQ qildi.
      * "Tasdiqlangan buyurtma" (shu davrda allaqachon berilgan
        qism) bu yerda ayirilmaydi -- bu ma'lumot ham Excelda yo'q.
    Formula manbasi -- kamomat_engine.zanjir_sim bilan bir xil konstantalar
    (common.py + Generate_Asosiy_order.calculate()dagi override'lar):
      kanal=="sex":     kunlik_divisor=30 (TSEX_ESKI_DIVISOR), horizon=55 (o'zgarmagan)
      boshqa (asosiy/osh): kunlik_divisor=45 (KUNLIK_SOTUV_BOLISH), horizon=70
                            (2026-07-23dan, YANGI_BUYURTMA_HORIZON)
      kunlik          = H/kunlik_divisor
      min_nuqta       = G - H*horizon/kunlik_divisor
      qoldiq_gorizont = MAX(0, min_nuqta) + F
      nishon          = H + H*30/kunlik_divisor
      taklif_raw      = agar min_nuqta < H bo'lsa MAX(0, nishon-qoldiq_gorizont), aks holda 0
      taklif_yaxlit   = CEILING(taklif_raw, 50)   (yoki 1 -- qalin Лист/yirik tovar uchun)
      Buyurtma        = MAX(0, taklif_yaxlit - E)   (E = 🇨🇳 Ostatka)
    """
    if kanal == "sex":
        kunlik_divisor, horizon = 30, 55
    else:
        kunlik_divisor, horizon = 45, 70
    e, f, g, h = f"E{row}", f"F{row}", f"G{row}", f"H{row}"
    unit = 1 if not yaxlitla else 50
    min_nuqta   = f"({g}-{h}*{horizon}/{kunlik_divisor})"
    qol_goriz   = f"(MAX(0,{min_nuqta})+{f})"
    nishon      = f"({h}*(1+30/{kunlik_divisor}))"
    taklif_raw  = f"MAX(0,{nishon}-{qol_goriz})"
    taklif_50   = f"IF({taklif_raw}>0,CEILING({taklif_raw},{unit}),0)"
    return (f"=IF({h}<=0,0,"
            f"IF({min_nuqta}<{h},"
            f"MAX(0,{taklif_50}-{e}),"
            f"0))")


def write_product(ws, row, r, kanal: str = "asosiy") -> int:
    bg   = get_row_bg(row)
    fill = _fill(bg)
    brd  = _border()

    uzunlik  = get_length(r["tovar"])

    # A — Tovar nomi: 15pt bold qora
    ca = ws.cell(row=row, column=1)
    ca.value     = rich_tovar_name(r["tovar"])
    ca.fill      = fill
    ca.border    = brd
    ca.alignment = _align(indent=1)
    # Asl inventar nomini comment sifatida yozamiz — lookup collision uchun
    # (ст 1,35 va ст 1,4 ikkalasi ham Xitoy Excelda "ст 1,35" ko'rinadi)
    ca.comment   = Comment(str(r["tovar"]), "bot", width=280, height=40)
    ws.row_dimensions[row].height = 26

    # B — Uzunlik: qizil bold
    cb = ws.cell(row=row, column=2, value=uzunlik)
    cb.font      = Font(name=FONT_NAME, bold=True, size=FONT_SZ, color=RED)
    cb.fill      = fill
    cb.border    = brd
    cb.alignment = _align(center=True)

    # E — Zakaz (K): Xitoy ostatkasi, FAQAT ko'rsatish uchun EMAS endi --
    # C ustunidagi formula buni ham hisobga oladi (pastga qarang).
    # Ma'lumot yo'q bo'lsa (masalan ostatka fayli hali yuklanmagan) --
    # 0 ko'rsatiladi, bo'sh emas.
    try:
        zakaz_v = int(float(r.get("zakaz", 0) or 0))
    except (ValueError, TypeError):
        zakaz_v = 0

    ce = ws.cell(row=row, column=EXTRA_COL_ZAKAZ, value=zakaz_v)
    ce.font      = Font(name=FONT_NAME, size=FONT_SZ - 1, color="1A5276")
    ce.fill      = fill
    ce.border    = brd
    ce.alignment = _align(center=True)

    # F/G/H — Yo'lda / Qoldiq / Min Zaxira (2026-07-21 qayta tartiblandi:
    # Ostatka -> Yo'lda -> Qoldiq -> Min Zaxira). C ustunidagi formula
    # aynan shu uch katakdan (F/G/H) + E dan foydalanadi.
    def _int0(v):
        try:
            return int(float(v or 0))
        except (ValueError, TypeError):
            return 0

    cf = ws.cell(row=row, column=EXTRA_COL_YOLDA, value=_int0(r.get("yoldagi", 0)))
    cf.font      = Font(name=FONT_NAME, size=FONT_SZ - 1, color="555555")
    cf.fill      = fill
    cf.border    = brd
    cf.alignment = _align(center=True)

    cg = ws.cell(row=row, column=EXTRA_COL_QOLDIQ, value=_int0(r.get("qoldiq", 0)))
    cg.font      = Font(name=FONT_NAME, size=FONT_SZ - 1, color="555555")
    cg.fill      = fill
    cg.border    = brd
    cg.alignment = _align(center=True)

    ch = ws.cell(row=row, column=EXTRA_COL_MINZ, value=_int0(r.get("min_zaxira", 0)))
    ch.font      = Font(name=FONT_NAME, size=FONT_SZ - 1, bold=True, color="1A5276")
    ch.fill      = fill
    ch.border    = brd
    ch.alignment = _align(center=True)
    # 2026-07-21 (Huzayfa: "boshqa ustunlarga qo'l tegib o'zgarib
    # ketmasin"): FAQAT H (Min Zaxira) qulfdan OCHIQ -- boshqa hamma
    # katak (A/B/C/E/F/G) standart qulflangan holatda qoladi (openpyxl
    # default: Protection(locked=True)). Qulf ws.protection.sheet=True
    # bilan fill_sheet() oxirida yoqiladi (parol YO'Q -- xohlasa Excelda
    # "Himoyani bekor qilish" bilan istalgan payt ocha oladi).
    ch.protection = Protection(locked=False)

    # C — Buyurtma: qizil bold, REAL zanjir_sim natijasi (STATIK son --
    # 2026-07-23: JONLI FORMULA bu yerdan OLIB TASHLANDI, chunki u har
    # konteynerning O'Z SANASINI hisobga OLMAYDI (hammasi 55-kunda bir
    # yo'la keladi deb soddalashtiradi) va real natijadan bir necha
    # baravar farq qilishi mumkin edi (haqiqiy voqea: Ф-51 uchun real 8600,
    # formula 1650 ko'rsatgan). Buyurtma qarori shu ustunga -- ANIQ,
    # backend natijasiga -- asoslanishi kerak.
    cc = ws.cell(row=row, column=3, value=int(r["buyurtma"]))
    cc.font      = Font(name=FONT_NAME, bold=True, size=FONT_SZ, color=RED)
    cc.fill      = fill
    cc.border    = brd
    cc.alignment = _align(center=True)

    return row + 1


def write_nobuy_section(ws, row, nobuy_df, kanal: str = "asosiy") -> int:
    """
    2026-07-22 (Huzayfa: "byurtma berilmaydigan tovarlar bor -- yo'lda va
    qoldiq norm bo'lsa berilmaydi va byurtma varag'ida bo'lmaydi, shularni
    alohida ro'yxat qilish kerak, ro'yxat oxirida ajralib tursin"):
    shu kategoriyada HOZIRCHA buyurtma berilmagan (zaxira+yo'lda yetarli
    deb hisoblangan, YOKI Xitoy ostatkasi/tasdiqlangan buyurtma yoki mayda
    limit tufayli 0ga tushgan) tovarlar -- asosiy buyurtma ro'yxatidan
    KEYIN, aniq ajratilgan alohida bo'lim sifatida bir-in ketin yoziladi.
    Ustunlar bir xil (E/F/G/H, Buyurtma(C) ham doimgidek jonli formula --
    Min Zaxirani qo'lda o'zgartirib sinab ko'rish mumkin, xuddi yuqoridagi
    qatorlar kabi).
    """
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    c = ws.cell(row=row, column=1,
                value="  ⛔ BUYURTMA BERILMAGAN  (zaxira/yo'lda hozircha yetarli)")
    c.font      = Font(name=FONT_NAME, bold=True, size=FONT_SZ, color="FFFFFF")
    c.fill      = _fill("6B6B6B")
    c.alignment = _align(indent=1)
    c.border    = _border()
    ws.row_dimensions[row].height = 28
    row += 1

    write_col_headers(ws, row)
    row += 1

    sort_cols = [sc for sc in ("surf_ord", "mark_ord", "size_ord", "_st_f", "tovar")
                 if sc in nobuy_df.columns]
    if not sort_cols:
        sort_cols = ["tovar"]
    for _, r in nobuy_df.sort_values(sort_cols).iterrows():
        row = write_product(ws, row, r, kanal)

    return row


def write_size_separator(ws, row, next_size_ord: int = 1):
    """
    Kichik format (≤1250) va katta format (>1250) orasiga
    aniq sarlavha satr qo'shadi.
    """
    if next_size_ord == 1:
        label = "── КАТТА ФОРМАТ  (>1250 мм) ──"
        bg    = "1A3A5C"   # to'q ko'k
        fg    = "FFFFFF"
    else:
        label = "── КИЧИК ФОРМАТ  (≤1250 мм) ──"
        bg    = "2C4A6E"
        fg    = "FFFFFF"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    c = ws.cell(row=row, column=1, value=label)
    c.font      = Font(name=FONT_NAME, bold=True, size=FONT_SZ - 1, color=fg)
    c.fill      = _fill(bg)
    c.alignment = _align(indent=2)
    med = Side(style="medium", color="444444")
    c.border    = Border(left=med, right=med, top=med, bottom=med)
    ws.row_dimensions[row].height = 22
    return row + 1


def fill_sheet(ws, cat_data, sorted_mode: bool, nobuy_data=None,
                cat_name: str | None = None, kanal: str = "asosiy"):
    today = datetime.now().strftime("%d.%m.%Y")
    # 2026-07-22: cat_name endi tashqaridan (build()) uzatiladi -- cat_data
    # BO'SH bo'lishi mumkin (bu kategoriyada faqat "buyurtma berilmagan"
    # tovarlar bor, hech biriga buyurtma yo'q), o'sha holatda .iloc[0]
    # xato berardi.
    cat = cat_name if cat_name else cat_data["kategoriya"].iloc[0]

    set_cols(ws)
    row = 1

    write_category_banner(ws, row, f"{cat}  \u2014  {today}")
    row += 1

    write_col_headers(ws, row)
    row += 1

    # 2026-07-14 (Huzayfa: "pastga tushursam ustun nomlari yo'q bo'lib
    # ketmoqda"): sarlavha (banner + ustun nomlari) muzlatiladi — Excel'ni
    # qancha pastga aylantirsangiz ham 1-2 qatorlar ko'rinib turadi.
    ws.freeze_panes = f"A{row}"

    if sorted_mode:
        active_surfs = [s for s in SURFACE_ORDER
                        if not cat_data[cat_data["surface"] == s].empty]
        show_surf_sep = len(active_surfs) > 1

        for surf in SURFACE_ORDER:
            sg = cat_data[cat_data["surface"] == surf]
            if sg.empty:
                continue

            if cat == "Лист":
                # ── Лист: maxsus guruhlash ────────────────────────────────────
                if surf == "":
                    # Plain/rangli listlar: agar boshqa surface ham bo'lsa separator
                    if show_surf_sep:
                        write_surface_row(ws, row, surf)
                        row += 1
                    # Kichik format avval, katta keyin; КИЧИК/КATTA yozuvi YO'Q
                    for _, r in sg.sort_values(["size_ord", "_st_f", "tovar"]).iterrows():
                        row = write_product(ws, row, r, kanal)
                else:
                    # Матовый / Глянцевый / Чёрный / Голд / Цветной:
                    # Har (marka, o'lcham) juftligi uchun: ── Матовый 201 (1220х2440) ──
                    sg = sg.copy()
                    sg["_size_dims"] = sg["tovar"].apply(get_size_dims)
                    groups = []
                    seen = set()
                    for marka in MARKA_ORDER + [m for m in sg["marka"].unique()
                                                if m not in MARKA_ORDER]:
                        dims_list = sorted(
                            sg[sg["marka"] == marka]["_size_dims"].unique(),
                            key=lambda d: (_size_dim_first(d), d)
                        )
                        for dims in dims_list:
                            key = (marka, dims)
                            if key in seen:
                                continue
                            seen.add(key)
                            gdf = sg[(sg["marka"] == marka) & (sg["_size_dims"] == dims)]
                            if not gdf.empty:
                                groups.append((marka, dims, gdf))

                    for marka, dims, gdf in groups:
                        parts = [surf]
                        if marka:
                            parts.append(marka)
                        if dims:
                            parts.append(f"({dims})")
                        label = "── " + " ".join(parts) + " ──"
                        row = write_list_group_row(ws, row, label)
                        for _, r in gdf.sort_values(["_st_f", "tovar"]).iterrows():
                            row = write_product(ws, row, r, kanal)

            else:
                # ── Труба / Профиль ───────────────────────────────────────────
                if show_surf_sep:
                    write_surface_row(ws, row, surf)
                    row += 1

                use_marka_sep = cat in {"Труба", "Профиль"}
                active_markas = [m for m in MARKA_ORDER if not sg[sg["marka"] == m].empty]

                for marka in MARKA_ORDER:
                    mg = sg[sg["marka"] == marka]
                    if mg.empty:
                        continue
                    if use_marka_sep and len(active_markas) > 1:
                        row = write_marka_row(ws, row, marka)

                    if cat == "Труба":
                        for _, r in mg.sort_values(["_dia_f", "_st_f", "tovar"]).iterrows():
                            row = write_product(ws, row, r, kanal)
                    elif cat == "Профиль":
                        for _, r in mg.sort_values(["_d1_f", "_d2_f", "_st_f", "tovar"]).iterrows():
                            row = write_product(ws, row, r, kanal)
    elif cat == "Чашка":
        chashka_rows = cat_data[cat_data["tovar"].str.match(r'^Чашк', case=False, na=False)]
        quz_rows     = cat_data[~cat_data["tovar"].str.match(r'^Чашк', case=False, na=False)]
        for _, r in chashka_rows.sort_values("tovar").iterrows():
            row = write_product(ws, row, r, kanal)
        if not quz_rows.empty:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
            c = ws.cell(row=row, column=1, value='── Қўзиқорин ──')
            c.font      = Font(name=FONT_NAME, bold=True, size=FONT_SZ - 1, color='222222')
            c.fill      = _fill(SURFACE_BG)
            c.alignment = _align(indent=1)
            c.border    = _border()
            ws.row_dimensions[row].height = 22
            row += 1
            for _, r in quz_rows.sort_values('tovar').iterrows():
                row = write_product(ws, row, r, kanal)
    elif not cat_data.empty:
        for _, r in cat_data.sort_values('tovar').iterrows():
            row = write_product(ws, row, r, kanal)

    # 2026-07-22: shu kategoriyada buyurtma berilmagan (zaxira/yo'lda
    # yetarli) tovarlar -- asosiy ro'yxatdan KEYIN, aniq ajratilgan
    # alohida bo'lim. Qarang: write_nobuy_section().
    if nobuy_data is not None and not nobuy_data.empty:
        row = write_nobuy_section(ws, row, nobuy_data, kanal)

    # 2026-07-21 (Huzayfa: "E/F/G ustunlar qulf holda tursin, faqat Min
    # Zaxira ochiq bo'lsin, qo'l tegib o'zgarib ketmasin"): varaq
    # himoyasi yoqiladi -- write_product() da H (Min Zaxira) allaqachon
    # Protection(locked=False) qilingan, qolgan hamma katak standart
    # qulflangan. Parol YO'Q -- Excel'ning "Review > Unprotect Sheet"
    # bilan istalgan payt ochish mumkin, bu faqat tasodifiy bosilib
    # ketishdan himoya, xavfsizlik devori emas.
    ws.protection.sheet = True


def build(df, meyor_yoq=None, nobuy=None, kanal: str = "asosiy"):
    """
    nobuy (2026-07-22, Huzayfa: "buyurtma berilmagan tovarlarni alohida
    ko'rsatish kerak"): df bilan bir xil tuzilishdagi DataFrame -- shu
    kanalda buyurtma=0 (hozircha kerak emas) bo'lgan tovarlar. Berilgan
    bo'lsa, HAR BIR kategoriya varag'ining OXIRIDA, aniq ajratilgan
    bo'lim sifatida ko'rsatiladi (write_nobuy_section()). Faqat df'da
    o'sha kategoriyadan buyurtma bo'lmasa ham (butun kategoriya
    "kerak emas"), varaq baribir yaratiladi -- shu kategoriya ham
    ko'rinishi kerak.
    """
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)
    cats_buy    = set(df['kategoriya'].dropna().astype(str)) if not df.empty else set()
    cats_nobuy  = (set(nobuy['kategoriya'].dropna().astype(str))
                   if nobuy is not None and not nobuy.empty else set())
    available   = cats_buy | cats_nobuy
    ordered     = [cat for cat in CATEGORIES if cat in available]
    used_titles = set()
    for cat in ordered:
        cd    = df[df['kategoriya'] == cat] if not df.empty else df.iloc[0:0]
        cd_nb = (nobuy[nobuy['kategoriya'] == cat]
                 if nobuy is not None and not nobuy.empty else None)
        if cd.empty and (cd_nb is None or cd_nb.empty):
            continue
        title = cat[:31]
        base  = title
        idx   = 1
        while title in used_titles:
            title = f"{base[:28]}_{idx}"
            idx  += 1
        used_titles.add(title)
        ws          = wb.create_sheet(title=title)
        sorted_mode = cat in SORTED_CATS
        fill_sheet(ws, cd, sorted_mode, nobuy_data=cd_nb, cat_name=cat, kanal=kanal)

    # ── "Меъёр йўқ" varag'i (2026-07-14, Huzayfa so'rovi) ────────────────
    # Min zaxirasi belgilanmagan tovarlar buyurtma hisobiga KIRMAYDI va
    # ilgari hech qayerda ko'rinmasdi ("nega falon tovar ro'yxatda yo'q?"
    # savolining sababi). Endi ular oxirgi varaqda ro'yxat bo'lib chiqadi —
    # ko'rib chiqib Min_Zaxira.xlsx da meyor belgilash mumkin.
    # DIQQAT: sarlavha ataylab "Товар" (buyurtma varaqlaridagi "Tovar nomi"
    # emas) — buyurtma_tekshir() bu varaqni buyurtma varag'i deb adashib
    # o'qimasligi uchun.
    if meyor_yoq is not None and len(meyor_yoq):
        ws = wb.create_sheet(title="Меъёр йўқ")
        ws.append(["Товар", "Қолдиқ"])
        for c in ws[1]:
            c.font = Font(bold=True, size=11)
            c.fill = _fill("FFF2CC")
            c.border = _border()
        for _, r in meyor_yoq.iterrows():
            ws.append([str(r["tovar"]), int(r["qoldiq"])])
        ws.column_dimensions["A"].width = 48
        ws.column_dimensions["B"].width = 12
        ws.freeze_panes = "A2"
    return wb


def main():
    from pathlib import Path
    import sys
    BASE = Path(__file__).resolve().parent
    src  = BASE / "chiqish" / "NEJAVIYKA_POWER_BI.xlsx"
    if not src.exists():
        print(f"Fayl topilmadi: {src}")
        return
    df_raw = pd.read_excel(src, sheet_name="Инвентар")
    df_raw.columns = [str(c).strip() for c in df_raw.columns]
    col_map = {
        "Товар": "tovar", "Қолдиқ": "qoldiq",
        "Йўлда_Жами": "yoldagi", "Мин_Захира": "min_zaxira",
    }
    df_raw = df_raw.rename(columns={k: v for k, v in col_map.items() if k in df_raw.columns})
    for c in ["qoldiq", "yoldagi", "min_zaxira"]:
        if c not in df_raw.columns:
            df_raw[c] = 0
    df_full = calculate(df_raw)
    # 2026-07-22: calculate() endi filtrlamaydi (buyurtma==0 qatorlar ham
    # qaytadi) -- standalone skript uchun eski xatti-harakatni saqlab
    # qolish uchun shu yerda ajratamiz (nobuy bo'limi bu yerda ko'rsatilmaydi,
    # faqat services.py::asosiy_styled_excel_yarat -- bot yo'li -- buni
    # qo'llaydi).
    df   = df_full[df_full["buyurtma"] > 0].copy()
    wb   = build(df)
    from datetime import datetime
    out  = BASE / "chiqish" / f"Buyurtma_taklif_asosiy_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    print('Saqlandi: ' + str(out))
    return str(out)


if __name__ == '__main__':
    main()
