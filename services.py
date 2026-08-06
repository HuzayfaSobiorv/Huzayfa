"""
services.py — Biznes logika: inventar, kamomat, buyurtma, konteyner
"""
import difflib
import json
import logging
import re
from datetime import datetime
from io import BytesIO
from pathlib import Path

import pandas as pd

logger = logging.getLogger(__name__)

from config import (
    BASE_DIR, DATA_FILE, BOT_HOLAT_DIR, VARAQLAR,
    CAT_SHEET, AKSESSUAR_KATS, get_inv, get_kont,
    KONTEYNER_TARIX_FILE, XITOY_PARSED_DIR, CH_KEY,
)
from common import normalize_product_name, atomic_json_write
# 2026-07-29 (Huzayfa: "Oldingi tasdiq: name 't' is not defined" xato bilan
# duch keldi): `zakaz_preview_text()` `t(lang, ...)` chaqirardi, lekin bu
# fayl uni HECH QACHON import qilmagan edi — funksiya chaqirilgan ZAHOTI
# (kanal/fayldan qat'i nazar) NameError berardi, handlers.py'dagi try/except
# buni yutib, o'rniga xato matnining o'zini "preview" sifatida ko'rsatardi.
from texts import t

# Lokal aliaslar — config cache funksiyalari
_get_inv  = get_inv
_get_kont = get_kont


# ── Og'irlik lookup (vazn_lookup.xlsx — Power BI dan mustaqil) ───────────────
# Fayl: BASE_DIR / "vazn_lookup.xlsx"
# Sheet: "Вазн" | Ustunlar: Товар номи (A) | 1 дона вазни кг (B)
# Qoida: stenka -0.05 (Xitoy), ISTISNO: 1.35 va 1.45 stenka ayrilmaydi
_vazn_cache: dict | None = None
VAZN_FILE = BASE_DIR / "vazn_lookup.xlsx"

def vazn_map_yuklash(force: bool = False) -> dict:
    """
    vazn_lookup.xlsx fayldan tovar_nomi → og'irlik (kg) lug'atini qaytaradi.
    Power BI yangilanganda ham bu fayl o'zgarmaydi — mustaqil saqlanadi.
    """
    global _vazn_cache
    if _vazn_cache is not None and not force:
        return _vazn_cache
    try:
        import openpyxl
        if not VAZN_FILE.exists():
            logger.warning(f"vazn_lookup.xlsx topilmadi: {VAZN_FILE}")
            _vazn_cache = {}
            return _vazn_cache
        wb = openpyxl.load_workbook(VAZN_FILE, data_only=True, read_only=True)
        ws = wb.active
        result = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            tovar = str(row[0]).strip()
            vazn  = row[1]
            if isinstance(vazn, (int, float)) and vazn > 0:
                result[tovar] = float(vazn)
        wb.close()
        _vazn_cache = result
        logger.info(f"vazn_map_yuklash: {len(result)} ta tovar og'irligi yuklandi.")
        return result
    except Exception as e:
        logger.error(f"vazn_map_yuklash xato: {e}")
        _vazn_cache = {}
        return {}


def tovar_vazni_pb(tovar_nomi: str) -> float | None:
    """
    vazn_lookup.xlsx dan tovar og'irligini qaytaradi. Topilmasa — None.
    """
    vmap = vazn_map_yuklash()
    return vmap.get(str(tovar_nomi).strip())


# ── Whitelist ────────────────────────────────────────────────────────────────
_WHITELIST_FILE = BOT_HOLAT_DIR / "whitelist.json"


def whitelist_yuklash() -> set[int]:
    """Ruxsat berilgan foydalanuvchilar ID larini qaytaradi."""
    if not _WHITELIST_FILE.exists():
        return set()
    try:
        return set(json.loads(_WHITELIST_FILE.read_text(encoding="utf-8")))
    except Exception:
        return set()


def whitelist_saqlash(ids: set[int]) -> None:
    atomic_json_write(_WHITELIST_FILE, sorted(ids))


def whitelist_qosh(uid: int) -> bool:
    """ID ni whitelist ga qo'shadi. True — yangi qo'shildi, False — allaqachon bor."""
    wl = whitelist_yuklash()
    if uid in wl:
        return False
    wl.add(uid)
    whitelist_saqlash(wl)
    return True


def whitelist_ochir(uid: int) -> bool:
    """ID ni whitelist dan o'chiradi. True — o'chirildi, False — topilmadi."""
    wl = whitelist_yuklash()
    if uid not in wl:
        return False
    wl.discard(uid)
    whitelist_saqlash(wl)
    return True


# ── Filial biriktirish (2026-07-24, "har filial o'z qoldig'ini ko'rsin") ──────
# Ikki bosqichli: (1) foydalanuvchi /start bosib filial tanlaganda —
# PENDING (hali admin tasdiqlamagan) sifatida saqlanadi; (2) admin
# /adduser bilan qabul qilganda — YAKUNIY (committed) fayliga ko'chadi.
_FILIAL_SOROV_FILE    = BOT_HOLAT_DIR / "filial_sorovlar.json"
_USER_FILIAL_FILE     = BOT_HOLAT_DIR / "user_filiallar.json"
FILIAL_ESKI_DEFAULT   = "Сергили база"   # 2026-07-24 Huzayfa qarori: eski
                                          # userlar migratsiyada shu filialga


def _json_dict_yuklash(path) -> dict:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def filial_sorov_saqla(uid: int, filial: str, ism: str = "", username: str = "") -> None:
    """Foydalanuvchi /start bosib filial tanlagach — admin hali tasdiqlamagan
    holatda PENDING ro'yxatga yozadi. adduser_cmd shu yerdan o'qib, whitelist
    ga qo'shilgach YAKUNIY joyga ko'chiradi."""
    d = _json_dict_yuklash(_FILIAL_SOROV_FILE)
    d[str(uid)] = {"filial": filial, "ism": ism, "username": username,
                    "vaqt": datetime.now().strftime("%d.%m.%Y %H:%M")}
    atomic_json_write(_FILIAL_SOROV_FILE, d, indent=2)


def filial_sorov_olish(uid: int) -> dict | None:
    return _json_dict_yuklash(_FILIAL_SOROV_FILE).get(str(uid))


def filial_sorov_ochir(uid: int) -> None:
    d = _json_dict_yuklash(_FILIAL_SOROV_FILE)
    if str(uid) in d:
        del d[str(uid)]
        atomic_json_write(_FILIAL_SOROV_FILE, d, indent=2)


def filial_biriktir(uid: int, filial: str, ism: str = "", username: str = "") -> None:
    """Userga YAKUNIY (tasdiqlangan) filialni biriktiradi."""
    d = _json_dict_yuklash(_USER_FILIAL_FILE)
    mavjud = d.get(str(uid), {})
    d[str(uid)] = {
        "filial":   filial,
        "ism":      ism or mavjud.get("ism", ""),
        "username": username or mavjud.get("username", ""),
    }
    atomic_json_write(_USER_FILIAL_FILE, d, indent=2)


def user_filiali_olish(uid: int) -> str | None:
    """Bitta userning filialini qaytaradi (topilmasa None — qidiruvda
    filial qatori shunchaki chiqmaydi).

    2026-07-24 TUZATISH: ilgari faqat user_filiallar.json'dan o'qir edi —
    eski (whitelist'da bor-u hali filial yozuvi bo'lmagan) userlar uchun
    bu HAR DOIM None qaytarardi, chunki avtomatik "Сергили база" migratsiyasi
    faqat admin "Foydalanuvchilar ro'yxati" Excel tugmasini bosgandagina
    (user_filiallari_yuklash() orqali) ishga tushardi — Qidiruvda emas.
    Huzayfa buni sinab ko'rib, eski userlarga filial qatori umuman
    chiqmayotganini aniqladi. Endi shu yerda ham fallback bor: agar aniq
    yozuv topilmasa-yu, user whitelist'da bo'lsa — FILIAL_ESKI_DEFAULT
    qaytariladi (diskka yozmasdan, faqat shu chaqiruv uchun — arzon va
    har doim yangi natija beradi)."""
    row = _json_dict_yuklash(_USER_FILIAL_FILE).get(str(uid))
    if row and row.get("filial"):
        return row["filial"]
    if uid in whitelist_yuklash():
        return FILIAL_ESKI_DEFAULT
    return None


def user_ism_username_yangilash(uid: int, ism: str = "", username: str = "") -> None:
    """2026-07-24 (Huzayfa: "tizim doimiy ishlashi kerak — yangi userlar
    qo'shilganda ham ismi/username avtomatik yozilishi kerak"): har safar
    whitelist'dagi user botga biror narsa yozganda yoki tugma bossa
    (handlers.py::text_keldi / callback_handler dan) chaqiriladi — Ism/
    Username maydonlarini eng so'nggi Telegram ma'lumoti bilan yangilab
    turadi (filial/boshqa maydonlarga tegmaydi). Bo'sh qiymat yuborilsa
    yoki hech narsa o'zgarmagan bo'lsa — diskka yozilmaydi (arzon,
    idempotent chaqiriq)."""
    if not ism and not username:
        return
    d = _json_dict_yuklash(_USER_FILIAL_FILE)
    mavjud = d.get(str(uid), {})
    yangi_ism      = ism or mavjud.get("ism", "")
    yangi_username = username or mavjud.get("username", "")
    if yangi_ism == mavjud.get("ism", "") and yangi_username == mavjud.get("username", ""):
        return
    d[str(uid)] = {
        "filial":   mavjud.get("filial", FILIAL_ESKI_DEFAULT),
        "ism":      yangi_ism,
        "username": yangi_username,
    }
    atomic_json_write(_USER_FILIAL_FILE, d, indent=2)


async def eski_userlar_ism_toldir(bot) -> int:
    """2026-07-24: Ism/Username hali bo'sh bo'lgan whitelist userlar uchun
    Telegram'ning o'zidan (bot.get_chat) haqiqiy ism/username so'rab,
    topilsa saqlaydi — FAQAT bot ular bilan avval kontaktda bo'lgan
    bo'lsa ishlaydi (get_chat shuni talab qiladi). Har bir user uchun
    xato bo'lsa jimgina o'tkazib yuboriladi (bloklangan/hech qachon
    yozmagan va h.k.). Qaytaradi: nechta user yangilanganini (son).

    2026-07-27 TUZATISH (Huzayfa: taqdimot oldidan kamida 30 ta user
    qo'shilishi kutilmoqda — bu funksiya "Userlar ro'yxati" tugmasi
    bosilganda ishga tushadi, HAR BIR whitelist user uchun `await
    bot.get_chat(...)` qiladi, ya'ni ko'p soniyaga cho'zilishi mumkin.
    ESKI kod faylni FAQAT bitta marta, tsikl BOSHIDA o'qib, tsikl
    OXIRIDA o'sha eskirgan nusxa ustiga yozardi — shu paytda (masalan
    30 ta userdan biri tugma bossa/yozsa) `user_ism_username_yangilash`
    yoki yangi user tasdiqlanganda `filial_biriktir` yozgan har qanday
    o'zgarish shu yerda JIMGINA YO'QOLIB KETARDI (lost update race).
    Endi: get_chat natijalari alohida to'planadi, DISKKA FAQAT BITTA
    marta — tsikl TUGAGACH, faylni QAYTA (yangi holatda) o'qib, faqat
    o'zimiz hisoblagan ism/username maydonlarini shu YANGI nusxaga
    qo'shib yozamiz. Bu race oynasini soniyalardan bitta sinxron
    yozish amaliga qisqartiradi."""
    d  = _json_dict_yuklash(_USER_FILIAL_FILE)
    wl = whitelist_yuklash()
    natijalar: dict[str, tuple[str, str]] = {}
    for uid in wl:
        row = d.get(str(uid), {})
        if row.get("ism") and row.get("username"):
            continue  # ikkalasi ham allaqachon bor — o'tkazib yuboramiz
        try:
            chat = await bot.get_chat(uid)
            full_name = " ".join(filter(None, [getattr(chat, "first_name", None),
                                                 getattr(chat, "last_name", None)])).strip()
            username = f"@{chat.username}" if getattr(chat, "username", None) else ""
        except Exception:
            continue
        if not full_name and not username:
            continue
        natijalar[str(uid)] = (full_name, username)
    if natijalar:
        d2 = _json_dict_yuklash(_USER_FILIAL_FILE)  # tsikl paytidagi o'zgarishlarni yo'qotmaslik uchun QAYTA o'qiladi
        for k, (full_name, username) in natijalar.items():
            mavjud = d2.get(k, {})
            d2[k] = {
                "filial":   mavjud.get("filial", FILIAL_ESKI_DEFAULT),
                "ism":      full_name or mavjud.get("ism", ""),
                "username": username or mavjud.get("username", ""),
            }
        atomic_json_write(_USER_FILIAL_FILE, d2, indent=2)
    return len(natijalar)


def user_filiallari_yuklash() -> dict:
    """{uid_str: {"filial":.., "ism":.., "username":..}} — barcha
    tasdiqlangan userlar. Whitelist'da bor-u filiali hali yo'q eski
    userlarni avtomatik FILIAL_ESKI_DEFAULT ga biriktirib qo'yadi
    (Huzayfa qarori, 2026-07-24: "hammasini Sergili filialiga o'tkazamiz
    — barchasi Sergili sklad sotuvchilari") — idempotent, xavfsiz."""
    d  = _json_dict_yuklash(_USER_FILIAL_FILE)
    wl = whitelist_yuklash()
    ozgardi = False
    for uid in wl:
        if str(uid) not in d:
            d[str(uid)] = {"filial": FILIAL_ESKI_DEFAULT, "ism": "", "username": ""}
            ozgardi = True
    if ozgardi:
        atomic_json_write(_USER_FILIAL_FILE, d, indent=2)
    return d


def userlar_excel_yarat() -> BytesIO | None:
    """Admin uchun: barcha whitelist foydalanuvchilarning Ism/Username/ID/
    Filial ro'yxatini Excel qilib qaytaradi. Bo'sh bo'lsa None."""
    wl = whitelist_yuklash()
    if not wl:
        return None
    filiallar = user_filiallari_yuklash()
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Userlar"
    sarlavha = ["Ism", "Username", "ID", "Filial"]
    ws.append(sarlavha)
    hdr_fill = PatternFill("solid", fgColor="305496")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col_i, _ in enumerate(sarlavha, start=1):
        c = ws.cell(row=1, column=col_i)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center")

    for uid in sorted(wl):
        row = filiallar.get(str(uid), {})
        ws.append([
            row.get("ism") or "—",
            row.get("username") or "—",
            uid,
            row.get("filial") or "—",
        ])

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ── Konteyner qo'shish tarixi (fayl o'chsa ham unutilmasin) ───────────────────
# MUHIM: kalit ISO emas, ISO+sana ("ISO|sana" shaklida) — chunki jismoniy
# konteyner raqamlari (ISO) dunyoda qayta-qayta ishlatiladi va bitta ISO
# boshqa sanada butunlay YANGI yuk bilan qaytib kelishi mumkin. Faqat aynan
# bir xil ISO+sana (ya'ni xuddi shu yetkazib berish) qayta bloklanadi.

def konteyner_tarix_kalit(iso: str, sana: str) -> str:
    return f"{iso}|{sana}"


def konteyner_tarix_olish() -> set[str]:
    """Bir marta tasdiqlangan konteynerlar (ISO+sana kalit shaklida)
    to'plamini qaytaradi. xitoy_parsed papkasidagi fayl keyinchalik
    o'chirilgan bo'lsa ham, shu kalit qayta "yangi" deb qo'shilib
    ketmasligi uchun ishlatiladi.

    Birinchi chaqiriqda (tarix fayli hali yo'q bo'lsa) hozirgi
    xitoy_parsed papkasidagi BARCHA konteynerlar bilan bir martalik
    "boshlang'ich to'ldirish" qilinadi — shu orqali oldindan (bu tarix
    tizimi yaratilishidan oldin) qo'shilgan konteynerlar ham himoyalanadi."""
    if not KONTEYNER_TARIX_FILE.exists():
        boshlangich = set()
        if XITOY_PARSED_DIR.exists():
            for f in XITOY_PARSED_DIR.glob("*.xlsx"):
                stem = f.stem[:-2] if f.stem.endswith("_D") else f.stem
                # 2026-07-13: "F_" (tezkor/aksessuar) prefiksi ham olib
                # tashlanadi -- aks holda (masalan "F_AksessuarKont_26.06.2026")
                # noto'g'ri iso="F" bo'lib chiqib, bu konteyner keyinchalik
                # dedup tekshiruvida "ko'rinmas" bo'lib qolardi (konteyner_qosh.py
                # dagi bir xil tuzatish bilan izchil).
                if stem.startswith("F_"):
                    stem = stem[2:]
                # 2026-08-06 (Huzayfa: "66 kun kechikkan konteyner hali ham
                # yo'lda ko'rsatilmoqda, u allaqachon kelgan"): ILGARI
                # `stem.partition("_")` ishlatilardi -- BIRINCHI "_"dan
                # bo'lardi. Pseudo-ID nomlar ("NORAQAM_07_04_2026_list_11")
                # ko'p "_" ichiga oladi -- natijada iso="NORAQAM" (kesilgan)
                # bo'lib, sana maydoniga esa "07_04_2026_list_11_07.04.2026"
                # kabi parslanmaydigan chalkash matn tushib qolardi. Bu ISO
                # keyinchalik dedup tekshiruvida (iso_boyicha_yangilarini_ajrat)
                # HAQIQIY to'liq pseudo-ID bilan MOS KELMASDI -- konteyner
                # doim "hali umuman ko'rilmagan" deb topilib, qayta-qayta YANGI
                # sifatida qo'shilib turardi (garchi allaqachon KELDI qilingan
                # bo'lsa ham) -- xuddi shu naqsh main.py/_iso_from_stem'da
                # oldin tuzatilgan edi, bu yerda unutilgan ekan. Endi OXIRGI
                # "_"dan bo'linadi (main.py bilan izchil).
                if "_" in stem:
                    iso, sana = stem.rsplit("_", 1)
                else:
                    iso, sana = stem, ""
                if iso:
                    boshlangich.add(konteyner_tarix_kalit(iso, sana))
        try:
            atomic_json_write(KONTEYNER_TARIX_FILE, sorted(boshlangich))
        except Exception as e:
            logger.error(f"konteyner_tarix boshlang'ich to'ldirishda xato: {e}")
        return boshlangich
    try:
        tarix = set(json.loads(KONTEYNER_TARIX_FILE.read_text(encoding="utf-8")))
    except Exception:
        return set()

    # 2026-08-06: yuqoridagi bug bilan ILGARI yozilgan buzilgan yozuvlarni
    # (masalan "NORAQAM|07_04_2026_list_11_07.04.2026" -- sana maydoni
    # haqiqiy sana emas) avtomatik aniqlab, bir martalik tuzatib qo'yamiz --
    # aks holda kod tuzatilsa ham, allaqachon diskka yozilgan buzuq
    # yozuvlar muammoni davom ettiraveradi.
    from konteyner_qosh import _sana_parse
    tuzatilgan = set()
    ozgardimi = False
    for kalit in tarix:
        iso, ajratuvchi, sana = str(kalit).partition("|")
        if not ajratuvchi or _sana_parse(sana) is not None:
            tuzatilgan.add(kalit)
            continue
        stem = f"{iso}_{sana}" if iso else sana
        if "_" in stem:
            iso2, sana2 = stem.rsplit("_", 1)
        else:
            iso2, sana2 = stem, ""
        tuzatilgan.add(konteyner_tarix_kalit(iso2, sana2))
        ozgardimi = True
    if ozgardimi:
        try:
            atomic_json_write(KONTEYNER_TARIX_FILE, sorted(tuzatilgan))
            logger.info("konteyner_tarix: eski buzilgan yozuvlar avtomatik tuzatildi")
        except Exception as e:
            logger.error(f"konteyner_tarix tuzatishda xato: {e}")
        return tuzatilgan
    return tarix


def konteyner_tarix_qoshish(konteynerlar: list) -> None:
    """Tasdiqlangan konteynerlarni doimiy tarixga qo'shadi.
    `konteynerlar` — [{"iso":.., "sana":..}, ...] yoki (iso, sana) juftliklari."""
    if not konteynerlar:
        return
    tarix = konteyner_tarix_olish()
    for k in konteynerlar:
        if isinstance(k, dict):
            iso, sana = k["iso"], k["sana"]
        else:
            iso, sana = k
        tarix.add(konteyner_tarix_kalit(iso, sana))
    try:
        atomic_json_write(KONTEYNER_TARIX_FILE, sorted(tarix))
    except Exception as e:
        logger.error(f"konteyner_tarix_qoshish yozishda xato: {e}")


# ── Qo'shimcha adminlar (dinamik, .env kerak emas) ────────────────────────────
# 2026-07-16 (Huzayfa: serverga jismonan/AnyDesk orqali kirmasdan admin
# holatini boshqarish kerak bo'ldi — eski ADMIN_IDS .env faylida edi, uni
# o'zgartirish har safar serverga kirishni talab qilardi). Shu nuqtadan
# boshlab "kim admin" degan savolga javob IKKI manbadan keladi:
# SUPER_ADMIN_ID (.env, o'zgarmas, yagona bosh admin) + shu fayl (dinamik,
# /addadmin va /removeadmin komandalari orqali, Telegram ichidan
# boshqariladi). Eski .env'dagi ko'p qiymatli ADMIN_IDS ENDI avtorizatsiya
# uchun ishlatilmaydi (config.ADMIN_IDS hali mavjud, lekin faqat tarixiy/
# backup sifatida qoladi) — shuning uchun admin_idlari() ishlatilishi kerak,
# config.ADMIN_IDS TO'G'RIDAN-TO'G'RI EMAS.
_ADMIN_FILE = BOT_HOLAT_DIR / "qoshimcha_adminlar.json"


def qoshimcha_admin_yuklash() -> set[int]:
    if not _ADMIN_FILE.exists():
        return set()
    try:
        return set(json.loads(_ADMIN_FILE.read_text(encoding="utf-8")))
    except Exception:
        return set()


def qoshimcha_admin_saqlash(ids: set[int]) -> None:
    atomic_json_write(_ADMIN_FILE, sorted(ids))


def qoshimcha_admin_qosh(uid: int) -> bool:
    """ID ni admin qilib tayinlaydi. True — yangi qo'shildi, False — allaqachon bor."""
    ids = qoshimcha_admin_yuklash()
    if uid in ids:
        return False
    ids.add(uid)
    qoshimcha_admin_saqlash(ids)
    return True


def qoshimcha_admin_ochir(uid: int) -> bool:
    """ID ni admin ro'yxatidan chiqaradi. True — chiqarildi, False — topilmadi."""
    ids = qoshimcha_admin_yuklash()
    if uid not in ids:
        return False
    ids.discard(uid)
    qoshimcha_admin_saqlash(ids)
    return True


def admin_idlari() -> set[int]:
    """Hozirgi HAQIQIY admin ID'lar to'plami — SUPER_ADMIN_ID + dinamik
    ro'yxat (/addadmin bilan tayinlanganlar). Har qanday avtorizatsiya
    tekshiruvi shu funksiyadan foydalanishi kerak."""
    from config import SUPER_ADMIN_ID
    ids = qoshimcha_admin_yuklash()
    if SUPER_ADMIN_ID:
        ids.add(SUPER_ADMIN_ID)
    return ids


# ── Foydalanuvchi so'rovlari (buyurtma so'rovi va h.k.) ───────────────────────
# 2026-07-16 (Huzayfa): filial foydalanuvchilari mahsulot kamayganini yozib
# yuborishi mumkin bo'lgan kanal. Har bir xabar SHU YERGA (doimiy jurnalga)
# yoziladi — adminga darhol xabar SHAKLIDA yubormaymiz (20+ kishi bo'lsa,
# uzluksiz xabar oqimi bo'lib qoladi, deb Huzayfa aytdi) — buning o'rniga
# admin xohlagan vaqtida "📊 Foydalanuvchi so'rovlari" tugmasi bilan
# to'plangan hammasini Excel qilib oladi. Jurnal HECH QACHON tozalanmaydi
# (Huzayfa: "eski so'rovlar Excelda pastga tushib ketaversin") — Excel har
# safar ENG YANGISI TEPADA bo'ladigan tartibda tuziladi.
_SOROV_FILE = BOT_HOLAT_DIR / "foydalanuvchi_sorovlari.json"


def sorov_qoshish(uid: int, ism: str, matn: str) -> None:
    try:
        d = []
        if _SOROV_FILE.exists():
            d = json.loads(_SOROV_FILE.read_text(encoding="utf-8"))
        d.append({
            "vaqt":    datetime.now().strftime("%d.%m.%Y %H:%M"),
            "user_id": uid,
            "ism":     ism,
            "matn":    matn,
        })
        atomic_json_write(_SOROV_FILE, d, indent=2)
    except Exception as e:
        logger.error(f"sorov_qoshish xato: {e}")


def sorovlar_olish() -> list:
    if not _SOROV_FILE.exists():
        return []
    try:
        return json.loads(_SOROV_FILE.read_text(encoding="utf-8"))
    except Exception:
        return []


def sorovlar_excel_yarat() -> BytesIO | None:
    """To'plangan foydalanuvchi so'rovlarini Excel qilib qaytaradi (eng
    yangisi tepada). Hech narsa bo'lmasa None qaytaradi."""
    sorovlar = sorovlar_olish()
    if not sorovlar:
        return None
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "So'rovlar"
    sarlavha = ["Vaqt", "Foydalanuvchi", "ID", "Xabar"]
    ws.append(sarlavha)
    hdr_fill = PatternFill("solid", fgColor="305496")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col_i, _ in enumerate(sarlavha, start=1):
        c = ws.cell(row=1, column=col_i)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center")

    for row in reversed(sorovlar):   # eng yangisi tepada
        ws.append([row.get("vaqt", ""), row.get("ism", ""),
                   row.get("user_id", ""), row.get("matn", "")])

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 60
    for r in ws.iter_rows(min_row=2):
        r[3].alignment = Alignment(wrap_text=True, vertical="top")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def rasm_pending_iso_royxati() -> set[str]:
    """Rasm allaqachon guruhga yuborilgan, lekin hali qo'lda KELDI
    qilinmagan konteynerlarning ISO raqamlari. 2026-07-16 (Huzayfa):
    "Yo'ldagi yuklar" Excel'idan shu konteynerlarni chiqarib tashlash
    uchun — aks holda foydalanuvchi (kimga rasmi allaqachon guruhga
    ketgan bo'lsa ham) bu yukni "hali yo'lda" deb ko'radi. Buyurtma
    hisob-kitobiga (zanjir_sim) BU TA'SIR QILMAYDI — u main.py orqali,
    fayl statusidan (hali _D ga o'zgarmagan) hisoblanadi, alohida.
    Fayl nomi qoidasi handlers.py::_iso_from_stem bilan bir xil bo'lishi
    kerak (F_ prefiksi tashlanadi, oxirgi "_sana" qismi olib tashlanadi)."""
    f = BOT_HOLAT_DIR / "rasm_yuborilgan.json"
    if not f.exists():
        return set()
    try:
        d = json.loads(f.read_text(encoding="utf-8"))
    except Exception:
        return set()
    isos = set()
    for fname in d.keys():
        stem = fname[:-5] if fname.lower().endswith(".xlsx") else fname
        if stem.startswith("F_"):
            stem = stem[2:]
        iso = stem.rsplit("_", 1)[0]
        if iso:
            isos.add(iso)
    return isos


def kirish_ruxsati(uid: int) -> bool:
    """Foydalanuvchi kirish huquqiga ega ekanligini tekshiradi."""
    if uid in admin_idlari():
        return True
    return uid in whitelist_yuklash()


def vazn_lookup_yangilash(yangi_vazn: dict) -> int:
    """
    Xitoy faylidan olingan yangi tovar vaznlarini vazn_lookup.xlsx ga qo'shadi.
    FAQAT mavjud bo'lmagan tovarlar qo'shiladi — eski yozuvlar o'ZGARMAYDI.
    Qaytaradi: qo'shilgan tovarlar soni.
    """
    if not yangi_vazn:
        return 0
    global _vazn_cache
    try:
        import openpyxl
        mavjud = vazn_map_yuklash()
        yangilar = {k: v for k, v in yangi_vazn.items()
                    if str(k).strip() and k not in mavjud
                    and isinstance(v, (int, float)) and v > 0}
        if not yangilar:
            return 0
        if VAZN_FILE.exists():
            wb = openpyxl.load_workbook(VAZN_FILE)
        else:
            wb = openpyxl.Workbook()
            wb.active.append(["Товар номи", "1 дона вазни кг"])
        ws = wb.active
        for nom, vazn in sorted(yangilar.items()):
            ws.append([nom, round(float(vazn), 3)])
            logger.info(f"vazn_lookup: yangi tovar qo'shildi — {nom}: {vazn} kg")
        wb.save(VAZN_FILE)
        # Keshni yangilaymiz
        _vazn_cache = None
        vazn_map_yuklash()
        return len(yangilar)
    except Exception as e:
        logger.error(f"vazn_lookup_yangilash xato: {e}")
        return 0


def xlsx_mavjud() -> bool:
    return DATA_FILE.exists()


def _num(df: pd.DataFrame, col: str):
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)


def inventar_olish(kanal: str | None = None) -> pd.DataFrame:
    from common import keraksizmi
    try:
        df = _get_inv().copy()
        for col in ["Қолдиқ", "Мин_Захира", "Фарқ", "Кун_Етади", "Йўлда_Жами"]:
            _num(df, col)
        if kanal and "Тур" in df.columns:
            if kanal == "sex":
                df = df[df["Тур"] == "ЦЕХ🏭"]
            elif kanal in ("asosiy", "osh"):
                df = df[df["Тур"] != "ЦЕХ🏭"]
        if "Товар" in df.columns:
            df = df[~df["Товар"].apply(keraksizmi)].copy()
        return df.reset_index(drop=True)
    except Exception as e:
        logger.error(f"inventar_olish: {e}")
        return pd.DataFrame()


def _kamomat_filter(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "Холат" not in df.columns:
        return pd.DataFrame()
    return df[df["Холат"].isin(["🔴 КРИТИК", "🟡 ПАСТ"])].copy()


def kamomat_olish(kanal: str) -> pd.DataFrame:
    try:
        return _kamomat_filter(inventar_olish(kanal)).reset_index(drop=True)
    except Exception as e:
        logger.error(f"kamomat_olish: {e}")
        return pd.DataFrame()


def kritiklar_olish(kanal: str, limit: int = 15) -> pd.DataFrame:
    df = inventar_olish(kanal)
    if df.empty or "Холат" not in df.columns:
        return pd.DataFrame()
    df = df[df["Холат"] == "🔴 КРИТИК"].copy()
    if df.empty:
        return df
    sort_cols = [c for c in ["Кун_Етади", "Фарқ", "Товар"] if c in df.columns]
    return df.sort_values(sort_cols).head(limit).reset_index(drop=True)


def kritiklar_text(df: pd.DataFrame, lang: str) -> str:
    if df.empty:
        return "Bugun kritik tovar yo'q." if lang == "lat" else "Бугун критик товар йўқ."
    title = "Bugungi eng xavfli kritiklar:" if lang == "lat" else "Бугунги энг хавфли критиклар:"
    lines = [title]
    for i, row in df.iterrows():
        tovar = str(row.get("Товар", "?"))
        qoldiq = int(row.get("Қолдиқ", 0))
        min_z = int(row.get("Мин_Захира", 0))
        kun = row.get("Кун_Етади", "")
        farq = int(row.get("Фарқ", 0))
        lines.append(
            f"{i + 1}. {tovar}\n"
            f"   Qoldiq: {qoldiq} | Min: {min_z} | Farq: {farq} | Kun: {kun}"
        )
    return "\n".join(lines)


# ── Kirill/lotin bir xil deb hisoblovchi normalizatsiya (umumiy qidiruv) ──────
_CYR2LAT = {
    'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'yo',
    'ж': 'j', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
    'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
    'ф': 'f', 'х': 'x', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'sh',
    'ъ': '', 'ы': 'i', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya',
    'қ': 'q', 'ғ': 'g', 'ў': 'o', 'ҳ': 'h',
}


def _translit(s: str) -> str:
    """Kirillni lotinga o'giradi (harf-baharf); lotin harflar o'zgarishsiz qoladi."""
    return "".join(_CYR2LAT.get(ch, ch) for ch in s.lower())


def _qidiruv_normalize(s: str) -> str:
    """Qidiruv uchun: kirill/lotin, katta/kichik harf va '.'/',' farqini
    yo'qotadi; so'z chegaralarini (bo'sh joy) saqlab qoladi — pastda
    tokenlarga ajratish (fuzzy qidiruv) uchun kerak."""
    s = _translit(str(s))
    s = s.replace(".", ",")
    s = re.sub(r"[\'ʻʼ`’]", "", s)
    s = re.sub(r"[^a-z0-9,]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _term_pattern(term: str) -> str:
    """Qidiruv so'zi -> regex. 2026-07-14: raqam bilan boshlanib/tugasa
    ANCHOR qo'yiladi -- "0,6" endi "0,65" ga, "19" esa "190" ga mos
    kelmaydi (grafik qidiruvdagi 2026-07-11 tuzatishning umumiy qidiruvga
    ko'chirilishi -- bu yerda o'sha bug qolib ketgan edi)."""
    pat = re.escape(term)
    if term and term[0].isdigit():
        pat = r"(?<!\d)" + pat
    if term and term[-1].isdigit():
        pat = pat + r"(?!\d)"
    return pat


def _fuzzy_score(terms: list, name_tokens: list) -> float:
    """Har bir so'rov so'zi uchun nom ichidagi eng yaqin so'zni topib,
    o'rtacha o'xshashlik darajasini qaytaradi (0..1)."""
    if not terms or not name_tokens:
        return 0.0
    total = 0.0
    for term in terms:
        best = max(
            (difflib.SequenceMatcher(None, term, tok).ratio() for tok in name_tokens),
            default=0.0,
        )
        total += best
    return total / len(terms)


def qidiruv_olish(query: str, kanal: str | None = None, limit: int = 10) -> pd.DataFrame:
    """
    Umumiy (kategoriyasiz) qidiruv — butun inventar bo'yicha.
    Kirill va lotin yozuvi farqi hisobga olinmaydi (masalan "труба" va "truba"
    bir xil natija beradi), '.' va ',' ham bir xil qabul qilinadi.
    Aniq mos kelmasa — xato/yozuv farqiga chidamli (taxminiy/fuzzy) eng
    yaqin natijalarni qaytaradi.
    """
    # 2026-07-14: kanal filtri OLIB TASHLANDI — ilgari asosiy kanaldan
    # qidirilganda ЦЕХ tovarlari (19 ta) umuman TOPILMASdi ("Natija
    # topilmadi" chiqardi). Endi butun inventar qidiriladi, ЦЕХ tovar
    # natijada 🏭 belgisi bilan ko'rinadi (qidiruv_text).
    df = inventar_olish(None)
    if df.empty or "Товар" not in df.columns:
        return pd.DataFrame()

    terms = [_qidiruv_normalize(t) for t in re.split(r"\s+", query.strip()) if t.strip()]
    terms = [t for t in terms if t]
    if not terms:
        return pd.DataFrame()

    names_norm = df["Товар"].astype(str).apply(_qidiruv_normalize)
    mask = pd.Series(True, index=df.index)
    for term in terms:
        mask &= names_norm.str.contains(_term_pattern(term), na=False, regex=True)

    out = df[mask].copy()

    # ── Aniq mos kelmasa — so'z darajasida taxminiy (fuzzy) qidiruv ─────────
    if out.empty:
        name_tokens = names_norm.str.split(" ")
        scores = name_tokens.apply(lambda toks: _fuzzy_score(terms, toks))
        best = scores[scores >= 0.6].sort_values(ascending=False)
        if not best.empty:
            out = df.loc[best.index[:limit]].copy()

    if out.empty:
        return out
    jami = len(out)
    status_rank = {"🔴 КРИТИК": 1, "🟡 ПАСТ": 2, "🟢 НОРМА": 3, "МЕЁР ЙЎҚ": 4}
    if "Холат" in out.columns:
        out["_rank"] = out["Холат"].map(status_rank).fillna(9)
    else:
        out["_rank"] = 9
    sort_cols = [c for c in ["_rank", "Кун_Етади", "Товар"] if c in out.columns]
    natija = (
        out.sort_values(sort_cols)
        .head(limit)
        .drop(columns=["_rank"], errors="ignore")
        .reset_index(drop=True)
    )
    natija.attrs["jami"] = jami   # limit'dan tashqarida qolganlar haqida xabar uchun
    return natija


def qidiruv_text(query: str, df: pd.DataFrame, lang: str) -> str:
    if df.empty:
        return f"Natija topilmadi: {query}" if lang == "lat" else f"Натижа топилмади: {query}"
    title = f"Qidiruv: {query}" if lang == "lat" else f"Қидирув: {query}"
    lines = [title]
    for i, row in df.iterrows():
        tovar = str(row.get("Товар", "?"))
        holat = str(row.get("Холат", ""))
        qoldiq = int(row.get("Қолдиқ", 0))
        yolda = int(row.get("Йўлда_Жами", 0))
        min_z = int(row.get("Мин_Захира", 0))
        kun = row.get("Кун_Етади", "")
        # 2026-07-14: ЦЕХ tovarlari endi qidiruvda chiqadi — belgi bilan
        cex = " 🏭ЦЕХ" if "ЦЕХ" in str(row.get("Тур", "")) else ""
        lines.append(
            f"{i + 1}. {tovar}{cex}\n"
            f"   {holat} | Qoldiq: {qoldiq} | Yo'lda: {yolda} | Min: {min_z} | Kun: {kun}"
        )
    jami = df.attrs.get("jami", len(df))
    if jami > len(df):
        lines.append(
            f"\n... yana {jami - len(df)} ta natija bor — aniqroq yozing"
            if lang == "lat" else
            f"\n... яна {jami - len(df)} та натижа бор — аниқроқ ёзинг"
        )
    return "\n".join(lines)


def grafik_qidirish(query: str, kat: str, kanal: str) -> pd.DataFrame:
    """
    "51 -> 0,9 -> 5,8 -> 201" yoki "51 0,9 5,8 201" formatini parse qilib tovar topadi.
    kat: truba | profil | list | bal
    '.' va ',' bir xil, lotin 'x'/'X' va kirill 'х' bir xil qabul qilinadi.
    """
    # Normalize: nuqtani vergulga, lotin x ni kirill х ga
    query = query.replace('.', ',').replace('x', 'х').replace('X', 'х')

    # -> bilan yoki bo'sh joy bilan ham ishlaydi
    if '->' in query:
        parts = [p.strip() for p in query.split('->')]
    else:
        parts = [p.strip() for p in query.split()]
    # 2026-07-18 (Huzayfa: "ko'plab tsex tovarlarini qidiruvdan topib
    # bo'lmayapti", masalan Ж-1 lar): kanal filtri OLIB TASHLANDI — ilgari
    # asosiy kanaldan qidirilganda ЦЕХ🏭 tovarlari (Ж-1 va h.k.) umuman
    # topilmasdi. Umumiy qidiruvda bu 2026-07-14 da tuzatilgandi — endi
    # kategoriyali (grafik) qidiruvda ham butun inventar qidiriladi.
    df = inventar_olish(None)
    if df.empty or "Товар" not in df.columns:
        return pd.DataFrame()

    names = df["Товар"].astype(str)
    mask  = pd.Series(True, index=df.index)

    def _q(s):
        return re.escape(s)  # allaqachon vergul formatida

    # 2026-07-11 (tuzatildi): raqamli qidiruvlar oldin ANCHOR'siz substring
    # edi -- "0,6" qidirilsa "0,65"/"0,61" ham, "Ф-19" qidirilsa "Ф-190" ham
    # (bo'lsa) aralashib chiqardi, chunki "0,6" harfma-harf "0,65" ichida
    # ham bor. Endi raqamdan keyin YANA raqam kelmasligi tekshiriladi
    # (?!\d) -- shunda "0,6" faqat "0,6" bilan tugagan joyda mos keladi,
    # "0,65" ga mos kelmaydi.
    def _qn(s):
        return re.escape(s.strip()) + r'(?!\d)'

    # 2026-07-14 (Huzayfa: "20 10 5,8 desam topilmadi deydi"): qism ROLI
    # endi o'rniga qarab emas, QIYMATIGA qarab aniqlanadi — foydalanuvchi
    # stenkani tashlab "51 5,8" (diametr+uzunlik) yozsa ham ishlaydi:
    #   201/304/430/316/321  -> marka
    #   4 dan katta son      -> uzunlik (м)   (stenka hech qachon 4+ emas)
    #   4 dan kichik son     -> stenka
    #   son bo'lmasa         -> matn (голд, бесшовный, ...)
    def _rol(p: str) -> str:
        if p in ('201', '304', '430', '316', '321'):
            return 'marka'
        try:
            v = float(p.replace(',', '.'))
        except ValueError:
            return 'matn'
        return 'uzunlik' if v >= 4 else 'stenka'

    def _rol_mask(p: str):
        r = _rol(p)
        if r == 'marka':
            return names.str.contains(f'{p} марка', na=False, case=False)
        if r == 'uzunlik':
            # "(5,8 м)" — chapda raqam/vergul bo'lmasin ("5,8" "5,6 м"dagi
            # "6 м"ga o'xshab adashmasin), o'ngda "м" kelsin
            return names.str.contains(
                r'(?<![\d,])' + re.escape(p) + r'\s*м', na=False, case=False, regex=True)
        if r == 'stenka':
            return names.str.contains(f'ст {_qn(p)}', na=False, case=False)
        return names.str.contains(re.escape(p), na=False, case=False)

    # 2026-07-18 (Huzayfa: "25 25 0,9 5,8 Ж 1 topilmayapti, ketma-ketlik
    # muhim emas"): rol-mask juda qat'iy edi — masalan "Ж 1" dagi "1"
    # tokeni STENKA deb qabul qilinib "ст 1" qidirilardi va hech narsa
    # topilmasdi. Endi har token YO o'z roli bo'yicha, YO umumiy
    # (raqam-chegarali substring) bo'yicha mos kelsa yetadi — "1" tokeni
    # "Ж-1" dagi 1 ga ham mos keladi, "201"/"1220" larga esa adashmaydi.
    def _token_mask(p: str):
        p = p.strip()
        pat = re.escape(p)
        if p and p[0].isdigit():
            pat = r'(?<!\d)' + pat
        if p and p[-1].isdigit():
            pat = pat + r'(?!\d)'
        generic = names.str.contains(pat, na=False, case=False, regex=True)
        return _rol_mask(p) | generic

    if kat == "truba":
        if len(parts) >= 1:
            mask &= names.str.contains(f'Ф-{_qn(parts[0])}', na=False, case=False)
        for p in parts[1:]:
            mask &= _token_mask(p)

    elif kat == "profil":
        # "20 20 0,7 6 201" yoki "20х20 0,7 6 201" — ikkalasi ishlaydi
        # Birinchi ikki qism raqam bo'lsa (masalan "20" "20") → "20х20" deb birlashtir
        idx = 0
        if len(parts) >= 2:
            try:
                float(parts[0].replace(',', '.')); float(parts[1].replace(',', '.'))
                # Ikkalasi raqam → o'lcham ikki alohida qism
                olcham = f"{parts[0]}х{parts[1]}"
                idx = 2
            except ValueError:
                olcham = parts[0].strip()
                idx = 1
        elif len(parts) >= 1:
            olcham = parts[0].strip()
            idx = 1
        else:
            olcham = ""
        # 2026-07-14: Профиль kategoriyasida faqat haqiqiy profillar chiqsin —
        # ilgari "20 20" qidirilsa "Пласт-20х20", "Трубогиб 20х20", "Кольцо"
        # kabi begona tovarlar ham aralashib chiqardi (Huzayfa 3-rasm).
        mask &= names.str.contains(r'Пр\.', na=False, regex=True)
        if olcham:
            mask &= names.str.contains(_qn(olcham), na=False, case=False)
        # 2026-07-18: qolgan qismlar — rol YOKI umumiy moslik (_token_mask)
        for p in parts[idx:]:
            mask &= _token_mask(p)

    elif kat == "list":
        if len(parts) >= 1:
            mask &= names.str.contains(f'Лист.*{_qn(parts[0])}', na=False, case=False)
        # 2026-07-18: qolgan qismlar tartib-erkin, raqam-chegarali moslik
        for p in parts[1:]:
            mask &= _token_mask(p)

    elif kat == "bal":
        q = query.strip()
        if re.match(r'^\d+$', q):
            num = q.zfill(2)
            mask &= names.str.contains(f'Баласина-{num}', na=False, case=False)
        else:
            mask &= names.str.lower().str.contains(re.escape(q.lower()), na=False)

    elif kat == "stoyka":
        mask &= names.str.contains('Стойка', na=False, case=False)
        q = query.strip()
        if q:
            # "01" → "№01", "34" → "№34"
            num = re.sub(r'[^0-9,]', '', q)
            if num:
                mask &= names.str.contains(re.escape(num), na=False, case=False)

    elif kat == "chas":
        mask &= names.str.contains('Чашка', na=False, case=False)
        if len(parts) == 1:
            p0 = parts[0].strip()
            # Boshidagi raqam/o'lcham bo'yicha filtr
            mask &= names.str.contains(f'^{re.escape(p0)}', na=False, regex=True, case=False)
        elif len(parts) >= 2:
            # "40 40" → profil chashkasi "40х40-Чашка"
            olcham = f"{parts[0]}х{parts[1]}"
            mask &= names.str.contains(re.escape(olcham), na=False, case=False)

    elif kat == "kuz":
        mask &= names.str.contains('Кузикорин', na=False, case=False)
        q = query.strip()
        if q:
            mask &= names.str.contains(f'^{re.escape(q)}', na=False, regex=True, case=False)

    elif kat == "shar":
        mask &= names.str.contains('Шар', na=False, case=False)
        mask &= ~names.str.contains('Шаркона|Шарнир|Шаршара', na=False, case=False)
        q = query.strip()
        if q:
            if len(parts) >= 2:
                # "4 4" yoki "4х4" → profil sharlari
                olcham = f"{parts[0]}х{parts[1]}"
                mask &= names.str.contains(f'^{re.escape(olcham)}', na=False, regex=True, case=False)
            else:
                # "51", "76" — diametr bo'yicha
                mask &= names.str.contains(f'^{re.escape(q)}', na=False, regex=True, case=False)

    elif kat == "sokka":
        mask &= names.str.contains('Сокка', na=False, case=False)
        mask &= ~names.str.contains('Баласина', na=False, case=False)
        if len(parts) == 1:
            p0 = parts[0].strip()
            mask &= names.str.contains(f'^{re.escape(p0)}', na=False, regex=True, case=False)
        elif len(parts) >= 2:
            # "90 25" → "Сокка 90х25" yoki "90х25"
            olcham = f"{parts[0]}х{parts[1]}"
            mask &= names.str.contains(re.escape(olcham), na=False, case=False)

    elif kat == "oyna":
        mask &= names.str.contains('Ойна держатель', na=False, case=False)
        q = query.strip()
        num = re.sub(r'[^0-9]', '', q)
        if num:
            mask &= names.str.contains(f'№-{num.zfill(2)}', na=False, case=False)

    out = df[mask].copy()
    if not out.empty:
        # 2026-07-14 (Huzayfa: "Голд/Гулли birinchi chiqmasin, xodovoylar
        # birinchi chiqsin"): natijalar saralanadi —
        #   1) nomi qavs bilan BOSHLANMAGANLAR avval ((Голд)/(Янги)/(Кора)
        #      kabi kam yuradigan variantlar pastga),
        #   2) keyin Мин_Захира KATTAROQLARI avval (min katta = xodovoy tovar).
        nm = out["Товар"].astype(str)
        out["_pfx"] = nm.str.match(r"^\(").astype(int)
        out["_min"] = (pd.to_numeric(out["Мин_Захира"], errors="coerce").fillna(0)
                       if "Мин_Захира" in out.columns else 0)
        out = (out.sort_values(["_pfx", "_min"], ascending=[True, False])
                  .drop(columns=["_pfx", "_min"]))
    return out.reset_index(drop=True)


def buyurtma_yuklash(kanal: str) -> dict | None:
    p = BOT_HOLAT_DIR / f"buyurtma_{kanal}.json"
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        logger.exception(f"buyurtma_yuklash: {p} o'qib bo'lmadi (buzilgan JSON?)")
        return None


def buyurtma_saqlash(kanal: str, items: list):
    """
    Yangi tasdiqlangan buyurtmani mavjud tasdiqlangan bilan BIRLASHTIRADI.
    Bir tovar uchun miqdorlar yig'iladi — ustiga yozilmaydi.
    Masalan: 1-kun 619 ta + 3-kun 50 ta = jami 669 ta saqlanadi.
    """
    p = BOT_HOLAT_DIR / f"buyurtma_{kanal}.json"

    # Mavjud tasdiqlangan buyurtmalarni yuklash
    mavjud = buyurtma_yuklash(kanal)
    eski_map: dict[str, float] = {}
    if mavjud and mavjud.get("buyurtmalar"):
        for eski in mavjud["buyurtmalar"]:
            tov = str(eski.get("tovar", "")).strip()
            miq = float(eski.get("miqdor", 0))
            if tov:
                eski_map[tov] = eski_map.get(tov, 0) + miq

    # Yangilarini qo'shish (yig'ish)
    for yangi in items:
        tov = str(yangi.get("tovar", "")).strip()
        miq = float(yangi.get("miqdor", 0))
        if tov and miq > 0:
            eski_map[tov] = eski_map.get(tov, 0) + miq

    # Birlashtirilgan ro'yxat
    jami_items = [
        {"tovar": tov, "miqdor": miq}
        for tov, miq in eski_map.items()
        if miq > 0
    ]

    atomic_json_write(p,
        {"sana":        datetime.now().strftime("%d.%m.%Y %H:%M"),
         "kanal":       kanal,
         "buyurtmalar": jami_items},
        indent=2)


def pending_saqlash(kanal: str, user_id: int, items: list):
    """Tasdiqlash kutilayotgan buyurtmani diskka saqlaydi (bot qayta ishga tushsa yo'qolmasin)."""
    p = BOT_HOLAT_DIR / f"pending_{kanal}_{user_id}.json"
    atomic_json_write(p,
        {"kanal": kanal, "user_id": user_id, "items": items,
         "sana": datetime.now().strftime("%d.%m.%Y %H:%M")},
        indent=2)


def pending_yuklash(kanal: str, user_id: int) -> list | None:
    """Diskdan pending zakaz yuklaydi. Topilmasa None."""
    p = BOT_HOLAT_DIR / f"pending_{kanal}_{user_id}.json"
    if not p.exists():
        return None
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
        return data.get("items")
    except (json.JSONDecodeError, OSError):
        logger.exception(f"pending_yuklash: {p} o'qib bo'lmadi (buzilgan JSON?)")
        return None


def pending_tozala(kanal: str, user_id: int):
    """Pending zakaz faylini o'chiradi."""
    p = BOT_HOLAT_DIR / f"pending_{kanal}_{user_id}.json"
    if p.exists():
        p.unlink()


def draft_saqlash(kanal: str, tovarlar: list):
    """Excel yaratilganda tovar → to'liq nom mappingini saqlaydi.
    buyurtma_tekshir shu fayldan lookup qiladi — inventar o'zgarsa ham ishlaydi."""
    p = BOT_HOLAT_DIR / f"draft_{kanal}.json"
    atomic_json_write(p,
        {"sana": datetime.now().strftime("%d.%m.%Y %H:%M"),
         "tovarlar": tovarlar},
        indent=2)


def draft_yuklash(kanal: str) -> list | None:
    p = BOT_HOLAT_DIR / f"draft_{kanal}.json"
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8")).get("tovarlar")
    except (json.JSONDecodeError, OSError):
        logger.exception(f"draft_yuklash: {p} o'qib bo'lmadi (buzilgan JSON?)")
        return None


def _kun_oldin(sana_str: str) -> str:
    """'10.06.2026 14:23' → 'Bugun', 'Kecha', '3 kun oldin', yoki '15 kun oldin ⚠️'"""
    try:
        d = datetime.strptime(sana_str, "%d.%m.%Y %H:%M")
        delta = (datetime.now() - d).days
        if delta == 0:
            return "Bugun"
        elif delta == 1:
            return "Kecha"
        elif delta <= 13:
            return f"{delta} kun oldin"
        else:
            return f"{delta} kun oldin ⚠️"
    except Exception:
        return sana_str


def xitoy_yuklash(kanal: str) -> dict | None:
    p = BOT_HOLAT_DIR / f"xitoy_{kanal}.json"
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        logger.exception(f"xitoy_yuklash: {p} o'qib bo'lmadi (buzilgan JSON?)")
        return None


def xitoy_saqlash(kanal: str, xitoy_map: dict, ombor_map: dict | None = None,
                   vazn_map: dict | None = None):
    """
    xitoy_map  — K ustun (jami zakaz, buyurtma hisoblash uchun)
    ombor_map  — L ustun (tayyor/ready, yuklatish rejasi uchun)
    vazn_map   — {tovar_nomi: 1_dona_kg} Xitoy faylidan olingan haqiqiy vazn
    """
    p = BOT_HOLAT_DIR / f"xitoy_{kanal}.json"
    data = {
        "sana":     datetime.now().strftime("%d.%m.%Y %H:%M"),
        "kanal":    kanal,
        "tovarlar": xitoy_map,
    }
    if ombor_map:
        data["ombor"] = ombor_map
    if vazn_map:
        data["vazn"] = vazn_map
    atomic_json_write(p, data, indent=2)


def buyurtma_tozala(kanal: str):
    """Xitoy ostatka yuklanganda chaqiriladi — tasdiqlangan buyurtma endi
    K ustunda aks etgan, shu sababli alohida hisobga olish kerak emas."""
    p = BOT_HOLAT_DIR / f"buyurtma_{kanal}.json"
    if p.exists():
        p.unlink()


def zakaz_preview_text(kanal: str, items: list, lang: str) -> str:
    ch = t(lang, CH_KEY[kanal])
    by_varaq: dict = {}
    for i in items:
        v = i.get("varaq", "?")
        by_varaq[v] = by_varaq.get(v, 0) + 1
    tafsilot = "".join(f"\n  • {v}: {c} ta" for v, c in by_varaq.items())

    # KRITIK tovarlar soni — inventardan cross-reference
    kritik_satri = ""
    try:
        inv_df = inventar_olish(kanal)
        if not inv_df.empty and "Товар" in inv_df.columns and "Холат" in inv_df.columns:
            kritik_set = set(inv_df[inv_df["Холат"] == "🔴 КРИТИК"]["Товар"].astype(str))
            n_kritik   = sum(1 for i in items if str(i.get("tovar","")).strip() in kritik_set)
            if n_kritik > 0:
                if lang == "cyr":
                    kritik_satri = f"\n🔴 Шундан *{n_kritik} та* КРИТИК ҳолатдаги товар."
                else:
                    kritik_satri = f"\n🔴 Shundan *{n_kritik} ta* KRITIK holatdagi tovar."
    except Exception:
        pass

    # Oldingi tasdiqlangan buyurtma borligini tekshirish — akkumulyatsiya ogohlantirish
    oldingi = buyurtma_yuklash(kanal)
    ogoh = ""
    if oldingi and oldingi.get("buyurtmalar"):
        n_oldin = len(oldingi["buyurtmalar"])
        sana    = oldingi.get("sana", "?")
        if lang == "cyr":
            ogoh = f"\n\n⚠️ *Диққат:* {kanal} каналида аввал *{n_oldin} та* товар тасдиқланган ({sana}).\nЯнгиси устига *қўшилади* — бекор қилиш учун Созламалар → Буюртмани тозалаш."
        else:
            ogoh = f"\n\n⚠️ *Diqqat:* {kanal} kanalida avval *{n_oldin} ta* tovar tasdiqlangan ({sana}).\nYangisi ustiga *qo'shiladi* — bekor qilish uchun Sozlamalar → Buyurtmani tozalash."

    return t(lang, "zakaz_preview").format(
        ch=ch, n=len(items), tafsilot=tafsilot
    ) + kritik_satri + ogoh


def kamomat_stats(kanal: str) -> dict:
    df  = kamomat_olish(kanal)
    if df.empty:
        return {"n": 0, "b": 0, "p": 0}
    buy     = buyurtma_yuklash(kanal)
    ordered = {i["tovar"] for i in buy.get("buyurtmalar",[])} if buy else set()
    tovs    = df["Товар"].tolist() if "Товар" in df.columns else []
    b       = sum(1 for x in tovs if x in ordered)
    return {"n": len(df), "b": b, "p": len(df) - b}



def asosiy_styled_excel_yarat(xitoy_ostatka: dict | None = None,
                               kanal: str = "asosiy") -> BytesIO:
    from Generate_Asosiy_order import load_data, calculate, build, strip_length, get_length

    # Agar xitoy_ostatka berilmagan bo'lsa — JSON fayldan avtomatik o'qiymiz
    mavjud = xitoy_yuklash(kanal)
    if xitoy_ostatka is None:
        if mavjud and mavjud.get("tovarlar"):
            xitoy_ostatka = mavjud["tovarlar"]
    # 2026-07-13 (Huzayfa: "K va L ustunlarini Excelda ham ko'rsatib
    # turish kerak"): Tayyor (L) ustuni ham xuddi shu xitoy_{kanal}.json
    # faylida ("ombor" kaliti) saqlanadi -- ILGARI bu yerda umuman
    # o'qilmasdi, faqat Zakaz (K, "tovarlar") buyurtmani kamaytirish
    # uchun ishlatilardi. Endi ikkalasi ham Excelga (E=Zakaz, F=Tayyor)
    # yoziladi -- admin har bir tovar uchun Xitoydagi ostatkani buyurtma
    # qatori yonida to'g'ridan-to'g'ri ko'radi, alohida fayl ochib
    # qidirishga hojat qolmaydi.
    ombor_map = (mavjud or {}).get("ombor") or {}

    _df, _kont_map = load_data(kanal=kanal)
    df_calc = calculate(_df, _kont_map, kanal=kanal)
    # 2026-07-22 (Huzayfa: "buyurtma berilmagan tovarlarni alohida
    # ko'rsatish kerak"): calculate() endi buyurtma==0 qatorlarni ham
    # qaytaradi -- to'liq (filtrlanmagan) nusxani shu yerda saqlab
    # qolamiz, pastda "nobuy" (buyurtma berilmagan) ro'yxatini shundan
    # hisoblaymiz. df_calc esa pastda eskisidek buyurtma>0 ga filtrlanadi.
    df_all_unfiltered = df_calc.copy()

    # ── Xitoy nomi to'qnashgan qatorlarni birlashtirish (2026-07-14) ────────
    # Inventarda IKKI XIL nom (masalan "ст 1,4" va "ст 1,35") Xitoy qoidasi
    # (stenka −0,05; 1,35/1,45 istisno) bilan Excel'da BIR XIL ko'rinishga
    # tushib, "bitta tovar ikki qator" bo'lib chiqardi (Huzayfa shikoyati).
    # Xitoy uchun bu bitta tovar — miqdorlar qo'shiladi, kanonik nom sifatida
    # min_zaxirasi kattasi olinadi. Ayirishlarda (K ustuni, tasdiqlangan)
    # IKKALA nom ham tekshiriladi — buning uchun merge_nomlar lug'ati.
    try:
        from vazn_hisobla import xitoy_nomi as _xn
    except ImportError:
        def _xn(n): return n

    merge_nomlar: dict = {}
    if not df_calc.empty:
        df_calc["_xkey"] = df_calc["tovar"].apply(
            lambda t: (strip_length(_xn(str(t))).strip(), get_length(str(t)).strip())
        )
        if df_calc["_xkey"].duplicated().any():
            yangi_rows = []
            for _, grp in df_calc.groupby("_xkey", sort=False):
                if len(grp) == 1:
                    yangi_rows.append(grp.iloc[0])
                    continue
                grp  = grp.sort_values("min_zaxira", ascending=False)
                bosh = grp.iloc[0].copy()
                bosh["buyurtma"] = int(grp["buyurtma"].sum())
                for col in ("qoldiq", "yoldagi"):
                    if col in grp.columns:
                        bosh[col] = grp[col].sum()
                # 2026-07-29: birlashtirilgan qatorda C jonli formulasi
                # ishlatilmaydi (interpolatsiya nuqtalarini turli min-panjarali
                # tovarlar bo'yicha qo'shish chalkash) — bp tozalanadi, C statik
                # (yig'ilgan buyurtma) bo'ladi.
                if "bp_x" in bosh.index:
                    bosh["bp_x"] = []
                    bosh["bp_y"] = []
                merge_nomlar[str(bosh["tovar"]).strip()] = [
                    str(t).strip() for t in grp["tovar"]
                ]
                logger.info(
                    f"[{kanal}] Xitoy-nom to'qnashuvi birlashtirildi: "
                    f"{list(grp['tovar'])} -> {bosh['tovar']}"
                )
                yangi_rows.append(bosh)
            df_calc = pd.DataFrame(yangi_rows).reset_index(drop=True)
        df_calc = df_calc.drop(columns=["_xkey"])

    # 2026-07-22: calculate() endi filtrlamagani uchun -- eski (har doim
    # amal qilgan) kafolatni shu yerda tiklaymiz: bundan keyingi hamma
    # mantiq (Xitoy ostatka ayirish, tasdiqlangan ayirish, mayda filtri)
    # ILGARIDEK faqat buyurtma>0 qatorlar bilan ishlaydi.
    df_calc = df_calc[df_calc["buyurtma"] > 0].copy()

    def _nomlar(tovar) -> list:
        """Kanonik nom -> ayirishda tekshiriladigan barcha nomlar."""
        tovar = str(tovar).strip()
        return merge_nomlar.get(tovar, [tovar])

    # 1. Xitoy ostatka K ustuni ayiriladi (Xitoyda buyurtma berilgan + tayyorlanayotgan)
    if xitoy_ostatka and not df_calc.empty:
        def _adjust_xitoy(row):
            ayir = sum(float(xitoy_ostatka.get(n, 0)) for n in _nomlar(row["tovar"]))
            return max(0, int(row["buyurtma"]) - int(ayir))
        df_calc["buyurtma"] = df_calc.apply(_adjust_xitoy, axis=1)
        df_calc = df_calc[df_calc["buyurtma"] > 0].copy()

    # 2. Tasdiqlangan buyurtma ayiriladi (allaqachon buyurilgan, lekin yo'lda emas)
    #    Agar tasdiqlangan hamma yetishmovchilikni qoplasa → buyurtma = 0 → Excel bo'sh
    tasdiq = buyurtma_yuklash(kanal)
    tasdiq_map = {}
    if tasdiq and tasdiq.get("buyurtmalar") and not df_calc.empty:
        for item in tasdiq["buyurtmalar"]:
            tov = str(item.get("tovar", "")).strip()
            miq = float(item.get("miqdor", 0))
            if tov:
                tasdiq_map[tov] = tasdiq_map.get(tov, 0) + miq
        if tasdiq_map:
            oldin = len(df_calc)
            def _adjust_tasdiq(row):
                ayir = sum(tasdiq_map.get(n, 0) for n in _nomlar(row["tovar"]))
                return max(0, int(row["buyurtma"]) - int(ayir))
            df_calc["buyurtma"] = df_calc.apply(_adjust_tasdiq, axis=1)
            df_calc = df_calc[df_calc["buyurtma"] > 0].copy()
            keyin = len(df_calc)
            logger.info(
                f"[{kanal}] tasdiqlangan ayirish: {oldin} → {keyin} ta "
                f"({oldin-keyin} ta o'chirildi, {len(tasdiq_map)} ta tasdiqlangan)"
            )

    # 2026-07-29: C ustuni INTERPOLATSIYA nuqtalari (bp_y) ham xuddi "buyurtma"
    # kabi Xitoy_K + tasdiqlangan miqdorga kamaytirilishi SHART — aks holda
    # jonli C formulasi bu ayirishlarni hisobga olmay, Min o'zgarganda haqiqiydan
    # KATTA son ko'rsatardi (buyurtma_logika.md 2.3). Ayirma har qator uchun
    # KONSTANTA (min'ga bog'liq emas), shuning uchun har bir nuqtadan ayiriladi
    # (0 dan past bo'lmasin). H0 (joriy min) nuqtasi shundan keyin AYNAN
    # yakuniy "buyurtma" ga teng bo'ladi — yashirin fallback bilan mos.
    if "bp_y" in df_calc.columns and not df_calc.empty:
        def _ayirma(tovar) -> int:
            s = 0.0
            if xitoy_ostatka:
                s += sum(float(xitoy_ostatka.get(n, 0)) for n in _nomlar(tovar))
            s += sum(tasdiq_map.get(n, 0) for n in _nomlar(tovar))
            return int(s)
        df_calc["bp_y"] = df_calc.apply(
            lambda r: [max(0, int(y) - _ayirma(r["tovar"])) for y in (r["bp_y"] or [])],
            axis=1,
        )

    # ── Mayda truba/profil filtri (2026-07-14, Huzayfa qoidasi) ─────────
    # Ф<51 truba va <50х50 profil buyurtmasi MAYDA_LIMIT(200) dan oshmasa
    # Excelga chiqmaydi — ehtiyoj yig'ilib limitdan oshganda o'zi chiqadi.
    # Безшовный (alohida kategoriya), Лист va boshqalarga tegilmaydi.
    if not df_calc.empty:
        from Generate_Asosiy_order import mayda_buyurtma_limiti
        _mask = df_calc.apply(
            lambda r: int(r["buyurtma"]) > mayda_buyurtma_limiti(r["tovar"], r["kategoriya"]),
            axis=1,
        )
        if (~_mask).any():
            logger.info(
                f"[{kanal}] mayda filtri: {(~_mask).sum()} ta tovar limitga "
                f"yetmagani uchun chiqarilmadi"
            )
        df_calc = df_calc[_mask].copy()

    if df_calc.empty:
        return None  # Buyurtma kerak emas — chaqiruvchi xabar beradi

    # E/F ustunlari uchun — xuddi shu tovar nomi bo'yicha Zakaz/Tayyor
    # qiymatlarini qo'shib qo'yamiz (FAQAT ko'rsatish uchun, hisoblashga
    # ta'sir qilmaydi — "buyurtma" ustuni yuqorida allaqachon tuzatilgan).
    df_calc["zakaz"] = df_calc["tovar"].apply(
        lambda t: sum(xitoy_ostatka.get(n, 0) for n in _nomlar(t)) if xitoy_ostatka else 0
    )
    df_calc["tayyor"] = df_calc["tovar"].apply(
        lambda t: sum(ombor_map.get(n, 0) for n in _nomlar(t)) if ombor_map else 0
    )

    # Draft tovarlarni saqlaymiz — buyurtma_tekshir shu ro'yxatdan lookup qiladi.
    # Inventar keyinchalik o'zgarsa ham xato bo'lmaydi.
    draft_saqlash(kanal, df_calc["tovar"].tolist())

    # "Меъёр йўқ" varaq (2026-07-14, Huzayfa so'rovi): min zaxirasi
    # belgilanmagan tovarlar hisobga KIRMAYDI va ilgari hech qayerda
    # ko'rinmasdi — "nega falon tovar ro'yxatda yo'q?" savoliga sabab
    # shu edi. Endi ular Excel oxirida alohida varaqda chiqadi.
    myoq_df = _df[_df["min_zaxira"] <= 0][["tovar", "qoldiq"]].copy()
    myoq_df = myoq_df.sort_values("qoldiq", ascending=False)

    # "Buyurtma berilmagan" ro'yxati (2026-07-22, Huzayfa: "yo'lda va
    # qoldiq norm bo'lsa berilmaydi va byurtma varag'ida bo'lmaydi,
    # shularni alohida ko'rsatish kerak"): min_zaxira>0 bo'lgan-u, biror
    # sababdan (zanjir_sim taklif=0 berdi, Xitoy ostatkasi qopladi,
    # tasdiqlangan buyurtma qopladi yoki mayda limitga yetmadi) YAKUNIY
    # df_calc'ga TUSHMAGAN tovarlar. "Меъёр йўқ" (min_zaxira<=0) bilan
    # ARALASHMASIN -- ular alohida, o'z varag'ida qoladi.
    ordered_nomlar = set(str(t).strip() for t in df_calc["tovar"])
    for nomlar in merge_nomlar.values():
        ordered_nomlar.update(str(t).strip() for t in nomlar)
    nobuy_df = df_all_unfiltered[
        (df_all_unfiltered["min_zaxira"] > 0)
        & (~df_all_unfiltered["tovar"].astype(str).str.strip().isin(ordered_nomlar))
    ].copy()
    if not nobuy_df.empty:
        nobuy_df["buyurtma"] = 0
        # 2026-07-29: "buyurtma berilmagan" bo'limida C statik 0 (jonli emas) —
        # bp tozalanadi (write_product h0>0 va bp bo'sh bo'lsa statik yozadi).
        if "bp_x" in nobuy_df.columns:
            nobuy_df["bp_x"] = [[] for _ in range(len(nobuy_df))]
            nobuy_df["bp_y"] = [[] for _ in range(len(nobuy_df))]
        nobuy_df["zakaz"] = nobuy_df["tovar"].apply(
            lambda t: sum(xitoy_ostatka.get(n, 0) for n in _nomlar(t)) if xitoy_ostatka else 0
        )
        nobuy_df["tayyor"] = nobuy_df["tovar"].apply(
            lambda t: sum(ombor_map.get(n, 0) for n in _nomlar(t)) if ombor_map else 0
        )

    wb  = build(df_calc, meyor_yoq=myoq_df, nobuy=nobuy_df, kanal=kanal)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def draft_excel_yarat(kanal: str, xitoy_ostatka: dict | None = None):
    """Barcha kanallar uchun bir xil styled Excel. None → buyurtma kerak emas."""
    return asosiy_styled_excel_yarat(xitoy_ostatka, kanal=kanal)


def buyurtma_tekshir(fayl_bytes: bytes, kanal: str = "asosiy"):
    """
    Tasdiqlangan buyurtma Excel'ini tekshiradi.
    Generate_Asosiy_order.build() formati: "Tovar nomi" | "Uzunlik" | "Buyurtma"

    Qaytaradi: (ok, xato, items, tanilmagan)
      - ok=False bo'lsa: items=None, tanilmagan=[] — FAYL DARAJASIDA jiddiy
        xato (ochilmadi, "Tovar" ustuni topilmadi, raqam bo'lmagan qiymat).
      - ok=True bo'lsa: items — tanilgan qatorlar, tanilmagan — nomi/soni bor
        bo'lsa-da HECH QANDAY manbada topilmagan qatorlar nomlari ro'yxati
        (chaqiruvchi buni ogohlantirish sifatida ko'rsatishi kerak).

    2026-07-29 (Huzayfa: "Tsex va Oshda tasdiqlashni qabul qilmayapti,
    Asosiy qabul qildi lekin tanimaganlari bor-yo'qligini aytmayapti"):
    ANIQLANGAN ikki mustaqil bug:
      1) Lookup manbasi (draft yoki inventar_olish) topilmasa/bo'sh bo'lsa,
         FUNKSIYA BUTUNLAY "Инвентар ma'lumoti topilmadi" bilan rad etardi —
         BUTUN faylni HATTO O'QIB KO'RMASDAN. Aslida har qatordagi haqiqiy
         tovar nomi deyarli DOIM cell comment'da bor (write_product() buni
         SHART qilib yozadi) — lookup faqat ESKI, commentsiz fayllar uchun
         zaxira. Tsex/Osh kanallari uchun `draft_{kanal}.json` hali umuman
         yaratilmagan (faqat Asosiy'niki mavjud edi) va shu ikki kanalning
         `inventar_olish()` natijasi bo'sh/aniq bo'lmasligi mumkin — shu
         sabab BUTUNLAY ISHLAMAY qolgan, garchi fayl o'zi to'g'ri bo'lsa ham.
         ENDI: lookup manbasi topilmasa ham funksiya DAVOM ETADI (bo'sh
         lookup bilan) — comment orqali topiladigan qatorlar baribir ishlaydi.
      2) Nomi/comment ORQALI ham topilmagan qatorlar JIMGINA o'tkazib
         yuborilardi (`continue`) — foydalanuvchiga hech qanday xabar
         berilmasdi, Xitoy ostatka oqimidagi "tanilgan/tanilmagan" hisoboti
         bu yerda YO'Q edi. ENDI: bunday qatorlar `tanilmagan` ro'yxatiga
         yoziladi va chaqiruvchi (handlers.py) buni ko'rsatadi.
    """
    import openpyxl
    from Generate_Asosiy_order import (
        strip_length, get_length, COL_B0BACK, get_category, get_marka,
        get_surface, get_size_dims, _dia_f, _d1_f, _d2_f, _stenka_f,
    )

    try:
        # C (Buyurtma) ustuni jonli formula — Min Zaxira o'zgarsa qayta
        # hisoblaydi. data_only=True: foydalanuvchi Excelda ochib/saqlaganda
        # formula KESHLANGAN qiymatini o'qiymiz. Fayl umuman ochilmagan
        # bo'lsa C=None keladi -> pastda yashirin statik B0 (COL_B0BACK)
        # ustunidan olinadi.
        wb = openpyxl.load_workbook(BytesIO(fayl_bytes), data_only=True)
    except Exception as e:
        return False, f"Faylni ochishda xato: {type(e).__name__}", None, []

    try:
        from vazn_hisobla import xitoy_nomi as _xitoy_nomi
    except ImportError:
        def _xitoy_nomi(n): return n   # fallback

    # 2026-07-29 (Huzayfa: "o'lchamlarning ketma-ketligi emas, o'lchamlarning
    # o'zini olish kerak — qidiruv tizimi BIR XIL bo'lsin"): botning o'zi
    # yaratgan faylda tovar nomi comment'da (har doim mos keladi), lekin
    # BOSHQA (bot yaratmagan) fayllarda -- (a) marka/uzunlik matn ichida
    # BOSHQA TARTIBDA kelishi mumkin ("(201 марка) (6 м)" o'rniga "(6 м)
    # (201 марка)"), (b) "Uzunlik" ustuni (B) botning o'zinikidan farqli
    # ishlatilishi mumkin. ESKI (nomi, uzun) satr-solishtirish TARTIBGA
    # SEZGIR edi. ENDI — kategoriya + har bir o'lcham (diametr/profil
    # o'lchami/qalinlik/uzunlik/marka/sirt) ALOHIDA-ALOHIDA regex bilan
    # olinadi (tartibdan mustaqil) va shulardan KALIT quriladi. Stenka
    # (qalinlik) ikki xil yozilishi mumkin (haqiqiy YOKI Xitoy -0,05
    # uslubida) — shu sabab HAR BIR katalog tovari uchun IKKALA variant
    # ham kalitlanadi.
    def _olcham_kaliti(name: str, stenka_override: float | None = None):
        cat = get_category(name)
        if cat not in ("Труба", "Профиль", "Лист"):
            return None
        marka = get_marka(name)
        surf  = get_surface(name)
        stenka = round(stenka_override, 3) if stenka_override is not None \
                 else round(_stenka_f(name), 3)
        if stenka >= 9999:
            return None
        if cat == "Труба":
            dia = round(_dia_f(name), 2)
            if dia >= 9999:
                return None
            return (cat, dia, stenka, get_length(name), marka, surf)
        if cat == "Профиль":
            d1, d2 = round(_d1_f(name), 1), round(_d2_f(name), 1)
            if d1 >= 9999:
                return None
            return (cat, min(d1, d2), max(d1, d2), stenka, get_length(name), marka, surf)
        # Лист — uzunlik tushunchasi yo'q, o'lcham (dims) bor
        dims = get_size_dims(name)
        if not dims:
            return None
        return (cat, dims, stenka, marka, surf)

    # Lookup manbalari: AVVALO draft faylidan (Excel yaratilganda saqlangan),
    # yo'q/bo'sh bo'lsa inventardan. Ikkalasi ham bo'lmasa — bo'sh lookup,
    # LEKIN funksiya endi davom etadi (comment yoki o'lcham-kaliti orqali
    # topish YETARLI bo'lishi kerak — pastga qarang).
    lookup      = {}   # eski (nomi, uzun) — moslik uchun saqlanadi
    dim_lookup  = {}   # yangi, tartibdan mustaqil o'lcham-kaliti
    norm_lookup = {}   # umumiy normalize_product_name fallback (aksessuar va h.k.)

    # 2026-07-30 (Huzayfa savoli: "kunlik qoldiqda ham yo'qmikan?"): draft_{kanal}.json
    # faqat SHU KUNGI hisoblangan buyurtmaga (buyurtma>0, mayda-filtrdan o'tgan)
    # kirgan tovarlarni saqlaydi — qoldig'i yetarli (buyurtma=0) yoki mayda-limitdan
    # past bo'lgan tovarlar Excelga chiqmagani uchun draft'da HAM yo'q bo'lib chiqadi,
    # garchi Min_Zaxira'da haqiqatda mavjud bo'lsa ham. Tasdiqlangan buyurtma esa
    # ADMIN tomonidan o'zgartirilgan/qo'shilgan bo'lishi mumkin — shuning uchun
    # lookup manbai FAQAT draft emas, balki draft + TO'LIQ inventar (Min_Zaxira)
    # birlashmasi bo'lishi SHART. Tasdiqlandi: (Кора) Пр. 25х25/40х40/50х50 ст 0,9
    # va Пр. 20х20/40х20 ст 0,7 (5,6/5,8 м) kabi item'lar aynan shu sabab
    # ("tanilmagan" chiqishi) — Min_Zaxira'da bor, draft'da yo'q edi.
    draft = draft_yuklash(kanal) or []
    inv = inventar_olish(kanal)
    inv_nomlari = list(inv["Товар"].dropna()) if not inv.empty and "Товар" in inv.columns else []
    manba_nomlari = []
    _seen = set()
    for x in list(draft) + inv_nomlari:
        s = str(x).strip()
        if s and s not in _seen:
            _seen.add(s)
            manba_nomlari.append(s)

    for full_name in manba_nomlari:
        excel_nomi = strip_length(_xitoy_nomi(full_name)).strip()
        uzun       = get_length(full_name).strip()
        lookup[(excel_nomi, uzun)] = full_name
        if (excel_nomi, "") not in lookup:
            lookup[(excel_nomi, "")] = full_name

        real_stenka = _stenka_f(full_name)
        if real_stenka < 9999:
            k1 = _olcham_kaliti(full_name, stenka_override=real_stenka)
            if k1:
                dim_lookup[k1] = full_name
            xitoy_stenka = _stenka_f(_xitoy_nomi(full_name))
            if xitoy_stenka < 9999:
                k2 = _olcham_kaliti(full_name, stenka_override=xitoy_stenka)
                if k2:
                    dim_lookup[k2] = full_name

        norm_lookup[normalize_product_name(full_name)] = full_name

    items = []
    tanilmagan = []
    found_any = False

    for ws in wb.worksheets:
        header_row = None
        for r in range(1, min(ws.max_row, 5) + 1):
            v = ws.cell(row=r, column=1).value
            if v is not None and str(v).strip().lower() == "tovar nomi":
                header_row = r
                break
        if header_row is None:
            continue
        found_any = True

        for r in range(header_row + 1, ws.max_row + 1):
            cell1 = ws.cell(row=r, column=1)
            nomi = cell1.value
            if nomi is None or str(nomi).strip() == "":
                continue
            nomi = str(nomi).strip()
            uzun_val = ws.cell(row=r, column=2).value
            uzun = "" if uzun_val is None else str(uzun_val).strip()
            buy_val = ws.cell(row=r, column=3).value
            if buy_val in (None, ""):
                # C formulasining keshi yo'q (fayl Excelda ochilmagan)
                # -> yashirin statik B0 fallback ustunidan o'qiymiz.
                buy_val = ws.cell(row=r, column=COL_B0BACK).value
            if buy_val in (None, ""):
                continue
            # 1) Avvalo cell comment dan asl inventar nomini o'qiymiz —
            # bot o'zi yaratgan faylda BU DOIM ishlaydi (write_product()
            # har doim yozadi), qolgan usullar faqat comment YO'Q bo'lsa
            # (bot yaratmagan/eski fayl) kerak bo'ladi.
            tovar = None
            if cell1.comment and cell1.comment.text and cell1.comment.text.strip():
                tovar = cell1.comment.text.strip()
            if tovar is None:
                # 2) Eski (nomi, uzun) aniq mos kelish (B ustunidagi qiymat bilan)
                tovar = lookup.get((nomi, uzun)) or lookup.get((nomi, ""))
            if tovar is None:
                # 3) YANGI: o'lcham-asosli, TARTIBDAN MUSTAQIL kalit —
                # faqat nomi matnining o'zidan (B ustuniga qaramasdan)
                # kategoriya/diametr/qalinlik/uzunlik/marka/sirt ajratiladi.
                dk = _olcham_kaliti(nomi)
                if dk:
                    tovar = dim_lookup.get(dk)
            if tovar is None:
                # 4) Umumiy normalize (aksessuar/Баласина va h.k. — stenka
                # tushunchasi bo'lmagan tovarlar uchun oxirgi zaxira).
                tovar = norm_lookup.get(normalize_product_name(nomi))
            try:
                m = float(buy_val)
            except (ValueError, TypeError):
                return False, f"'{ws.title}' varaqida raqam bo'lmagan buyurtma qiymati", None, []
            if tovar is None:
                if m > 0:
                    # Miqdori bor, lekin tovar HECH QAYERDA topilmadi —
                    # jim o'tkazib yuborish o'rniga ogohlantirishga yoziladi.
                    tanilmagan.append(f"{nomi} ({uzun})" if uzun else nomi)
                continue
            if m > 0:
                items.append({"tovar": tovar, "miqdor": m, "varaq": ws.title})

    if not found_any:
        return False, (
            "Hech qaysi varaqda 'Tovar' ustuni topilmadi.\n"
            "Faylda quyidagi varaqlar bor: "
            + ", ".join(str(s) for s in wb.sheetnames)
        ), None, []
    return True, None, items, tanilmagan
