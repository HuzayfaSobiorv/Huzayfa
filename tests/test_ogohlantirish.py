# -*- coding: utf-8 -*-
"""
Excel "Ogohlantirish" varag'i uchun testlar.

2026-08-08 (Huzayfa: "tanimagan yoki noma'lum bo'lgan ... shunday
mahsulotlarni shundayligicha skip qivormayaptimi yo'qmi bilishim kerak").
Yuklatish hisobida tovar bir necha xil yo'l bilan rejadan tushib qolardi va
bir qismi HECH QAYERDA ko'rinmasdi (ayniqsa kategoriyasi аксессуар/бошқа
bo'lganlar — ular butunlay jim `continue` bilan tashlanardi).

Ishga tushirish (loyiha papkasida):  python -m pytest tests/ -v
"""
import os
import sys

import pandas as pd
import pytest
from openpyxl import Workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from Yuklama_optimal import optimallashtir
from yuklatish_rejasi import _ogohlantirish_varaq


def _varaq(**kw):
    """Ogohlantirish varag'ini yasab, undagi matnlar ro'yxatini qaytaradi."""
    wb = Workbook()
    _ogohlantirish_varaq(
        wb,
        kw.get("qolgan", []), kw.get("vazn_yoq", []),
        kw.get("boshqa_kat", []), kw.get("yangi", []), kw.get("notanish", []),
    )
    ws = wb["Ogohlantirish"]
    matnlar = [
        str(c) for r in ws.iter_rows(values_only=True) for c in r
        if c not in (None, "")
    ]
    return ws, matnlar


# ── Varaqning o'zi ────────────────────────────────────────────────────
class TestVaraq:

    def test_har_doim_yaratiladi(self):
        """Muammo bo'lmasa ham varaq MAVJUD bo'lishi kerak — admin
        "ogohlantirish bormi?" degan savolga doim bir joydan javob olsin."""
        wb = Workbook()
        _ogohlantirish_varaq(wb, [], [], [], [], [])
        assert "Ogohlantirish" in wb.sheetnames

    def test_muammo_yoq_bolsa_bosh(self):
        """Huzayfa talabi: "hech qanday muammo bo'lmasa ogohlantirish
        stranitsasi bo'm-bo'sh bo'lsin"."""
        _, matnlar = _varaq()
        assert matnlar == [], f"bo'sh bo'lishi kerak edi: {matnlar}"

    def test_bosh_bloklar_chizilmaydi(self):
        """Faqat TO'LDIRILGAN toifa ko'rinadi — bo'sh toifa sarlavhasi ham
        chiqmasligi kerak (aks holda varaq "muammo bordek" ko'rinadi)."""
        _, matnlar = _varaq(vazn_yoq=[{"tovar": "Баласина 32х32", "dona": 40}])
        birlashgan = " ".join(matnlar)
        assert "VAZNI ANIQLANMADI" in birlashgan
        assert "ИНВЕНТАРДА ТОПИЛМАДИ" not in birlashgan
        assert "АКСЕССУАР" not in birlashgan


# ── Toifalar ──────────────────────────────────────────────────────────
class TestToifalar:

    @pytest.mark.parametrize("kalit,tovar,sarlavha_bolagi", [
        ("vazn_yoq",   {"tovar": "Баласина 32х32", "dona": 40},   "VAZNI ANIQLANMADI"),
        ("boshqa_kat", {"tovar": "Отвод Ф-51", "dona": 60},       "АКСЕССУАР"),
        ("qolgan",     {"tovar": "Ф-16 ст 0,7", "dona": 8, "vazn_kg": 11.2}, "BU SAFAR YUKLANMADI"),
        ("yangi",      {"tovar": "Ф-25 ст 1,2", "dona": 799},     "ИНВЕНТАРДА ТОПИЛМАДИ"),
    ])
    def test_toifa_korinadi(self, kalit, tovar, sarlavha_bolagi):
        _, matnlar = _varaq(**{kalit: [tovar]})
        birlashgan = " ".join(matnlar)
        assert sarlavha_bolagi in birlashgan
        assert tovar["tovar"] in birlashgan
        assert str(tovar["dona"]) in birlashgan

    def test_notanish_spec_korinadi(self):
        _, matnlar = _varaq(notanish=["Φ 999 XX 1.0 [5.8M]"])
        birlashgan = " ".join(matnlar)
        assert "O'QIB BO'LMADI" in birlashgan
        assert "Φ 999 XX 1.0 [5.8M]" in birlashgan

    def test_sarlavhada_soni_korsatiladi(self):
        _, matnlar = _varaq(boshqa_kat=[
            {"tovar": "Отвод Ф-51", "dona": 60},
            {"tovar": "Шар d-50", "dona": 120},
        ])
        assert any("(2 ta)" in m for m in matnlar), matnlar


# ── Manba: optimallashtir "Бошқа" toifani qayd qiladimi ───────────────
class TestBoshqaKategoriya:

    def test_aksessuar_jim_tashlanmaydi(self):
        """ILGARIGI BUG: kategoriyasi Труба/Профиль/Лист bo'lmagan tovar
        `continue` bilan JIMGINA tashlab ketilardi — hech qanday ro'yxatga
        tushmasdi, foydalanuvchi tovarning yo'qolganini bilolmasdi."""
        tovar = "Отвод Ф-51 (304 марка)"
        kerak = pd.DataFrame([{"Товар": tovar, "Холат": "🔴 КРИТИК",
                               "Кам": 60, "urg_kun": 0}])
        mavjud = pd.DataFrame([{"Товар": tovar, "Миқдор": 60}])
        yuklar, _, _, _, boshqa_kat = optimallashtir(
            kerak, mavjud, abc_map={}, max_yuklar=20)
        assert not yuklar, "аксессуар yuklanmasligi kerak (mantiq o'zgarmadi)"
        assert [b["tovar"] for b in boshqa_kat] == [tovar]
        assert boshqa_kat[0]["dona"] == 60

    def test_oddiy_tovar_boshqaga_tushmaydi(self):
        tovar = "Ф-16 ст 0,7 (5,8 м) (201 марка)"
        kerak = pd.DataFrame([{"Товар": tovar, "Холат": "🔴 КРИТИК",
                               "Кам": 500, "urg_kun": 0}])
        mavjud = pd.DataFrame([{"Товар": tovar, "Миқдор": 500}])
        yuklar, _, _, _, boshqa_kat = optimallashtir(
            kerak, mavjud, abc_map={}, max_yuklar=20)
        assert yuklar
        assert boshqa_kat == []
