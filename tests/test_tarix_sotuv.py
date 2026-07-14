# -*- coding: utf-8 -*-
"""
Tarix fayllardan kunlik sotuvni hisoblash — TEST.

Mantiq:
  1. Tarix/DD.MM.YYYY.xlsx fayllaridan har bir tovar qoldig'i olinadi (I ustun, index 8).
  2. Ketma-ket ikki sana orasidagi farq:
       - kamaygan  -> sotuv (farq / kunlar soni = kunlik sotuv)
       - oshgan    -> YUK KELGAN (bu interval sotuv o'rtachasiga kirmaydi,
                      lekin kelgan miqdor qayd etiladi)
  3. Qoldiq 0 bo'lgan kunlar (tugagan davr) o'rtacha hisobiga KIRMAYDI —
     "sotuv yo'q" emas, "tovar yo'q" deb qaraladi. O'rtacha faqat tovar
     mavjud bo'lgan kunlar bo'yicha chiqadi, keyingi oylar uchun shu
     tezlik asos qilib olinadi.
"""
import os
import re
import sys
from datetime import datetime, date

import openpyxl

BAZA = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TARIX = os.path.join(BAZA, "Tarix")

DATE_RE = re.compile(r"^(\d{2})\.(\d{2})\.(\d{4})\.xlsx$")


def tarix_fayllar():
    """[(sana, yo'l), ...] o'sish tartibida."""
    out = []
    for f in os.listdir(TARIX):
        m = DATE_RE.match(f)
        if m:
            d, mo, y = map(int, m.groups())
            out.append((date(y, mo, d), os.path.join(TARIX, f)))
    return sorted(out)


def qoldiq_seriya(nomlar):
    """
    nomlar: {kalit: nom_matcher(str)->bool}
    return: {kalit: [(sana, qoldiq), ...]}
    """
    seriya = {k: [] for k in nomlar}
    for sana, yol in tarix_fayllar():
        wb = openpyxl.load_workbook(yol, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        topildi = {k: None for k in nomlar}
        for row in ws.iter_rows(values_only=True):
            nom = row[0]
            if not isinstance(nom, str):
                continue
            for k, match in nomlar.items():
                if topildi[k] is None and match(nom):
                    q = row[8] if len(row) > 8 and row[8] is not None else 0
                    try:
                        topildi[k] = float(q)
                    except (TypeError, ValueError):
                        topildi[k] = 0.0
        wb.close()
        for k, q in topildi.items():
            seriya[k].append((sana, 0.0 if q is None else q))
    return seriya


def tahlil(seriya):
    """
    Bitta tovar seriyasi bo'yicha:
      - jami sotuv, sotuv kunlari, kunlik o'rtacha
      - yuk kelgan sanalar (+miqdor, taxminiy)
      - tugagan (qoldiq=0) davrlar
    """
    sotuv_jami = 0.0
    sotuv_kunlar = 0
    kelishlar = []      # (sana, taxminiy miqdor)
    nol_kunlar = 0
    intervallar = []    # debug uchun

    for (s1, q1), (s2, q2) in zip(seriya, seriya[1:]):
        kunlar = (s2 - s1).days
        if kunlar <= 0:
            continue
        farq = q2 - q1
        if q1 <= 0 and q2 <= 0:
            # butun interval tovar yo'q
            nol_kunlar += kunlar
            intervallar.append((s2, kunlar, 0.0, "TUGAGAN"))
        elif farq < 0:
            sotuv_jami += -farq
            sotuv_kunlar += kunlar
            intervallar.append((s2, kunlar, -farq, "sotuv"))
        elif farq > 0:
            # yuk kelgan: kelgan >= farq (o'sha kunlardagi sotuv noma'lum)
            kelishlar.append((s2, farq))
            intervallar.append((s2, kunlar, farq, "YUK KELDI"))
        else:
            sotuv_kunlar += kunlar
            intervallar.append((s2, kunlar, 0.0, "o'zgarishsiz"))

    kunlik = sotuv_jami / sotuv_kunlar if sotuv_kunlar else 0.0
    return {
        "sotuv_jami": sotuv_jami,
        "sotuv_kunlar": sotuv_kunlar,
        "kunlik": kunlik,
        "kelishlar": kelishlar,
        "nol_kunlar": nol_kunlar,
        "intervallar": intervallar,
    }


MISOLLAR = {
    "Ф-51 ст 0,9 (5,8 м) (201 марка)":
        lambda n: n.strip() == "Ф-51 ст 0,9 (5,8 м) (201 марка)",
    "Баласина-02":
        lambda n: n.strip() == "Баласина-02",
    "Лист-0,8 (1220х2440) (Глянцевый) (201 марка)":
        lambda n: n.strip() == "Лист-0,8 (1220х2440) (Глянцевый) (201 марка)",
}


def main():
    seriya = qoldiq_seriya(MISOLLAR)
    for nom, s in seriya.items():
        r = tahlil(s)
        print("=" * 70)
        print(nom)
        print(f"  Davr: {s[0][0]} .. {s[-1][0]}  ({len(s)} ta fayl)")
        print(f"  Boshl. qoldiq: {s[0][1]:,.0f}   Oxirgi qoldiq: {s[-1][1]:,.0f}")
        print(f"  Jami sotuv: {r['sotuv_jami']:,.0f}  ({r['sotuv_kunlar']} kun)")
        print(f"  KUNLIK SOTUV: {r['kunlik']:,.1f}")
        if r["nol_kunlar"]:
            print(f"  Tugagan (qoldiq=0) kunlar: {r['nol_kunlar']} — o'rtachaga kirmadi")
        for sana, miq in r["kelishlar"]:
            print(f"  YUK KELDI: {sana} atrofida +{miq:,.0f}")


if __name__ == "__main__":
    main()
