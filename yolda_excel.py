"""
yolda_excel.py
==============
Power BI "Контейнерлар" varaqidan yo'ldagi konteynerlarni
chiroyli blokli Excel formatda tayyorlaydi.

Har bir konteyner = bitta blok:
  - Sarlavha (1 qator): konteyner raqami | Yuklangan sana | Kelish sanasi | Kun qoldi / Kechikdi
  - Ustun sarlavhalari (1 qator)
  - Tovar qatorlari
  - 2 bo'sh qator ajratuvchi sifatida

Tartib: Сана_Тартиб bo'yicha (kelish sanasiga qarab)
Rang: КЕЧИКДИ -> to'q sariq/to'q qizil, ЙЎЛДА -> ko'k
"""

import io
import re
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from vazn_hisobla import tovar_vazni

# ── Ranglar ──────────────────────────────────────────────────────────────────
CLR_KECHIKDI_BG   = "C0392B"   # qizil (КЕЧИКДИ sarlavha fon)
CLR_KECHIKDI_TEXT = "FFFFFF"

CLR_YOLDA_BG      = "1A5276"   # to'q ko'k (ЙЎЛДА sarlavha fon)
CLR_YOLDA_TEXT    = "FFFFFF"
CLR_YOLDA_SUB     = "2980B9"   # ustun sarlavhasi (har doim ko'k)

CLR_COL_HDR_TEXT  = "FFFFFF"
CLR_TOTAL_BG      = "D5DBDB"   # jami qator

# Hamma kategoriya — bir xil och ko'k, oddiy alternating
ROW_CLR_DARK  = "D6EAF8"
ROW_CLR_LIGHT = "EBF5FB"
CAT_COLORS = {k: (ROW_CLR_DARK, ROW_CLR_LIGHT, "2E86C1") for k in
              ["Лист", "Труба", "Профиль", "Баласина", "Стойка", "_other"]}

# Tovarlar tartibi (kategoriya bo'yicha): Лист → Труба → Профиль → boshqalar
CAT_ORDER = {"Лист": 0, "Труба": 1, "Профиль": 2, "Баласина": 3, "Стойка": 4}

THIN = Side(style="thin", color="BDBDBD")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=11) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size)


def _align(h="left", v="center", wrap=False, indent=0) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)


def _get_marka(name: str) -> str:
    m = re.search(r'\((\d{3})\s*марка\)', str(name))
    if m:
        return m.group(1)
    m2 = re.search(r'\b(201|304|316|321|430)\b', str(name))
    return m2.group(1) if m2 else ""


def _set_row(ws, row_i: int, values: list, fills: list = None,
             fonts: list = None, aligns: list = None,
             height: float = None, border: bool = True):
    """Qatorga qiymatlar yozadi, uslublar qo'shadi."""
    for col_i, val in enumerate(values, start=1):
        cell = ws.cell(row=row_i, column=col_i, value=val)
        if fills and fills[col_i - 1]:
            cell.fill = fills[col_i - 1]
        if fonts and fonts[col_i - 1]:
            cell.font = fonts[col_i - 1]
        if aligns and aligns[col_i - 1]:
            cell.alignment = aligns[col_i - 1]
        if border:
            cell.border = BORDER_THIN
    if height:
        ws.row_dimensions[row_i].height = height


def yolda_excel(data_file: str | Path) -> io.BytesIO | None:
    """
    data_file — NEJAVIYKA_POWER_BI.xlsx yo'li.
    Qaytaradi BytesIO (Excel fayli) yoki None (ma'lumot yo'q).
    """
    try:
        df = pd.read_excel(data_file, sheet_name="Контейнерлар")
    except Exception:
        return None

    # КЕЛДИ ✅ ni olib tashlash
    df = df[df["Холат"] != "КЕЛДИ ✅"].copy()
    if df.empty:
        return None

    # Tartib
    if "Сана_Тартиб" in df.columns:
        df["_tartib"] = pd.to_numeric(df["Сана_Тартиб"], errors="coerce").fillna(999)
    elif "Кун_Қолди" in df.columns:
        df["_tartib"] = pd.to_numeric(df["Кун_Қолди"], errors="coerce").fillna(999)
        df["_tartib"] = -df["_tartib"]   # kam kun qoldi -> oldin
    else:
        df["_tartib"] = 0

    # КЕЧИКДИ ni oldin ko'rsatamiz (Сана_Тартиб allaqachon bunga mos)
    df = df.sort_values(["_tartib", "Контейнер"])

    # Konteynerlar tartibini aniqlash
    container_order = list(dict.fromkeys(df["Контейнер"].tolist()))

    wb = Workbook()
    ws = wb.active
    ws.title = "Йўлдаги контейнерлар"

    # Ustun kengliklari: A=30(tovar), B=12(miqdor), C=14(turkum)
    ws.column_dimensions["A"].width = 52   # Tovar nomi
    ws.column_dimensions["B"].width = 12   # Miqdor
    ws.column_dimensions["C"].width = 16   # Kategoriya
    ws.column_dimensions["D"].width = 14   # Yuklangan sana
    ws.column_dimensions["E"].width = 14   # Kelish sanasi
    ws.column_dimensions["F"].width = 16   # Holat
    # Sarlavha uchun A-F merge ishlatamiz

    cur_row = 1
    umumiy_tonna_kg = 0.0

    for cont_no, cont_id in enumerate(container_order):
        grp = df[df["Контейнер"] == cont_id]
        first = grp.iloc[0]

        holat      = str(first.get("Холат", ""))
        yukl_sana  = str(first.get("Юкланган_Сана", ""))[:10]
        kelish     = str(first.get("Келиш_Санаси", ""))[:10]
        kun_qoldi  = first.get("Кун_Қолди", None)
        kechik_kun = first.get("Кечикиш_Кун", 0)
        turi       = str(first.get("Тури", ""))

        kechikdi = "КЕЧИКДИ" in holat

        # ── Konteynerning umumiy tonnasi ─────────────────────────────────────
        # Birinchi navbatda "Вазн_кг" ustunidagi OLDINDAN HISOBLANGAN
        # (konteyner qo'shilgan paytda, bir marta) qiymatlar yig'iladi —
        # bu yerda tovar nomidan qayta hisoblash SHART EMAS. Faqat bu
        # ustun umuman yo'q yoki bo'sh bo'lgan (masalan eski, bu funksiya
        # qo'shilishidan OLDINGI konteynerlar) qatorlar uchun — orqaga
        # moslashuvchan fallback sifatida — tovar nomidan taxminiy
        # hisoblanadi.
        has_vazn_col = "Вазн_кг" in grp.columns
        jami_tonna_kg = 0.0
        for _, trow in grp.iterrows():
            vazn_kg = trow.get("Вазн_кг") if has_vazn_col else None
            if pd.notna(vazn_kg) and vazn_kg not in (None, ""):
                jami_tonna_kg += float(vazn_kg)
            else:
                vazn = tovar_vazni(str(trow.get("Товар", "")))
                miq_t = trow.get("Миқдор", 0)
                if vazn and pd.notna(miq_t):
                    jami_tonna_kg += vazn * miq_t
        jami_tonna = round(jami_tonna_kg / 1000, 2)
        umumiy_tonna_kg += jami_tonna_kg

        # ── Rang tanlash ──────────────────────────────────────────────────
        # Faqat sarlavha qatori farqlanadi; ustun sarlavha va tovarlar — har doim ko'k
        if kechikdi:
            hdr_bg  = CLR_KECHIKDI_BG
            hdr_txt = CLR_KECHIKDI_TEXT
        else:
            hdr_bg  = CLR_YOLDA_BG
            hdr_txt = CLR_YOLDA_TEXT
        sub_bg = CLR_YOLDA_SUB   # ustun sarlavhasi har doim ko'k

        # ── 1-qator: Konteyner sarlavhasi (A:F merge) ────────────────────
        if kechikdi:
            kk = int(kechik_kun) if pd.notna(kechik_kun) else 0
            vaqt_info = f"⚠️  {kk} кун кечикди"
        else:
            kq = int(float(kun_qoldi)) if pd.notna(kun_qoldi) else "?"
            vaqt_info = f"🕐 {kq} кун қолди"

        hdr_text = (
            f"🚢  {cont_id}    │    "
            f"Юкланган: {yukl_sana}    │    "
            f"Келиш: {kelish}    │    "
            f"{vaqt_info}    │    "
            f"⚖️ {jami_tonna} т"
            + (f"    │    {turi}" if turi and turi != "STANDART" else "")
        )

        hdr_cell = ws.cell(row=cur_row, column=1, value=hdr_text)
        hdr_cell.fill   = _fill(hdr_bg)
        hdr_cell.font   = Font(name="Calibri", bold=True, color=hdr_txt, size=12)
        hdr_cell.alignment = _align(h="left", v="center")
        hdr_cell.border = BORDER_THIN
        ws.merge_cells(
            start_row=cur_row, start_column=1,
            end_row=cur_row, end_column=6
        )
        ws.row_dimensions[cur_row].height = 24
        cur_row += 1

        # ── 2-qator: Ustun sarlavhalari ───────────────────────────────────
        col_headers = ["Товар", "Миқдор", "Категория", "", "", ""]
        col_fills   = [_fill(sub_bg)] * 3 + [_fill(sub_bg)] * 3
        col_fonts   = [_font(bold=True, color=CLR_COL_HDR_TEXT)] * 6
        col_aligns  = [
            _align(h="left",   v="center", indent=1),
            _align(h="center", v="center"),
            _align(h="left",   v="center", indent=1),
        ] + [_align()] * 3
        _set_row(ws, cur_row, col_headers, fills=col_fills,
                 fonts=col_fonts, aligns=col_aligns, height=18)
        ws.row_dimensions[cur_row].height = 18
        cur_row += 1

        # ── Tovar qatorlari (Лист → Труба → Профиль tartibida) ───────────
        grp_sorted = grp.copy()
        if "Категория" in grp_sorted.columns:
            grp_sorted["_ord"] = grp_sorted["Категория"].apply(
                lambda c: CAT_ORDER.get(str(c).strip(), 99)
            )
            grp_sorted = grp_sorted.sort_values(["_ord", "Товар"]).drop(columns=["_ord"])

        jami_miq    = 0
        row_counter = 0   # umumiy alternating uchun
        blok_start  = cur_row - 2  # sarlavha + ustun header (oldingi 2 qator)

        for _, trow in grp_sorted.iterrows():
            # DIQQAT: bu yer — "Yo'ldagi konteynerlarni ko'rish" (filyal
            # boshqaruvchilari ko'radigan) varaq. Bu yerda tovar nomi HAR
            # DOIM inventar formatida (masalan "ст 0,9", "ст 1,0", "ст
            # 2,0", "ст 3,0" — yaxlitlangan) ko'rsatilishi SHART, chunki
            # inventar tizimidagi haqiqiy nom bilan bir xil bo'lishi kerak.
            # Avval shu yerda xitoy_nomi() chaqirilib, nomni Xitoyning O'Z
            # (kamaytirilgan, masalan 0,85/0,95/1,95/2,95) formatiga
            # aylantirib yuborardi — bu XATO edi: xitoy_nomi() faqat
            # Xitoyga YUBORILADIGAN "Yuklangan ro'yxat" (yuklatish_rejasi.py)
            # uchun mo'ljallangan, bu yerga aloqasi yo'q edi.
            tovar = str(trow.get("Товар", ""))
            miq   = trow.get("Миқдор", 0)
            kat   = str(trow.get("Категория", "")).strip()

            jami_miq += miq if pd.notna(miq) else 0

            # Alternating: hamma kategoriya bir xil och ko'k
            row_clr = ROW_CLR_DARK if row_counter % 2 == 0 else ROW_CLR_LIGHT
            row_counter += 1

            row_bg = _fill(row_clr)
            values = [tovar, miq, kat, "", "", ""]
            fills  = [row_bg] * 6
            fonts  = [
                _font(size=10),
                _font(size=10, bold=True),
                _font(size=10, color="444444"),
            ] + [_font(size=10)] * 3
            aligns = [
                _align(h="left",   v="center", wrap=True, indent=1),
                _align(h="center", v="center"),
                _align(h="left",   v="center", indent=1),
            ] + [_align()] * 3
            _set_row(ws, cur_row, values, fills=fills,
                     fonts=fonts, aligns=aligns, height=17)
            cur_row += 1

        # ── Jami qator (Kategoriya ustunida — chekkasida — tonna ham) ─────
        jami_vals  = ["ЖАМИ:", jami_miq, f"⚖️ {jami_tonna} т", "", "", ""]
        jami_fills = [_fill(CLR_TOTAL_BG)] * 6
        jami_fonts = [
            _font(bold=True, size=10),
            _font(bold=True, size=10),
            _font(bold=True, size=10, color="444444"),
        ] + [_font(size=10)] * 3
        jami_aligns = [
            _align(h="right",  v="center"),
            _align(h="center", v="center"),
            _align(h="left",   v="center", indent=1),
        ] + [_align()] * 3
        _set_row(ws, cur_row, jami_vals, fills=jami_fills,
                 fonts=jami_fonts, aligns=jami_aligns, height=16)
        blok_end = cur_row
        cur_row += 1

        # ── Blok atrofini medium border bilan o'rash ──────────────────────
        brd_s  = Side(style="medium", color="1A3A5C")
        brd_i  = Side(style="thin",   color="CCCCCC")
        n_cols = 6
        for r in range(blok_start, blok_end + 1):
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=r, column=c)
                top    = brd_s if r == blok_start else brd_i
                bottom = brd_s if r == blok_end   else brd_i
                left   = brd_s if c == 1          else brd_i
                right  = brd_s if c == n_cols      else brd_i
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

        # ── Ajratuvchi chiziq + bo'sh qator ──────────────────────────────
        sep_fill = _fill("BDC3C7")
        for col_i in range(1, 7):
            cell = ws.cell(row=cur_row, column=col_i, value="")
            cell.fill = sep_fill
        ws.row_dimensions[cur_row].height = 3
        cur_row += 1
        ws.row_dimensions[cur_row].height = 8
        cur_row += 1

    # ── 2-varaq: Xulosa ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Хулоса")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 22
    ws2.column_dimensions["D"].width = 14

    total_yolda   = df[df["Холат"] == "ЙЎЛДА 🚢"]["Контейнер"].nunique()
    total_kechikdi = df[df["Холат"] == "КЕЧИКДИ ⚠️"]["Контейнер"].nunique()
    total_cont    = df["Контейнер"].nunique()

    xulosa_rows = [
        ("Жами йўлда",  total_cont,    "ЙЎЛДА 🚢",    total_yolda),
        ("КЕЧИКДИ",     total_kechikdi, "",            ""),
        ("Жами тонна",  f"{round(umumiy_tonna_kg / 1000, 2)} т", "", ""),
    ]
    ws2.cell(row=1, column=1, value="📊  Хулоса").font = _font(bold=True, size=13)
    ws2.merge_cells("A1:D1")
    ws2.row_dimensions[1].height = 22

    ws2.cell(row=2, column=1, value="Ҳолат").font  = _font(bold=True)
    ws2.cell(row=2, column=2, value="Сони").font   = _font(bold=True)
    ws2.cell(row=2, column=3, value="Ҳолат").font  = _font(bold=True)
    ws2.cell(row=2, column=4, value="Сони").font   = _font(bold=True)
    for c in range(1, 5):
        ws2.cell(row=2, column=c).fill   = _fill("1A5276")
        ws2.cell(row=2, column=c).font   = _font(bold=True, color="FFFFFF")
        ws2.cell(row=2, column=c).border = BORDER_THIN

    for ri, (a, b, c, d) in enumerate(xulosa_rows, start=3):
        for ci, val in enumerate([a, b, c, d], start=1):
            cell = ws2.cell(row=ri, column=ci, value=