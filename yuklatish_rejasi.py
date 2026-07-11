"""
yuklatish_rejasi.py
====================
Xitoyda tayyor tovarlardan (L ustun = Ombor) optimal yuklatish rejasini tuzadi.
Chiqish: chiqish/Yuklatish_Rejasi_YYYY-MM-DD.xlsx

Ma'lumot manbalari:
  1. Power BI Excel (chiqish/NEJAVIYKA_POWER_BI.xlsx) — Инвентар varaqi
     Ustunlar: Товар, Қолдиқ, Йўлда_Жами, Мин_Захира, Холат
  2. Xitoy Ombor xaritasi {tovar: tayyor_miqdor}
     — Bot orqali: bot_holat/xitoy_{kanal}.json -> "ombor" kaliti
     — Yoki to'g'ridan-to'g'ri dict sifatida uzatiladi (main_with_data)

Qoidalar:
  - Jami yuk  <= 28 000 kg
  - Труба+Профиль <= 11 000 kg
  - Лист <= 18 000 kg
  - Max 4 konteyner/kun, 20 konteyner/oy
  - Ustunlik: Холат (🔴КРИТИК > 🟡ПАСТ > 🟢НОРМА) → urgentlik (kun)

Ishlatish (standalone):
  python yuklatish_rejasi.py [kanal]

Bot orqali (ombor_map mavjud bo'lsa):
  from yuklatish_rejasi import main_with_data
  xlsx_path = main_with_data("asosiy", ombor_map)
"""

import json
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

# Bot boshqa papkadan ishga tushganda ham modul topilsin
_THIS_DIR = Path(__file__).resolve().parent
if str(_THIS_DIR) not in sys.path:
    sys.path.insert(0, str(_THIS_DIR))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from vazn_hisobla import xitoy_nomi, XITOY_DELTA
from Yuklama_optimal import optimallashtir, LIMIT_TOTAL, LIMIT_TRUBA_PROFIL, LIMIT_LIST, LIMITS_BY_TYPE

# ── Sozlamalar ───────────────────────────────────────────────────────────────
BASE_DIR      = Path(__file__).resolve().parent
POWER_BI_FILE = BASE_DIR / "chiqish" / "NEJAVIYKA_POWER_BI.xlsx"
BOT_HOLAT_DIR = BASE_DIR / "bot_holat"
CHIQISH_DIR   = BASE_DIR / "chiqish"
CHIQISH_DIR.mkdir(exist_ok=True)

MAX_PER_DAY   = 4       # bir kunda max konteyner
MAX_PER_MONTH = 20      # bir oyda max konteyner
DELIVERY_DAYS = 55      # kunalik sotuv hisobi uchun

PB_SHEET = "Инвентар"
PB_COLS  = {
    "Товар":       "tovar",
    "Қолдиқ":      "qoldiq",
    "Йўлда_Жами":  "yolda",
    "Мин_Захира":  "min_zaxira",
    "Холат":       "holat",
    "Етишмайди":   "etishmaydi",   # 2026-07-11: zanjir-hisobdan haqiqiy kerak miqdor
}

# ── Ranglar ──────────────────────────────────────────────────────────────────
C_HDR_MAIN  = "1A3A5C"  # asosiy sarlavha
C_HDR_SUB   = "1F618D"  # vazn satri
C_COL_HDR   = "2C4A6E"  # ustun sarlavhalari
C_GRAY      = "EBEBEB"  # jami satri
C_WHITE     = "FFFFFF"
C_RED       = "C0392B"
C_BLUE      = "1F618D"
C_GREEN     = "1E8449"

# Hamma kategoriya — bir xil och ko'k (Лист rangi), alternating toq/yengil
ROW_CLR_DARK  = "D6EAF8"   # toq qator
ROW_CLR_LIGHT = "EBF5FB"   # yengil qator
CAT_COLORS = {k: (ROW_CLR_DARK, ROW_CLR_LIGHT, "2E86C1") for k in
              ["Лист", "Труба", "Профиль", "Баласина", "Стойка", "_other"]}
# Orqaga moslik uchun (qolgan blokda ishlatiladi)
MARKA_BG = {"201":"E8F4FD","304":"E8F8E8","430":"FFFBEA","316":"F3E8FF","321":"FFE8E8","":"F8F8F8"}
YUK_TURI_COLOR = {
    "6m":  C_BLUE,    # ko'k — 6m konteyner
    "12m": C_GREEN,   # yashil — 12m mashina
}
HOLAT_COLOR = {
    "🔴 КРИТИК": C_RED,
    "🟡 ПАСТ":   "E67E22",
    "🟢 НОРМА":  "27AE60",
}


# ── Openpyxl yordamchilari ───────────────────────────────────────────────────
def _fill(hex_: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_)


def _border(style: str = "thin", color: str = "CCCCCC") -> Border:
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _outer_border(color: str = "444444") -> Border:
    s = Side(style="medium", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _font(bold=False, size=11, color=C_WHITE, name="Calibri", italic=False) -> Font:
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)


def _align(h="left", v="center", wrap=False, indent=0) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)


# ── Parsing yordamchilari ────────────────────────────────────────────────────
def get_marka(name: str) -> str:
    m = re.search(r'\((\d{3})\s*марка\)', name)
    if m:
        return m.group(1)
    m2 = re.search(r'\b(201|304|430|316|321)\b', name)
    if m2:
        return m2.group(1)
    return ""


def get_uzunlik(name: str) -> str:
    m = re.search(r'\(([\d,\.]+)\s*м\)', name)
    return m.group(0) if m else ""


def strip_uzunlik(name: str) -> str:
    return re.sub(r'\s*\([\d,\.]+\s*м\)', '', name).strip()


def holat_color(holat: str) -> str:
    for k, v in HOLAT_COLOR.items():
        if k in str(holat):
            return v
    return "555555"


# ── Ma'lumot yuklash ──────────────────────────────────────────────────────────
def power_bi_yuklash() -> pd.DataFrame:
    if not POWER_BI_FILE.exists():
        raise FileNotFoundError(f"Power BI fayli topilmadi:\n{POWER_BI_FILE}")

    df = pd.read_excel(POWER_BI_FILE, sheet_name=PB_SHEET)
    rename = {k: v for k, v in PB_COLS.items() if k in df.columns}
    df = df.rename(columns=rename)

    needed = ["tovar", "qoldiq", "min_zaxira"]
    missing = [c for c in needed if c not in df.columns]
    if missing:
        raise ValueError(f"Inventar varaqida ustunlar yo'q: {missing}")

    if "yolda" not in df.columns:
        df["yolda"] = 0
    if "holat" not in df.columns:
        df["holat"] = "🟢 НОРМА"
    if "etishmaydi" not in df.columns:
        df["etishmaydi"] = 0

    keep = [v for v in PB_COLS.values() if v in df.columns]
    df = df[keep].dropna(subset=["tovar"]).copy()
    # (ЯНГИ)/(Янги) prefiksli tovarlarni chiqarib tashlaymiz — eskirgan nom
    _yangi_mask = df["tovar"].str.contains(r'^\s*\(ЯНГИ\)|\(Янги\)', regex=True, na=False)
    if _yangi_mask.any():
        print(f"⚠️  (ЯНГИ) prefiksli {_yangi_mask.sum()} ta tovar o'chirildi")
    df = df[~_yangi_mask].copy()

    for col in ["qoldiq", "yolda", "min_zaxira", "etishmaydi"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    return df


def xitoy_ombor_yuklash(kanal: str = "asosiy") -> tuple[dict, dict]:
    """
    bot_holat/xitoy_{kanal}.json dan ombor va vazn xaritasini yuklaydi.
    Qaytaradi: (ombor_map, vazn_map)
      ombor_map — L ustun (tayyor/ready, yuklatish uchun)
      vazn_map  — {tovar_nomi: 1_dona_kg} Xitoy faylidan
    """
    p = BOT_HOLAT_DIR / f"xitoy_{kanal}.json"
    if not p.exists():
        raise FileNotFoundError(
            f"Xitoy ombor ma'lumoti topilmadi: {p}\n"
            f"Avval botdan xitoy ostatka Excelini yuklang."
        )
    data = json.loads(p.read_text(encoding="utf-8"))
    # Avval L ustun (ombor = tayyor/ready), aks holda K ustun (tovarlar = jami zakaz)
    ombor = data.get("ombor") or data.get("tovarlar", {})
    if not ombor:
        raise ValueError("Xitoy ombor bo'sh.")
    vazn = data.get("vazn", {})
    return ombor, vazn


# ── Kerak miqdori hisoblash ───────────────────────────────────────────────────
def kerak_hisob(pb_df: pd.DataFrame) -> pd.DataFrame:
    """
    2026-07-11 (tuzatildi): Кам endi ESKI statik "sotuv55" formuladan
    HISOBLANMAYDI -- to'g'ridan-to'g'ri Power BI Инвентар'ning
    "Етишмайди" ustunidan olinadi. Bu ustun main.py'ning kun-ma-kun
    zanjir-simulyatsiyasi natijasi (Мин_Захира - Якуний_Қолдиқ, agar
    manfiy bo'lmasa 0) -- ya'ni real ravishda "hozirgi qoldiq + yo'ldagi
    konteynerlar kelguncha kunlik sotuv bilan qancha kamayishi"ni
    hisobga oladi. Eski statik formula buni yakuniy holatga qarab taxmin
    qilardi (kamroq aniq, konteynerlar orasidagi bo'shliqni ko'rmasdi).
    """
    df = pb_df.copy()
    df["Кам"] = pd.to_numeric(df["etishmaydi"], errors="coerce").fillna(0).round(0).astype(int)

    # Urgentlik (faqat axborot/tartib uchun -- hisoblashga ta'sir qilmaydi)
    sotuv_k  = (df["min_zaxira"] / 30).clip(lower=0.1)
    df["urg_kun"] = ((df["qoldiq"] + df["yolda"] - df["min_zaxira"]) / sotuv_k).round(0).astype(int)

    return df.rename(columns={"tovar": "Товар", "holat": "Холат"})[
        ["Товар", "Холат", "Кам", "urg_kun"]
    ]


# ── Sana belgilash ────────────────────────────────────────────────────────────
def sana_belgi(yuklar: list[dict], start_date: datetime) -> list[dict]:
    """Har bir yukga yuklatish kunini belgilaydi (max 4/kun)."""
    d     = start_date
    count = 0
    for yuk in yuklar:
        if count >= MAX_PER_DAY:
            d     += timedelta(days=1)
            count  = 0
        yuk["yuklatish_sana"] = d.strftime("%d.%m.%Y")
        count += 1
    return yuklar


# ── Excel yozish ──────────────────────────────────────────────────────────────
# 2026-07-11: F -- bo'sh ajratuvchi ustun, G/H -- Қолдиқ/Йўлда (Power BI dan)
COLS    = ["Товар номи", "Узунлик", "Миқдор", "Вазн (кг)", "Холат", "", "Қолдиқ", "Йўлда"]
WIDTHS  = [50, 10, 10, 12, 14, 3, 11, 11]
NC      = len(COLS)
SEP_COL = 6   # F ustun -- faqat vizual ajratuvchi
# 2026-07-11 (tuzatildi): sarlavha/жами satrlari va blok bordери ESKI
# holatdagidek FAQAT A:E (5 ustun) bilan tugashi kerak -- F butunlay
# bo'm-bo'sh qoladi, G/H alohida (mishka bilan A:E tortib ko'chirish
# ishi buzilmasin -- Huzayfa talabi).
CORE_NC = 5   # A..E -- eski "asosiy blok" kengligi


def _set_widths(ws):
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _yuk_header(ws, row: int, num: int, yuk: dict) -> int:
    """Konteyner asosiy sarlavhasi — sticker ko'rinishi."""
    color  = YUK_TURI_COLOR.get(yuk["turi"], C_HDR_MAIN)
    sana   = yuk.get("yuklatish_sana", "—")
    jami   = yuk["jami_kg"]
    tp     = yuk["truba_profil_kg"]
    lst    = yuk["list_kg"]
    turi_l = "📦  6m Контейнер" if yuk["turi"] == "6m" else "🚛  12m Машина"

    # ── 1-qator: asosiy sarlavha (sticker) ──────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=CORE_NC)
    c = ws.cell(row=row, column=1)
    c.value = (
        f"  ■  ЮК №{num}  —  {turi_l}"
        f"          📅  Юкланиш: {sana}"
        f"          ⚖️  {jami:,.0f} кг / 28 000 кг"
    )
    c.font      = _font(bold=True, size=13, color=C_WHITE)
    c.fill      = _fill(color)
    c.alignment = _align(h="left", indent=1)
    c.border    = _border("medium", "FFFFFF")
    ws.row_dimensions[row].height = 34
    row += 1

    # ── 2-qator: vazn tafsiloti ──────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=CORE_NC)
    c2 = ws.cell(row=row, column=1)
    _lims = LIMITS_BY_TYPE.get(yuk["turi"], LIMITS_BY_TYPE["6m"])
    _tp_lim  = _lims["truba_profil"]
    _lst_lim = _lims["list"]
    c2.value = (
        f"       Труба+Профиль: {tp:,.0f} / {_tp_lim:,} кг"
        + (f"     │     Лист: {lst:,.0f} / {_lst_lim:,} кг" if _lst_lim > 0 else "")
        + f"     │     Бўш жой: {LIMIT_TOTAL - jami:,.0f} кг"
    )
    c2.font      = _font(bold=False, size=10, color="BBCCDD")
    c2.fill      = _fill(C_HDR_SUB)
    c2.alignment = _align(h="left", indent=1)
    c2.border    = _border("thin", "334466")
    ws.row_dimensions[row].height = 18
    row += 1

    # ── 3-qator: ustun sarlavhalari ─────────────────────────────────────────
    for i, name in enumerate(COLS, 1):
        c3 = ws.cell(row=row, column=i, value=name)
        if i == SEP_COL:
            c3.fill = _fill("BDC3C7")   # ajratuvchi ustun -- boshqacha rang
        else:
            c3.font      = _font(bold=True, size=11, color=C_WHITE)
            c3.fill      = _fill(C_COL_HDR)
            c3.alignment = _align(h="center" if i > 1 else "left", indent=1 if i == 1 else 0)
        c3.border    = _border()
        ws.row_dimensions[row].height = 24
    row += 1

    return row


def _kat(name: str) -> str:
    """Tovar nomidan kategoriya (katta/kichik harf ahamiyatsiz)."""
    import re as _re
    n = str(name)
    nl = n.lower()
    if _re.match(r'^(\([^)]*\)\s*)?ф-\d+', nl):
        return "Труба"
    if _re.search(r'пр\.\s*\d+', nl):
        return "Профиль"
    if nl.startswith("лист"):
        return "Лист"
    if nl.startswith("балас"):
        return "Баласина"
    if nl.startswith("стойк"):
        return "Стойка"
    return "_other"


def _product_row(ws, row: int, item: dict, holat: str, bg: str = "",
                 qoldiq_yolda: tuple | None = None) -> int:
    """Bitta tovar satri. bg berilmasa kategoriya rangidan olinadi.
    qoldiq_yolda: (qoldiq, yolda) -- Power BI Инвентар'dan (G/H ustunlari uchun).
    """
    name  = xitoy_nomi(item["tovar"])
    uzunl = get_uzunlik(name)
    base  = strip_uzunlik(name)
    if not bg:
        bg = CAT_COLORS.get(_kat(name), CAT_COLORS["_other"])[0]
    brd   = _border()

    # A — Tovar nomi
    ca = ws.cell(row=row, column=1, value=base)
    ca.font      = _font(bold=True, size=11, color="111111")
    ca.fill      = _fill(bg)
    ca.border    = brd
    ca.alignment = _align(h="left", indent=1)

    # B — Uzunlik
    cb = ws.cell(row=row, column=2, value=uzunl)
    cb.font      = _font(bold=True, size=11, color=C_RED)
    cb.fill      = _fill(bg)
    cb.border    = brd
    cb.alignment = _align(h="center")

    # C — Miqdor
    cc = ws.cell(row=row, column=3, value=int(item["dona"]))
    cc.font      = _font(bold=True, size=11, color=C_RED)
    cc.fill      = _fill(bg)
    cc.border    = brd
    cc.alignment = _align(h="center")

    # D — Vazn
    cd = ws.cell(row=row, column=4, value=round(item["vazn_kg"], 1))
    cd.font      = _font(bold=False, size=11, color=C_BLUE)
    cd.fill      = _fill(bg)
    cd.border    = brd
    cd.alignment = _align(h="center")

    # E — Holat
    hc = holat_color(holat)
    ce = ws.cell(row=row, column=5, value=holat)
    ce.font      = _font(bold=True, size=10, color=hc)
    ce.fill      = _fill(bg)
    ce.border    = brd
    ce.alignment = _align(h="center")

    # F — bo'sh ajratuvchi (kulrang ingichka chiziq ko'rinishi)
    cf = ws.cell(row=row, column=SEP_COL, value="")
    cf.fill   = _fill("BDC3C7")
    cf.border = Border()

    qoldiq, yolda = qoldiq_yolda if qoldiq_yolda else (None, None)

    # G — Қолдиқ
    cg = ws.cell(row=row, column=7, value=int(qoldiq) if qoldiq is not None else "")
    cg.font      = _font(bold=False, size=10, color="444444")
    cg.fill      = _fill(bg)
    cg.border    = brd
    cg.alignment = _align(h="center")

    # H — Йўлда
    ch = ws.cell(row=row, column=8, value=int(yolda) if yolda is not None else "")
    ch.font      = _font(bold=False, size=10, color="444444")
    ch.fill      = _fill(bg)
    ch.border    = brd
    ch.alignment = _align(h="center")

    ws.row_dimensions[row].height = 24
    return row + 1


def _jami_row(ws, row: int, yuk: dict, n_items: int) -> int:
    """Jami qatori."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=CORE_NC)
    c = ws.cell(row=row, column=1)
    c.value = (
        f"  ✅  Жами: {n_items} хил товар"
        f"  │  {yuk['jami_kg']:,.0f} кг юкланди"
        f"  │  Бўш жой: {LIMIT_TOTAL - yuk['jami_kg']:,.0f} кг"
    )
    c.font      = _font(bold=True, size=11, color="222222")
    c.fill      = _fill(C_GRAY)
    c.alignment = _align(h="left", indent=1)
    c.border    = _border("medium", "AAAAAA")
    ws.row_dimensions[row].height = 22
    return row + 1


def _empty_rows(ws, row: int, count: int = 2) -> int:
    for _ in range(count):
        ws.row_dimensions[row].height = 10
        row += 1
    return row


def _separator(ws, row: int) -> int:
    """Konteyner bloklari orasiga ingichka chiziq + bo'sh qator."""
    sep_fill = PatternFill("solid", fgColor="BDC3C7")
    for col_i in range(1, CORE_NC + 1):
        cell = ws.cell(row=row, column=col_i, value="")
        cell.fill = sep_fill
    ws.row_dimensions[row].height = 3
    row += 1
    ws.row_dimensions[row].height = 8
    row += 1
    return row


def _blok_border(ws, row_start: int, row_end: int, n_cols: int):
    """Yuk bloki (sarlavhadan jamigacha) atrofini medium border bilan o'raydi."""
    brd_top    = Side(style="medium", color="1A3A5C")
    brd_bottom = Side(style="medium", color="1A3A5C")
    brd_side   = Side(style="medium", color="1A3A5C")
    brd_inner  = Side(style="thin",   color="CCCCCC")

    for r in range(row_start, row_end + 1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            top    = brd_top    if r == row_start else brd_inner
            bottom = brd_bottom if r == row_end   else brd_inner
            left   = brd_side   if c == 1         else brd_inner
            right  = brd_side   if c == n_cols     else brd_inner
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def _qolgan_blok(ws, row: int, qolgan: list[dict]) -> int:
    """Oylik limit yetmagan tovarlar bloki."""
    if not qolgan:
        return row
    row = _empty_rows(ws, row, 3)

    # Sarlavha
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC)
    c = ws.cell(row=row, column=1,
                value="  ⏭️   KEYINGI OYGA QOLGAN — oylik limit (20 ta) yetmadi")
    c.font      = _font(bold=True, size=12, color=C_WHITE)
    c.fill      = _fill("7F8C8D")
    c.alignment = _align(h="left", indent=1)
    c.border    = _border()
    ws.row_dimensions[row].height = 28
    row += 1

    for it in qolgan:
        marka = get_marka(it["tovar"])
        bg    = MARKA_BG.get(marka, "F0F0F0")

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c1 = ws.cell(row=row, column=1, value=f"  {it['tovar']}")
        c1.font  = _font(bold=False, size=10, color="555555")
        c1.fill  = _fill(bg)
        c1.border = _border()
        c1.alignment = _align(h="left", indent=1)

        cd = ws.cell(row=row, column=3, value=int(it["dona"]))
        cd.font  = _font(bold=False, size=10, color="555555")
        cd.fill  = _fill(bg); cd.border = _border(); cd.alignment = _align(h="center")

        cv = ws.cell(row=row, column=4, value=round(it["vazn_kg"], 1))
        cv.font  = _font(bold=False, size=10, color="555555")
        cv.fill  = _fill(bg); cv.border = _border(); cv.alignment = _align(h="center")

        ws.row_dimensions[row].height = 20
        row += 1

    return row


def _xulosa_varaq(wb: Workbook, yuklar: list[dict], qolgan: list[dict],
                  kanal: str, start_date: str):
    """Xulosa varaqi."""
    ws = wb.create_sheet("Xulosa")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18

    rows = [
        ("📅 Yuklatish boshlanishi",      start_date),
        ("📦 Jami konteynerlar",           len(yuklar)),
        ("🔢 Jami tovar xillari",          sum(len(y["items"]) for y in yuklar)),
        ("⚖️  Jami og'irlik (kg)",          f"{sum(y['jami_kg'] for y in yuklar):,.0f}"),
        ("🔩 Труба+Профиль jami (kg)",      f"{sum(y['truba_profil_kg'] for y in yuklar):,.0f}"),
        ("🗂️  Лист jami (kg)",              f"{sum(y['list_kg'] for y in yuklar):,.0f}"),
        ("⏭️  Keyingi oyga qolgan (xil)",  len(qolgan)),
        ("🏭 Kanal",                        kanal.upper()),
    ]

    for r, (label, val) in enumerate(rows, 1):
        ca = ws.cell(row=r, column=1, value=label)
        ca.font = Font(name="Arial", size=11, bold=True)
        ca.alignment = Alignment(vertical="center")
        cb = ws.cell(row=r, column=2, value=val)
        cb.font = Font(name="Arial", size=11)
        cb.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 22


def excel_yaz(yuklar: list[dict], qolgan: list[dict],
              holat_map: dict, kanal: str = "asosiy",
              qoldiq_yolda_map: dict | None = None) -> Path:
    """
    Yuklatish rejasi Excel faylini yozadi.
    holat_map: {tovar_nomi: holat_str}  (Power BI dan)
    qoldiq_yolda_map: {tovar_nomi: (qoldiq, yolda)}  (Power BI dan, G/H ustunlari)
    """
    qoldiq_yolda_map = qoldiq_yolda_map or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Yuklatish Rejasi"
    ws.sheet_view.showGridLines = False
    _set_widths(ws)

    # ── Asosiy sarlavha ──────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NC)
    c_main = ws.cell(row=1, column=1)
    c_main.value = (
        f"  🏭  NEJAVIYKA — Yuklatish Rejasi"
        f"    │    Kanal: {kanal.upper()}"
        f"    │    {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        f"    │    Jami: {len(yuklar)} konteyner"
    )
    c_main.font      = _font(bold=True, size=14, color=C_WHITE)
    c_main.fill      = _fill(C_HDR_MAIN)
    c_main.alignment = _align(h="left", indent=1)
    c_main.border    = _border("medium", C_WHITE)
    ws.row_dimensions[1].height = 40

    row = 3  # 2-qator bo'sh

    start_sana_str = yuklar[0].get("yuklatish_sana", datetime.now().strftime("%d.%m.%Y")) if yuklar else "—"

    for i, yuk in enumerate(yuklar, 1):
        row = _yuk_header(ws, row, i, yuk)

        # Kategoriya bo'yicha tartiblash va alternating rang
        cat_order = {"Лист": 0, "Труба": 1, "Профиль": 2, "Баласина": 3, "Стойка": 4}
        items_sorted = sorted(yuk["items"],
                              key=lambda it: (cat_order.get(_kat(it["tovar"]), 9), it["tovar"]))
        cat_counters: dict = {}
        row_counter  = 0   # umumiy alternating uchun
        blok_start   = row  # blok border uchun

        for item in items_sorted:
            holat = holat_map.get(item["tovar"], "🟢 НОРМА")
            bg    = ROW_CLR_DARK if row_counter % 2 == 0 else ROW_CLR_LIGHT
            row_counter += 1
            qy    = qoldiq_yolda_map.get(item["tovar"])
            row   = _product_row(ws, row, item, holat, bg=bg, qoldiq_yolda=qy)

        jami_row_num = row
        row = _jami_row(ws, row, yuk, len(yuk["items"]))
        blok_end = row - 1

        # Blok atrofini medium border bilan o'rash
        _blok_border(ws, blok_start - 3, blok_end, CORE_NC)

        row = _separator(ws, row)

    _qolgan_blok(ws, row, qolgan)

    # Xulosa varaqi
    _xulosa_varaq(wb, yuklar, qolgan, kanal, start_sana_str)

    out = CHIQISH_DIR / f"Yuklatish_Rejasi_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    wb.save(out)
    return out


# ── Asosiy funksiyalar ────────────────────────────────────────────────────────
def main_with_data(kanal: str, ombor_map: dict,
                   start_date: datetime | None = None,
                   xitoy_vazn: dict | None = None) -> str | None:
    """
    Bot tomonidan chaqiriladi — ombor_map tayyordan uzatiladi.
    ombor_map:  {tovar_nomi: tayyor_miqdor}  (L ustun)
    xitoy_vazn: {tovar_nomi: 1_dona_kg} — Xitoy faylidan olingan vazn (ixtiyoriy)
    Qaytaradi: Excel fayl yo'li yoki None (kerak yo'q)
    """
    if start_date is None:
        start_date = datetime.now() + timedelta(days=1)

    print("📊 Power BI ma'lumotlari yuklanmoqda...")
    try:
        pb_df = power_bi_yuklash()
    except FileNotFoundError as e:
        print(f"⚠️  {e}")
        return None

    print("🔢 Kerak miqdori hisoblanmoqda...")
    from common import normalize_product_name as _norm

    kerak_df = kerak_hisob(pb_df)
    # Barcha Power BI tovarlarni saqlaymiz — Кам=0 bo'lsa ham.
    # Buning sababi: Xitoydan kelgan tovar inventarda bo'lsa ЯНГИ bo'lmasin,
    # to'g'ri holat (НОРМА/ПАСТ/КРИТИК) ko'rsatilsin.
    # Inventar nomlarini ham normalize qilamiz (Лист-0,8 → Лист- 0,8)
    if not kerak_df.empty:
        kerak_df["Товар"] = kerak_df["Товар"].apply(_norm)

    holat_map = dict(zip(kerak_df["Товар"], kerak_df["Холат"])) if not kerak_df.empty else {}
    kerak_set = set(kerak_df["Товар"]) if not kerak_df.empty else set()

    # 2026-07-11: G/H ustunlari (Қолдиқ/Йўлда) uchun -- pb_df dan (kerak_hisob
    # bu ustunlarni tashlab yuboradi, shu sabab alohida map qurilyapti).
    # Nom normalize qilinadi -- kerak_df/mavjud_df bilan bir xil kalit bo'lsin.
    qoldiq_yolda_map = {}
    if not pb_df.empty:
        for _, _r in pb_df.iterrows():
            _key = _norm(_r.get("tovar"))
            qoldiq_yolda_map[_key] = (
                float(_r.get("qoldiq", 0) or 0),
                float(_r.get("yolda", 0) or 0),
            )

    # Xitoy ombor → mavjud_df (nomlar allaqachon normalize qilingan)
    mavjud_df = pd.DataFrame(list(ombor_map.items()), columns=["Товар", "Миқдор"])
    mavjud_df["Миқдор"] = pd.to_numeric(mavjud_df["Миқдор"], errors="coerce").fillna(0)
    mavjud_df = mavjud_df[mavjud_df["Миқдор"] > 0]

    if mavjud_df.empty:
        print("⚠️  Xitoy omborida tayyor tovar yo'q.")
        return "OMBOR_BOʻSH"

    mavjud_set = set(mavjud_df["Товар"])

    # YANGI tovarlar: xitoyda bor lekin inventarda yo'q
    yangi_tovarlar = mavjud_set - kerak_set

    # kerak_df ga yangi tovarlarni qo'shamiz (yuklash uchun kerak, inventarda yo'q bo'lsa ham)
    if yangi_tovarlar:
        yangi_rows = []
        for tov in yangi_tovarlar:
            miq = int(mavjud_df[mavjud_df["Товар"] == tov]["Миқдор"].iloc[0])
            yangi_rows.append({"Товар": tov, "Холат": "🆕 ЯНГИ", "Кам": miq, "urg_kun": 999})
            holat_map[tov] = "🆕 ЯНГИ"
        yangi_df = pd.DataFrame(yangi_rows)
        kerak_df = pd.concat([kerak_df, yangi_df], ignore_index=True) if not kerak_df.empty else yangi_df
        print(f"Yangi tovarlar qoshildi: {len(yangi_tovarlar)} ta")

    if kerak_df.empty:
        print("Power BI da hozircha kamomati bor tovar yoq.")
        return "KERAK_YOQ"

    print(f"Kerak: {len(kerak_df)} ta, Xitoyda tayyor: {len(mavjud_set)} ta, Yangi: {len(yangi_tovarlar)} ta")
    yuklar, qolgan = optimallashtir(kerak_df, mavjud_df, max_yuklar=MAX_PER_MONTH,
                                     xitoy_vazn=xitoy_vazn or {})

    if not yuklar:
        print("Mos tovar topilmadi yoki xitoyda yetarli tovar yoq.")
        return "YUKLAR_YOQ"

    print(f"Sanalar belgilanmoqda (max {MAX_PER_DAY}/kun)...")
    yuklar = sana_belgi(yuklar, start_date)

    print(f"Excel yozilmoqda ({len(yuklar)} konteyner)...")
    out = excel_yaz(yuklar, qolgan, holat_map, kanal, qoldiq_yolda_map=qoldiq_yolda_map)

    n_konteyner    = len(yuklar)
    yuklangan_kg   = sum(y["jami_kg"] for y in yuklar)
    yuklangan_xil  = sum(len(y["items"]) for y in yuklar)
    qolgan_xil     = len(qolgan)
    qolgan_kg      = sum(it["vazn_kg"] for it in qolgan)
    n_yangi        = len(yangi_tovarlar) if yangi_tovarlar else 0

    print(f"Saqlandi: {out}")
    print(f"   {n_konteyner} konteyner | {yuklangan_kg:,.0f} kg | {yuklangan_xil} xil")
    if qolgan_xil:
        print(f"   Yuklanmadi: {qolgan_xil} xil | {qolgan_kg:,.0f} kg")

    # STATS:konteyner|yuklangan_kg|yuklangan_xil|qolgan_xil|qolgan_kg|path
    return f"STATS:{n_konteyner}|{yuklangan_kg:.0f}|{yuklangan_xil}|{qolgan_xil}|{qolgan_kg:.0f}|{out}"


def main(kanal: str = "asosiy", start_date=None):
    """Standalone ishlatish."""
    try:
        ombor_map, vazn_map = xitoy_ombor_yuklash(kanal)
    except FileNotFoundError as e:
        print(e)
        return None
    return main_with_data(kanal, ombor_map, start_date, xitoy_vazn=vazn_map)


if __name__ == "__main__":
    main()
